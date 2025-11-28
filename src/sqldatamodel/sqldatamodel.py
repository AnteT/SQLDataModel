from __future__ import annotations
import sqlite3, os, csv, sys, datetime, pickle, re, shutil, datetime, json, random, urllib.request
from collections.abc import Callable, Iterator, Iterable
from collections import namedtuple
from typing import Literal, Any, Type, NamedTuple
from io import StringIO

from . import optionals
from . import utils
from .exceptions import DimensionError, SQLProgrammingError, ErrorFormat, WarnFormat
from .standarddeviation import StandardDeviation
from .jsonencoder import DataTypesEncoder
from .htmlparser import HTMLParser
from .ansicolor import ANSIColor

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import numpy as _np
    import pandas as _pd
    import polars as _pl
    import pyarrow as _pa

class SQLDataModel:
    """
    ------------
    SQLDataModel
    ------------

    Primary class for the package of the same name. Its meant to provide a fast & light-weight alternative to the common pandas, numpy and sqlalchemy setup for moving data in a source/destination agnostic manner. It is not an ORM, any modifications outside of basic joins, group bys and table alteration requires knowledge of SQL. The primary use-case envisaged by the package is one where a table needs to be ETL'd from location A to destination B with arbitrary modifications made if needed:
    
    Summary
    -------

        - Extract your data from SQL, websites or HTML, parquet, JSON, CSV, pandas, numpy, pickle, python dictionaries, lists, etc.
        - Transform your data using raw SQL or any number of built-in methods covering some of the most used pandas data methods.
        - Load your data to any number of sources including popular SQL databases, CSV files, JSON, HTML, parquet, pickle, etc.
    
    Usage
    -----

    ```python
        import sqldatamodel as sdm
        
        # Lets grab a random table from Wikipedia
        df = sdm.from_html("https://en.wikipedia.org/wiki/FIFA_World_Cup", table_identifier=7)

        # Lets see what we found
        print(df)
    ```

    This will output:

    ```shell
        ┌───────────────┬──────┬──────┬──────────┬──────────┬──────┬──────┬───────┐
        │ Confederation │  AFC │  CAF │ CONCACAF │ CONMEBOL │  OFC │ UEFA │ Total │
        ├───────────────┼──────┼──────┼──────────┼──────────┼──────┼──────┼───────┤
        │ Teams         │   43 │   49 │       46 │       89 │    4 │  258 │   489 │
        │ Top 16        │    9 │   11 │       15 │       37 │    1 │   99 │   172 │
        │ Top 8         │    2 │    4 │        5 │       36 │    0 │  105 │   152 │
        │ Top 4         │    1 │    1 │        1 │       23 │    0 │   62 │    88 │
        │ Top 2         │    0 │    0 │        0 │       15 │    0 │   29 │    44 │
        │ 4th           │    1 │    1 │        0 │        5 │    0 │   15 │    22 │
        │ 3rd           │    0 │    0 │        1 │        3 │    0 │   18 │    22 │
        │ 2nd           │    0 │    0 │        0 │        5 │    0 │   17 │    22 │
        │ 1st           │    0 │    0 │        0 │       10 │    0 │   12 │    22 │
        └───────────────┴──────┴──────┴──────────┴──────────┴──────┴──────┴───────┘
        [9 rows x 8 columns]
    ```
    
    Example::

        import sqldatamodel as sdm

        # For example, setup a source connection
        source_db_conn = pyodbc.connect(...)

        # A destination connection
        destination_db_conn = sqlite3.connect(...)

        # Grab your source table
        df = sdm.from_sql("select * from source_table", source_db_conn)

        # Modify it however you want, whether through plain SQL
        df = df.execute_fetch('select "whatever", "i", "want" from "wherever_i_want" where "what_i_need" is not null ')

        # Or through any number of built-in methods like filtering
        df = df[df['create_date'] >= '2023-01-01']

        # Or creating new columns
        df['new_date'] = datetime.now()

        # Or modifying existing ones
        df['salary'] = df['salary'] * 2

        # Or applying functions
        df['user_id'] = df['user_id'].apply(lambda x: x**2)

        # Or deduplicating
        df = df.deduplicate(subset=['user_id','user_name'])

        # Or iterate through it row-by-row and modify it
        for idx, row in df.iter_tuples(index=True):
            if row['number'] % 2 == 0:
                row[idx,'odd_even'] = 'even'
            else:
                row[idx,'odd_even'] = 'odd'

        # Or join it using any of the standard join operations
        df = df_left.merge(df_right, how='left', left_on='id', right_on='id')

        # Or group or aggregate the data:
        df_agg = df.group_by(["first", "last", "position"])            

        # Or have your data imported and described for you
        df = sdm.from_parquet('titanic.parquet').describe()

        # View result
        print(df)

    This will output:
        
    ```shell
        ┌────────┬─────────────┬──────────┬────────┬────────┬───────┬────────┐
        │ metric │ passengerid │ survived │ pclass │    sex │   age │   fare │
        ├────────┼─────────────┼──────────┼────────┼────────┼───────┼────────┤
        │ count  │         891 │      891 │    891 │    891 │   714 │    891 │
        │ unique │         891 │        2 │      3 │      2 │    88 │    248 │
        │ top    │         891 │        0 │      3 │   male │    24 │   8.05 │
        │ freq   │           1 │      549 │    491 │    577 │    30 │     43 │
        │ mean   │         446 │        0 │      2 │    NaN │  29.7 │   32.2 │
        │ std    │         257 │        0 │      0 │    NaN │ 14.53 │  49.69 │
        │ min    │           1 │        0 │      1 │ female │  0.42 │      0 │
        │ p25    │         223 │        0 │      2 │    NaN │     6 │    7.9 │
        │ p50    │         446 │        0 │      3 │    NaN │    24 │  14.45 │
        │ p75    │         669 │        1 │      3 │    NaN │    35 │     31 │
        │ max    │         891 │        1 │      3 │   male │    80 │ 512.33 │
        │ dtype  │         int │      int │    int │    str │ float │  float │
        └────────┴─────────────┴──────────┴────────┴────────┴───────┴────────┘
        [12 rows x 7 columns]
    ```

    Move data quickly from one source or format to another:

    ```python
        # Load it to your destination database:
        df.to_sql("new_table", destination_db_conn)

        # Or any number of formats including:
        df.to_csv("output.csv")
        df.to_html("output.html")
        df.to_json("output.json")
        df.to_latex("output.tex")
        df.to_markdown("output.md")
        df.to_parquet("output.parquet")
        df.to_pickle("output.sdm")
        df.to_text("output.txt")
        df.to_local_db("output.db")

        # Reload it back again from more formats:
        df = sdm.from_csv("output.csv")
        df = sdm.from_dict(py_dict)
        df = sdm.from_html("output.html")
        df = sdm.from_json("output.json")
        df = sdm.from_latex("output.tex")
        df = sdm.from_markdown("output.md")
        df = sdm.from_numpy(np_arr)
        df = sdm.from_pandas(pd_df)
        df = sdm.from_polars(pl_df)
        df = sdm.from_parquet("output.parquet")
        df = sdm.from_pickle("output.sdm")
        df = sdm.from_sql("output", sqlite3.connect('output.db'))
    ```
    
    Data Formats
    ------------

    ``SQLDataModel`` seamlessly interacts with a wide range of data formats providing a versatile platform for data extraction, conversion, and writing. Supported formats include:

        - ``Arrow``: Convert to and from Apache Arrow format, ``pyarrow`` required.
        - ``CSV``: Extract from and write to comma separated value, ``.csv``, files.
        - ``Excel``: Extract from and write to Excel ``.xlsx`` files, ``openpyxl`` required.
        - ``HTML``: Extract from web and write to and from ``.html`` files including formatted string literals.
        - ``JSON``: Extract from and write to ``.json`` files, JSON-like objects, or JSON formatted sring literals.
        - ``LaTeX``: Extract from and write to ``.tex`` files, LaTeX formatted string literals.
        - ``Markdown``: Extract from and write to ``.MD`` files, Markdown formatted string literals.
        - ``Numpy``: Convert to and from ``numpy.ndarray`` objects, ``numpy`` required.
        - ``Pandas``: Convert to and from ``pandas.DataFrame`` objects, ``pandas`` required.
        - ``Parquet``: Extract from and write to ``.parquet`` files, ``pyarrow`` required.
        - ``Pickle``: Extract from and write to ``.pkl`` files, package uses ``.sdm`` extension when pickling for ``SQLDataModel`` metadata.
        - ``Polars``: Convert to and from ``polars.DataFrame`` objects, ``polars`` required.
        - ``SQL``: Extract from and write to the following popular SQL databases:

          - ``SQLite``: Using the built-in ``sqlite3`` module.
          - ``PostgreSQL``: Using the ``psycopg2`` package.
          - ``SQL Server``: Using the ``pyodbc`` package.
          - ``Oracle``: Using the ``cx_Oracle`` package.
          - ``Teradata``: Using the ``teradatasql`` package.

        - ``Text``: Write to and from ``.txt`` files including other ``SQLDataModel`` string representations.
        - ``TSV or delimited``: Write to and from files delimited by:

          - ``\\t``: Tab separated values or ``.tsv`` files.
          - ``\\s``: Single space or whitespace separated values.
          - ``;``: Semicolon separated values.
          - ``|``: Pipe separated values.
          - ``:``: Colon separated values.
          - ``,``: Comma separated values or ``.csv`` files.

        - ``Python objects``:

          - ``dictionaries``: Convert to and from collections of python ``dict`` objects.
          - ``lists``: Convert to and from collections of python ``list`` objects.
          - ``tuples``: Convert to and from collections of python ``tuple`` objects.
          - ``namedtuples``: Convert to and from collections of ``namedtuples`` objects.
          
    Pretty Printing
    ---------------

    SQLDataModel also pretty prints your table in any color you specify, use :meth:`SQLDataModel.set_display_color()` and provide either a hex value or a tuple of rgb and print the table, example output:

    ```shell
        ┌───┬─────────────────────┬────────────┬─────────────┬────────┬─────────┐
        │   │ full_name           │ date       │ country     │    pin │ service │
        ├───┼─────────────────────┼────────────┼─────────────┼────────┼─────────┤
        │ 0 │ Pamela Berg         │ 2024-09-15 │ New Zealand │   3010 │    3.02 │
        │ 1 │ Mason Hoover        │ 2024-01-23 │ Australia   │   6816 │    5.01 │
        │ 2 │ Veda Suarez         │ 2023-09-04 │ Ukraine     │   1175 │    4.65 │
        │ 3 │ Guinevere Cleveland │ 2024-03-22 │ New Zealand │   4962 │    3.81 │
        │ 4 │ Vincent Mccoy       │ 2023-09-16 │ France      │   4446 │    2.95 │
        │ 5 │ Holmes Kemp         │ 2024-11-13 │ Germany     │   9396 │    4.61 │
        │ 6 │ Donna Mays          │ 2023-06-06 │ Costa Rica  │   8153 │    5.34 │
        │ 7 │ Rama Galloway       │ 2023-09-22 │ Italy       │   3384 │    3.87 │
        │ 8 │ Lucas Rodriquez     │ 2024-03-16 │ New Zealand │   3278 │    2.73 │
        │ 9 │ Hunter Donaldson    │ 2023-06-30 │ Belgium     │   1593 │    4.58 │
        └───┴─────────────────────┴────────────┴─────────────┴────────┴─────────┘
    ```

    Note:
        - No additional dependencies are installed with this package, however you will obviously need to have pandas or numpy to create pandas or numpy objects.
        - Use :meth:`SQLDataModel.set_display_color()` to modify the terminal color of the table, by default no color styling is applied.
        - Use :meth:`SQLDataModel.get_supported_sql_connections()` to view supported SQL connection packages, please reach out with any issues or questions, thanks!
    """
    __slots__ = ('sql_idx','sql_model','display_max_rows','min_column_width','max_column_width','column_alignment','display_color','display_index','row_count','headers','column_count','static_py_to_sql_map_dict','static_sql_to_py_map_dict','sql_db_conn','display_float_precision','header_master','indicies','dtypes','shape','table_style')
    
    def __init__(self, data:list[list]=None, headers:list[str]=None, dtypes:dict[str,str]=None, display_max_rows:int=None, min_column_width:int=3, max_column_width:int=38, column_alignment:Literal['dynamic','left','center','right']='dynamic', display_color:str=None, display_index:bool=True, display_float_precision:int=2, infer_types:bool=False, table_style:Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple']='default'):
        """
        Initializes a new instance of ``SQLDataModel``.

        Parameters:
            ``data`` (list[list]): The data to populate the model. Should be a list of lists or a list of tuples or a dictionary orientated by rows or columns.
            ``headers`` (list[str], optional): The column headers for the model. If not provided, default headers will be used.
            ``dtypes`` (dict, optional): A dictionary specifying the data types for each column. Format: {'column': 'dtype'}.
            ``display_max_rows`` (int, optional): The maximum number of rows to display. Default is None, using terminal height to format number of rows.
            ``min_column_width`` (int, optional): The minimum width for each column. Default is 3.
            ``max_column_width`` (int, optional): The maximum width for each column. Default is 38.
            ``column_alignment`` (str, optional): The alignment for columns, must be 'dynamic', 'left', 'center' or 'right'). Default is 'dynamic'.
            ``display_color`` (str|tuple, optional): The color for display as hex code string or rgb tuple.
            ``display_index`` (bool, optional): Whether to display row indices. Default is True.
            ``display_float_precision`` (int, optional): The number of decimal places to display for float values. Default is 2.
            ``infer_types`` (bool, optional): Whether to infer the data types based on a randomly selected sample. 
                Default is False, using first row to derive the corresponding type directly.
            ``table_style`` (str, optional): The styling to use when representing the table in textual formats. 
                Must be 'ascii', 'bare', 'dash', 'default', 'double', 'list', 'markdown', 'outline', 'pandas', 'polars', 'postgresql', 'rst-grid', 'rst-simple' or 'round'.

        Raises:
            ``ValueError``: If ``data`` and ``headers`` are not provided, or if ``data`` is of insufficient length.
            ``TypeError``: If ``data`` or ``headers`` is not a valid type (list or tuple), or if ``dtypes`` is not a dictionary.
            ``DimensionError``: If the length of ``headers`` does not match the implied column count from the data.
            ``SQLProgrammingError``: If there's an issue with executing SQL statements during initialization.

        Example::

            import sqldatamodel as sdm

            # Create sample data
            data = [('Alice', 20, 'F'), ('Bob', 25, 'M'), ('Gerald', 30, 'M')]

            # Create the model with custom headers
            df = sdm.SQLDataModel(data, headers=['Name','Age','Sex'])
            
            # Display the model
            print(df)
        
        This will output the SQLDataModel formatted to fit within the current terminal:
        
        ```shell
            ┌────────┬──────┬──────┐
            │ Name   │  Age │ Sex  │
            ├────────┼──────┼──────┤
            │ Alice  │   20 │ F    │
            │ Bob    │   25 │ M    │
            │ Gerald │   30 │ M    │
            └────────┴──────┴──────┘
            [3 rows x 3 columns]
        ```

        A ``SQLDataModel`` can be initialized from dozens of data formats, including python dictionaries:

        ```python
            import sqldatamodel as sdm
            
            # Dictionary with sample data
            data = {
                'Name': ['Ali', 'Bob', 'Chris'],
                'Role': ['Judge', 'Pilot', 'Nurse'],
                'Height': [174.2, 180.9, 173.4],
            }

            # Create the model and set a new style
            df = sdm.SQLDataModel(data, table_style='list')

            # View it
            print(df)
        ```

        This will output the SQLDataModel using the 'list' styling:

        ```text
            Name   Role    Height
            -----  -----  -------
            Ali    Judge   174.20
            Bob    Pilot   180.90
            Chris  Nurse   173.40
        ```

        Note:
            - If ``data`` is not provided, an empty model is created with headers, at least one of ``data``, ``headers`` or ``dtypes`` are required to instantiate the model.
            - If ``headers`` are not provided, default headers will be generated using the the format ``'0', '1', ..., N`` where ``N`` is the column count.
            - If ``dtypes`` is provided, it must be a dictionary with column names as keys and Python data types as string values, e.g., `{'first_name': 'str', 'weight': 'float'}`
            - If ``infer_types = True`` and ``dtypes`` are provided, the order will be resolved by first inferring the types, then overriding the inferred types for each ``{col:type}`` provided in the ``dtypes`` argument. If one is not provided, then the inferred type will be used as a fallback.
            - For creating ``SQLDataModel`` from file formats like CSV, Markdown, LaTeX, Excel, Parquet or Text files, see :meth:`SQLDataModel.from_data()` or go to format specific constructor.
            - For creating ``SQLDataModel`` from object formats like Pyarrow, JSON, HTML, Pandas, Numpy or Polars, see format specific constructor like :meth:`SQLDataModel.from_pandas()` or :meth:`SQLDataModel.from_numpy()`.
            - Use :meth:`SQLDataModel.set_table_style()` to change the format and styling used when displaying the model.
            - Use :meth:`SQLDataModel.set_display_index()` to toggle inclusion of index column in table representations.
            - Use :meth:`SQLDataModel.set_display_color()` to modify the terminal color used to style the model. 
            - Use :meth:`SQLDataModel.set_display_max_rows()` to modify the number of rows output in the representations.
        
        Changelog:
            - Version 0.12.0 (2024-07-06):
                - Modified the default minimum number of displayed rows from 1 to 4 when :py:attr:`SQLDataModel.display_max_rows` is None.
            - Version 0.11.0 (2024-07-05):
                - Added additional option 'latex' for ``table_style`` parameter.            
            - Version 0.9.3 (2024-06-28):
                - Added additional options 'rst-simple' and 'rst-grid' for ``table_style`` parameter.            
        """
        if data is None:
            if headers is None:
                if dtypes is None:
                    msg = ErrorFormat(f"ValueError: insufficient data, an empty header-less model cannot be created, to create a model with zero rows a `headers` or `dtypes` argument is required")
                    raise ValueError(msg)
                else:
                    if not isinstance(dtypes,dict):
                        msg = ErrorFormat(f"TypeError: invalid type '{type(dtypes).__name__}', `dtypes` must be of type 'dict' with format of `{{'column':'dtype'}}` where 'dtype' must be a string representing a valid python datatype")
                        raise TypeError(msg)
                    headers = list(dtypes.keys())
            data = [tuple(None for _ in range(len(headers)))]
            had_data = False
        else:
            had_data = True
        if not isinstance(data, (list,tuple,dict)) and had_data:  
            msg = ErrorFormat(f"TypeError: type mismatch, '{type(data).__name__}' is not a valid type for data, which must be of type list, tuple or dict")
            raise TypeError(msg)
        if len(data) < 1 and had_data:
            msg = ErrorFormat(f"ValueError: data not found, data of length '{len(data)}' is insufficient to construct a valid model, additional rows of data required")
            raise ValueError(msg)
        if had_data:
            if isinstance(data, dict) or isinstance(data[0], dict):
                if isinstance(data, list):
                    data = utils.flatten_json(data)
                rowwise = True if all(isinstance(x, int) for x in data.keys()) else False
                if rowwise:
                    col_count = len(data[next(iter(data))])
                    if headers is None:
                        headers = ['idx',*[f'{i}' for i in range(col_count)]] # get column count from first key value pair in provided dict
                    elif (col_count) == len(headers):
                        headers = ['idx',*headers]
                    data = [tuple([k,*v]) for k,v in data.items()]
                else:
                    first_key_val = data[next(iter(data))]
                    if isinstance(first_key_val, dict):
                        inferred_headers = list(data.keys())
                        data = [[data[col][val] for col in inferred_headers] for val in data.keys()]
                    elif isinstance(first_key_val, (list, tuple)):
                        inferred_headers = list(data.keys())
                        data = list(data.values())
                        data = [tuple(data[j][i] for j in range(len(inferred_headers))) for i in range(len(first_key_val))]
                        headers = inferred_headers if headers is None else ['idx', *headers] if len(headers) + 1 == len(inferred_headers) else headers
                    else:
                        msg = ErrorFormat(f"TypeError: invalid dict values, received type '{type(first_key_val).__name__}' but expected dict values as one of type 'list', 'tuple' or 'dict'")
                        raise TypeError(msg)
            try:
                _ = data[0]
            except Exception as e:
                msg = ErrorFormat(f"IndexError: data index error, data index provided does not exist for length '{len(data)}' due to '{e}'")
                raise IndexError(msg) from None
            if not isinstance(data[0], (list,tuple)):
                if type(data[0]).__module__ != 'pyodbc': # check for pyodbc.Row which is acceptable
                    msg = ErrorFormat(f"TypeError: type mismatch, '{type(data[0]).__name__}' is not a valid type for data rows, which must be of type list or tuple")
                    raise TypeError(msg)
            if len(data[0]) < 1:
                msg = ErrorFormat(f"ValueError: data rows not found, data rows of length '{len(data[0])}' are insufficient to construct a valid model, at least one row is required")
                raise ValueError(msg)
        if headers is not None:
            if not isinstance(headers, (list,tuple)):
                msg = ErrorFormat(f"TypeError: invalid header types, '{type(headers).__name__}' is not a valid type for headers, please provide a tuple or list type")
                raise TypeError(msg)
            if (len(headers) != len(data[0])) and had_data:
                msg = ErrorFormat(f"DimensionError: shape mismatch '{len(headers)} != {len(data[0])}', provided headers dim, '{len(headers)}', does not match data dim '{len(data[0])}', compatible dimensions are required")
                raise DimensionError(msg)                
            if isinstance(headers,tuple):
                try:
                    headers = list(headers)
                except:
                    msg = ErrorFormat(f"TypeError: failed header conversion, unable to convert provided headers tuple to list type, please provide headers as a list type")
                    raise TypeError(msg) from None
            if not all(isinstance(x, str) for x in headers):
                try:
                    headers = [str(x) for x in headers]
                except:
                    msg = ErrorFormat(f"TypeError: invalid header values, all headers provided must be of type string")
                    raise TypeError(msg) from None
            for col in headers:
                if "'" in col:
                    msg = ErrorFormat(f"ValueError: invalid character in column '{col}', headers must be of type 'str' consisting of valid SQL column identifiers")
                    raise ValueError(msg)
        else:
            headers = list(dtypes.keys()) if dtypes is not None else [f"{x}" for x in range(len(data[0]))]
        self.sql_idx = "idx"
        """``str``: The index column name applied to the sqlite3 in-memory representation of the model. Default is ``'idx'``"""
        self.sql_model = "sdm"
        """``str``: The table name applied to the sqlite3 in-memory representation of the model. Default is ``'sdm'``"""
        self.display_max_rows = display_max_rows
        """``int``: The maximum number of rows to display. Default is 1,000 rows."""
        self.min_column_width = min_column_width
        """``int``: The minimum column width in characters to use for string representations of the data. Default is 3."""
        self.max_column_width = max_column_width
        """``int``: The maximum column width in characters to use for string representations of the data. Default is 38."""
        self.column_alignment = column_alignment # 'dynamic','left','center','right'
        """``str``: The column alignment to use for string representations of the data, value must be one of ``['dynamic','left','center','right']`` Default is ``'dynamic'``, using right-alignment for numeric columns and left-aligned for all others."""
        self.display_index = display_index
        """``bool``: Determines whether the index column is displayed when string representations of the table are generated. Default is True."""
        self.display_float_precision = display_float_precision
        """``int``: The floating point precision to use for string representations of the table, does not affect the actual floating point values stored in the model. Default is 2."""
        self.row_count = len(data) if had_data else 0
        """``int``: The current row count of the model."""
        had_idx = True if headers[0] == self.sql_idx else False
        dyn_idx_offset,dyn_idx_bind,dyn_add_idx_insert = (1, "?,", f'"{self.sql_idx}",') if had_idx else (0, "", "")
        headers = headers[dyn_idx_offset:]
        self.headers = headers
        """``list[str]``: The current column names of the model. If not provided, default column names will be used."""
        self.column_count = len(self.headers)
        """``int``: The current column count of the model."""
        self.shape = (self.row_count, self.column_count)
        """``tuple[int, int]``: The current dimensions of the model as a tuple of ``(rows, columns)``."""
        self.display_color = ANSIColor(display_color) if isinstance(display_color, (str,tuple)) else display_color if isinstance(display_color,ANSIColor) else None
        """``ANSIColor``: The display color to use for string representations of the model. Default is ``None``, using the standard terminal color."""
        self.static_py_to_sql_map_dict = {'None': 'TEXT','int': 'INTEGER','float': 'REAL','str': 'TEXT','bytes': 'BLOB', 'date': 'DATE', 'datetime': 'TIMESTAMP', 'NoneType':'TEXT', 'bool':'INTEGER'}
        """``dict``: The data type mapping to use when converting python types to SQL column types."""
        self.static_sql_to_py_map_dict = {'NULL': 'None','INTEGER': 'int','REAL': 'float','TEXT': 'str','BLOB': 'bytes', 'DATE': 'date', 'TIMESTAMP': 'datetime','':'str'}
        """``dict``: The data type mapping to use when converting SQL column types to python types."""
        self.table_style = table_style
        """``str``: The table style used for string representations of the model. Available styles are ``'ascii'``, ``'bare'``, ``'dash'``, ``'default'``, ``'double'``, ``'list'``, ``'markdown'``, ``'outline'``, ``'pandas'``, ``'polars'``, ``'postgresql'`` or ``'round'``. Defaults to ``'default'`` table style."""
        if infer_types and self.row_count > 0:
            inferred_dtypes = utils.infer_types_from_data(input_data=random.sample(data, min(self.row_count, 16)))
            headers_to_py_dtypes_dict = {self.headers[i]:inferred_dtypes[i+dyn_idx_offset] for i in range(self.column_count)}
        else:
            headers_to_py_dtypes_dict = {self.headers[i]:type(data[0][i+dyn_idx_offset]).__name__ for i in range(self.column_count)}        
        if dtypes is not None:
            [(headers_to_py_dtypes_dict.__setitem__(col,dtype)) for col,dtype in dtypes.items() if col in self.headers and dtype in self.static_py_to_sql_map_dict]
        try:
            headers_with_sql_dtypes_str = ",".join(f'"{col}" {self.static_py_to_sql_map_dict[headers_to_py_dtypes_dict[col]]}' for col in self.headers)
        except KeyError as e:
            msg = ErrorFormat(f"TypeError: invalid data type {e}, values in `data` must be a list of lists comprised of types 'str', 'int', 'float', 'bytes', 'datetime' or 'bool' ")
            raise TypeError(msg) from None
        self.sql_db_conn = sqlite3.connect(":memory:", uri=True, detect_types=sqlite3.PARSE_DECLTYPES)
        """``sqlite3.Connection``: The in-memory sqlite3 connection object in use by the model."""
        self.dtypes = headers_to_py_dtypes_dict
        """``dict[str, str]``: The current model data types mapped to each column in the format of ``{'col': 'dtype'}`` where ``'dtype'`` is a string representing the corresponding python type."""
        sql_create_stmt = f"""create table if not exists "{self.sql_model}" ("{self.sql_idx}" INTEGER PRIMARY KEY,{headers_with_sql_dtypes_str})"""
        sql_insert_params = ",".join([utils.sqlite_cast_type_format(dtype=headers_to_py_dtypes_dict[col], as_binding=True) for col in self.headers])        
        sql_insert_stmt = f"""insert into "{self.sql_model}" ({dyn_add_idx_insert}{','.join([f'"{col}"' for col in self.headers])}) values ({dyn_idx_bind}{sql_insert_params})"""
        try:
            self.sql_db_conn.execute(sql_create_stmt)
        except sqlite3.OperationalError as e:
            msg = ErrorFormat(f"SQLProgrammingError: invalid model structure, failed to create table due to '{e}'")
            raise SQLProgrammingError(msg) from None           
        self.header_master = None
        """``dict[str, tuple]``: Maps the current model's column metadata in the format of ``'column_name': ('sql_dtype', 'py_dtype', is_regular_column, 'default_alignment')``, updated by :meth:`SQLDataModel._update_model_metadata`."""
        self._update_model_metadata()
        if not had_data:
            trig_zero_init = f"""CREATE TRIGGER 'zero_init' AFTER INSERT 
            ON "{self.sql_model}" WHEN (select count("{self.sql_idx}") from "{self.sql_model}") = 1 
            BEGIN update "{self.sql_model}" set "{self.sql_idx}" = 0 where "{self.sql_idx}" = 1; END;"""
            self.sql_db_conn.execute(trig_zero_init)
            self.indicies = tuple()
            """``tuple``: The current valid row indicies of the model."""
            return
        if not had_idx:
            first_row_insert_stmt = f"""insert into "{self.sql_model}" ("{self.sql_idx}",{','.join([f'"{col}"' for col in self.headers])}) values (?,{sql_insert_params})"""
            try:
                self.sql_db_conn.execute(first_row_insert_stmt, (0,*data[0]))
            except sqlite3.ProgrammingError as e:
                msg = ErrorFormat(f"SQLProgrammingError: invalid or inconsistent data, failed with '{e}'")
                raise SQLProgrammingError(msg) from None   
            self.indicies = tuple(range(self.row_count))
            data = data[1:] # remove first row from remaining data
        else:
            self.indicies = tuple(sorted([row[0] for row in data]))
        try:
            self.sql_db_conn.executemany(sql_insert_stmt,data)
        except sqlite3.ProgrammingError as e:
            msg = ErrorFormat(f"SQLProgrammingError: invalid or inconsistent data, failed with '{e}'")
            raise SQLProgrammingError(msg) from None  

################################################################################################################
################################################ static methods ################################################
################################################################################################################

#############################################################################################################
######################################### columns & display params ##########################################
#############################################################################################################

    def drop_column(self, column:int|str|list, inplace:bool=True) -> None|SQLDataModel:
        """
        Drops the specified column(s) from the ``SQLDataModel``. Values for ``column`` can be a single column name or index, or a list of multiple column names or indicies to drop from the model.

        Parameters:
            ``column`` (int | str | list): The index, name, or list of indices/names of the column(s) to drop.
            ``inplace`` (bool): If True, drops the column(s) in-place and updates the model metadata. If False, returns a new ``SQLDataModel`` object without the dropped column(s) and does not modify the original object. Default is True.

        Returns:
            ``None | SQLDataModel``: If inplace is True, returns None. Otherwise, returns a new ``SQLDataModel`` object without the dropped column(s).

        Raises:
            ``TypeError``: If the column parameter is not of type 'int', 'str', or a list containing equivalent types.
            ``IndexError``: If any provided column index is outside the current column range.
            ``ValueError``: If any provided column name is not found in the model's headers.
        
        Examples::

            import sqldatamodel as sdm
            
            # Sample data
            headers = ['Name', 'Age', 'Gender', 'City']
            data = [
                ('Alice', 30, 'Female', 'Milwaukee'),
                ('Sarah', 35, 'Female', 'Houston'),
                ('Mike', 28, 'Male', 'Atlanta'),
                ('John', 25, 'Male', 'Boston'),
                ('Bob', 22, 'Male', 'Chicago'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data,headers)

            # Drop the 'Gender' column
            df.drop_column('Gender')

            # View updated model
            print(df)

        This will output:

        ```shell
            ┌───────┬──────┬───────────┐
            │ Name  │  Age │ City      │
            ├───────┼──────┼───────────┤
            │ Alice │   30 │ Milwaukee │
            │ Sarah │   35 │ Houston   │
            │ Mike  │   28 │ Atlanta   │
            │ John  │   25 │ Boston    │
            │ Bob   │   22 │ Chicago   │
            └───────┴──────┴───────────┘
            [5 rows x 3 columns]
        ```

        Dropping multiple columns:

        ```python
            # Drop first and last columns by index
            df.drop_column([0,-1])

            # View updated model
            print(df)
        ```

        This will output:

        ```shell
            ┌──────┬────────┐
            │  Age │ Gender │
            ├──────┼────────┤
            │   30 │ Female │
            │   35 │ Female │
            │   28 │ Male   │
            │   25 │ Male   │
            │   22 │ Male   │
            └──────┴────────┘
            [5 rows x 2 columns]
        ```

        Drop columns and return as a new ``SQLDataModel``:

        ```python
            # Drop the multiple columns and return as a new model
            df = df.drop_column(['Age','Gender'], inplace=False)

            # View updated model    
            print(df)
        ```

        This will output:

        ```shell
            ┌───────┬───────────┐
            │ Name  │ City      │
            ├───────┼───────────┤
            │ Alice │ Milwaukee │
            │ Sarah │ Houston   │
            │ Mike  │ Atlanta   │
            │ John  │ Boston    │
            │ Bob   │ Chicago   │
            └───────┴───────────┘
            [5 rows x 2 columns]
        ```

        Note:
            - Arguments for ``column`` can be a single ``str`` or ``int`` or ``list[str]`` containing ``str`` or ``list[int]`` containing ``int`` representing column names or column indicies, respectively, but they cannot be combined and provided together. For example, passing ``columns = ['First Name', 3]`` will raise a ``TypeError`` exception.
            - The equivalent of this method can also be achieved by simply indexing the required rows and columns using ``sdm[rows, column]`` notation, see :meth:`SQLDataModel.__getitem__()` for additional details.

        Changelog:
            - Version 0.2.3 (2024-03-28):
                - New method.            
        """
        column = self._validate_column(column, unmodified=False) # +VALCOL
        if inplace:
            drop_cols = ";".join([f'''alter table "{self.sql_model}" drop column "{col}"''' for col in column])
            sql_stmt = f"""begin transaction;{drop_cols}; end transaction;"""
            try:
                self.sql_db_conn.executescript(sql_stmt)
                self.sql_db_conn.commit()
            except Exception as e:
                self.sql_db_conn.rollback()
                msg = ErrorFormat(f'SQLProgrammingError: unable to rename columns, SQL execution failed with: "{e}"')
                raise SQLProgrammingError(msg) from None
            self._update_model_metadata()
            return
        else:
            keep_headers = [col for col in self.headers if col not in column]
            sql_stmt = self._generate_sql_stmt(columns=keep_headers)
            return self.execute_fetch(sql_stmt)    

    def drop_row(self, row:int|Iterable[int], inplace:bool=True, ignore_index:bool=False) -> None|SQLDataModel:
        """
        Drops the specified row(s) indicies from the ``SQLDataModel``. Values for ``row`` can be a single row index, or an iterable collection of multiple row indicies to drop.

        Parameters:
            ``row`` (int | Iterable[int]): The row index or row indicies to drop.
            ``inplace`` (bool, optional): If True, drops the rows(s) in-place and updates the model metadata. If False, returns a new ``SQLDataModel`` object without the dropped row(s). Default is True.
            ``ignore_index`` (bool, optional): If True, drops the row(s) and ignores the index for the resulting model. Default is False, keeping original indicies in new model.

        Returns:
            ``None | SQLDataModel``: If in-place is True, returns None. Otherwise, returns a new ``SQLDataModel`` object without the dropped rows(s).

        Raises:
            ``TypeError``: If the row parameter is not of type 'int' or an iterable collection of type 'int' representing the row indicies to drop.
            ``IndexError``: If any provided row index is outside the current row range determined by the values at :py:attr:`SQLDataModel.indicies`.
        
        Example::

            import sqldatamodel as sdm

            headers = ['Rank', 'Location', 'Population']
            data = [(1, "Tokyo, Japan", 37.4),
                    (2, "Delhi, India", 31.0),
                    (3, "Shanghai, China", 27.1),
                    (4, "São Paulo, Brazil", 22.0),
                    (5, "Mexico City, Mexico", 21.8),
                    (6, "Cairo, Egypt", 21.3),
                    (7, "Dhaka, Bangladesh", 21.0),
                    (8, "Mumbai, India", 20.7),
                    (9, "Beijing, China", 20.5),
                    (10,"Osaka, Japan", 19.1)]

            # Create the sample model
            df = sdm.SQLDataModel(data, headers)

            # Drop the last row
            df.drop_row(-1)

            # Drop rows based on condition of less than 25 Million population
            df.drop_row(df['Population'] < 25.0)

            # View result
            print(df)  

        This will output:

        ```shell
            ┌──────┬─────────────────┬────────────┐
            │ Rank │ Location        │ Population │
            ├──────┼─────────────────┼────────────┤
            │    1 │ Tokyo, Japan    │       37.4 │
            │    2 │ Delhi, India    │       31.0 │
            │    3 │ Shanghai, China │       27.1 │
            └──────┴─────────────────┴────────────┘
            [3 rows x 3 columns]
        ```

        Dropping multiple rows and returning a new model:

        ```python
            # Create a new model using the same sample data
            df = SQLDataModel(data, headers)

            # Set row indicies to drop
            row_indices = range(0, 5) # or [0, 1, 2, 3, 4]

            # Drop top 5 cities and return as a new model
            df_new = df.drop_row(row_indices, inplace=False)

            # View new model
            print(df_new)
        ```

        This will output:

        ```shell
            ┌──────┬───────────────────┬────────────┐
            │ Rank │ Location          │ Population │
            ├──────┼───────────────────┼────────────┤
            │    6 │ Cairo, Egypt      │       21.3 │
            │    7 │ Dhaka, Bangladesh │       21.0 │
            │    8 │ Mumbai, India     │       20.7 │
            │    9 │ Beijing, China    │       20.5 │
            │   10 │ Osaka, Japan      │       19.1 │
            └──────┴───────────────────┴────────────┘
            [5 rows x 3 columns]
        ```

        Important:
            Rows are referenced by their integer index, and not by their value. 
            This means that row index ``0`` will always refer to the first row in the model, and ``-1`` will always refer to the last. 
            This distinction is usually irrelevant when the two are aligned, however this is no longer the case when row(s) are dropped from anywhere except the very last row.

        Note:
            - Row indicies are retained after being deleted by default, provide ``ignore_index=True`` to reset row indicies if required.
            - The equivalent of this method can also be achieved by simply indexing the required rows and columns using ``sdm[rows, column]`` notation, see :meth:`SQLDataModel.__getitem__()` for additional details.

        Changelog:
            - Version 0.7.4 (2024-06-13):
                - New method.
        """
        row = self._validate_row(row, unmodified=False) # +VALROW
        row = row if len(row) != 1 else (row[0], row[0])
        if inplace:
            sql_stmt = f"""DELETE FROM "{self.sql_model}" WHERE "{self.sql_idx}" IN {row}"""
            self.execute_statement(sql_stmt)
            if ignore_index:
                self.reset_index()
            return
        else:
            columns = self.headers if ignore_index else [self.sql_idx, *self.headers] # Keep original index if not ignoring
            cols_str = ",".join([f'"{col}" as "{col}"' for col in columns])
            sql_stmt = " ".join(("SELECT ",cols_str,f'FROM "{self.sql_model}" WHERE "{self.sql_idx}" NOT IN {row}'))
            return self.execute_fetch(sql_stmt)

    def rename_column(self, column:int|str, new_column_name:str) -> None:
        """
        Renames a column in the ``SQLDataModel`` at the specified index or using the old column name into the new value provided in ``new_column_name``.

        Parameters:
            ``column`` (int|str): The index or current str value of the column to be renamed.
            ``new_column_name`` (str): The new name as a str value for the specified column.

        Raises:
            ``TypeError``: If the ``column`` or ``new_column_name`` parameters are invalid types.
            ``IndexError``: If the provided column index is outside the current column range.
            ``SQLProgrammingError``: If there is an issue with the SQL execution during the column renaming.

        Example::

            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age']
            data = [
                (0, 'john', 'smith', 27)
                ,(1, 'sarah', 'west', 29)
                ,(2, 'mike', 'harlin', 36)
                ,(3, 'pat', 'douglas', 42)
            ]

            # Create the model with sample data
            df = sdm.SQLDataModel(data, headers)

            # Example: Rename the column at index 1 to 'first_name'
            df.rename_column(1, 'first_name')

            # Get current values
            new_headers = df.get_headers()

            # Outputs ['first_name', 'last', 'age']
            print(new_headers)

        Note:
            - The method allows renaming a column identified by its index in the SQLDataModel.
            - Handles negative indices by adjusting them relative to the end of the column range.
            - If an error occurs during SQL execution, it rolls back the changes and raises a SQLProgrammingError with an informative message.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        column = self._validate_column(column, unmodified=False)[0] # +VALCOL
        rename_stmts = f"""alter table "{self.sql_model}" rename column "{column}" to "{new_column_name}" """
        full_stmt = f"""begin transaction; {rename_stmts}; end transaction;"""
        try:
            self.sql_db_conn.executescript(full_stmt)
            self.sql_db_conn.commit()
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f'SQLProgrammingError: unable to rename columns, SQL execution failed with: "{e}"')
            raise SQLProgrammingError(msg) from None
        self.headers[self.headers.index(column)] = new_column_name # replace old column with new column in same position as original column
        self._update_model_metadata()

    def rename_header(self, header:int|str, new_header_name:str) -> None:
        """
        Renames a header in the ``SQLDataModel`` at the specified index or using the old header name into the new value provided in ``new_header_name``.

        Parameters:
            ``header`` (int|str): The index or current str value of the header to be renamed.
            ``new_header_name`` (str): The new name as a str value for the specified header.

        Raises:
            ``TypeError``: If the ``header`` or ``new_header_name`` parameters are invalid types.
            ``IndexError``: If the provided header index is outside the current column range.
            ``SQLProgrammingError``: If there is an issue with the SQL execution during the header renaming.

        Example::

            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age']
            data = [
                (0, 'john', 'smith', 27)
                ,(1, 'sarah', 'west', 29)
                ,(2, 'mike', 'harlin', 36)
                ,(3, 'pat', 'douglas', 42)
            ]

            # Create the model with sample data
            df = sdm.SQLDataModel(data, headers)

            # Example: Rename the column at index 1 to 'first_name'
            df.rename_column(1, 'first_name')

            # Get current values
            new_headers = df.get_headers()

            # Outputs ['first_name', 'last', 'age']
            print(new_headers)

        Note:
            - The method allows renaming a column identified by its index in the SQLDataModel.
            - Handles negative indices by adjusting them relative to the end of the column range.
            - If an error occurs during SQL execution, it rolls back the changes and raises a SQLProgrammingError with an informative message.

        Changelog:
            - Version 1.2.0 (2025-01-28):
                - New convenience method wrapping :meth:`SQLDataModel.rename_column()` to reflect additional naming convention.       
        """        
        self.rename_column(header, new_header_name)

    def rename_headers(self, new_headers:list[str] | Callable[[list[str]], list[str]]) -> None:
        """
        Renames the current ``SQLDataModel`` headers to values provided in ``new_headers``. Headers must match the existing column count.

        Parameters:
            ``new_headers`` (list[str] or Callable[[list[str]], list[str]]): A sequence (e.g., list, tuple) of new header names
                or a callable that takes the existing headers and returns a new list of header names.
                
        Raises:
            ``TypeError``: If the ``new_headers`` type is not a valid type (list or tuple) or contains instances that are not of type 'str'.
            ``DimensionError``: If the length of ``new_headers`` does not match the column count.
            ``TypeError``: If the type of the first element in ``new_headers`` is not a valid type (str, int, or float).

        Returns:
            ``None``

        Example::
            
            import sqldatamodel as sdm

            # Create model
            df = sdm.from_csv('example.csv', headers=['First Name', 'Last Name', 'Salary'])

            # Set new headers
            df.rename_headers(['First_Name', 'Last_Name', 'Payment'])

            # Alternatively, provide a callable argument to transform headers using existing names
            df.rename_headers(lambda headers: [header.replace(' ', '_') for header in headers])
        
        Changelog:
            - Version 1.2.0 (2025-01-28):
                - New convenience method wrapping :meth:`SQLDataModel.set_headers()` to reflect additional naming convention.             
        """
        self.set_headers(new_headers)
        
    def replace(self, pattern:str, replacement:str, inplace:bool=False, **kwargs) -> SQLDataModel:
        """
        Replaces matching occurrences of a specified pattern with a replacement value in the ``SQLDataModel`` instance. 
        If inplace is True, the method updates the existing SQLDataModel; otherwise, it returns a new ``SQLDataModel`` with the replacements applied.

        Parameters:
            ``pattern`` (str): The substring or regular expression pattern to search for in each column.
            ``replacement`` (str): The string to replace the matched pattern with.
            ``inplace`` (bool, optional): If True, modifies the current SQLDataModel instance in-place. Default is False.
            ``**kwargs``: Additional keyword arguments to be passed to the ``execute_fetch`` method when not in-place.
        
        Raises:
            ``TypeError``: If the ``pattern`` or ``replacement`` parameters are invalid types.

        Returns:
            ``SQLDataModel``: If ``inplace=True``, modifies the current instance in-place and returns ``None``. Otherwise, returns a new SQLDataModel with the specified replacements applied.

        Example::
            
            import sqldatamodel as sdm

            headers = ['first', 'last', 'age', 'service']
            data = [
                ('John', 'Smith', 27, 1.22),
                ('Sarah', 'West', 39, 0.7),
                ('Mike', 'Harlin', 36, 3),
                ('Pat', 'Douglas', 42, 11.5)
            ]        

            # Create the model
            df = sdm.SQLDataModel(data, headers,display_float_precision=2, display_index=False)

            # Replace 'John' in the 'first' column
            df['first'] = df['first'].replace("John","Jane")
            
            # View model
            print(df)

        This will output:

        ```shell
            ┌───────┬─────────┬──────┬─────────┐
            │ first │ last    │  age │ service │
            ├───────┼─────────┼──────┼─────────┤
            │ Jane  │ Smith   │   27 │    1.22 │
            │ Sarah │ West    │   39 │    0.70 │
            │ Mike  │ Harlin  │   36 │    3.00 │
            │ Pat   │ Douglas │   42 │   11.50 │
            └───────┴─────────┴──────┴─────────┘   
            [4 rows x 4 columns] 
        ```
        
        Note:
            - See :meth:`SQLDataModel.contains()` for identifying substring presence.
            - See :meth:`SQLDataModel.endswith()` for checking end of value for substring.
            - See :meth:`SQLDataModel.startswith()` for checking beggining of value for substring.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(pattern, str):
            msg = ErrorFormat(f"TypeError: invalid argument type '{type(pattern).__name__}', method argument `pattern` must be of type 'str' for `replace()` method")
            raise TypeError(msg)
        if not isinstance(replacement, str):
            msg = ErrorFormat(f"TypeError: invalid argument type '{type(pattern).__name__}', method argument `replacement` must be of type 'str' for `replace()` method")
            raise TypeError(msg)
        if inplace:
            replace_stmt = " ".join((f"""update "{self.sql_model}" set""",",".join([f""" "{col}"=replace("{col}",'{pattern}','{replacement}') """ for col in self.headers])))
            self.sql_db_conn.execute(replace_stmt)
            return
        else:
            replace_stmt = " ".join(("select",",".join([f""" replace("{col}",'{pattern}','{replacement}') as "{col}" """ for col in self.headers]),f'from "{self.sql_model}"'))
            return self.execute_fetch(replace_stmt, **kwargs)

    def get_headers(self) -> list[str]:
        """
        Returns the current ``SQLDataModel`` headers.

        Returns:
            ``list``: A list of strings representing the headers.

        Example::
        
            import sqldatamodel as sdm

            # Create model
            df = sdm.from_csv('example.csv', headers=['First Name', 'Last Name', 'Salary'])

            # Get current model headers
            headers = df.get_headers()

            # Display values
            print(headers) # outputs: ['First Name', 'Last Name', 'Salary']

        Changelog:
            - Version 0.1.5 (2023-11-24):
                - New method.        
        """
        return self.headers

    def set_headers(self, new_headers:list[str] | dict[str|int, str] | Callable[[list[str]], list[str]]) -> None:
        """
        Renames the current ``SQLDataModel`` headers to values in ``new_headers``, which can be provided as:

        1. A sequence (list/tuple) of strings matching the full column count.
        2. A dictionary mapping existing indices (int) or names (str) to new names (str).
        3. A callable that transforms the existing header list.

        Parameters:
            ``new_headers`` (list[str] | dict[str|int, str] | Callable): The new header configuration.
                - If **list/tuple**: Must be same length as column count.
                - If **dict**: Keys can be current column names or indices (supporting negative indexing). Values are new names. Unspecified columns remain unchanged.
                - If **callable**: Receives current headers list, returns new headers list.
                
        Raises:
            ``TypeError``: If ``new_headers`` contains invalid types.
            ``DimensionError``: If the length of the resulting header list does not match the column count.
            ``IndexError``: If a dictionary key is an integer index out of bounds.
            ``ValueError``: If a dictionary key is a string name not found in current headers.
            ``SQLProgrammingError``: If an error is encountered while attempting to rename the SQL columns.

        Returns:
            ``None``

        Example::
            
            import sqldatamodel as sdm

            # Create model
            df = sdm.from_csv('example.csv', headers=['First Name', 'Last Name', 'Salary'])

            # 1. Rename all using list
            df.set_headers(['First_Name', 'Last_Name', 'Payment'])

            # 2. Rename specific columns using dict (mixed keys allowed)
            # Rename 'Payment' (index -1) to 'Annual_Salary' and 'First_Name' to 'FName'
            df.set_headers({-1: 'Annual_Salary', 'First_Name': 'FName'})

            # 3. Transform using callable
            df.set_headers(lambda headers: [h.upper() for h in headers])
        
        Changelog:
            - Version 2.2.0 (2025-11-28):
                - Added support for ``dict`` input to allow partial renaming and mixed index/name keys.
            - Version 2.1.1 (2025-11-25):
                - Modified to retain original column ordering when existing model headers match a subset of ``new_headers``.
            - Version 1.2.0 (2025-01-28):
                - Added ability to provide a callable for ``new_headers``.
            - Version 0.1.5 (2023-11-24):
                - New method.                
        """
        # 1. Handle Callable
        if callable(new_headers):
            new_headers = new_headers(self.headers)
        # 2. Handle Dictionary (Partial/Mixed Renaming)
        if isinstance(new_headers, dict):
            # Start with a copy of current headers
            final_headers = list(self.headers)
            for key, val in new_headers.items():
                # Validate the new name is a string
                if not isinstance(val, str):
                    msg = ErrorFormat(f"TypeError: invalid new header value type '{type(val).__name__}' for key '{key}', value must be of type 'str'")
                    raise TypeError(msg)
                target_idx = None
                # key is index (int)
                if isinstance(key, int):
                    # Validate bounds (handle negative indexing standard logic)
                    if key < -self.column_count or key >= self.column_count:
                        msg = ErrorFormat(f"IndexError: header index '{key}' out of range for model with {self.column_count} columns, use `get_headers()` to veiw current valid arguments")
                        raise IndexError(msg)
                    target_idx = key % self.column_count
                # key is column name (str)
                elif isinstance(key, str):
                    try:
                        target_idx = self.headers.index(key)
                    except ValueError:
                        msg = ErrorFormat(f"ValueError: header '{key}' not found, use `.get_headers()` to view current valid model headers")
                        raise ValueError(msg) from None
                else:
                    msg = ErrorFormat(f"TypeError: invalid key type '{type(key).__name__}', dictionary keys must be 'int' (index) or 'str' (current header name)")
                    raise TypeError(msg)
                final_headers[target_idx] = val
            new_headers = final_headers
        # 3. Validate Sequence Type
        if not isinstance(new_headers, Iterable) or isinstance(new_headers, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(new_headers).__name__}', `new_headers` must be a list, tuple, dict, or callable returning a sequence of strings")
            raise TypeError(msg)
        # Ensure it is subscriptable list for comparisons
        new_headers = list(new_headers)
        # 4. Validate Dimensions
        if len(new_headers) != self.column_count:
            msg = ErrorFormat(f"DimensionError: invalid header dimensions, resulting headers length '{len(new_headers)}' != '{self.column_count}' column count")
            raise DimensionError(msg)
        # 5. Validate Content Types
        if not all(isinstance(header, str) for header in new_headers):
            msg = ErrorFormat("TypeError: invalid new header type(s), all values in `new_headers` must be of type 'str'")
            raise TypeError(msg)        
        # 6. Generate SQL (Only for columns that actually changed)
        rename_ops = []
        for i in range(self.column_count):
            if self.headers[i] != new_headers[i]:
                # Escape double quotes in column names if necessary, though simpler standard is usually enough
                rename_ops.append(f"""alter table "{self.sql_model}" rename column "{self.headers[i]}" to "{new_headers[i]}" """)
        # If nothing changed, just return
        if not rename_ops:
            return
        sql_stmt = ";".join(rename_ops)
        try:
            self.sql_db_conn.executescript(f"begin transaction; {sql_stmt}; end transaction;")
            self.sql_db_conn.commit()
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f'SQLProgrammingError: unable to update headers, SQL execution failed with: "{e}"')
            raise SQLProgrammingError(msg) from None
        self.headers = new_headers
        self._update_model_metadata(update_row_meta=False)

    def normalize_headers(self, apply_function:Callable=None) -> None:
        """
        Reformats the current ``SQLDataModel`` headers into an uncased normalized form using alphanumeric characters only.
        Wraps :meth:`SQLDataModel.set_headers()`.

        Parameters:
            ``apply_function`` (Callable, optional): Specify an alternative normalization pattern. When ``None``, the pattern
                ``'[^0-9a-z _]+'`` will be used on uncased values.

        Returns:
            ``None``

        Example::
            
            import sqldatamodel as sdm

            # Create model
            df = sdm.from_csv('example.csv', headers=['First Name', 'Last Name', 'Salary'])

            # Use default normalization scheme, uncased and strips invalid SQL identifiers
            df.normalize_headers()

            # Get renamed headers after default normalization
            df.get_headers() # now outputs ['first_name', 'last_name', 'salary']

            # Or use custom renaming scheme
            df.normalize_headers(lambda x: x.upper())

            # Get renamed headers again
            df.get_headers() # now outputs ['FIRST_NAME', 'LAST_NAME', 'SALARY']
        
        Changelog:
            - Version 1.2.0 (2025-01-28):
                - Added duplicate aliasing to prevent post-normalization name collisions using :meth:`utils.alias_duplicates`
                - Modified default normalization function to better handle occurrences of multiple invalid characters.
            - Version 0.1.5 (2023-11-24):
                - New method.            
        """
        if apply_function is None:
            apply_function = lambda x: re.sub(r'[^0-9a-z]+', '_', x.lower().strip()).strip('_')
        new_headers = list(utils.alias_duplicates([apply_function(x) for x in self.get_headers()]))
        self.set_headers(new_headers)

    def get_display_max_rows(self) -> int|None:
        """
        Retrieves the current value at :py:attr:`SQLDataModel.display_max_rows`, which determines the maximum rows displayed for the ``SQLDataModel``.

        Returns:
            ``int`` or ``None``: The current value set at :py:attr:`SQLDataModel.display_max_rows`.

        Example::

            import sqldatamodel as sdm

            # Create model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Get current value
            display_max_rows = df.get_display_max_rows()

            # By default rows will be limited by current terminal height
            print(display_max_rows) # None
        
        Note:
            - This does not affect the actual number of rows in the model, only the maximum **displayed**.
            - Use :meth:`SQLDataModel.set_display_max_rows()` to explicitly set a max row limit instead of using terminal height.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        return self.display_max_rows
    
    def set_display_max_rows(self, rows:int|None) -> None:
        """
        Sets value at :py:attr:`SQLDataModel.display_max_rows` to limit maximum rows displayed when ``repr`` or ``print`` is called. Use ``rows = None`` to derive max number to display from the current terminal height.

        Parameters:
            ``rows`` (int): The maximum number of rows to display.

        Raises:
            ``TypeError``: If the provided argument is not ``None`` or is not an integer.
            ``IndexError``: If the provided value is an integer less than or equal to 0.

        Returns:
            ``None``

        Example::
        
            import sqldatamodel as sdm

            # Create model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Any call to `print` or `repr` will be restricted to 500 max rows
            df.set_display_max_rows(500)

            # Alternatively, auto-detect dimensions by setting to `None`
            df.set_display_max_rows(None)
        
        Note:
            - Modifying :py:attr:`SQLDataModel.display_max_rows` does not affect the actual number of rows in the model, only the maximum rows **displayed**.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(rows, (int,type(None))):
            msg = ErrorFormat(f'TypeError: invalid argument type "{type(rows).__name__}", please provide an integer value to set the maximum rows attribute...')
            raise TypeError(msg)
        if isinstance(rows, int) and rows <= 0:
            msg = ErrorFormat(f'IndexError: invalid value "{rows}", please provide an integer value >= 1 to set the maximum rows attribute...')
            raise IndexError(msg)
        self.display_max_rows = rows

    def get_min_column_width(self) -> int:
        """
        Returns the current ``min_column_width`` property value.

        Returns:
            ``int``: The current value of the ``min_column_width`` property.

        Example::
            
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Get and save the current value
            min_width = df.get_min_column_width()

            # Output
            print(min_width)  # 6
        
        """
        return self.min_column_width
    
    def set_min_column_width(self, width:int) -> None:
        """
        Set ``min_column_width`` as the minimum number of characters per column when ``repr`` or ``print`` is called.

        Parameters:
            ``width`` (int): The minimum width for each column.

        Returns:
            ``None``: Sets the ``min_column_width`` property.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Set a new minimum column width value
            df.set_min_column_width(8)

            # Check updated value
            print(df.min_column_width) # 8
        
        Note:
            - If ``min_column_width`` is set to a value below the current ``max_column_width`` property, the maximum width will override the minimum width.
            - The minimum required width is ``2``, when ``min_column_width < 2``, ``2`` will be used regardless of the ``width`` provided.
            - See :meth:`SQLDataModel.set_max_column_width()` to set maximum column width for table representations.
        """
        self.min_column_width = width if width >= 2 else 2

    def get_max_column_width(self) -> int:
        """
        Returns the current ``max_column_width`` property value.

        Returns:
            ``int``: The current value of the ``max_column_width`` property.

        Example::
        
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Get the current max column width value
            max_width = df.get_max_column_width()

            # Output
            print(max_width)  # 32
        
        """
        return self.max_column_width
    
    def set_max_column_width(self, width:int) -> None:
        """
        Set ``max_column_width`` as the maximum number of characters per column when ``repr`` or ``print`` is called.

        Parameters:
            ``width`` (int): The maximum width for each column.

        Returns:
            ``None``: Sets the ``max_column_width`` property.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Change the max column width for the table representation
            df.set_max_column_width(20)
        
        Note:
            - If ``max_column_width`` is set to a value below the current ``min_column_width`` property, the maximum width will override the minimum width.
            - The minimum required width is ``2``, when ``max_column_width < 2``, ``2`` will be used regardless of the ``width`` provided.        
            - See :meth:`SQLDataModel.set_min_column_width()` to set minimum column width for table representations.        
        """
        self.max_column_width = width if width >= 2 else 2

    def get_column_alignment(self) -> str:
        """
        Returns the current ``column_alignment`` property value, ``dynamic`` by default.

        Returns:
            ``str``: The current value of the ``column_alignment`` property.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Get the current alignment value
            alignment = df.get_column_alignment()

            # Outputs 'dynamic'
            print(alignment)

        Note:
            - Use :meth:`SQLDataModel.set_column_alignment()` to modify column alignment.
        """
        return self.column_alignment
    
    def set_column_alignment(self, alignment:Literal['dynamic', 'left', 'center', 'right']='dynamic') -> None:
        """
        Sets the default alignment behavior for ``SQLDataModel`` when ``repr`` or ``print`` is called, modifies ``column_alignment`` attribute.
        Default behavior set to ``'dynamic'``, which right-aligns numeric data types, left-aligns all other types, with headers matching value alignment.
        
        Parameters:
            ``alignment`` (str): The column alignment setting to use.
                ``'dynamic'``: Default behavior, dynamically aligns columns based on column data types.
                ``'left'``: Left-align all column values.
                ``'center'``: Center-align all column values.
                ``'right'``: Right-align all column values.
        
        Raises:
            ``TypeError``: If the argument for alignment is not of type 'str'.
            ``ValueError``: If the provided alignment is not one of 'dynamic', 'left', 'center', 'right'.

        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Set to right-align columns
            df.set_column_alignment('right')

            # Output
            print(df)
        
        This will output the model with values right-aligned:

        ```shell
            ┌───┬────────┬─────────┬────────┬─────────┐
            │   │  first │    last │    age │ service │
            ├───┼────────┼─────────┼────────┼─────────┤
            │ 0 │   john │   smith │     27 │    1.22 │
            │ 1 │  sarah │    west │     39 │    0.70 │
            │ 2 │   mike │  harlin │     36 │    3.00 │
            │ 3 │    pat │ douglas │     42 │   11.50 │
            └───┴────────┴─────────┴────────┴─────────┘        
        ```

        Setting columns to be left-aligned:

        ```python            
            # Set to left-align
            sdm.set_column_alignment('left')

            # Output
            print(df)
        ```

        This will output the model with left-aligned values instead:

        ```text            
            ┌───┬────────┬─────────┬────────┬─────────┐
            │   │ first  │ last    │ age    │ service │
            ├───┼────────┼─────────┼────────┼─────────┤
            │ 0 │ john   │ smith   │  27    │  1.22   │
            │ 1 │ sarah  │ west    │  39    │  0.70   │
            │ 2 │ mike   │ harlin  │  36    │  3.00   │
            │ 3 │ pat    │ douglas │  42    │  11.50  │
            └───┴────────┴─────────┴────────┴─────────┘        
        ```

        Note:
            - Use :meth:`SQLDataModel.get_column_alignment()` to return the current column alignment setting.
            - When using 'center', if the column contents cannot be perfectly centralized, the left side will be favored.
            - Use 'dynamic' to return to default column alignment, which is right-aligned for numeric types and left-aligned for others.
            - See :meth:`SQLDataModel.set_table_style()` for modifying table format and available styles.

        Changelog:
            - Version 0.1.80 (2024-02-24):
                - Changed expected values for ``alignment`` parameter from f-string modifiers to more descriptive values 'dynamic', 'left', 'center' or 'right'.
        """
        if not isinstance(alignment, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(alignment).__name__}', expected type for `alignment` to be type 'str', setting the alignment style to use")
            raise TypeError(msg)
        if alignment not in ('dynamic', 'left', 'center', 'right'):
            msg = ErrorFormat(f"ValueError: invalid value '{alignment}', argument for `alignment` must be one of 'dynamic', 'left', 'center', 'right' representing the column alignment setting, use 'dynamic' for default behaviour")
            raise ValueError(msg)
        self.column_alignment = alignment
        return

    def get_display_index(self) -> bool:
        """
        Returns the current value set at :py:attr:`SQLDataModel.display_index`, which determines whether or not the index is displayed in the ``SQLDataModel`` representation.

        Returns:
            ``bool``: The current value of the ``display_index`` property.

        Example::
        
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Get the current value for displaying the index
            display_index = df.get_display_index()

            # Output: True
            print(display_index)
        
        Note:
            - Use :meth:`SQLDataModel.set_display_index()` to modify this property and toggle index display visibility.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        return self.display_index

    def set_display_index(self, display_index:bool) -> None:
        """
        Sets the value for :py:attr:`SQLDataModel.display_index` to enable or disable the inclusion of the
        ``SQLDataModel`` index value in print or repr calls.

        Parameters:
            ``display_index`` (bool): Whether or not to include the index in ``SQLDataModel`` representations.

        Raises:
            ``TypeError``: If the provided argument is not a boolean value.

        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Disable displaying index
            df.set_display_index(False)

        Note:
            - Use :meth:`SQLDataModel.set_table_style()` to more broadly modify the appearance and formatting style of ``SQLDataModel`` string representations.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(display_index, bool):
            msg = ErrorFormat(f"TypeError: invalid type '{type(display_index).__name__}', argument for `display_index` must be of type 'bool' representating whether or not the index should be displayed")
            raise TypeError(msg)
        self.display_index = display_index
    
    def get_shape(self) -> tuple[int, int]:
        """
        Returns the current shape of the ``SQLDataModel`` as a tuple of ``(rows x columns)``.

        Returns:
            ``tuple[int, int]``: A tuple representing the current dimensions of rows and columns in the ``SQLDataModel``.

        Example::

            import sqldatamodel as sdm

            # Create the model
            sdm = SQLDataModel([[1,2,3],
                                [4,5,6],
                                [7,8,9]])

            # Get the current shape
            shape = df.get_shape()

            # View it
            print("shape:", shape)

        This will output:

        ```text
            shape: (3, 3)
        ```

        The shape can also be seen when printing the model:

        ```python
            import sqldatamodel as sdm

            # Create the model
            df = sdm.SQLDataModel([[1,2,3],
                                [4,5,6],
                                [7,8,9]])

            # View it and the shape
            print(sdm, "<-- shape is also visible here")
        ```

        This will output:

        ```text
            ┌───┬───────┬───────┬───────┐
            │   │ col_0 │ col_1 │ col_2 │
            ├───┼───────┼───────┼───────┤
            │ 0 │     1 │     2 │     3 │
            │ 1 │     4 │     5 │     6 │
            │ 2 │     7 │     8 │     9 │
            └───┴───────┴───────┴───────┘
            [3 rows x 3 columns] <-- shape is also visible here
        ```

        Note:
            - If an empty model is initialized, the :py:attr:`SQLDataModel.row_count` will be 0 until the first row is inserted.
            - Using the :meth:`SQLDataModel.__getitem__()` syntax of ``df[row, col]`` returns a new model instance with the corresponding shape.

        Changelog:
            - Version 0.3.6 (2024-04-09):
                - Returns the new :py:attr:`SQLDataModel.shape` directly, making this method redundant.
        """
        return self.shape
    
    def get_display_float_precision(self) -> int:
        """
        Retrieves the current float display precision used exclusively for representing the values of real numbers
        in the ``repr`` method for the ``SQLDataModel``. Default value is set to 4 decimal places of precision.

        Returns:
            ``int``: The current float display precision.

        Note:
            - The float display precision is the number of decimal places to include when displaying real numbers in the string representation of the ``SQLDataModel``.
            - This value is utilized in the ``repr`` method to control the precision of real number values.
            - The method does not affect the actual value of float dtypes in the underlying ``SQLDataModel``

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        return self.display_float_precision
    
    def set_display_float_precision(self, float_precision:int) -> None:
        """
        Sets the current float display precision to the specified value for use in the ``repr`` method of the ``SQLDataModel``
        when representing float data types. Note that this precision limit is overridden by the ``max_column_width`` value
        if the precision limit exceeds the specified maximum width.

        Parameters:
            ``float_precision`` (int): The desired float display precision to be used for real number values.

        Raises:
            ``TypeError``: If the ``float_precision`` argument is not of type 'int'.
            ``ValueError``: If the ``float_precision`` argument is a negative value, as it must be a valid f-string precision identifier.

        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age', 'service_time']
            data = [
                (0, 'john', 'smith', 27, 1.22)
                ,(1, 'sarah', 'west', 0.7)
                ,(2, 'mike', 'harlin', 3)
                ,(3, 'pat', 'douglas', 11.5)
            ]

            # Create the model with sample data
            df = sdm.SQLDataModel(data,headers)

            # Example: Set the float display precision to 2
            df.set_display_float_precision(2)

            # View model
            print(df)

        This will output:
        
        ```shell
            ┌───┬────────┬─────────┬────────┬──────────────┐
            │   │ first  │ last    │    age │ service_time │
            ├───┼────────┼─────────┼────────┼──────────────┤
            │ 0 │ john   │ smith   │     27 │         2.10 │
            │ 1 │ sarah  │ west    │     29 │         0.80 │
            │ 2 │ mike   │ harlin  │     36 │         1.30 │
            │ 3 │ pat    │ douglas │     42 │         7.02 │
            └───┴────────┴─────────┴────────┴──────────────┘
        ```

        Use :meth:`SQLDataModel.get_display_float_precision()` to get the current value set:

        ```python
            # Get the updated float display precision
            updated_precision = sdm.get_display_float_precision()

            # Outputs 2
            print(updated_precision)
        ```

        Note:
            - The ``display_float_precision`` attribute only affects the precision for displaying real or floating point values.
            - The actual precision of the stored value in the model is unaffected by the value set.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(float_precision, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(float_precision).__name__}' received for `float_precision` argument, expected type 'int'")
            raise TypeError(msg)
        if float_precision < 0:
            msg = ErrorFormat(f"ValueError: invalid value '{float_precision}' received for `float_precision` argument, value must be valid f-string precision identifier")
            raise ValueError(msg)
        self.display_float_precision = float_precision
        
    def describe(self, exclude_columns:str|list=None, exclude_dtypes:list[Literal["str","int","float","date","datetime","bool"]]=None, ignore_na:bool=True, **kwargs) -> SQLDataModel:
        """
        Generates descriptive statistics for columns in the ``SQLDataModel`` instance based on column dtype including count, unique values, top value, frequency, mean, standard deviation, minimum, 25th, 50th, 75th percentiles, maximum and dtype for specified column.

        Parameters:
            ``exclude_columns`` (str | list, optional): Columns to exclude from the analysis. Default is None.
            ``exclude_dtypes`` (list[Literal["str", "int", "float", "date", "datetime", "bool"]], optional): Data types to exclude from the analysis. Default is None.
            ``ignore_na`` (bool, optional): If True, ignores NA like values ('NA', ' ', 'None') when computing statistics. Default is True.
            ``**kwargs``: Additional keyword arguments to be passed to the ``execute_fetch`` method.

        Statistics Described:
            - ``count``: Total number of non-null values for specified column
            - ``unique``: Total number of unique values for specified column
            - ``top``: Top value represented for specified column, ties broken arbitrarily
            - ``freq``: Frequency of corresponding value represented in 'top' metric
            - ``mean``: Mean as calculated by summing all values and dividing by 'count'
            - ``std``: Standard Deviation for specified column

              - Uncorrected sample standard deviation for ``int``, ``float`` dtypes
              - Mean time difference represented in number of days for ``date``, ``datetime`` dtypes
              - 'NaN' for all other dtypes

            - ``min``: Minimum value for specified column

              - Least value for ``int``, ``float`` dtypes
              - Least value sorted by alphabetical ascending for ``str`` dtypes
              - Earliest date or datetime for ``date``, ``datetime`` dtypes

            - ``p25``: Percentile, 25th

              - Max first bin value as determined by quartered binning of values for ``int``, ``float`` dtypes
              - 'NaN' for all other dtypes

            - ``p50``: Percentile, 50th

              - Max second bin value as determined by quartered binning of values for ``int``, ``float`` dtypes
              - 'NaN' for all other dtypes

            - ``p75``: Percentile, 75th

              - Max third bin value as determined by quartered binning of values for ``int``, ``float`` dtypes
              - 'NaN' for all other dtypes     

            - ``max``: Maximum value for specified column

              - Greatest value for ``int``, ``float`` dtypes
              - Greatest value sorted by alphabetical ascending for ``str`` dtypes
              - Latest date or datetime for ``date``, ``datetime`` dtypes  

            - ``dtype``: Datatype of specified column

              - Python datatype as determined by relevant class ``__name__`` attribute, e.g. 'float' or 'int'
              - dtypes can be excluded by using ``exclude_dtypes`` parameter

        Returns:
            ``SQLDataModel``: A new SQLDataModel containing a comprehensive set of descriptive statistics for selected columns.

        Note:
            - Standard deviation is calculated using uncorrected sample standard deviation for numeric dtypes, and timediff in days for datetime dtypes
            - Ties in unique, top and freq columns are broken arbitrarily as determined by first ordering of values prior to calling ``describe()``
            - Ties encountered when binning for p25, p50, p75 will favor lower bins for data that cannot be quartered cleanly
            - Metrics for count, min, p25, p50, p75 and max include non-null values only
            - Using ``ignore_na=True`` only affects inclusion of 'NA like' values such as empty strings
            - Floating point precision determined by :py:attr:`SQLDataModel.display_float_precision` attribute
        
        Example::
            
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('employees.csv')

            # View all 10 rows
            print(df)

        This will output:
        
        ```shell            
            ┌───┬──────────────────┬────────────┬─────────────┬───────────────┬────────┬─────────────────────┐
            │   │ name             │ hire_date  │ country     │ service_years │    age │ last_update         │
            ├───┼──────────────────┼────────────┼─────────────┼───────────────┼────────┼─────────────────────┤
            │ 0 │ Pamela Berg      │ 2007-06-06 │ New Zealand │          3.02 │     56 │ 2023-08-12 17:13:46 │
            │ 1 │ Mason Hoover     │ 2009-04-19 │ Australia   │          5.01 │     41 │ 2023-05-18 01:29:44 │
            │ 2 │ Veda Suarez      │ 2007-07-02 │ Ukraine     │          4.65 │     26 │ 2023-12-09 15:38:01 │
            │ 3 │ John Smith       │ 2017-08-12 │ New Zealand │          3.81 │     35 │ 2023-03-10 18:23:56 │
            │ 4 │ Xavier McCoy     │ 2021-04-03 │ France      │          2.95 │     42 │ 2023-09-27 11:39:08 │
            │ 5 │ John Smith       │ 2020-10-11 │ Germany     │          4.61 │     56 │ 2023-12-09 18:41:52 │
            │ 6 │ Abigail Mays     │ 2021-07-25 │ Costa Rica  │          5.34 │     50 │ 2023-02-11 16:43:07 │
            │ 7 │ Rama Galloway    │ 2009-02-09 │ Italy       │          3.87 │     24 │ 2023-03-13 16:08:48 │
            │ 8 │ Lucas Rodriquez  │ 2018-06-19 │ New Zealand │          2.73 │     28 │ 2023-03-17 01:45:22 │
            │ 9 │ Hunter Donaldson │ 2015-12-18 │ Belgium     │          4.58 │     43 │ 2023-04-06 03:22:54 │
            └───┴──────────────────┴────────────┴─────────────┴───────────────┴────────┴─────────────────────┘      
            [10 rows x 6 columns]  
        ```
        
        Now that we have our ``SQLDataModel``, we can generate some statistics:

        ```python
            # Generate statistics
            df_described = df.describe()

            # View stats
            print(df_described)
        ```

        This will output:
        
        ```shell
            ┌────────┬──────────────┬─────────────┬─────────────┬───────────────┬────────┬─────────────────────┐
            │ metric │         name │   hire_date │     country │ service_years │    age │         last_update │
            ├────────┼──────────────┼─────────────┼─────────────┼───────────────┼────────┼─────────────────────┤
            │ count  │           10 │          10 │          10 │            10 │     10 │                  10 │
            │ unique │            9 │          10 │           8 │            10 │      9 │                  10 │
            │ top    │   John Smith │  2021-07-25 │ New Zealand │          5.34 │     56 │ 2023-12-09 18:41:52 │
            │ freq   │            2 │           1 │           3 │             1 │      2 │                   1 │
            │ mean   │          NaN │  2014-11-24 │         NaN │          4.06 │     40 │ 2023-06-16 19:18:39 │
            │ std    │          NaN │ 2164.4 days │         NaN │          0.92 │     11 │         117.58 days │
            │ min    │ Abigail Mays │  2007-06-06 │   Australia │          2.73 │     24 │ 2023-02-11 16:43:07 │
            │ p25    │          NaN │  2009-02-09 │         NaN │          3.02 │     28 │ 2023-03-13 16:08:48 │
            │ p50    │          NaN │  2017-08-12 │         NaN │          4.58 │     42 │ 2023-05-18 01:29:44 │
            │ p75    │          NaN │  2020-10-11 │         NaN │          4.65 │     50 │ 2023-09-27 11:39:08 │
            │ max    │ Xavier McCoy │  2021-07-25 │     Ukraine │          5.34 │     56 │ 2023-12-09 18:41:52 │
            │ dtype  │          str │        date │         str │         float │    int │            datetime │
            └────────┴──────────────┴─────────────┴─────────────┴───────────────┴────────┴─────────────────────┘
            [12 rows x 7 columns]    
        ```

        Specific columns or data types can be excluded from result:

        ```python
            # Set filters to exclude all str dtypes and the 'hire_date' column:
            df_describe = df.describe(exclude_dtypes=['str'], exclude_columns=['hire_date'])

            # View statistics
            print(df_described)
        ```

        This will output:

        ```shell
            ┌────────┬───────────────┬────────┬─────────────────────┐
            │ metric │ service_years │    age │         last_update │
            ├────────┼───────────────┼────────┼─────────────────────┤
            │ count  │            10 │     10 │                  10 │
            │ unique │            10 │      9 │                  10 │
            │ top    │          5.34 │     56 │ 2023-10-28 05:42:43 │
            │ freq   │             1 │      2 │                   1 │
            │ mean   │          4.06 │     40 │ 2023-08-11 23:18:12 │
            │ std    │          0.92 │     11 │          73.15 days │
            │ min    │          2.73 │     24 │ 2023-04-07 23:56:06 │
            │ p25    │          3.02 │     28 │ 2023-06-02 14:36:19 │
            │ p50    │          4.58 │     42 │ 2023-09-09 19:18:38 │
            │ p75    │          4.65 │     50 │ 2023-10-09 19:34:55 │
            │ max    │          5.34 │     56 │ 2023-10-28 05:42:43 │
            │ dtype  │         float │    int │            datetime │
            └────────┴───────────────┴────────┴─────────────────────┘
            [12 rows x 4 columns]
        ```
        
        Important:
            - Generally, do not rely on ``SQLDataModel`` to do statistics, use ``NumPy`` or a real scientific computing library instead.
            
        Note:
            - Use :meth:`SQLDataModel.infer_dtypes()` to cast columns to their apparent data type, or set it manually with :meth:`SQLDataModel.set_column_dtypes()` to convert columns to different data types.
            - Statistics for ``date`` and ``datetime`` can be unpredictable if formatting used is inconsistent with conversion to Julian days or if column data type is incorrect.

        Changelog:
            - Version 0.6.3 (2024-05-16):
                - Modified model to output values as string data types and set columns to right-aligned if arguments are not present in ``kwargs`` to retain metric resolution while having numeric alignment.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if exclude_columns is None:
            exclude_columns = []
        elif isinstance(exclude_columns, (str)):
            exclude_columns = [exclude_columns]
        if exclude_dtypes is None:
            exclude_dtypes = []
        elif isinstance(exclude_dtypes, (str)):
            exclude_dtypes = [exclude_dtypes]
        desc_cols = [col for col in self.headers if ((col not in exclude_columns) and (self.header_master[col][1] not in exclude_dtypes))]
        if (num_cols :=len(desc_cols)) < 1:
            msg = ErrorFormat(f"ValueError: invalid number of columns '{num_cols}', at least '1' column is required for the ``describe()`` method")
            raise ValueError(msg)
        self.sql_db_conn.create_aggregate("stdev", 1, StandardDeviation) # register stdev to calculate standard deviation in sqlite        
        has_numeric_dtype = any(map(lambda v: v in ('float','int','date','datetime'), [self.header_master[col][1] for col in desc_cols])) # ('float','int','date','datetime')
        headers_select_literal = [f""" "'{col}'" as "{col}" """ for col in desc_cols]
        headers_select = """ "'metric'" as "metric",""" + ",".join(headers_select_literal)
        headers_sub_literal = [f"'{col}'" for col in desc_cols]
        headers_subselect = " select 'metric'," + ",".join(headers_sub_literal)
        count_subselect = "select 'count'," + ",".join([f""" count("{col}") """ if not ignore_na else f"""(select count("{col}") from "{self.sql_model}" where trim(upper("{col}")) not in (' ', '', 'NA', 'NONE','NULL')) """ for col in desc_cols])
        count_subselect = f"""{count_subselect} {f'from "{self.sql_model}"' if not ignore_na else ''}"""
        unique_subselect = "select 'unique'," + ",".join([f""" count(distinct "{col}") """ if not ignore_na else f"""(select count(distinct "{col}") from "{self.sql_model}" where trim(upper("{col}")) not in (' ', '', 'NA', 'NONE','NULL')) """ for col in desc_cols])
        unique_subselect = f"""{unique_subselect} {f'from "{self.sql_model}"' if not ignore_na else ''}"""
        top_subselect = "select 'top'," + ",".join([f"""(select max("{col}") from "{self.sql_model}" group by "{col}" order by count(*) desc limit 1) """ if not ignore_na else f"""(select max("{col}") from "{self.sql_model}" where trim(upper("{col}")) not in (' ', '', 'NA', 'NONE','NULL') group by "{col}" order by count(*) desc limit 1) """ for col in desc_cols])
        freq_subselect = "select 'freq'," + ",".join([f"""(select count("{col}") from "{self.sql_model}" group by "{col}" order by count(*) desc limit 1) """ if not ignore_na else f"""(select count("{col}") from "{self.sql_model}" where trim(upper("{col}")) not in (' ', '', 'NA', 'NONE','NULL') group by "{col}" order by count(*) desc limit 1) """ for col in desc_cols])
        pcte_stmt = f"""with pcte as (select {','.join([f'"{col}",ntile(4) over (order by "{col}") as "p{col}"' for col in desc_cols if self.header_master[col][1] in ('float','int','date','datetime')])} from "{self.sql_model}")""" if has_numeric_dtype else ""
        mean_subselect = "select * from (select 'mean'," + ",".join([f"""round(avg("{col}"),{self.display_float_precision}) """ if self.header_master[col][1] == 'float' else f"""printf('%d',avg("{col}"))""" if self.header_master[col][1] == 'int' else f"""date(avg(julianday("{col}")))""" if self.header_master[col][1] == "date" else f"""datetime(avg(julianday("{col}")))""" if self.header_master[col][1] == "datetime" else "'NaN'" for col in desc_cols]) + f"""from "{self.sql_model}" limit 1) """
        std_subselect = "select * from (select 'std'," + ",".join([f"""round(stdev("{col}"),{self.display_float_precision}) """ if self.header_master[col][1] == 'float' else f"""printf('%d',stdev("{col}"))""" if self.header_master[col][1] == 'int' else f"""round(stdev(julianday("{col}")),{self.display_float_precision})||' days'""" if self.header_master[col][1] in ("date","datetime") else "'NaN'" for col in desc_cols]) + f"""from "{self.sql_model}" limit 1) """
        min_subselect = "select * from (select 'min'," + ",".join([f"""round(min("{col}"),{self.display_float_precision}) """ if self.header_master[col][1] in ('float','int') else f"""min("{col}")""" if (not ignore_na or self.header_master[col][1] in ("date","datetime")) else f"""(select min("{col}") from "{self.sql_model}" where trim(upper("{col}")) not in (' ', '', 'NA', 'NONE','NULL')) """ for col in desc_cols]) + f"""from "{self.sql_model}" limit 1) """
        p25_subselect = "select 'p25'," + ",".join([f"""(select round(max("{col}"),{self.display_float_precision}) as "max25{col}" from pcte where "p{col}" = 1 group by "p{col}") as "p25{col}" """ if self.header_master[col][1] in ('float','int') else f"""(select max("{col}") as "max25{col}" from pcte where "p{col}" = 1 group by "p{col}") as "p25{col}" """ if self.header_master[col][1] in ('date','datetime') else "'NaN'" for col in desc_cols])
        p50_subselect = "select 'p50'," + ",".join([f"""(select round(max("{col}"),{self.display_float_precision}) as "max50{col}" from pcte where "p{col}" = 2 group by "p{col}") as "p50{col}" """ if self.header_master[col][1] in ('float','int') else f"""(select max("{col}") as "max50{col}" from pcte where "p{col}" = 2 group by "p{col}") as "p50{col}" """ if self.header_master[col][1] in ('date','datetime') else "'NaN'" for col in desc_cols])
        p75_subselect = "select 'p75'," + ",".join([f"""(select round(max("{col}"),{self.display_float_precision}) as "max75{col}" from pcte where "p{col}" = 3 group by "p{col}") as "p75{col}" """ if self.header_master[col][1] in ('float','int') else f"""(select max("{col}") as "max75{col}" from pcte where "p{col}" = 3 group by "p{col}") as "p75{col}" """ if self.header_master[col][1] in ('date','datetime') else "'NaN'" for col in desc_cols])
        max_subselect = "select * from (select 'max'," + ",".join([f"""round(max("{col}"),{self.display_float_precision}) """ if self.header_master[col][1] in ('float','int') else f"""max("{col}")""" if (not ignore_na or self.header_master[col][1] in ("date","datetime")) else f"""(select max("{col}") from "{self.sql_model}" where trim(upper("{col}")) not in (' ', '', 'NA', 'NONE','NULL')) """ for col in desc_cols]) + f"""from "{self.sql_model}" limit 1) """
        dtype_subselect = "select 'dtype'," + ",".join([f"""'{self.header_master[col][1]}'""" for col in desc_cols])
        full_script = f"""{pcte_stmt} select {headers_select} from ({headers_subselect} UNION ALL {count_subselect} UNION ALL {unique_subselect} UNION ALL {top_subselect} UNION ALL {freq_subselect} UNION ALL {mean_subselect} UNION ALL {std_subselect} UNION ALL {min_subselect} UNION ALL {p25_subselect} UNION ALL {p50_subselect} UNION ALL {p75_subselect} UNION ALL {max_subselect} UNION ALL {dtype_subselect}) limit -1 offset 1"""
        if 'column_alignment' not in kwargs:
            kwargs['column_alignment'] = 'right'
        if 'dtypes' not in kwargs:
            kwargs['dtypes'] = {col:'str' for col in desc_cols}
        return self.execute_fetch(full_script, display_index=False, **kwargs)

    def sample(self, n_samples:float|int=0.05, **kwargs) -> SQLDataModel:
        """
        Return a random sample of size ``n_samples`` as a new ``SQLDataModel``.

        Parameters:
            ``n_samples`` (float | int): Number of rows or proportion of rows to sample. Default set to ``0.05``, proportional to 5% of the current :py:attr:`SQLDataModel.row_count`.
                If ``n_samples`` is an integer, it represents the exact number of rows to sample where ``0 < n_samples <= row_count``.
                If ``n_samples`` is a float, it represents the proportion of rows to sample where ``0.0 < n_samples <= 1.0``.

        Returns:
            ``SQLDataModel``: A new SQLDataModel instance containing the sampled rows.

        Raises:
            ``TypeError``: If the ``n_samples`` parameter is not of type 'int' or 'float'.
            ``ValueError``: If the ``n_samples`` value is invalid or out of range.

        This method generates a random sample of rows from the current SQLDataModel. The number of rows to sample
        can be specified either as an integer representing the exact number of rows or as a float representing
        the proportion of rows to sample. The sampled rows are returned as a new SQLDataModel instance.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Amount'])
            
            # Example 1: Sample 10 random rows
            sample_result = df.sample(n_samples=10)

            # Create the model
            df2 = sdm.from_csv('another_example.csv', headers=['Code', 'Description', 'Price'])
            
            # Example 2: Sample 20% of rows
            sample_result2 = df2.sample(n_samples=0.2)

        Note:
            - If the current model's :py:attr:`SQLDataModel.row_count` value is less than the sample size, the current row count will be used instead.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(n_samples, (float,int)):
            msg = ErrorFormat(f"TypeError: invalid `n_samples` type '{type(n_samples).__name__}', `n_samples` parameter type must be one of 'int', 'float' as number of rows or proportion of rows, respectively")
            raise TypeError(msg)            
        if isinstance(n_samples, float):
            if (n_samples <= 0.0) or (n_samples > 1.0):
                msg = ErrorFormat(f"ValueError: invalid `n_samples` value '{n_samples}', expected value in range '0.0 < n_samples <= 1.0' when using proportional value for `n_samples`")
                raise ValueError(msg)
            n_samples = round(self.row_count * n_samples)
        n_samples = min(n_samples, self.row_count)
        if n_samples <= 0:
            msg = ErrorFormat(f"ValueError: invalid `n_samples` value '{n_samples}', expected value within current row range '0 < n_samples <= {self.row_count}' when using integer value for `n_samples`")
            raise ValueError(msg) 
        row_indicies = tuple(random.sample(self.indicies, n_samples))
        return self.execute_fetch(self._generate_sql_stmt(rows=row_indicies), **kwargs)

    def infer_dtypes(self, n_samples:int=16, date_format:str='%Y-%m-%d', datetime_format:str='%Y-%m-%d %H:%M:%S') -> None:
        """
        Infer and set data types for columns based on a random subset of ``n_samples`` from the current model. 
        The ``dateutil`` library is required for complex date and datetime parsing, if the module is not found then ``date_format`` and ``datetime_format`` will be used for dates and datetimes respectively.

        Parameters:
            ``n_samples`` (int): The number of random samples to use for data type inference. Default set to `16`.
            ``date_format`` (str): The format string to use for parsing date values if ``dateutil`` library is not found. Default is `'%Y-%m-%d'`.
            ``datetime_format`` (str): The format string to use for parsing datetime values if ``dateutil`` library is not found. Default is `'%Y-%m-%d %H:%M:%S'`.
        
        Raises:
            ``TypeError``: If argument for ``n_samples`` is not of type ``int`` or if argument for ``date_format`` or ``datetime_format`` is not of type 'str'.
            ``ValueError``: If the current model contains zero columns from which to infer types from.
            ``DimensionError``: If the current model contains insufficient rows to sample from.

        Returns:
            ``None``: Inferred column types are updated and ``None`` is returned.

        Example::
        
            import sqldatamodel as sdm

            # Sample data of ``str`` containing probable datatypes
            headers = ['first', 'last', 'age', 'service', 'hire_date']
            data = [
                ('John', 'Smith', '27', '1.22', '2023-02-01'),
                ('Sarah', 'West', '39', '0.7', '2023-10-01'),
                ('Mike', 'Harlin', '36', '3.9', '2020-08-27'),
                ('Pat', 'Douglas', '42', '11.5', '2015-11-06'),
                ('Kelly', 'Lee', '32', '8.0', '2016-09-18')
            ]     

            # Create the model
            df = sdm.SQLDataModel(data, headers)
            
            # Get current column dtypes for reference
            dtypes_before = df.get_column_dtypes()

            # Infer and set data types based on 10 random samples
            df.infer_dtypes(n_samples=10)

            # View updated model
            print(df)

        This will output data with dtypes correctly aligned:
        
        ```shell            
            ┌───────┬─────────┬──────┬─────────┬────────────┐
            │ first │ last    │  age │ service │ hire_date  │
            ├───────┼─────────┼──────┼─────────┼────────────┤
            │ John  │ Smith   │   27 │    1.22 │ 2023-02-01 │
            │ Sarah │ West    │   39 │    0.70 │ 2023-10-01 │
            │ Mike  │ Harlin  │   36 │    3.90 │ 2020-08-27 │
            │ Pat   │ Douglas │   42 │   11.50 │ 2015-11-06 │
            │ Kelly │ Lee     │   32 │    8.00 │ 2016-09-18 │
            └───────┴─────────┴──────┴─────────┴────────────┘
            [5 rows x 5 columns]
        ```

        Use :meth:`SQLDataModel.get_column_dtypes()` or :py:attr:`SQLDataModel.dtypes` to view current types:
        
        ```python
            # Get new column types to confirm
            dtypes_after = df.get_column_dtypes()

            # View updated dtypes
            for col in df.headers:
                print(f"{col:<10} {dtypes_before[col]} -> {dtypes_after[col]}")
        ```

        This will output:
            
        ```shell
            first:      str -> str
            last:       str -> str
            age:        str -> int
            service:    str -> float
            hire_date:  str -> date 
        ```

        Note:
            - If a single ``str`` instance is found in the samples, the corresponding column dtype will remain as ``str`` to avoid data loss.
            - Co-occurences of ``int`` & ``float``, or ``date`` & ``datetime`` will favor the superset dtype after ``infer_threshold`` is met, so ``float`` and ``datetime`` respectively.
            - If a single ``datetime`` instance is found amongst a higher proportion of ``date`` dtypes, ``datetime`` will be used according to second rule.
            - If a single ``float`` instance is found amongst a higher proportion of ``int`` dtypes, ``float`` will be used according to second rule.
            - Ties between dtypes are broken according to `current type` < ``str`` < ``float`` < ``int`` < ``datetime`` < ``date`` < ``bytes`` < ``None``
            - This method calls :meth:`SQLDataModel.set_column_dtypes()` once the column dtypes have been inferred if they differ from the current dtype.
            - See :meth:`SQLDataModel.infer_str_type()` for type determination process.
            - See :meth:`utils.infer_types_from_data()` for type voting scheme used for inference.

        Changelog:
            - Version 0.2.0 (2024-03-19):
                - Increased sampling size for inference from ``n_samples=10`` to ``n_samples=16`` for better resolution.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(n_samples, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(n_samples).__name__}', argument for `n_samples` must be of type 'int' representing number of samples to use for inference")
            raise TypeError(msg)
        if not isinstance(date_format, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(date_format).__name__}', argument for `date_format` must be of type 'str' representing date format to use if `dateutil` is not found")
            raise TypeError(msg)  
        if not isinstance(datetime_format, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(datetime_format).__name__}', argument for `datetime_format` must be of type 'str' representing datetime format to use if `dateutil` is not found")
            raise TypeError(msg)                
        if self.row_count < 1:
            msg = ErrorFormat(f"DimensionError: invalid row count '{self.row_count}', at least 1 row is required for sampling when using `infer_dtypes()`")
            raise DimensionError(msg)
        str_columns = [col for col in self.headers if self.header_master[col][1] == 'str']
        if len(str_columns) < 1:
            msg = ErrorFormat(f"ValueError: zero inferrable columns '{len(str_columns)}', at least 1 column containing values of type 'str' is required to infer types from")
            raise ValueError(msg)
        n_samples = min(n_samples, self.row_count)
        row_targets = tuple(random.sample(self.indicies, n_samples))
        row_targets = row_targets if len(row_targets) > 1 else f"({row_targets[0]})"
        fetch_str_dtype_stmt = " ".join((f"""select""", ",".join([f'trim("{col}")' for col in str_columns]),f"""from "{self.sql_model}" where "{self.sql_idx}" in {row_targets} """))
        sample_data = self.sql_db_conn.execute(fetch_str_dtype_stmt).fetchall()
        sample_types = utils.infer_types_from_data(input_data=sample_data, date_format=date_format, datetime_format=datetime_format)
        for col, dtype in zip(str_columns, sample_types):
            if dtype != 'str':
                self.set_column_dtypes(column=col, dtype=dtype)

#############################################################################################################
############################################ class constructors #############################################
#############################################################################################################

    @classmethod
    def from_shape(cls, shape:tuple[int, int], fill:Any=None, headers:list[str]=None, dtype:Literal['bytes','date','datetime','float','int','str']=None, **kwargs) -> SQLDataModel:
        """
        Returns a SQLDataModel from shape ``(N rows, M columns)`` as a convenience method to quickly build a model through an iterative approach. 
        By default, no particular data type is assigned given the flexibility of ``sqlite3``, however one can be inferred by providing an initial ``fill`` value or explicitly by providing the ``dtype`` argument.

        Parameters:
            ``shape`` (tuple[int, int]): The shape to initialize the SQLDataModel with as ``(M, N)`` where ``M`` is the number of rows and ``N`` is the number of columns.
            ``fill`` (Any, optional): The scalar fill value to populate the new SQLDataModel with. Default is None, using SQL null values or deriving from ``dype`` if provided.
            ``headers`` (list[str], optional): The headers to use for the model. Default is None, incrementing headers ``0, 1, ..., N`` where ``N`` is the number of columns.
            ``dtype`` (str, optional): A valid python or SQL datatype to initialize the n-dimensional model with. Default is None, using the SQL text type.
            ``**kwargs``: Additional keyword arguments to pass to the ``SQLDataModel`` constructor.            

        Raises:
            ``TypeError``: If ``M`` or ``N`` are not of type 'int' representing a valid shape to initialize a SQLDataModel with.
            ``ValueError``: If ``M`` or ``N`` are not positive integer values representing valid nonzero row and column dimensions.
            ``ValueError``: If ``dtype`` is not a valid python or SQL convertible datatype to initialize the model with.

        Returns:
            ``SQLDataModel``: Instance with the specified number of rows and columns, initialized with by ``dtype`` fill values or with ``None`` values (default).

        Example::
        
            import sqldatamodel as sdm

            # Create a 3x3 model filled by 'X'
            df = sdm.from_shape((3,3), fill='X')            

            # View it
            print(df)
            
        This will output a 3x3 grid of 'X' characters:

        ```text
            ┌───┬─────┬─────┬─────┐
            │   │ 0   │ 1   │ 2   │
            ├───┼─────┼─────┼─────┤
            │ 0 │ X   │ X   │ X   │
            │ 1 │ X   │ X   │ X   │
            │ 2 │ X   │ X   │ X   │
            └───┴─────┴─────┴─────┘
            [3 rows x 3 columns]
        ```

        We can iteratively build the model from the shape dimensions:

        ```python
            import sqldatamodel as sdm

            # Define shape
            shape = (6,6)
            
            # Initialize the multiplcation table with integer dtypes
            mult_table = sdm.from_shape(shape=shape, dtype='int')

            # Construct the table values
            for x in range(shape[0]):
                for y in range(shape[1]):
                    mult_table[x, y] = x * y
            
            # View the multiplcation table
            print(mult_table)
        ```
        
        This will output our 6x6 multiplication table:

        ```text
            ┌───┬─────┬─────┬─────┬─────┬─────┬─────┐
            │   │   0 │   1 │   2 │   3 │   4 │   5 │
            ├───┼─────┼─────┼─────┼─────┼─────┼─────┤
            │ 0 │   0 │   0 │   0 │   0 │   0 │   0 │
            │ 1 │   0 │   1 │   2 │   3 │   4 │   5 │
            │ 2 │   0 │   2 │   4 │   6 │   8 │  10 │
            │ 3 │   0 │   3 │   6 │   9 │  12 │  15 │
            │ 4 │   0 │   4 │   8 │  12 │  16 │  20 │
            │ 5 │   0 │   5 │  10 │  15 │  20 │  25 │
            └───┴─────┴─────┴─────┴─────┴─────┴─────┘
            [6 rows x 6 columns]
        ```

        Note:
            - If both ``fill`` and ``dtype`` are provided, the data type will be derived from ``type(fill)`` overriding or ignoring the specified ``dtype``.
            - If only ``dtype`` is provided, sensible default initialization fill values will be used to populate the model such as 0 or 0.0 for numeric and empty string or null for others.
            - For those data types not natively implemented by ``sqlite3`` such as ``date`` and ``datetime``, today's date and now's datetime will be used respectively for initialization values.

        Changelog:
            - Version 0.5.2 (2024-05-13):
                - Added ``shape`` parameter in lieu of separate ``n_rows`` and ``n_cols`` arguments.
                - Added ``fill`` parameter to populate resulting SQLDataModel with values to override type-specific initialization defaults.
                - Added ``headers`` parameter to explicitly set column names when creating the SQLDataModel.
                - Added ``**kwargs`` parameter to align more closely with usage patterns of other model initializing constructor methods.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        try:
            n_rows, n_cols = shape
        except ValueError:
            msg = ErrorFormat(f"TypeError: invalid length '{len(shape)}', argument for `shape` must be a tuple or list with 2 elements of type 'int' representing `(n_rows, n_cols)`")
            raise ValueError(msg) from None            
        except TypeError:
            msg = ErrorFormat(f"TypeError: invalid type '{type(shape).__name__}', argument for `shape` must be a tuple or list with 2 elements of type 'int' representing `(n_rows, n_cols)`")
            raise TypeError(msg) from None
        if not isinstance(n_rows, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(n_rows).__name__}', argument of type 'int' expected for `n_rows` parameter")
            raise TypeError(msg)
        if not isinstance(n_cols, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(n_cols).__name__}', argument of type 'int' expected for `n_cols` parameter")
            raise TypeError(msg)
        if fill is None and dtype is not None:
            if dtype not in ('bytes','date','datetime','float','int','str'):
                msg = ErrorFormat(f"ValueError: invalid argument '{dtype}', `dtype` must be one of 'bytes','datetime','float','int','str'")
                raise ValueError(msg)
            else:
                if dtype == 'bytes':
                    fill_value = b''
                elif dtype == 'date':
                    fill_value = datetime.date.today()
                elif dtype == 'datetime':
                    fill_value = datetime.datetime.now()
                elif dtype == 'float':
                    fill_value = 0.0
                elif dtype == 'int':
                    fill_value = 0
                else:
                    fill_value = ''
        else:
            fill_value = fill
        return cls(data=[[fill_value for _ in range(n_cols)] for _ in range(n_rows)], headers=headers, **kwargs)
        
    @classmethod
    def from_csv(cls, csv_source:str, infer_types:bool=True, encoding:str = 'Latin1', delimiter:str = ',', quotechar:str = '"', headers:list[str] = None, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` generated from the provided CSV source, which can be either a file path or a raw delimited string.

        Parameters:
            ``csv_source`` (str): The path to the CSV file or a raw delimited string.
            ``infer_types`` (bool, optional): Infer column types based on random subset of data. Default is True, when False, all columns are str type.
            ``encoding`` (str, optional): The encoding used to decode the CSV source if it is a file. Default is 'Latin1'.
            ``delimiter`` (str, optional): The delimiter to use when parsing CSV source. Default is ``,``.
            ``quotechar`` (str, optional): The character used for quoting fields. Default is ``"``.
            ``headers`` (List[str], optional): List of column headers. If None, the first row of the CSV source is assumed to contain headers.
            ``**kwargs``: Additional keyword arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the provided CSV source.

        Raises:
            ``ValueError``: If no delimited data is found in ``csv_source`` or if parsing with delimiter does not yield valid tabular data.
            ``Exception``: If an error occurs while attempting to read from or process the provided CSV source.

        Examples:

        From CSV File
        -------------

        ```python
            import sqldatamodel as sdm

            # CSV file path or raw CSV string
            csv_source = "/path/to/data.csv"

            # Create the model using the CSV file, providing custom headers
            df = sdm.from_csv(csv_source, headers=['ID', 'Name', 'Value'])
        ```

        From CSV Literal
        ----------------

        ```python
            import sqldatamodel as sdm

            # CSV data
            data = '''
            A, B, C
            1a, 1b, 1c
            2a, 2b, 2c
            3a, 3b, 3c
            '''

            # Create the model
            df = sdm.from_csv(data)

            # View result
            print(df)
        ```

        This will output:

        ```shell
            ┌──────┬──────┬──────┐
            │ A    │ B    │ C    │
            ├──────┼──────┼──────┤
            │ 1a   │ 1b   │ 1c   │
            │ 2a   │ 2b   │ 2c   │
            │ 3a   │ 3b   │ 3c   │
            └──────┴──────┴──────┘
            [3 rows x 3 columns]
        ```

        Note:
            - If ``csv_source`` is delimited by characters other than those specified, use :meth:`SQLDataModel.from_delimited()` and provide delimiter to ``delimiters``.
            - If ``headers`` are provided, the first row parsed from source will be the first row in the table and not discarded.
            - The ``infer_types`` argument can be used to infer the appropriate data type for each column:

              - When ``infer_types = True``, a random subset of the data will be used to infer the correct type and cast values accordingly
              - When ``infer_types = False``, values from the first row only will be used to assign types, almost always 'str' when reading from CSV.

        Changelog:
            - Version 0.4.0 (2024-04-23):
                - Modifed to only parse CSV files and removed all delimiter sniffing with introduction of new method :meth:`SQLDataModel.from_delimited()` to handle other delimiters.
                - Renamed ``delimiters`` parameter to ``delimiter`` with ``,`` set as new default to reflect revised focus on CSV files only.
        """
        if os.path.exists(csv_source):
            try:
                with open(csv_source, encoding=encoding) as csvfile:
                    tmp_all_rows = list(csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar))
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `csv_source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        else:
            csv_source = csv_source.strip()
            try:
                tmp_all_rows = list(csv.reader(csv_source.splitlines(), delimiter=delimiter, quotechar=quotechar))
            except ValueError as e:
                msg = ErrorFormat(f"{e}") # using existing formatting from validation
                raise ValueError(msg) from None
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to parse the provided raw CSV string")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        if not tmp_all_rows:
            msg = ErrorFormat(f"ValueError: no delimited tabular data found in provided `csv_source`, ensure content contains delimited tabular data")
            raise ValueError(msg)
        if headers is None:
            headers = tmp_all_rows.pop(0)
        return cls(data=tmp_all_rows, headers=headers, infer_types=infer_types, **kwargs)
    
    @classmethod
    def from_data(cls, data:Any=None, **kwargs) -> SQLDataModel:
        """
        Convenience method to infer the source of ``data`` and return the appropriate constructor method to generate a new ``SQLDataModel`` instance.

        Parameters:
            ``data`` (Any, required): The input data from which to create the SQLDataModel object. 
            ``**kwargs``: Additional keyword arguments to be passed to the constructor method, see init method for arguments.
            
        Constructor methods are called according to the input type:
            - ``dict``: If all values are python datatypes, passed as ``dtypes`` to constructor, otherwise as ``data`` to :meth:`SQLDataModel.from_dict()`.
            - ``list``: If single dimension, passed as ``headers`` to constructor, otherwise as ``data`` containing list of lists.
            - ``tuple``: Same as with list, if single dimension passed as ``headers``, otherwise as ``data`` containing tuple of lists.
            - ``numpy.ndarray``: passed to :meth:`SQLDataModel.from_numpy()` as array data.
            - ``pandas.DataFrame``: passed to :meth:`SQLDataModel.from_pandas()` as dataframe data.
            - ``polars.DataFrame``: passed to :meth:`SQLDataModel.from_polars()` as dataframe data.
            - ``str``: If starts with 'http', passed to :meth:`SQLDataModel.from_html()` as url, otherwise:
        
              - ``'.csv'``: passed to :meth:`SQLDataModel.from_csv()` as csv source data.
              - ``'.html'``: passed to :meth:`SQLDataModel.from_html()` as html source data.
              - ``'.json'``: passed to :meth:`SQLDataModel.from_json()` as json source data.
              - ``'.md'``: passed to :meth:`SQLDataModel.from_markdown()` as markdown source data.
              - ``'.parquet'``: passed to :meth:`SQLDataModel.from_parquet()` as parquet source data.
              - ``'.pkl'``: passed to :meth:`SQLDataModel.from_pickle()` as pickle source data.
              - ``'.sdm'``: passed to :meth:`SQLDataModel.from_pickle()` as pickle source data.
              - ``'.tex'``: passed to :meth:`SQLDataModel.from_latex()` as latex source data.
              - ``'.tsv'``: passed to :meth:`SQLDataModel.from_csv()` as csv source data.
              - ``'.txt'``: passed to :meth:`SQLDataModel.from_text()` as text source data.
              - ``'.xlsx'``: passed to :meth:`SQLDataModel.from_excel()` as excel source data.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the provided data.

        Raises:
            ``TypeError``: If the type of ``data`` is not supported.
            ``ValueError``: If the file extension is not found, unsupported, or if the SQL extension is not supported.
            ``Exception``: If an OS related error occurs during file read operations if ``data`` is a filepath.

        Example::

            import sqldatamodel as sdm

            # Create SQLDataModel from a CSV file
            df_csv = sdm.from_data("data.csv", headers=['ID', 'Name', 'Value'])

            # Create SQLDataModel from a dictionary
            df_dict = sdm.from_data({"ID": int, "Name": str, "Value": float})

            # Create SQLDataModel from a list of tuples
            df_list = sdm.from_data([(1, 'Alice', 100.0), (2, 'Bob', 200.0)], headers=['ID', 'Name', 'Value'])

            # Create SQLDataModel from raw string literal
            delimited_literal = '''
            A, B, C
            1, 2, 3
            4, 5, 6
            7, 8, 9
            '''

            # Create the model by having correct constructor inferred
            df = sdm.from_data(delimited_literal)

            # View output
            print(df)
        
        This will output:
        
        ```shell            
            ┌────┬────┬────┐
            │ A  │ B  │ C  │
            ├────┼────┼────┤
            │ 1  │ 2  │ 3  │
            │ 4  │ 5  │ 6  │
            │ 7  │ 8  │ 9  │
            └────┴────┴────┘
            [3 rows x 3 columns]
        ```

        Note:
            - This method attempts to infer the correct method to call based on ``data`` argument, if one cannot be inferred an exception is raised.
            - For data type specific implementation or examples, see related method for appropriate data type.
        """
        if not isinstance(data, (list, tuple, str, dict)) and (type(data).__name__ not in ('ndarray','DataFrame')):
            msg = ErrorFormat(f"TypeError: invalid type '{type(data).__name__}', argument for ``data`` must be one of 'list', 'tuple', 'str', 'dict' or a supported external object type")
            raise TypeError(msg)
        supported_ext = ('.csv','.html','.json','.md','.parquet','.pkl','.sdm','.tex','.tsv','.txt','.xlsx')
        ext_operation = {
             '.csv': cls.from_csv
            ,'.html': cls.from_html
            ,'.json': cls.from_json
            ,'.md': cls.from_markdown
            ,'.parquet': cls.from_parquet
            ,'.pkl': cls.from_pickle
            ,'.sdm': cls.from_pickle
            ,'.tex': cls.from_latex
            ,'.tsv': cls.from_delimited
            ,'.txt': cls.from_text
            ,'.xlsx': cls.from_excel
        }
        if isinstance(data, dict):
            if all(value in ('None','int','float','str','bytes','date','datetime','NoneType','bool') for value in data.values()):
                return cls(dtypes=data, **kwargs)
            else:
                return cls.from_dict(data, **kwargs)            
        elif isinstance(data, (list,tuple)):
            if len(data) == 1 and not isinstance(data[0], dict):
                return cls(headers=data, **kwargs)
            elif len(data) >= 1 and isinstance(data[0], dict):
                return cls.from_json(json_source=data, **kwargs)
            else:
                return cls(data=data, **kwargs)
        elif isinstance(data, str):
            if data.startswith('http'):
                return ext_operation['.html'](data, **kwargs)
            if os.path.exists(data):
                ext = os.path.splitext(data)[-1]
                if not ext:
                    msg = ErrorFormat(f"ValueError: file extension not found, files without extensions cannot be parsed without further information")
                    raise ValueError(msg)
                if ext.lower() not in supported_ext:
                    if ext.lower() in ('.db', '.sqlite'):
                        msg = ErrorFormat(f"ValueError: sql extension '{ext}' not supported by `from_data()`, use specialized method `from_sql()` instead")
                        raise ValueError(msg)                        
                    else:
                        msg = ErrorFormat(f"ValueError: unsupported file extension '{ext}', see documentation for a list of supported file types")
                        raise ValueError(msg)
                return ext_operation[ext.lower()](data, **kwargs)
            html_pattern = r'</table>'
            latex_pattern = r'\\begin\{tabular\}'
            markdown_pattern = r'\|?:?-+:?\|:?-+:?\|?' # changed from: r'\| *(:?-{3,}:? *\|)+'
            json_pattern = r'[?\{.*\}]?'
            if bool(re.search(html_pattern, data)):
                return ext_operation['.html'](data, **kwargs)
            elif bool(re.search(latex_pattern, data)):
                return ext_operation['.tex'](data, **kwargs)
            elif bool(re.search(markdown_pattern, data)):
                return ext_operation['.md'](data, **kwargs)
            elif bool(re.search(json_pattern, data)):
                try:
                    json.loads(data)
                    is_json = True
                except Exception:
                    is_json = False
                if is_json:
                    return ext_operation['.json'](data, **kwargs)
            return cls.from_text(data, **kwargs)
        else:
            arg_type = type(data).__name__
            if arg_type == 'ndarray':
                return cls.from_numpy(data, **kwargs)
            elif arg_type == 'DataFrame':
                if 'pandas' in type(data).__module__:
                    return cls.from_pandas(data, **kwargs)
                else:
                    return cls.from_polars(data, **kwargs)
            elif arg_type == 'Table':
                return cls.from_pyarrow(data, **kwargs)
            else:
                msg = ErrorFormat(f"TypeError: unsupported type '{arg_type}', current supported external types are 'numpy.ndarray' or 'pandas.DataFrame' objects")
                raise TypeError(msg)
    
    @classmethod
    def from_delimited(cls, source:str, infer_types:bool=True, encoding:str='Latin1', delimiters:str=', \t;|:', quotechar:str='"', headers:list[str]=None, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` generated from the provided delimited source, which can be either a file path or a raw delimited string.

        Parameters:
            ``source`` (str): The path to the delimited file or a raw delimited string.
            ``infer_types`` (bool, optional): Infer column types based on random subset of data. Default is True, when False, all columns are str type.
            ``encoding`` (str, optional): The encoding used to decode the source if it is a file. Default is ``'Latin1'``.
            ``delimiters`` (str, optional): Possible delimiters. Default is ``\\s``, ``\\t``, ``;``, ``|``, ``:`` or ``,`` (space, tab, semicolon, pipe, colon or comma).
            ``quotechar`` (str, optional): The character used for quoting fields. Default is ``"``.
            ``headers`` (list[str], optional): List of column headers. If None, the first row of the delimited source is assumed to be the header row.
            ``**kwargs``: Additional keyword arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the provided CSV source.

        Raises:
            ``ValueError``: If no delimiter is found in ``source`` or if parsing with delimiter does not yield valid tabular data.
            ``Exception``: If an error occurs while attempting to read from or process the provided CSV source.

        Example:

        From Delimited Literal
        ----------------------

        ```python            
            import sqldatamodel as sdm

            # Space delimited literal
            source_data = '''
            Name Age Height
            Beth 27 172.4
            Kate 28 162.0
            John 30 175.3
            Will 35 185.8'''

            # Create the model
            df = sdm.from_delimited(source_data)

            # View output
            print(df)
        ```

        This will output:

        ```shell
            ┌──────┬─────┬─────────┐
            │ Name │ Age │  Height │
            ├──────┼─────┼─────────┤
            │ Beth │  27 │  172.40 │
            │ Kate │  28 │  162.00 │
            │ John │  30 │  175.30 │
            │ Will │  35 │  185.80 │
            └──────┴─────┴─────────┘
            [4 rows x 3 columns]
        ```

        From Delimited File
        -------------------

        ```python            
            import sqldatamodel as sdm

            # Tab separated file
            tsv_file = 'persons.tsv'

            # Create the model
            df = sdm.from_delimited(tsv_file)
        ```

        Note:
            - Use :meth:`SQLDataModel.from_csv()` if delimiter in source is already known and available as this method requires more compute to determine a plausible delimiter.
            - Use :meth:`SQLDataModel.from_text()` if data is not delimited but is a string representation such as an ASCII table or the output from another ``SQLDataModel`` instance.
            - If file is delimited by delimiters other than the default targets ``\\s``, ``\\t``, ``;``, ``|``, ``:`` or ``,`` (space, tab, semicolon, pipe, colon or comma) make sure they are provided as single character values to ``delimiters``.

        Changelog:
            - Version 0.4.0 (2024-04-23):
                - New method.
        """
        if os.path.exists(source):
            try:
                with open(source, 'r', encoding=encoding) as f:
                    source = f.read()
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None   
        else:
            source = source.strip()
        try:
            dialect = csv.Sniffer().sniff(source, delimiters=delimiters)
            delimiter = dialect.delimiter
            if delimiter is None:
                msg = ErrorFormat(f"ValueError: delimiter not found, ensure `source` contains data delimited by one of ` `, `\\t`, `;`, `|`, `:` or `,` or provide additional delimiters to search for" )
                raise ValueError(msg)
            tmp_all_rows = list(csv.reader(source.splitlines(), delimiter=delimiter, quotechar=quotechar, skipinitialspace=True))
        except ValueError as e:
            msg = ErrorFormat(f"{e}") # using existing formatting from validation
            raise ValueError(msg) from None
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to parse the provided delimited string literal")
            raise type(e)(msg).with_traceback(e.__traceback__) from None
        if not tmp_all_rows:
            msg = ErrorFormat(f"ValueError: no delimited tabular data found in provided `source`, ensure content contains delimited tabular data")
            raise ValueError(msg)
        if headers is None:
            headers = tmp_all_rows.pop(0)
        return cls(data=tmp_all_rows, headers=headers, infer_types=infer_types, **kwargs)          

    @classmethod
    def from_dict(cls, data:dict|list, **kwargs) -> SQLDataModel:
        """
        Create a new ``SQLDataModel`` instance from the provided dictionary.

        Parameters:
            ``data`` (dict): The dictionary or list of dictionaries to convert to SQLDataModel. If keys are of type int, they will be used as row indexes; otherwise, keys will be used as headers.
            ``**kwargs``: Additional arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the provided dictionary.

        Raises:
            ``TypeError``: If the provided dictionary values are not of type 'list', 'tuple', or 'dict'.
            ``ValueError``: If the provided data appears to be a list of dicts but is empty.

        Example::

            import sqldatamodel as sdm
            
            # Sample data with column orientation
            data = {
                'Name': ['Beth', 'John', 'Alice', 'Travis'], 
                'Height': [172.4, 175.3, 162.0, 185.8],
                'Age': [27, 30, 28, 35]
            }

            # Create the model
            df = sdm.from_dict(data)

            # View it
            print(df)
        
        This will output:

        ```shell
            ┌────────┬─────────┬─────┐
            │ Name   │  Height │ Age │
            ├────────┼─────────┼─────┤
            │ Beth   │  172.40 │  27 │
            │ John   │  175.30 │  30 │
            │ Alice  │  162.00 │  28 │
            │ Travis │  185.80 │  35 │
            └────────┴─────────┴─────┘
            [4 rows x 3 columns]
        ```

        We can also create a model using a dictionary with row orientation:
        
        ```python
            import sqldatamodel as sdm

            # Sample data with row orientation
            data = {
                 0: ['Mercury', 0.38]
                ,1: ['Venus', 0.91]
                ,2: ['Earth', 1.00]
                ,3: ['Mars', 0.38]
            }

            # Create the model with custom headers
            df = sdm.from_dict(data, headers=['Planet', 'Gravity'])

            # View output
            print(df)
        ```

        This will output the model created using row-wise dictionary data:
        
        ```shell            
            ┌─────────┬─────────┐
            │ Planet  │ Gravity │
            ├─────────┼─────────┤
            │ Mercury │    0.38 │
            │ Venus   │    0.91 │
            │ Earth   │    1.00 │
            │ Mars    │    0.38 │
            └─────────┴─────────┘
            [4 rows x 2 columns]
        ```

        Note:
            - If data orientation suggests JSON like structure, then :meth:`SQLDataModel.from_json()` will attempt to construct the model.
            - Dictionaries in list like orientation can also be used with structures similar to JSON objects.
            - The method determines the structure of the SQLDataModel based on the format of the provided dictionary.
            - If the keys are integers, they are used as row indexes; otherwise, keys are used as headers.
            - See :meth:`SQLDataModel.to_dict()` for converting existing instances of ``SQLDataModel`` to dictionaries.

        Changelog:
            - Version 0.6.3 (2024-05-16):
                - Modified to try parsing input data as JSON if initial inspection does not signify row or column orientation.
            - Version 0.1.5 (2023-11-24):
                - New method.            
        """
        if isinstance(data, list):
            if len(data) < 1:
                msg = ErrorFormat(f"ValueError: insufficient data length '{len(data)}', if `data` is of type 'list' at least 1 row is required for `from_dict()` method")
                raise ValueError(msg)
            if not isinstance(data[0], dict):
                msg = ErrorFormat(f"TypeError: invalid type in list '{type(data[0].__name__)}', if `data` is of type 'list' its items must be of type 'dict' to use the `from_dict()` method")
                raise TypeError(msg)
            return cls.from_json(data, **kwargs)
        rowwise = True if all(isinstance(x, int) for x in data.keys()) else False
        if rowwise:
            if 'headers' not in kwargs:
                kwargs['headers'] = ['idx',*[f'{i}' for i in range(len(data[next(iter(data))]))]] # get column count from first key value pair in provided dict
            return cls([tuple([k,*v]) for k,v in data.items()], **kwargs)
        else:
            first_key_val = data[next(iter(data))]
            if isinstance(first_key_val, dict):
                headers = list(data.keys())
                data = [[data[col][val] for col in headers] for val in data.keys()]
            elif isinstance(first_key_val, (list,tuple)):
                headers = [k for k in data.keys()]
                column_count = len(headers)
                row_count = len(first_key_val)
                data = [x for x in data.values()]
                data = [tuple([data[j][row] for j in range(column_count)]) for row in range(row_count)]
            else:
                return cls.from_json(data, **kwargs)
            if 'headers' not in kwargs:
                kwargs['headers'] = headers
            return cls(data=data, **kwargs) 

    @classmethod
    def from_excel(cls, filename:str, worksheet:int|str=0, min_row:int|None=None, max_row:int|None=None, min_col:int|None=None, max_col:int|None=None, headers:list[str]=None, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` instance from the specified Excel file.

        Parameters:
            ``filename`` (str): The file path to the Excel file, e.g., ``filename = 'titanic.xlsx'``.
            ``worksheet`` (int | str, optional): The index or name of the worksheet to read from. Defaults to 0, indicating the first worksheet.
            ``min_row`` (int | None, optional): The minimum row number to start reading data from. Defaults to None, indicating the first row.
            ``max_row`` (int | None, optional): Maximum row index (1-based) to import. Defaults to None, indicating all rows are read.
            ``min_col`` (int | None, optional): Minimum column index (1-based) to import. Defaults to None, indicating the first column.
            ``max_col`` (int | None, optional): Maximum column index (1-based) to import. Defaults to None, indicating all the columns are read.
            ``headers`` (List[str], optional): The column headers for the data. Default is None, using the first row of the Excel sheet as headers.
            ``**kwargs``: Additional keyword arguments to pass to the ``SQLDataModel`` constructor.

        Raises:
            ``ModuleNotFoundError``: If the required package ``openpyxl`` is not installed as determined by ``optionals._has_xl`` flag.
            ``TypeError``: If the ``filename`` argument is not of type 'str' representing a valid Excel file path.
            ``Exception``: If an error occurs during Excel read and write operations related to openpyxl processing.

        Returns:
            ``SQLDataModel``: A new instance of ``SQLDataModel`` created from the Excel file.

        Examples:

        We'll use this Excel file, ``data.xlsx``, as the source for the below examples:

        ```text
                ┌───────┬─────┬────────┬───────────┐
                │ A     │ B   │ C      │ D         │
            ┌───┼───────┼─────┼────────┼───────────┤
            │ 1 │ Name  │ Age │ Gender │ City      │
            │ 2 │ John  │ 25  │ Male   │ Boston    │
            │ 3 │ Alice │ 30  │ Female │ Milwaukee │
            │ 4 │ Bob   │ 22  │ Male   │ Chicago   │
            │ 5 │ Sarah │ 35  │ Female │ Houston   │
            │ 6 │ Mike  │ 28  │ Male   │ Atlanta   │
            └───┴───────┴─────┴────────┴───────────┘
            [ Sheet1 ]
        ```
        
        Example 1: Load Excel file with default parameters

        ```python
            import sqldatamodel as sdm

            # Create the model using the default parameters
            df = sdm.from_excel('data.xlsx')

            # View imported data
            print(df)
        ```

        This will output all of the data starting from 'A1':

        ```shell
            ┌───────┬──────┬────────┬───────────┐
            │ Name  │  Age │ Gender │ City      │
            ├───────┼──────┼────────┼───────────┤
            │ John  │   25 │ Male   │ Boston    │
            │ Alice │   30 │ Female │ Milwaukee │
            │ Bob   │   22 │ Male   │ Chicago   │
            │ Sarah │   35 │ Female │ Houston   │
            │ Mike  │   28 │ Male   │ Atlanta   │
            └───────┴──────┴────────┴───────────┘
            [5 rows x 4 columns]
        ```

        Example 2: Load Excel file from specific worksheet

        ```python
            import sqldatamodel as sdm

            # Create the model from 'Sheet2'
            df = sdm.from_excel('data.xlsx', worksheet='Sheet2')

            # View imported data
            print(df)
        ```
        
        This will output the contents of 'Sheet2':

        ```shell
            ┌────────┬───────┐
            │ Gender │ count │
            ├────────┼───────┤
            │ Male   │     3 │
            │ Female │     2 │
            └────────┴───────┘
            [2 rows x 2 columns]
        ```

        Example 3: Load Excel file with custom headers starting from different row

        ```python
            import sqldatamodel as sdm

            # Use our own headers instead of the Excel ones
            new_cols = ['Col A', 'Col B', 'Col C', 'Col D']

            # Create the model starting from the 2nd row to ignore the original headers
            df = sdm.from_excel('data.xlsx', min_row=2, headers=new_cols)

            # View the data
            print(df)
        ```

        This will output the data with our renamed headers:

        ```shell
            ┌───────┬───────┬────────┬───────────┐
            │ Col A │ Col B │ Col C  │ Col D     │
            ├───────┼───────┼────────┼───────────┤
            │ John  │    25 │ Male   │ Boston    │
            │ Alice │    30 │ Female │ Milwaukee │
            │ Bob   │    22 │ Male   │ Chicago   │
            │ Sarah │    35 │ Female │ Houston   │
            │ Mike  │    28 │ Male   │ Atlanta   │
            └───────┴───────┴────────┴───────────┘
            [5 rows x 4 columns]
        ```
        
        Example 4: Load Excel file with specific subset of columns

        ```python
            import sqldatamodel as sdm

            # Create the model using the middle two columns only
            df = sdm.from_excel('data.xlsx', min_col=2, max_col=3)

            # View the data
            print(df)
        ```        

        This will output only the middle two columns:

        ```shell
            ┌──────┬────────┐
            │  Age │ Gender │
            ├──────┼────────┤
            │   25 │ Male   │
            │   30 │ Female │
            │   22 │ Male   │
            │   35 │ Female │
            │   28 │ Male   │
            └──────┴────────┘
            [5 rows x 2 columns]
        ```

        Note:
            - This method entirely relies on ``openpyxl``, see their amazing documentation for further information on Excel file handling in python.
            - If custom ``headers`` are provided using the default ``min_row``, then the original headers, if present, will be duplicated.
            - All indicies for ``min_row``, ``max_row``, ``min_col`` and ``max_col`` are 1-based instead of 0-based, again see ``openpyxl`` for more details.
            - See related :meth:`SQLDataModel.to_excel()` for exporting an existing ``SQLDataModel`` to Excel.

        Changelog:
            - Version 0.2.2 (2024-03-26):
                - New method.
        """
        if not optionals._has_xl:
            msg = ErrorFormat(f"ModuleNotFoundError: required package not found, `openpyxl` must be installed in order to use `from_excel()` method")
            raise ModuleNotFoundError(msg)
        if not isinstance(filename, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid Excel file path")
            raise TypeError(msg)        
        try:
            _xl = optionals._get_xl()
            wb = _xl.load_workbook(filename=filename, read_only=True)
            ws = wb.worksheets[worksheet] if isinstance(worksheet, int) else wb[worksheet]
            data = [row for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)]
            headers = data.pop(0) if headers is None else headers
            wb.close()        
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided Excel file")
            raise type(e)(msg).with_traceback(e.__traceback__) from None
        return cls(data=data, headers=headers, **kwargs)

    @classmethod
    def from_json(cls, json_source:str|list|dict, encoding:str='utf-8', **kwargs) -> SQLDataModel:
        """
        Creates a new ``SQLDataModel`` instance from JSON file path or JSON-like source, flattening if required.

        Parameters:
            ``json_source`` (str | list | dict): The JSON source. If a string, it can represent a file path or a JSON-like object.
            ``encoding`` (str): The encoding to use when reading from a file. Defaults to 'utf-8'.
            ``**kwargs``: Additional keyword arguments to pass to the ``SQLDataModel`` constructor.

        Returns:
            ``SQLDataModel``: A new SQLDataModel instance created from the JSON source.

        Raises:
            ``TypeError``: If the ``json_source`` argument is not of type 'str', 'list', or 'dict'.
            ``OSError``: If related exception occurs when trying to open and read from ``json_source`` as file path.

        Examples:

        From JSON String Literal
        ------------------------

        ```python            
            import sqldatamodel as sdm

            # Sample JSON string
            json_data = '''[{
                "id": 1,
                "color": "red",
                "value": "#f00"
                },
                {   
                "id": 2,
                "color": "green",
                "value": "#0f0"
                },
                {
                "id": 3,
                "color": "blue",
                "value": "#00f"
            }]'''   

            # Create the model
            df = sdm.from_json(json_data)

            # View result
            print(df)
        ```

        This will output:

        ```shell
            ┌──────┬───────┬───────┐
            │   id │ color │ value │
            ├──────┼───────┼───────┤
            │    1 │ red   │ #f00  │
            │    2 │ green │ #0f0  │
            │    3 │ blue  │ #00f  │
            └──────┴───────┴───────┘
            [3 rows x 3 columns]

        ```
        From JSON-like Object
        ---------------------

        ```python            
            import sqldatamodel as sdm

            # JSON-like sample
            json_data = [{
                "alpha": "A",
                "value": "1"
            },  
            {
                "alpha": "B",
                "value": "2"
            },
            {
                "alpha": "C",
                "value": "3"
            }]

            # Create the model
            df = sdm.from_json(json_data)

            # Output
            print(df)
        ```

        This will output:

        ```shell            
            ┌───────┬───────┐
            │ alpha │ value │
            ├───────┼───────┤
            │ A     │ 1     │
            │ B     │ 2     │
            │ C     │ 3     │
            └───────┴───────┘
            [3 rows x 2 columns]

        ```

        From JSON file
        --------------

        ```python
            import sqldatamodel as sdm

            # JSON file path
            json_data = 'data/json-sample.json'

            # Create the model
            df = sdm.from_json(json_data, encoding='latin-1')

            # View output
            print(df)
        ```

        This will output:

        ```shell            
            ┌──────┬────────┬───────┬─────────┐
            │   id │ color  │ value │ notes   │
            ├──────┼────────┼───────┼─────────┤
            │    1 │ red    │ #f00  │ primary │
            │    2 │ green  │ #0f0  │         │
            │    3 │ blue   │ #00f  │ primary │
            │    4 │ cyan   │ #0ff  │         │
            │    5 │ yellow │ #ff0  │         │
            │    5 │ black  │ #000  │         │
            └──────┴────────┴───────┴─────────┘
            [6 rows x 4 columns]
        ```

        Note:
            - If ``json_source`` is deeply-nested it will be flattened according to the staticmethod :meth:`utils.flatten_json()`
            - If ``json_source`` is a JSON-like string object that is not an array, it will be wrapped according as an array.
        
        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(json_source, (str,list,dict)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(json_source).__name__}', expected `json_source` to be one of 'str', 'list' or 'dict' representing a JSON file path or JSON-like object")
            raise TypeError(msg)
        if isinstance(json_source, str):
            if os.path.exists(json_source):
                try:
                    with open(json_source, 'r', encoding=encoding) as f:
                        json_source = f.read()
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `json_source`")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None    
            json_source = json.loads(json_source)
        data_dict = utils.flatten_json(json_source)
        return SQLDataModel.from_dict(data_dict, **kwargs)

    @classmethod
    def from_html(cls, html_source:str, encoding:str='utf-8', table_identifier:int|str=1, infer_types:bool=True, **kwargs) -> SQLDataModel:
        """
        Parses HTML table element from one of three possible sources: web page at url, local file at path, raw HTML string literal.
        If ``table_identifier`` is not specified, the first <table> element successfully parsed is returned, otherwise if ``table_identifier`` is a ``str``, the parser will return the corresponding 'id' or 'name' HTML attribute that matches the identifier specified. 
        If ``table_identifier`` is an ``int``, the parser will return the table matched as a sequential index after parsing all <table> elements from the top of the page down, starting at '1', the first table found. 
        By default, the first <table> element found is returned if ``table_identifier`` is not specified.

        Parameters:
            ``html_source`` (str): The HTML source, which can be a URL, a valid path to an HTML file, or a raw HTML string.
                If starts with 'http', the argument is considered a url and the table will be parsed from returned the web request.
                If is a valid file path, the argument is considered a local file and the table will be parsed from its html.
                If is not a valid url or path, the argument is considered a raw HTML string and the table will be parsed directly from the input.
            ``encoding`` (str): The encoding to use for reading HTML when ``html_source`` is considered a valid url or file path (default is 'utf-8').
            ``table_identifier`` (int | str): An identifier to specify which table to parse if there are multiple tables in the HTML source. Default is 1, returning the first table element found.
            ``infer_types`` (bool, optional): If column data types should be inferred in the return model. Default is True, meaning column types will be inferred otherwise are returned as 'str' types.
                If is ``int``, identifier is treated as the indexed location of the <table> element on the page from top to bottom starting from zero and will return the corresponding position when encountered.
                If is ``str``, identifier is treated as a target HTML 'id' or 'name' attribute to search for and will return the first case-insensitive match when encountered.
            ``**kwargs``: Additional keyword arguments to pass when using ``urllib.request.urlopen`` to fetch HTML from a URL.

        Returns:
            ``SQLDataModel``: A new SQLDataModel instance containing the data from the parsed HTML table.

        Raises:
            ``TypeError``: If ``html_source`` is not of type ``str`` representing a possible url, filepath or raw HTML stream.
            ``HTTPError``: Raised from ``urllib`` when ``html_source`` is considered a url and an HTTP exception occurs.
            ``URLError``: Raised from ``urllib`` when ``html_source`` is considered a url and a URL exception occurs.
            ``ValueError``: If no <table> elements are found or if the targeted ``table_identifier`` is not found.
            ``OSError``: Related exceptions that may be raised when ``html_source`` is considered a file path.

        Examples:

        From Website URL
        ----------------

        ```python            
            import sqldatamodel as sdm

            # From URL
            url = 'https://en.wikipedia.org/wiki/1998_FIFA_World_Cup'
            
            # Lets get the 95th table from the 1998 World Cup
            df = sdm.from_html(url, table_identifier=95)

            # View result:
            print(df)
        ```

        This will output:

        ```shell            
            ┌────┬─────────────┬────┬────┬────┬────┬────┬────┬────┬─────┬──────┐
            │  R │ Team        │ G  │  P │  W │  D │  L │ GF │ GA │ GD  │ Pts. │
            ├────┼─────────────┼────┼────┼────┼────┼────┼────┼────┼─────┼──────┤
            │  1 │ France      │ C  │  7 │  6 │  1 │  0 │ 15 │  2 │ +13 │   19 │
            │  2 │ Brazil      │ A  │  7 │  4 │  1 │  2 │ 14 │ 10 │ +4  │   13 │
            │  3 │ Croatia     │ H  │  7 │  5 │  0 │  2 │ 11 │  5 │ +6  │   15 │
            │  4 │ Netherlands │ E  │  7 │  3 │  3 │  1 │ 13 │  7 │ +6  │   12 │
            │  5 │ Italy       │ B  │  5 │  3 │  2 │  0 │  8 │  3 │ +5  │   11 │
            │  6 │ Argentina   │ H  │  5 │  3 │  1 │  1 │ 10 │  4 │ +6  │   10 │
            │  7 │ Germany     │ F  │  5 │  3 │  1 │  1 │  8 │  6 │ +2  │   10 │
            │  8 │ Denmark     │ C  │  5 │  2 │  1 │  2 │  9 │  7 │ +2  │    7 │
            └────┴─────────────┴────┴────┴────┴────┴────┴────┴────┴─────┴──────┘
            [8 rows x 11 columns]
        ```

        From Local File 
        ---------------

        ```python            
            import sqldatamodel as sdm

            # From HTML file
            df = sdm.from_html('path/to/file.html')

            # View output
            print(df)
        ```

        This will output:

        ```shell            
            ┌─────────────┬────────┬──────┐
            │ Team        │ Points │ Rank │
            ├─────────────┼────────┼──────┤
            │ Brazil      │ 63.7   │ 1    │
            │ England     │ 50.7   │ 2    │
            │ Spain       │ 50.0   │ 3    │
            │ Germany [a] │ 49.3   │ 4    │
            │ Mexico      │ 47.3   │ 5    │
            │ France      │ 46.0   │ 6    │
            │ Italy       │ 44.3   │ 7    │
            │ Argentina   │ 44.0   │ 8    │
            └─────────────┴────────┴──────┘
            [8 rows x 3 columns]
        ```

        From Raw HTML
        -------------

        ```python        
            import sqldatamodel as sdm

            # Raw HTML
            raw_html = 
            '''<table id="find-me">
                <tr>
                    <th>Col 1</th>
                    <th>Col 2</th>
                </tr>    
                <tr>
                    <td>A</td>
                    <td>1</td>
                </tr>
                <tr>
                    <td>B</td>
                    <td>2</td>
                </tr>
                <tr>
                    <td>C</td>
                    <td>3</td>
                </tr>                
            </table>'''

            # Create the model and search for id attribute
            df = sdm.from_html(raw_html, table_identifier="find-me")

            # View output
            print(df)
        ```

        This will output:

        ```shell            
            ┌───┬───────┬───────┐
            │   │ Col 1 │ Col 2 │
            ├───┼───────┼───────┤
            │ 1 │ B     │ 2     │
            │ 2 │ C     │ 3     │
            └───┴───────┴───────┘
            [3 rows x 2 columns]
        ```

        Note:
            - ``**kwargs`` passed to method are used in ``urllib.request.urlopen`` if ``html_source`` is being considered as a web url.
            - ``**kwargs`` passed to method are used in ``open`` if ``html_source`` is being considered as a filepath.
            - The largest row size encountered will be used as the ``column_count`` for the returned ``SQLDataModel``, rows will be padded with ``None`` if less.
            - See :meth:`utils.generate_html_table_chunks()` for initial source chunking before content fed to :mod:`SQLDataModel.HTMLParser`.

        Changelog:
            - Version 0.9.0 (2024-06-26):
                - Modified ``table_identifier`` default value to 1, changing from zero-based to one-based indexing for referencing target table in source to align with similar extraction methods throughout package.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        if not isinstance(html_source, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(html_source).__name__}', argument for `html_source` must be of type 'str' representing a valid website url, HTML filepath or raw HTML string")
            raise TypeError(msg)
        if html_source.startswith("http"):
            try:
                html_source = urllib.request.urlopen(html_source, **kwargs).read().decode(encoding)
            except urllib.error.HTTPError as e:
                msg = ErrorFormat(f"HTTPError: HTTP Error {e.code}: {e.reason}, was encountered when trying to access url '{e.url}'")
                raise urllib.error.HTTPError(e.url, e.code, msg, e.headers, e.fp) from None                
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: encountered '{e}' when trying to request from provided `html_source`, check url parameters")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        elif os.path.exists(html_source):
            try:
                with open(html_source, 'r', encoding=encoding, **kwargs) as f:
                    html_source = f.read()
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `html_source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        tparser = HTMLParser(table_identifier=table_identifier)
        for c in utils.generate_html_table_chunks(html_source):
            if tparser._is_finished:
                break
            tparser.feed(c)
        data, headers = tparser.validate_table()
        tparser.close() 
        headers = list(utils.alias_duplicates(headers)) if headers is not None else headers
        return cls(data=data, headers=headers, infer_types=infer_types)

    @classmethod
    def from_latex(cls, latex_source:str, table_identifier:int=1, encoding:str='utf-8', **kwargs) -> SQLDataModel:
        """
        Creates a new ``SQLDataModel`` instance from the provided LaTeX file or raw literal.
        
        Parameters:
            ``latex_source`` (str): The LaTeX source containing one or more LaTeX tables.
                If ``latex_source`` is a valid system filepath, source will be treated as a ``.tex`` file and parsed.
                If ``latex_source`` is not a valid filepath, source will be parsed as raw LaTeX literal.
            ``table_identifier`` (int, optional): The index position of the LaTeX table to extract. Default is 1.
            ``encoding`` (str, optional): The file encoding to use if source is a LaTex filepath. Default is 'utf-8';.
            ``**kwargs``: Additional keyword arguments to be passed to the ``SQLDataModel`` constructor.

        Returns:
            ``SQLDataModel``: The ``SQLDataModel`` instance created from the parsed LaTeX table.            

        Raises:
            ``TypeError``: If the ``latex_source`` argument is not of type 'str', or if the ``table_identifier`` argument is not of type 'int'.
            ``ValueError``: If the ``table_identifier`` argument is less than 1, or if no tables are found in the LaTeX source.
            ``IndexError``: If the ``table_identifier`` is greater than the number of tables found in the LaTeX source.
        
        Table Indicies:    
            - In the last example, ``sdm`` will contain the data from the second table found in the LaTeX content.
            - Tables are indexed starting from index 1 at the top of the LaTeX content, incremented as they are found.
            - LaTeX parsing stops after the table specified at ``table_identifier`` is found without parsing the remaining content.    

        Examples:

        From LaTeX literal
        ------------------

        ```python            
            import sqldatamodel as sdm

            # Raw LaTeX literal
            latex_content = '''
            \\begin{tabular}{|l|r|r|}
            \\hline
                {Name} & {Age} & {Height} \\\\
            \\hline
                John    &   30 &  175.30 \\\\
                Alice   &   28 &  162.00 \\\\
                Michael &   35 &  185.80 \\\\
            \\hline
            \\end{tabular}
            '''

            # Create the model from the LaTeX
            df = sdm.from_latex(latex_content)

            # View result
            print(df)
        ```

        This will output:

        ```shell            
            ┌─────────┬──────┬─────────┐
            │ Name    │  Age │  Height │
            ├─────────┼──────┼─────────┤
            │ John    │   30 │  175.30 │
            │ Alice   │   28 │  162.00 │
            │ Michael │   35 │  185.80 │
            └─────────┴──────┴─────────┘
            [3 rows x 3 columns]
        ```

        From LaTeX file
        ---------------

        ```python            
            import sqldatamodel as sdm

            # Load LaTeX content from file
            latex_file = 'path/to/latex/file.tex'

            # Create the model using the path
            df = sdm.from_latex(latex_file)
        ```

        Specifying table identifier
        ---------------------------

        ```python            
            import sqldatamodel as sdm

            # Raw LaTeX literal with multiple tables
            latex_content = '''
            %% LaTeX with a Table

            \\begin{tabular}{|l|l|}
            \\hline
                {Header A} & {Header B} \\\\
            \\hline
                Value A1 & Value B1 \\\\
                Value A2 & Value B2 \\\\
            \\hline
            \\end{tabular}

            %% Then another Table

            \\begin{tabular}{|l|l|}
            \\hline
                {Header X} & {Header Y} \\\\
            \\hline
                Value X1 & Value Y1 \\\\
                Value X2 & Value Y2 \\\\
            \\hline
            \\end{tabular}
            '''

            # Create the model from the 2nd table
            df = sdm.from_latex(latex_content, table_identifier=2)

            # View output
            print(df)
        ```

        This will output:

        ```shell
            ┌──────────┬──────────┐
            │ Header X │ Header Y │
            ├──────────┼──────────┤
            │ Value X1 │ Value Y1 │
            │ Value X2 │ Value Y2 │
            └──────────┴──────────┘
            [2 rows x 2 columns]
        ```

        Note:
            - LaTeX tables are identified based on the presence of tabular environments: ``\\begin{tabular}...\\end{tabular}``.
            - The ``table_identifier`` specifies which table to extract when multiple tables are present, beginning at position '1' from the top of the source.
            - The provided ``kwargs`` are passed to the ``SQLDataModel`` constructor for additional parameters to the instance returned.           

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(latex_source, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(latex_source).__name__}', expected `latex_source` to be of type 'str', representing a LaTeX filepath or LaTeX formatted string literal")
            raise TypeError(msg)
        if not isinstance(table_identifier, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(table_identifier).__name__}', expected `table_identifier` to be of type 'int' representing the index position of the LaTeX table")
            raise TypeError(msg) 
        if table_identifier < 1:
            msg = ErrorFormat(f"ValueError: invalid value '{table_identifier}', argument for `table_identifier` must be an integer index for the table beginning at index '1' ")
            raise ValueError(msg)
        if os.path.exists(latex_source):
            try:
                with open(latex_source, 'r', encoding=encoding) as f:
                    latex_source = f.read()
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `latex_source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None          
        tables = re.findall(r'\\begin{tabular}.*?\\end{tabular}', latex_source, re.DOTALL)
        if not tables:
            msg = ErrorFormat("ValueError: no LaTeX tables found in `latex_source`, confirm correct filepath or that provided content contains valid tabular data")
            raise ValueError(msg)
        if table_identifier > len(tables):
            msg = ErrorFormat(f"IndexError: found '{len(tables)}' LaTeX tables in `latex_source`, however none were found at provided `table_identifier` index '{table_identifier}'")
            raise IndexError(msg)
        target_table = tables[table_identifier - 1]
        target_table = target_table.replace(r'\\','')
        table = []
        for line in target_table.split('\n'):
            line = line.strip()
            if line.startswith(r'\hline') or line.startswith(r'\begin') or line.startswith(r'\end'):
                continue
            row = [cell.strip().replace('{','').replace('}','') for cell in line.split('&')]
            table.append(row)
        if len(table) == 1:
            return cls(headers=table[0], **kwargs)
        return cls(data=table[1:], headers=table[0], **kwargs)

    @classmethod
    def from_markdown(cls, markdown_source: str, table_identifier:int=1, **kwargs) -> SQLDataModel:
        """
        Creates a new ``SQLDataModel`` instance from the provided Markdown source file or raw content.
        
        If ``markdown_source`` is a valid system path, the markdown file will be parsed. 
        Otherwise, the provided string will be parsed as raw markdown.

        Parameters:
            ``markdown_source`` (str): The Markdown source file path or raw content.
            ``table_identifier`` (int, optional): The index position of the markdown table to extract. Default is 1.
            ``**kwargs``: Additional keyword arguments to be passed to the ``SQLDataModel`` constructor.

        Raises:
            ``TypeError``: If the ``markdown_source`` argument is not of type 'str', or if the ``table_identifier`` argument is not of type 'int'.
            ``ValueError``: If the ``table_identifier`` argument is less than 1, or if no tables are found in the markdown source.
            ``IndexError``: If the ``table_identifier`` is greater than the number of tables found in the markdown source.
        
        Returns:
            ``SQLDataModel``: The SQLDataModel instance created from the parsed markdown table.

        Table indicies:    
            - In the last example, ``sdm`` will contain the data from the second table found in the markdown content.
            - Tables are indexed starting from index 1 at the top of the markdown content, incremented as they are found.
            - Markdown parsing stops after the table specified at ``table_identifier`` is found without parsing the remaining content.

        Examples:

        From Markdown Literal
        ---------------------

        ```python            
            import sqldatamodel as sdm

            # Raw markdown literal
            markdown_content = '''
            | Item          | Price | # In stock |
            |---------------|-------|------------|
            | Juicy Apples  | 1.99  | 37         |
            | Bananas       | 1.29  | 52         |
            | Pineapple     | 3.15  | 14         |
            '''

            # Create the model from the markdown
            df = sdm.from_markdown(markdown_content)

            # View result
            print(df)
        ```

        This will output:

        ```shell
            ┌──────────────┬───────┬────────────┐
            │ Item         │ Price │ # In stock │
            ├──────────────┼───────┼────────────┤
            │ Juicy Apples │ 1.99  │ 37         │
            │ Bananas      │ 1.29  │ 52         │
            │ Pineapple    │ 3.15  │ 14         │
            └──────────────┴───────┴────────────┘
            [3 rows x 3 columns]
        
        ```

        From Markdown File
        ------------------

        ```python
            import sqldatamodel as sdm

            # Load markdown content from file
            markdown_file_path = 'path/to/markdown_file.md'

            # Create the model using the path
            df = sdm.from_markdown(markdown_file_path)
        ```

        Specifying Table Identifier
        ---------------------------

        ```python            
            import sqldatamodel as sdm

            # Raw markdown literal with multiple tables
            markdown_content = '''
            ### Markdown with a Table

            | Header A | Header B |
            |----------|----------|
            | Value A1 | Value B1 |
            | Value A2 | Value B2 |

            ### Then another Table

            | Header X | Header Y |
            |----------|----------|
            | Value X1 | Value Y1 |
            | Value X2 | Value Y2 |

            '''
            # Create the model from the 2nd table
            df = sdm.from_markdown(markdown_content, table_identifier=2)

            # View output
            print(df)
        ```

        This will output:

        ```shell
            ┌──────────┬──────────┐
            │ Header X │ Header Y │
            ├──────────┼──────────┤
            │ Value X1 │ Value Y1 │
            │ Value X2 │ Value Y2 │
            └──────────┴──────────┘
            [2 rows x 2 columns]
        ```

        Note:
            - Markdown tables are identified based on the presence of pipe characters ``|`` defining table cells.
            - The ``table_identifier`` specifies which table to extract when multiple tables are present, beginning at position '1' from the top of the source.
            - Escaped pipe characters ``\\|`` within the markdown are replaced with the HTML entity reference ``&vert;`` for proper parsing.
            - The provided ``kwargs`` are passed to the ``SQLDataModel`` constructor for additional parameters to the instance returned.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(markdown_source, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(markdown_source).__name__}', expected `markdown_source` to be of type 'str', representing a markdown file path or markdown string literal")
            raise TypeError(msg)
        if not isinstance(table_identifier, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(table_identifier).__name__}', expected `table_identifier` to be of type 'int' representing the index position of the markdown table")
            raise TypeError(msg) 
        if table_identifier < 1:
            msg = ErrorFormat(f"ValueError: invalid value '{table_identifier}', argument for `table_identifier` must be an integer index for the table beginning at index '1' ")
            raise ValueError(msg)
        if os.path.exists(markdown_source):
            try:
                with open(markdown_source, 'r') as f:
                    markdown_source = f.read()
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `markdown_source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None  
        table = None
        in_table = False
        prev_line = None
        found_table = False
        tables_found = 0
        table_column_count = -1
        md_pattern = r'\|?:?-+:?\|:?-+:?\|?'
        markdown_source = markdown_source.replace('\\|','&vert;') # replace escaped pipes with wrapped unicode representation
        for md_line in markdown_source.splitlines():
            if in_table:
                row = [cell.strip() for cell in md_line.strip().strip('|').split('|')]
                if len(row) == table_column_count:
                    table.append(row)
                else:
                    tables_found += 1
                    if tables_found == table_identifier:
                        found_table = True
                        break
                    table = None
                    in_table = False
            if not in_table:
                if re.search(md_pattern,md_line.replace(' ','')):
                    in_table = True
                    table_column_count = len(md_line.strip().strip('|').split('|'))
                    headers = [cell.strip() for cell in prev_line.strip().strip('|').split('|')]
                    table = [headers]
            prev_line = md_line
        # Check if last table ended on last line of content 
        if in_table:
            tables_found += 1
            if tables_found == table_identifier:
                found_table = True            
        if not found_table:
            if (tables_found > 0) and (table_identifier > tables_found):
                msg = ErrorFormat(f"IndexError: found '{tables_found}' tables in `markdown_source` at positions '1..{tables_found}', however none were found at provided `table_identifier` index '{table_identifier}'")
                raise IndexError(msg)
            else:
                msg = ErrorFormat("ValueError: no tables found in `markdown_source`, confirm provided target is a valid markdown file or literal with table elements")
                raise ValueError(msg)
        if len(table) == 1:
            return cls(headers=table[0], **kwargs)
        return cls(data=table[1:],headers=table[0], **kwargs)

    @classmethod
    def from_numpy(cls, array, headers:list[str]=None, **kwargs) -> SQLDataModel:
        """
        Returns a ``SQLDataModel`` object created from the provided numpy ``array``.

        Parameters:
            ``array`` (numpy.ndarray): The numpy array to convert to a SQLDataModel.
            ``headers`` (list of str, optional): The list of headers to use for the SQLDataModel. If None, no headers will be used, and the data will be treated as an n-dimensional array. Default is None.
            ``**kwargs``: Additional arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the numpy array.

        Raises:
            ``ModuleNotFoundError``: If the required package ``numpy`` is not found.
            ``TypeError``: If ``array`` argument is not of type ``numpy.ndarray``.
            ``DimensionError``: If ``array.ndim != 2`` representing a `(row, column)` tabular array.

        Example::

            import numpy as np
            import sqldatamodel as sdm

            # Sample array
            arr = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]])

            # Create the model with custom headers
            df = sdm.from_numpy(arr, headers=['Col A', 'Col B', 'Col C])

            # View output
            print(df)
        
        This will output:

        ```shell            
            ┌───────┬───────┬───────┐
            │ Col A │ Col B │ Col C │
            ├───────┼───────┼───────┤
            │     1 │     2 │     3 │
            │     4 │     5 │     6 │
            │     7 │     8 │     9 │
            └───────┴───────┴───────┘
            [3 rows x 3 columns]
        ```

        Note:
            - Numpy array must have '2' dimensions, the first representing the rows, and the second the columns.
            - If no headers are provided, default headers will be generated as 'col_N' where N represents the column integer index.

        Changelog:
            - Version 0.1.3 (2023-10-15):
                - New method.
        """
        if not optionals._has_np:
            msg = ErrorFormat(f"""ModuleNotFoundError: required package not found, numpy must be installed in order to use the `from_numpy()` method""")
            raise ModuleNotFoundError(msg)
        if (obj_type := type(array).__name__) != 'ndarray':
            msg = ErrorFormat(f"TypeError: invalid type '{obj_type}', argument for `array` must be of type 'ndarray'")
            raise TypeError(msg)
        if (obj_ndim := array.ndim) != 2:
            msg = ErrorFormat(f"DimensionError: invalid dimension '{obj_ndim}', argument for `array` must have 2 dimensions representing `(rows, columns)`")
            raise DimensionError(msg)
        return cls(data=array.tolist(),headers=headers, **kwargs)

    @classmethod
    def from_pandas(cls, df, headers:list[str]=None, **kwargs) -> SQLDataModel:
        """
        Returns a ``SQLDataModel`` object created from the provided ``df`` representing a Pandas ``DataFrame`` object. Note that ``pandas`` must be installed in order to use this method.

        Parameters:
            ``df`` (pandas.DataFrame): The pandas DataFrame to convert to a SQLDataModel.
            ``headers`` (list[str], optional): The list of headers to use for the SQLDataModel. Default is None, using the columns from the ``df`` object.
            ``**kwargs``: Additional arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the pandas DataFrame.

        Raises:
            ``ModuleNotFoundError``: If the required package ``pandas`` is not found.
            ``TypeError``: If ``df`` argument is not of type ``pandas.DataFrame``.

        Example::

            import pandas as pd
            import sqldatamodel as sdm

            # Create a pandas DataFrame
            df_pd = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
            
            # Create the model
            df_sdm = sdm.from_pandas(df_pd)

        Note:
            - If ``headers`` are not provided, the existing pandas columns will be used as the new ``SQLDataModel`` headers.

        Changelog:
            - Version 0.1.3 (2023-10-15):
                - New method.
        """
        if not optionals._has_pd:
            msg = ErrorFormat(f"""ModuleNotFoundError: required package not found, Pandas must be installed in order to use the `from_pandas()` method""")
            raise ModuleNotFoundError(msg)
        if (obj_type := type(df).__name__) != 'DataFrame':
            msg = ErrorFormat(f"TypeError: invalid type '{obj_type}', argument for `df` must be of type 'DataFrame'")
            raise TypeError(msg)        
        data = [x[1:] for x in df.itertuples()]
        headers = df.columns.tolist() if headers is None else headers
        return cls(data=data,headers=headers, **kwargs)

    @classmethod
    def from_parquet(cls, filename:str, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` instance from the specified parquet file.

        Parameters:
            ``filename`` (str): The file path to the parquet file, e.g., ``filename = 'user/data/titanic.parquet'``.
            ``**kwargs``: Additional keyword arguments to pass to the pyarrow ``read_table`` function, e.g., ``filters = [('Name','=','Alice')]``.

        Returns:
            ``SQLDataModel``: A new instance of ``SQLDataModel`` created from the parquet file.

        Raises:
            ``ModuleNotFoundError``: If the required package ``pyarrow`` is not installed as determined by ``optionals._has_pa`` flag.
            ``TypeError``: If the ``filename`` argument is not of type 'str' representing a valid parquet file path.
            ``FileNotFoundError``: If the specified parquet ``filename`` is not found.
            ``Exception``: If any unexpected exception occurs during the file or parquet reading process.

        Example::

            import sqldatamodel as sdm

            # Sample parquet file
            pq_file = "titanic.parquet"

            # Create the model
            df = sdm.from_parquet(pq_file)

            # View column counts
            print(df.count())

        This will output:

        ```shell            
            ┌────┬─────────────┬──────┬────────┬───────┬───────┐
            │    │ column      │   na │ unique │ count │ total │
            ├────┼─────────────┼──────┼────────┼───────┼───────┤
            │  0 │ PassengerId │    0 │    891 │   891 │   891 │
            │  1 │ Survived    │    0 │      2 │   891 │   891 │
            │  2 │ Pclass      │    0 │      3 │   891 │   891 │
            │  3 │ Name        │    0 │    891 │   891 │   891 │
            │  4 │ Sex         │    0 │      2 │   891 │   891 │
            │  5 │ Age         │  177 │     88 │   714 │   891 │
            │  6 │ SibSp       │    0 │      7 │   891 │   891 │
            │  7 │ Parch       │    0 │      7 │   891 │   891 │
            │  8 │ Ticket      │    0 │    681 │   891 │   891 │
            │  9 │ Fare        │    0 │    248 │   891 │   891 │
            │ 10 │ Cabin       │  687 │    147 │   204 │   891 │
            │ 11 │ Embarked    │    2 │      3 │   889 │   891 │
            └────┴─────────────┴──────┴────────┴───────┴───────┘
            [12 rows x 5 columns]
        ```

        Note:
            - The pyarrow package is required to use this method as well as the :meth:`SQLDataModel.to_parquet()` method.
            - Once the file is read into pyarrow.parquet, the ``to_pydict()`` method is used to pass the data to this package's :meth:`SQLDataModel.from_dict()` method.
            - Titanic parquet data used in example available at https://www.kaggle.com/code/taruntiwarihp/titanic-dataset
        """
        if not optionals._has_pa:
            msg = ErrorFormat(f"ModuleNotFoundError: required package not found, pyarrow must be installed in order to use `.from_parquet()` method")
            raise ModuleNotFoundError(msg)
        _pa, _pq = optionals._get_pa_pq()
        if not isinstance(filename, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid parquet file path")
            raise TypeError(msg)
        try:
            pq_array = _pq.read_table(filename, **kwargs)
        except FileNotFoundError:
            msg = ErrorFormat(f"FileNotFoundError: file not found '{filename}' encountered when trying to open and read from parquet")
            raise FileNotFoundError (msg) from None
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from parquet")
            raise type(e)(msg).with_traceback(e.__traceback__) from None
        return SQLDataModel.from_dict(pq_array.to_pydict())

    @classmethod
    def from_pickle(cls, filename:str=None, **kwargs) -> SQLDataModel:
        """
        Returns the ``SQLDataModel`` object from the provided ``filename``. If ``None``, the current directory will be scanned for the default :meth:`SQLDataModel.to_pickle()` format.

        Parameters:
            ``filename`` (str, optional): The name of the pickle file to load. If None, the current directory will be scanned for the default filename. Default is None.
            ``**kwargs``: Additional arguments to be passed to the SQLDataModel constructor, these will override the properties loaded from ``filename``.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the loaded pickle file.

        Raises:
            ``TypeError``: If filename is provided but is not of type 'str' representing a valid pickle filepath.
            ``FileNotFoundError``: If the provided filename could not be found or does not exist.
        
        Example::

            import sqldatamodel as sdm

            headers = ['Name','Age','Sex']
            data = [('Alice', 20, 'F'), ('Bob', 25, 'M'), ('Gerald', 30, 'M')]

            # Create the model with sample data
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Filepath
            pkl_file = 'people.sdm'

            # Save the model
            df.to_pickle(filename=pkl_file)

            # Load it back from file
            df = sdm.from_pickle(filename=pkl_file)

        Note:
            - All data, headers, data types and display properties will be saved when pickling.
            - Any additional ``kwargs`` provided will override those saved in the pickled model.
        """
        if filename is None:
            filename = os.path.basename(sys.argv[0]).split(".")[0]+'.sdm'
        else:
            if not isinstance(filename, str):
                msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid pickle filepath")
                raise TypeError(msg)
        try:
            with open(filename, 'rb') as f:
                sdm_deserialized = pickle.load(f)
        except FileNotFoundError:
            msg = ErrorFormat(f"FileNotFoundError: no such file or directory '{filename}', please ensure the filename exists and is a valid path")
            raise FileNotFoundError(msg) from None            
        if kwargs:
            sdm_deserialized.update(**kwargs)
        return cls(**sdm_deserialized)
 
    @classmethod
    def from_polars(cls, df, headers:list[str]=None, **kwargs) -> SQLDataModel:
        """
        Returns a ``SQLDataModel`` object created from the provided ``df`` representing a Polars ``DataFrame`` object. Note that ``polars`` must be installed in order to use this method.

        Parameters:
            ``df`` (polars.DataFrame): The Polars DataFrame to convert to a SQLDataModel.
            ``headers`` (list[str], optional): The list of headers to use for the SQLDataModel. Default is None, using the columns from the ``df`` object.
            ``**kwargs``: Additional arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the Polars DataFrame.

        Raises:
            ``ModuleNotFoundError``: If the required package ``polars`` is not found.
            ``TypeError``: If ``df`` argument is not of type ``polars.DataFrame``.

        Example::

            import polars as pl
            import sqldatamodel as sdm

            # Sample data
            data = {
                'Name': ['Beth', 'John', 'Alice', 'Travis'], 
                'Age': [27, 30, 28, 35], 
                'Height': [172.4, 175.3, 162.0, 185.8]
            }

            # Create the polars DataFrame
            df_pl = pl.DataFrame(data)
            
            # Create a SQLDataModel object
            df_sdm = sdm.from_polars(df_pl)
        
            # View result
            print(df_sdm)

        This will output a ``SQLDataModel`` constructed from the Polars ``df_pl``:

        ```shell
            ┌────────┬─────┬─────────┐
            │ Name   │ Age │  Height │
            ├────────┼─────┼─────────┤
            │ Beth   │  27 │  172.40 │
            │ John   │  30 │  175.30 │
            │ Alice  │  28 │  162.00 │
            │ Travis │  35 │  185.80 │
            └────────┴─────┴─────────┘
            [4 rows x 3 columns]
        ```

        Note:
            - If ``headers`` are not provided, the columns from the provided DataFrame's columns will be used as the new ``SQLDataModel`` headers.
            - Polars uses different data types than those used by ``SQLDataModel``, see :meth:`SQLDataModel.set_column_dtypes()` for specific casting rules.
            - See related :meth:`SQLDataModel.to_polars()` for the inverse method of converting a ``SQLDataModel`` into a Polars ``DataFrame`` object.

        Changelog:
            - Version 0.3.8 (2024-04-12):
                - New method.
        """
        if not optionals._has_pl:
            msg = ErrorFormat(f"""ModuleNotFoundError: required package not found, Polars must be installed in order to use the `from_polars()` method""")
            raise ModuleNotFoundError(msg)
        if (obj_type := type(df).__name__) != 'DataFrame':
            msg = ErrorFormat(f"TypeError: invalid type '{obj_type}', argument for `df` must be of type 'DataFrame'")
            raise TypeError(msg)        
        return cls(data=df.rows(), headers=df.columns if headers is None else headers, **kwargs)

    @classmethod
    def from_pyarrow(cls, table, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` instance from the provided Apache Arrow object.

        Parameters:
            ``table`` (pyarrow.lib.Table): Apache Arrow object from which to construct a new ``SQLDataModel`` object.
            ``**kwargs``: Additional keyword arguments to pass to the SQLDataModel constructor.

        Raises:
            ``ModuleNotFoundError``: If the required package ``pyarrow`` is not installed.
            ``TypeError``: If the provided `table` argument is not of type 'pyarrow.lib.Table'.

        Returns:
            ``SQLDataModel``: A new SQLDataModel instance representing the data in the provided Apache Arrow object.

        Example::

            import pyarrow as pa
            import sqldatamodel as sdm

            # Sample data
            data = {
                'Name': ['Alice', 'Bob', 'Charlie'],
                'Age': [25, 30, 35],
                'Grade': [3.8, 3.9, 3.2],
            }

            # Create PyArrow table from data
            table = pa.Table.from_pydict(data)

            # Create model from PyArrow table
            df = sdm.from_pyarrow(table)

        This will output:

        ```shell
            ┌─────────┬──────┬───────┐
            │ Name    │  Age │ Grade │
            ├─────────┼──────┼───────┤
            │ Alice   │   25 │  3.80 │
            │ Bob     │   30 │  3.90 │
            │ Charlie │   35 │  3.20 │
            └─────────┴──────┴───────┘
            [3 rows x 3 columns]
        ```            

        Note:
            - To convert an existing ``SQLDataModel`` instance to Apache Arrow format, see :meth:`SQLDataModel.to_pyarrow()`.
            - This method is only for in-memory Apache Arrow table objects, for reading and writing parquet see :meth:`SQLDataModel.from_parquet()`.

        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.2.3 (2024-03-28):
                - New method.            
        """
        if not optionals._has_pa:
            msg = ErrorFormat(f"ModuleNotFoundError: required package not found, pyarrow must be installed in order to use `.from_pyarrow()` method")
            raise ModuleNotFoundError(msg)
        _pa, _pq = optionals._get_pa_pq()            
        if not isinstance(table,_pa.lib.Table):
            msg = ErrorFormat(f"TypeError: invalid type '{type(table).__name__}', argument for `table` must point to an Apache Arrow object of type 'pyarrow.lib.Table'")
            raise TypeError(msg)            
        return cls.from_dict(table.to_pydict(), **kwargs)  

    @classmethod
    def from_sql(cls, sql: str, con: sqlite3.Connection|Any, dtypes:dict=None, **kwargs) -> SQLDataModel:
        """
        Create a ``SQLDataModel`` object by executing the provided SQL query using the specified SQL connection.
        If a single word is provided as the ``sql``, the method wraps it and executes a select all treating the text as the target table.
        
        Supported Connection APIs:
            - SQLite using ``sqlite3`` or url with format ``'file:///path/to/database.db'``
            - PostgreSQL using ``psycopg2`` or url with format ``'postgresql://user:pass@hostname:port/db'``
            - SQL Server ODBC using ``pyodbc`` or url with format ``'mssql://user:pass@hostname:port/db'``
            - Oracle using ``cx_Oracle`` or url with format ``'oracle://user:pass@hostname:port/db'``
            - Teradata using ``teradatasql`` or url with format ``'teradata://user:pass@hostname:port/db'``

        Parameters:
            ``sql`` (str): The SQL query to execute and use to create the SQLDataModel.
            ``con`` (sqlite3.Connection | Any): The database connection object or url, supported connection APIs are ``sqlite3``, ``psycopg2``, ``pyodbc``, ``cx_Oracle``, ``teradatasql``.
            ``dtypes`` (dict, optional): A dictionary of the format ``'column': 'python dtype'`` to assign to values. Default is None, mapping types from source connection.
            ``**kwargs``: Additional arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the executed SQL query.

        Raises:
            ``TypeError``: If dtypes argument is provided and is not of type ``dict`` representing python data types to assign to values.
            ``SQLProgrammingError``: If the provided SQL connection is not opened or valid, or the SQL query is invalid or malformed.
            ``ModuleNotFoundError``: If ``con`` is provided as a connection url and the specified scheme driver module is not found.
            ``DimensionError``: If the provided SQL query returns no data.

        Examples:

        From SQL Table
        --------------
        
        ```python
            import sqldatamodel as sdm

            # Single word parameter
            df = sdm.from_sql("table_name", sqlite3.Connection)
            
            # Equilavent query executed
            df = sdm.from_sql("select * from table_name", sqlite3.Connection)
        ```

        From SQLite Database
        --------------------

        ```python            
            import sqlite3
            import sqldatamodel as sdm

            # Create connection object
            sqlite_db_conn = sqlite3.connect('./database/users.db')

            # Basic usage with a select query
            df = sdm.from_sql("SELECT * FROM my_table", sqlite_db_conn)

            # When a single word is provided, it is treated as a table name for a select all query
            df_table = df.from_sql("my_table", sqlite_db_conn)
        ```
            
        From PostgreSQL Database
        ------------------------

        ```python
            import psycopg2
            import sqldatamodel as sdm

            # Create connection object
            pg_db_conn = psycopg2.connect('dbname=users user=postgres password=postgres')
            
            # Basic usage with a select query
            df = sdm.from_sql("SELECT * FROM my_table", pg_db_conn)

            # When a single word is provided, it is treated as a table name for a select all query
            df_table = df.from_sql("my_table", pg_db_conn)
        ```

        From SQL Server Databse
        -----------------------

        ```python
            import pyodbc
            import sqldatamodel as sdm

            # Create connection object
            con = pyodbc.connect("DRIVER={SQL Server};SERVER=host;DATABASE=db;UID=user;PWD=pw;")
            
            # Basic usage with a select query
            df = sdm.from_sql("SELECT * FROM my_table", con)

            # When a single word is provided, it is treated as a table name for a select all query
            df_table = df.from_sql("my_table", con)
        ```

        Note:
            - When ``con`` is provided as a string a connection will be attempted using :meth:`utils._create_connection()` if the path does not exist, otherwise a ``sqlite3`` local connection will be attempted.
            - When ``con`` is provided as an object a connection is assumed to be open and valid, if a cursor cannot be created from the object an exception will be raised. 
            - Unsupported connection object will output a ``SQLDataModelWarning`` advising unstable or undefined behaviour.
            - The ``dtypes``, if provided, are only applied to ``sqlite3`` connection objects as remaining supported connections implement SQL to python adapters.
            - See related :meth:`SQLDataModel.to_sql()` for writing to SQL database connections.
            - See utility methods :meth:`utils._parse_connection_url()` and :meth:`utils._create_connection()` for implementation on creating database connections from urls.                      

        Changelog:
            - Version 0.9.1 (2024-06-27):
                - Modified handling of ``con`` parameter to allow database connection url to also be provided as ``'scheme://user:pass@host:port/db'``
            - Version 0.8.2 (2024-06-24):
                - Modified handling of ``con`` parameter to allow providing SQLite database filepath directly as string to instantiate connection.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``sql_query`` parameter to ``sql`` for consistency with similar method arguments.  
        """
        if isinstance(con, str):
            if os.path.exists(con): # Connection provided as SQLite database filepath
                try:
                    con = sqlite3.connect(con)
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open database connection '{con}'")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None     
            else: # Connection provided as url with format 'scheme://user:pass@host:port/path'
                con = utils._create_connection(con)        
        if dtypes is not None and not isinstance(dtypes, dict):
            msg = ErrorFormat(f"TypeError: invalid type '{type(dtypes).__name__}', argument for ``dtypes`` must be of type 'dict' representing 'column': 'python dtype' values to assign model")
            raise TypeError(msg)
        db_dialect = type(con).__module__.split('.')[0].lower()
        if db_dialect not in cls.get_supported_sql_connections():
            print(WarnFormat(f"""SQLDataModelWarning: provided SQL connection has not been tested, behavior for "{db_dialect}" may be unpredictable or unstable"""))
        if len(sql.split()) == 1:
            sql = f""" select * from "{sql}" """
        try:
            sql_c = con.cursor()
        except Exception as e:
            msg = ErrorFormat(f"SQLProgrammingError: provided SQL connection is not opened or valid, failed with: '{e}'")
            raise SQLProgrammingError(msg) from None
        if db_dialect == 'sqlite3' and dtypes is None:
            try:
                table_name = sql.lower().split("from",1)[-1].split()[0].replace('"','')
                sql_c.execute(f"""
                select "name" as "column_name"
                ,case upper(substr("type",1,3)) 
                    when 'TEX' then 'str'
                    when 'TIM' then 'datetime'
                    when 'REA' then 'float'
                    when 'INT' then 'int'
                    when 'DAT' then 'date'
                    when 'BLO' then 'bytes'
                    else 'str' end as "column_dtype"
                from pragma_table_info("{table_name}") """)
                dtypes = {res[0]: res[1] for res in sql_c.fetchall()}
            except:
                dtypes = None
        try:
            sql_c.execute(sql)
        except Exception as e:
            msg = ErrorFormat(f"SQLProgrammingError: provided SQL query is invalid or malformed, failed with: '{e}'")
            raise SQLProgrammingError(msg) from None
        data = sql_c.fetchall()
        if (len(data) < 1) or (data is None):
            msg = ErrorFormat("DimensionError: provided SQL query returned no data, please provide a valid query with sufficient data to construct a model")
            raise DimensionError(msg)
        headers = [x[0] for x in sql_c.description]
        return cls(data=data, headers=headers, dtypes=dtypes, **kwargs)

    @classmethod
    def from_text(cls, text_source:str, table_identifier:int=1, encoding:str='utf-8', headers:list[str]=None, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` generated from the provided ``text_source``, either as a file if the path exists, or from a raw string literal if the path does not exist.

        Parameters:
            ``text_source`` (str): The path to the tabular data file or a raw string literal containing tabular data.
            ``table_identifier`` (int, optional): The index position of the target table within the text source. Default is 1.
            ``encoding`` (str, optional): The encoding used to decode the text source if it is a file. Default is 'utf-8'.
            ``headers`` (list, optional): The headers to use for the provided data. Default is to use the first row.
            ``**kwargs``: Additional keyword arguments to be passed to the SQLDataModel constructor.

        Returns:
            ``SQLDataModel``: The SQLDataModel object created from the provided tabular data.

        Raises:
            ``TypeError``: If ``text_source`` is not a string or ``table_identifier`` is not an integer.
            ``ValueError``: If no tabular data is found in ``text_source``, if parsing fails to extract valid tabular data, or if the provided ``table_identifier`` is out of range.
            ``IndexError``: If the provided ``table_identifier`` exceeds the number of tables found in ``text_source``.
            ``Exception``: If an error occurs while attempting to read from or process the provided ``text_source``.

        Example::

            import sqldatamodel as sdm

            # Text source containing tabular data
            text_source = "/path/to/tabular_data.txt"

            # Create the model using the text source
            df = sdm.from_text(text_source, table_identifier=2)

        Note:
            - This method is made for parsing ``SQLDataModel`` formatted text, such as the kind generated with ``print(df)`` or the output created by the inverse method :meth:`SQLDataModel.to_text()`
            - For parsing other delimited tabular data, this method calls the related :meth:`SQLDataModel.from_csv()` method, which parses tabular data constructed with common delimiters.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(text_source, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(text_source).__name__}', expected `text_source` to be of type 'str', representing a tabular data filepath or tabular string literal")
            raise TypeError(msg)
        if not isinstance(table_identifier, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(table_identifier).__name__}', expected `table_identifier` to be of type 'int' representing the index position of the target table")
            raise TypeError(msg) 
        if table_identifier < 1:
            msg = ErrorFormat(f"ValueError: invalid value '{table_identifier}', argument for `table_identifier` must be an integer index for the table beginning at index '1' ")
            raise ValueError(msg)
        if os.path.exists(text_source):
            try:
                with open(text_source, 'r', encoding=encoding) as f:
                    text_source = f.read()
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and read from provided `text_source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None                  
        tables = re.findall(r'┌.*?┘', text_source, re.DOTALL)
        if not tables:
            try:
                return SQLDataModel.from_delimited(text_source, headers=headers, **kwargs)
            except Exception:
                msg = ErrorFormat("ValueError: no tabular data found in `text_source`, confirm correct filepath or that provided content contains valid tabular data")
                raise ValueError(msg)
        if table_identifier > len(tables):
            msg = ErrorFormat(f"IndexError: found '{len(tables)}' tables in `text_source`, however none were found at provided `table_identifier` index '{table_identifier}'")
            raise IndexError(msg)
        target_table = tables[table_identifier - 1]
        if '│' in target_table:
            delimiter = '│'
        else:
            try:
                delimiter = csv.Sniffer().sniff(text_source).delimiter
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to parse delimiter from `text_source`")
                raise type(e)(msg).with_traceback(e.__traceback__) from None 
        table = []
        for row in target_table.strip().split('\n'):
            if delimiter in row:
                row = [cell.strip() for cell in row.strip().strip(delimiter).strip().split(delimiter)]
                table.append(row)
        if len(table) < 1:
            msg = ErrorFormat(f"ValueError: failed to parse tabular data from provided `text_source`, confirm target contains valid delimiters in a tabular format")
            raise ValueError(msg)
        if not table[0][0]: # does not have sdm index, keep all elements
            table = [x[1:] for x in table]
        if len(table) == 1:
            return cls(headers=table[0], **kwargs)
        if headers is None:
            headers = table.pop(0)        
        return cls(data=table, headers=headers, **kwargs)

    @classmethod
    def get_supported_sql_connections(cls) -> tuple:
        """
        Returns the currently tested DB API 2.0 dialects for use with :meth:`SQLDataModel.from_sql()` method.

        Returns:
            ``tuple``: A tuple of supported DB API 2.0 dialects.

        Example::
        
            import sqldatamodel as sdm

            # Get supported dialects
            supported_dialects = sdmSQLDataModel.get_supported_sql_connections()

            # View details
            print(supported_dialects)

            # Outputs
            supported_dialects = ('sqlite3', 'psycopg2', 'pyodbc', 'cx_oracle', 'teradatasql')
        
        Changelog:
            - Version 0.1.5 (2023-11-24):
                - New method.            
        """
        return ('sqlite3', 'psycopg2', 'pyodbc', 'cx_oracle', 'teradatasql')
   
################################################################################################################
############################################## conversion methods ##############################################
################################################################################################################

    def data(self, index:bool=False, include_headers:bool=False, strict_2d:bool=False) -> list[tuple]:
        """
        Returns the ``SQLDataModel`` data as a list of tuples for multiple rows, a single tuple for individual rows, as a single item for individual cells. 
        Data is returned without index and headers by default, use ``include_headers=True`` or ``index=True`` to modify.

        Parameters:
            ``index`` (bool, optional): If True, includes the index in the result; if False, excludes the index. Default is False.
            ``include_headers`` (bool, optional): If True, includes column headers in the result; if False, excludes headers. Default is False.
            ``strict_2d`` (bool, optional): If True, returns data as a 2-dimensional list of tuples regardless of data dimension. Default is False.

        Returns:
            ``list[tuple]``: The data currently stored in the model as a list of tuples.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers, display_float_precision=2)

            # View full table
            print(df)
        
        This will output:

        ```shell
            ┌────────┬──────┬─────────┐
            │ Name   │  Age │  Height │
            ├────────┼──────┼─────────┤
            │ John   │   30 │  175.30 │
            │ Alice  │   28 │  162.00 │
            │ Travis │   35 │  185.80 │
            └────────┴──────┴─────────┘
            [3 rows x 3 columns]
        ```

        Get data for specific row:

        ```python
            # Grab data from single row
            row_data = df[0].data()

            # View it
            print(row_data)
        ```

        This will output the row as a tuple of values:

        ```text
            ('John', 30, 175.3)
        ```

        Get data for specific column:

        ```python
            # Grab data from single column
            col_data = df['Name'].data()
            
            # View it
            print(col_data)
        ```

        This will output the column values as a list of tuples:

        ```text        
            [('John',), ('Alice',), ('Travis',)]
        ```

        Note:
            - Many other ``SQLDataModel`` methods rely on this method, changing it will lead to undefined behavior.
            - See related :meth:`SQLDataModel.from_data()` for creating a new ``SQLDataModel`` from existing data sources.
            - Use ``strict_2d = True`` to always return data as a list of tuples regardless of data dimension.

        Changelog:
            - Version 0.10.0 (2024-06-29):
                - Modified to use :meth:`SQLDataModel._generate_sql_stmt_fetchall()` to leverage deterministic behavior of method.    
            - Version 0.5.0 (2024-05-09):
                - Added ``strict_2d`` parameter to allow predictable return type regardless of data dimension.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.5 (2023-11-24):
                - New method.            
        """
        res = self.sql_db_conn.execute(self._generate_sql_stmt_fetchall(index=index))
        data = res.fetchall()
        if not strict_2d:
            if (len(data) == 1) and (not include_headers): # if only single row
                data = data[0]
                if len(data) == 1: # if only single cell
                    data = data[0]
        return [tuple(x[0] for x in res.description),*data] if include_headers else data

    def to_csv(self, filename:str=None, delimiter:str=',', quotechar:str='"', lineterminator:str='\r\n', na_rep:str='None', encoding:str='utf-8', index:bool=False, **kwargs) -> str|None:
        """
        Writes ``SQLDataModel`` to the specified file if ``filename`` argument if provided, otherwise returns the model directly as a CSV formatted string literal.

        Parameters:
            ``filename`` (str): The name of the CSV file to which the data will be written. Default is None, returning as raw literal.
            ``delimiter`` (str, optional): The delimiter to use for separating values. Default is ','.
            ``quotechar`` (str, optional): The character used to quote fields. Default is '"'.
            ``lineterminator`` (str, optional): The character used to terminate the row and move to a new line. Default is '\\r\\n'.
            ``na_rep`` (str, optional): String representation to use for null or missing values. Default is 'None'.
            ``encoding`` (str, optional): The encoding to use when writing the model to a CSV file. Default is 'utf-8'.
            ``index`` (bool, optional): If True, includes the index in the CSV file; if False, excludes the index. Default is False.
            ``**kwargs``: Additional arguments to be passed to the ``csv.writer`` constructor.

        Returns:
            ``str`` | ``None``: If ``filename`` is None, returns the model as a delimited string literal, ``None`` if ``filename`` is provided, writing the model to the specified file as a CSV file.

        Example:

        Returning CSV Literal
        ---------------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Generate the literal using tab delimiter
            csv_literal = df.to_csv(delimiter='\\t')

            # View output
            print(csv_literal)
        ```

        This will output:

        ```shell
            Name    Age     Height
            John    30      175.3
            Alice   28      162.0
            Travis  35      185.8
        ```

        Write to File
        -------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # CSV filename
            csv_file = 'persons.csv'

            # Write to the file, keeping the index
            df.to_csv(filename=csv_file, index=True)
        ```

        Contents of ``persons.csv``:

        ```shell
            idx,Name,Age,Height
            0,John,30,175.3
            1,Alice,28,162.0
            2,Travis,35,185.8
        ```

        Note:
            - When ``index=True``, the ``sdm_index`` property determines the column name of the index in the result.
            - Modifying ``delimiter`` affects how the data is delimited when writing to ``filename`` and when returning as raw literal, any valid delimiter can be used.
            - Quoting behavior can be modified by providing an additional keywork arg such as ``quoting=1`` to wrap all values in quotes, or ``quoting=2`` to quote only non-numeric values, see ``csv.QUOTE_X`` enums for all options.
            - Use :meth:`SQLDataModel.to_text()` to pretty print table in specified style for visualizing output if strict delimiting is unnecessary.
            - See :meth:`SQLDataModel.from_csv()` for creating a new ``SQLDataModel`` from existing CSV data
        
        Changelog:
            - Version 0.6.4 (2024-05-17):
                - Added ``encoding`` parameter to pass to file handler when writing contents as CSV file and set default to ``utf-8`` to align with expected SQLite codec.
            - Version 0.4.0 (2024-04-23):
                - Modified quoting behavior to avoid redundant quoting and to closely mimic csv module from standard library.
                - Added ``na_rep`` to fill null or missing values when generating output, useful for space delimited data and minimal quoting.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
        """
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index, na_rep=na_rep))
        headers = [x[0] for x in res.description]
        if filename is not None:
            with open(filename, 'w', newline='', encoding=encoding) as file:
                csvwriter = csv.writer(file,delimiter=delimiter,lineterminator=lineterminator,quotechar=quotechar,**kwargs)
                csvwriter.writerow(headers)
                csvwriter.writerows(res.fetchall())
            return
        else:
            csv_repr = StringIO()
            csv_writer = csv.writer(csv_repr,delimiter=delimiter,lineterminator=lineterminator,quotechar=quotechar,**kwargs)
            csv_writer.writerow(headers)
            csv_writer.writerows(res.fetchall())
            return csv_repr.getvalue().strip()

    def to_dict(self, orient:Literal["rows","columns","list"]="rows", index:bool=None) -> dict|list[dict]:
        """
        Converts the ``SQLDataModel`` instance to a dictionary or a list of dictionaries based on the specified orientation.

        Parameters:
            ``orient`` (Literal["rows", "columns", "list"]): The orientation of the output, see examples for more detail. ``"rows"``: Returns a dictionary with index values as keys and row values as values. ``"columns"``: Returns a dictionary with column names as keys and column values as tuples. ``"list"``: Returns a list of dictionaries, where each dictionary represents a row.
            ``index`` (bool): Whether to include the index column in the output. Defaults to the display_index property.

        Raises:
            ``ValueError``: if value for ``orient`` is not one of "rows", "columns" or "list".

        Returns:
            ``dict`` | ``list[dict]``: The converted data structure based on the specified orientation.

        Examples:

        Orient by Rows
        --------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Col A','Col B', 'Col C']
            data = [
                ['A,0', 'A,1', 'A,2'],
                ['B,0', 'B,1', 'B,2'],
                ['C,0', 'C,1', 'C,2']
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Convert to dictionary with rows as keys and values
            rows_dict = df.to_dict(orient="rows")

            # View output
            for k, v in rows_dict.items():
                print(f"{k}: {v}")    
        ```

        This will output:

        ```shell
            0: ('A,0', 'B,0', 'C,0')
            1: ('A,1', 'B,1', 'C,1')
            2: ('A,2', 'B,2', 'C,2')
        ```
        
        Orient by Columns
        -----------------
        
        ```python
            # Convert to dictionary with columns as keys and rows as values
            columns_dict = df.to_dict(orient="columns")

            # View output
            for k, v in columns_dict.items():
                print(f"{k}: {v}") 
        ```

        This will output:

        ```shell            
            Col A: ('A,0', 'A,1', 'A,2')
            Col B: ('B,0', 'B,1', 'B,2')
            Col C: ('C,0', 'C,1', 'C,2')
        ```

        Orient by List
        --------------
        
        ```python
            # Convert to list of dictionaries with each dictionary representing a row with columns as keys
            list_dict = df.to_dict(orient="list")

            # View output
            for row in list_dict:
                print(row)
        ```

        This will output:

        ```shell
            {'Col A': 'A,0', 'Col B': 'B,0', 'Col C': 'C,0'}
            {'Col A': 'A,1', 'Col B': 'B,1', 'Col C': 'C,1'}
            {'Col A': 'A,2', 'Col B': 'B,2', 'Col C': 'C,2'}
        ```

        Note:
            - Use ``index`` to return index data, otherwise current instance ``display_index`` value will be used.
            - For ``'list'`` orientation, data returned is JSON-like in structure, where each row has its own "column": "value" data.
        
        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.5 (2023-11-24):
                - New method.                
        """   
        if orient not in ("rows", "columns", "list"):
            msg = ErrorFormat(f"ValueError: invalid argument '{orient}', value for `orient` must be one of 'rows', 'columns' or 'list' to determine the object returned for the `to_dict()` method")
            raise ValueError(msg)
        index = self.display_index if index is None else index
        if orient == "rows":
            return {row[0]:row[1:] for row in self.sql_db_conn.execute(self._generate_sql_stmt(index=True)).fetchall()}
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index))
        data, headers = res.fetchall(),[x[0] for x in res.description] 
        if orient == "columns":
            return {headers[i]:tuple([x[i] for x in data]) for i in range(len(headers))}    
        return [{col:row[i] for i,col in enumerate(headers)} for row in data]
        
    def to_excel(self, filename:str, worksheet:int|str=1, index:bool=False, if_exists:Literal['append','replace','fail']='replace') -> None:
        """
        Writes the current ``SQLDataModel`` to the specified Excel ``filename``.

        Parameters:
            ``filename`` (str): The file path to save the Excel file, e.g., ``filename = 'output.xlsx'``.
            ``worksheet`` (int | str, optional): The index or name of the worksheet to write to. Defaults to 1, indicating the first worksheet.
            ``index`` (bool, optional): If ``SQLDataModel`` index should be included in the output. Default is False.
            ``if_exists`` (Literal['append','replace','fail']): Action to take if file already exists. Default is 'replace', overwriting existing file.

        Raises:
            ``ModuleNotFoundError``: If the required package ``openpyxl`` is not installed as determined by ``optionals._has_xl`` flag.        
            ``TypeError``: If the ``filename`` argument is not of type 'str' representing a valid Excel file path to create or write to.
            ``ValueError``: If ``if_exists`` is not one of 'append', 'replace' or 'fail' representing action to take if file exists.
            ``IndexError``: If ``worksheet`` is provided as type 'int' but is out of range of the available worksheets.
            ``Exception``: If any unexpected exception occurs during the Excel writing and saving process.
        
        Returns:
            ``None``: If successful, a new Excel file ``filename`` is created and ``None`` is returned.   

        Example::

            import openpyxl
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Rate', 'Gender']
            data = [
                ('Alice', 25, 26.50, 'Female'),
                ('Bob', 30, 21.25, 'Male'),
                ('Will', 35, 24.00, 'Male'),
                ('Mary', 32, 23.75, 'Female')
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Export into a new Excel file
            df.to_excel('Team-Overview.xlsx')

            # Or append to existing Excel file as a new worksheet
            df.to_excel('Team.xlsx', worksheet='Demographics', if_exists='append')
        
        This will create a new Excel file ``Team-Overview.xlsx``:

        ```shell
                ┌───────┬──────┬────────┬────────┐
                │ A     │  B   │ C      │ D      │
            ┌───┼───────┼──────┼────────┼────────┤
            │ 1 │ Name  │  Age │ Gender │   Rate │
            │ 2 │ Alice │   25 │ Female │  26.50 │
            │ 3 │ Mary  │   32 │ Female │  23.75 │
            │ 4 │ Bobby │   30 │ Male   │  21.25 │
            │ 5 │ Will  │   35 │ Male   │  24.00 │
            └───┴───────┴──────┴────────┴────────┘
            [ Sheet1 ]
        ```

        Note:
            - Headers are dynamically inserted based on value for ``if_exists``, where using 'replace' will include headers and 'append' will ignore them unless worksheet creation occurred.
            - When providing a string argument for ``worksheet``, if the sheet does not exist, it will be created. However if providing an integer index for an out of range sheet, an ``IndexError`` will be raised.
            - See related :meth:`SQLDataModel.from_excel()` for creating a ``SQLDataModel`` from existing Excel content.
        
        Changelog:
            - Version 0.8.1 (2024-06-23):
                - Added ``if_exists`` parameter to provide the options to replace or append to existing file, as well as to fail if already exists.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.2.2 (2024-03-26):
                - New method.
        """        
        if not optionals._has_xl:
            msg = ErrorFormat(f"ModuleNotFoundError: required package not found, `openpyxl` must be installed in order to use `from_excel()` method")
            raise ModuleNotFoundError(msg)
        _xl = optionals._get_xl()
        if not isinstance(filename, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid Excel file path")
            raise TypeError(msg)
        if if_exists not in ('append','replace','fail'):
            msg = ErrorFormat(f"ValueError: invalid value '{if_exists}', argument for `if_exists` must be one of 'append', 'replace' or 'fail' representing action to take if file exists")
            raise ValueError(msg)
        if os.path.exists(filename):
            if if_exists == 'fail':
                msg = ErrorFormat(f"FileExistsError: '{filename}' already exists, use `if_exists='replace'` to overwrite or `if_exists='append'` to append to existing file")
                raise FileExistsError(msg)
            elif if_exists == 'replace':
                include_headers = True
                wb = _xl.Workbook()
                if isinstance(worksheet, int):
                    ws = wb.active
                    ws.title = f"Sheet{worksheet}"
                else:
                    ws = wb.create_sheet(title=worksheet)
            elif if_exists == 'append':
                include_headers = False
                wb = _xl.load_workbook(filename)
                if isinstance(worksheet, int):
                    try:
                        ws = wb.worksheets[worksheet-1]  # Index is 1-based for the parameter, but 0-based for list
                    except IndexError:
                        include_headers = True
                        ws = wb.create_sheet(f"Sheet{worksheet}")
                else:
                    if worksheet in wb.sheetnames:
                        ws = wb[worksheet]
                    else:
                        ws = wb.create_sheet(title=worksheet)
        else:
            include_headers = True
            wb = _xl.Workbook()
            if isinstance(worksheet, int):
                ws = wb.active
                ws.title = f"Sheet{worksheet}"
            else:
                ws = wb.create_sheet(title=worksheet)
        try:
            for row in self.data(strict_2d=True, include_headers=include_headers, index=index):
                ws.append(row)
            wb.save(filename=filename)
            wb.close()
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to write and save to the provided Excel file")
            raise type(e)(msg).with_traceback(e.__traceback__) from None   

    def to_html(self, filename:str=None, index:bool=None, encoding:str='utf-8', style_params:dict=None) -> str:
        """
        Returns the current SQLDataModel as a lightly formatted HTML <table> element as a string if ``filename`` is None.
        If ``filename`` is specified, writes the HTML to the specified file as .html and returns None.

        Parameters:
            ``filename`` (str): The file path to save the HTML content. If None, returns the HTML as a string (default is None).
            ``index`` (bool): Whether to include the index column in the HTML table (default is current ``display_index``).
            ``encoding`` (str): Character encoding to use when writing model to HTML file, default set to ``'utf-8'``.            
            ``style_params`` (dict): A dictionary representing CSS styles {property: value} to customize the appearance of the HTML table (default is None).

        Raises:
            ``TypeError``: If ``filename`` is not a valid string when specified or if ``style_params`` is not a dictionary when specified.
            ``OSError``: If encountered while trying to open and write the HTML to the file.

        Returns:
            ``str`` | ``None``: If ``filename`` is None, returns the HTML content as a string. If ``filename`` is specified, writes to the file and returns None.

        Example::
            
            import sqldatamodel as sdm

            # Create the model
            df = sdm.SQLDataModel(data=[(1, 'John'), (2, 'Doe')], headers=['ID', 'Name'])

            # Create and save as new html file
            df.to_html('output.html', style_params={'font-size': '12pt'})
            
            # Get HTML as a string
            html_string = df.to_html()

            # View output
            print(html_string)
        
        This will output:

        ```shell            
            <table>
                <tr>
                    <th>ID</th>
                    <th>Name</th>
                </tr>
                <tr>
                    <td>1</td>
                    <td>John</td>
                </tr>
                <tr>
                    <td>2</td>
                    <td>Doe</td>
                </tr>
            </table>
            <style>
                table {font:size: 12pt;}
            </style>
        ```

        Note:
            - Base styles are applied to reflect the styling of ``SQLDataModel`` in the terminal, including any ``display_color`` which is applied to the table CSS.
            - Table index is determined by the instance ``display_index`` attribute unless specified in the argument of the same name, overriding the instance attribute.
            - The default background-color is #E5E5E5, and the default font color is #090909, with 1 px solid border to mimic the ``repr`` for the instance.
        
        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(filename, str) and filename is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', if specified, `filename` must be of type 'str' representing a valid file path")
            raise TypeError(msg)
        if not isinstance(style_params, dict) and style_params is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(style_params).__name__}', if specified, `style_params` must be of type 'dict' representing CSS {{'property': 'value'}} styles")
            raise TypeError(msg)
        font_color, background_color = self.display_color.text_color_hex if self.display_color is not None else "#E5E5E5", "#090909"
        if index is None:
            index = self.display_index
        display_headers = [self.sql_idx,*self.headers] if index else self.headers
        html_headers = "\n".join(("\t<tr>",*tuple(f"""\t\t<th>{col}</th>""" for col in display_headers),"\t</tr>")) # replace `{col}` with `{col if col != self.sql_idx else " "}` to revert idx display
        html_body ="".join(["\n".join(("\n\t<tr>",*tuple(f"""\t\t<td>{cell}</td>""" for cell in tr),"\t</tr>")) for tr in self.iter_rows(index=index)]).strip('\n')
        col_styles = "\n".join([f"""th:nth-child({i+1}),td:nth-child({i+1}) {{{"text-align: right;" if self.header_master[col][3] == '>' else "text-align: left;"}}}""" for i,col in enumerate(display_headers)])
        base_styles = f"""html {{background-color: {background_color}}}\ntable,th {{border: 1px solid {font_color}; border-collapse: collapse; overflow-x: auto; background-color:{background_color}; color:{font_color}; margin:6px;}}\ntr,td,th {{padding: 1px 6px; border-right: 1px solid {font_color}; font-family: 'Consolas', 'Monaco', 'Lucida Console', monospace; font-size: 9pt; font-weight: normal; overflow-x: auto;}}\ntr:first-child > th {{padding: 4px 6px;}}\ntr:nth-child(2) > td {{padding-top: 4px;}}\ntr:last-child > td {{padding-bottom: 4px;}}"""
        cascade_styles = "".join(("\ntable,tr,td,th {",*tuple(f"""{attr}:{value};""" for attr,value in style_params.items()),"}")) if style_params is not None else ""
        html_styling = "\n".join(("<style>",f"{base_styles}{cascade_styles}",col_styles,"</style>"))
        html_table = f"""<!DOCTYPE html>\n<table>\n{html_headers}\n{html_body}\n</table>\n{html_styling}"""
        if filename is None:
            return html_table
        try:
            with open(filename, "w", encoding=encoding) as f:
                f.write(html_table)
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and write html")
            raise type(e)(msg).with_traceback(e.__traceback__) from None

    def to_json(self, filename:str=None, index:bool=None, **kwargs) -> list|None:
        """
        Converts the ``SQLDataModel`` instance to JSON format. If ``filename`` is specified, the JSON is written to the file;
        otherwise, a JSON-like object is returned.

        Parameters:
            ``filename`` (str): The path to the file where JSON will be written. If None, no file is created and JSON-like object is returned.
            ``index`` (bool): Whether to include the index column in the JSON. Defaults to the ``display_index`` property.
            ``**kwargs``: Additional keyword arguments to pass to the json.dump() method.

        Raises:
            ``TypeError``: If ``filename`` is not of type 'str'.
            ``Exception``: If there is an OS related error encountered when opening or writing to the provided ``filename``.

        Returns:
            ``list`` | ``None``: If ``filename`` is None, a list containing a JSON-like object is returned. Otherwise JSON file created and returns ``None``.

        Examples:

        To JSON Literal
        -----------------

        ```python
            import sqldatamodel as sdm

            # Sample JSON to first create model
            json_source = [
                {"id": 1, "color": "red", "value": "#f00", "notes": "primary"}
                ,{"id": 2, "color": "green", "value": "#0f0", "notes": None}
                ,{"id": 3, "color": "blue", "value": "#00f", "notes": "primary"}
            ]

            # Create the model
            df = sdm.from_json(json_source)

            # View current state
            print(df)
        ```

        This will output:

        ```shell            
            ┌─────┬───────┬───────┬─────────┐
            │  id │ color │ value │ notes   │
            ├─────┼───────┼───────┼─────────┤
            │   1 │ red   │ #f00  │ primary │
            │   2 │ green │ #0f0  │         │
            │   3 │ blue  │ #00f  │ primary │
            └─────┴───────┴───────┴─────────┘
            [3 rows x 4 columns]
        ```
            
        Write JSON File
        -----------------

        ```python
            # Write model to JSON file
            df.to_json('output.json')

            # Or convert to JSON-like object
            json_data = df.to_json()

            # View JSON object
            print(json_data)
        ```

        This will output:

        ```shell
            [{
                "id": 1,
                "color": "red",
                "value": "#f00",
                "notes": "primary"
            },
            {
                "id": 2,
                "color": "green",
                "value": "#0f0",
                "notes": null
            },
            {
                "id": 3,
                "color": "blue",
                "value": "#00f",
                "notes": "primary"
            }]
        ```

        Note:
            - When no filename is specified, JSON-like object will be returned as a rowwise array.
            - Any nested structure will be flattened by this method as well as the :meth:`SQLDataModel.from_json()` method.
        
        Changelog:
            - Version 0.3.2 (2024-04-02):
                - Changed return object to JSON string literal when ``filename=None`` to convert to valid literal object.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(filename, str) and filename is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', expected `filename` to be of type 'str' representing a valid file path to write json")
            raise TypeError(msg)
        index = self.display_index if index is None else index
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index))
        res_headers = [x[0] for x in res.description] 
        json_data = [{col:row[i] for i,col in enumerate(res_headers)} for row in res.fetchall()]
        if filename is not None:
            try:
                with open(filename, "w") as f:
                    json.dump(json_data, f, cls=DataTypesEncoder, **kwargs)
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and write json")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        else:
            return json.dumps(json_data, cls=DataTypesEncoder, **kwargs)

    def to_latex(self, filename:str=None, index:bool=False, bold_headers:bool=False, min_column_width:int=None, max_column_width:int=None, float_precision:int=None, horizontal_ellipses:str='⠤⠄', index_rep:str=None, format_output_as:Literal['table', 'document']='table', column_alignment:Literal['left', 'center', 'right', 'dynamic']=None) -> str | None:
        """
        Returns the current ``SQLDataModel`` as a LaTeX table string if ``filename`` is None, otherwise writes the table to the provided file as a LaTeX document.

        Parameters:
            ``filename`` (str, optional): The name of the file to write the LaTeX content. If not provided, the LaTeX content is returned as a string. Default is None.
            ``index`` (bool, optional): Whether to include the index column in the LaTeX output. Default is False.
            ``bold_headers`` (bool, optional): Whether the headers should be bolded in the LaTeX table. Default is False.
            ``min_column_width`` (int, optional): The minimum column width for table cells. Default is current value set on attribute :py:attr:`SQLDataModel.min_column_width`.
            ``max_column_width`` (int, optional): The maximum column width for table cells. Default is current value set on attribute :py:attr:`SQLDataModel.max_column_width`.
            ``float_precision`` (int, optional): The precision for floating-point values. Default is current value set on :py:attr:`SQLDataModel.display_float_precision`.
            ``horizontal_ellipses`` (str, optional): Characters to truncate column and cell values when horizontal truncation is required. Default is ``'⠤⠄'``.
            ``index_rep`` (str, optional): String representation for the index. Default is None, using value set on :py:attr:`SQLDataModel.sql_idx` to represent the index column.     
                Only used when generating index column, otherwise ignored when ``index = False``.          
            ``format_output_as`` (Literal['table', 'document']), optional): Whether the output should be formatted as a LaTeX table or as a standalone document. 
                Default is 'table', formatting output as a modular table element.
            ``column_alignment`` (Literal['left', 'center', 'right', 'dynamic'], optional): The column alignment to use. 
                Default is current value set on attribute :py:attr:`SQLDataModel.column_alignment`.

        Returns:
            ``str`` or ``None``: If ``filename`` is None, returns the LaTeX formatted table as a string, if ``filename`` is provided, writes the LaTeX table to the specified file and returns None.

        Raises:
            ``TypeError``: If the ``filename`` argument is not of type 'str', ``index`` argument is not of type 'bool', ``min_column_width`` or ``max_column_width`` argument is not of type 'int'.
            ``ValueError``: If ``format_output_as`` is not one of 'table', 'document', or ``column_alignment`` provided and is not one of 'left', 'center', 'right', 'dynamic'.
            ``Exception``: If there is an OS related error encountered when opening or writing to the provided ``filename``.

        LaTeX Formatting:
            - LaTeX output format that is generated can be set by ``format_output_as`` which provides one of two formats:

              - ``'table'``: Output formatted as insertable table, beginning and ending with LaTeX ``\\begin{table}`` and ``\\end{table}`` respectively.
              - ``'document'``: Output formatted as standalone document, beginning and ending with LaTeX ``\\begin{document}`` and ``\\end{document}`` respectively.

            - LaTeX table alignment will follow the ``SQLDataModel`` instance alignment, set by :meth:`SQLDataModel.set_column_alignment()`:

              - ``'dynamic'``: Dynamically aligns column content, right for numeric types and left for remaining types.
              - ``'left'``: Left-aligns all column content, equivalent to LaTeX column format: ``|l|``.
              - ``'center'``: Center-aligns all column content preferring left on uneven splits, equivalent to LaTeX column format: ``|c|``.
              - ``'right'``: Right-aligns all column content, equivalent to LaTeX column format: ``|r|``.

            - The LaTeX rows generated will use ``dynamic`` alignment regardless of ``column_alignment`` provided, this will not affect the rendered alignment but will maintain consistent format without affecting the actual alignment rendered by LaTeX.

        Examples:

        Returning LaTeX Literal
        -----------------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Michael', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Generate LaTeX table literal
            latex_output = df.to_latex()

            # View LaTeX output
            print(latex_output)
        ```

        This will output:
        
        ```shell            
            \\begin{tabular}{|l|r|r|}
            \\hline
                {Name} & {Age} & {Height} \\
            \\hline
                John    &   30 &  175.30 \\
                Alice   &   28 &  162.00 \\
                Michael &   35 &  185.80 \\
            \\hline
            \\end{tabular}
        ```

        Write to LaTeX File
        -------------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Michael', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Write the output to the file, formatting the output as a proper LaTeX document
            latex_table = df.to_latex(filename='Table.tex', format_output_as='document')      
        ```

        Contents of file ``Table.tex``:
        
        ```shell            
            \\documentclass{article}
            \\begin{document}
            \\begin{table}[h]
            \\centering
            \\begin{tabular}{|l|r|r|}
            \\hline
                {Name} & {Age} & {Height} \\
            \\hline
                John    &   30 &  175.30 \\
                Alice   &   28 &  162.00 \\
                Michael &   35 &  185.80 \\
            \\hline
            \\end{tabular}
            \\end{table}
            \\end{document}
        ```

        Note:
            - A ``\\centering`` command is included in the LaTeX output by default regardless of alignments specified.
            - LaTeX headers and rows are indented by four spaces to keep with conventional table syntax and to distinguish the table data from commands.
            - Table commands and headers are checked for invalid LaTeX characters and escaped such as ``'_'`` and ``'#'``, however the model data is not. 
                Accordingly, ensure any model content is valid LaTeX when rendering to PDF, or simply format content as valid LaTeX before exporting.
        
        Changelog:
            - Version 0.11.0 (2024-07-05):
                - Added ``float_precision`` parameter to align with similar format specific methods and provide additional formatting options.
                - Added ``horizontal_ellipses`` parameter to allow customizing truncation characters used when column or cell values exceed maximumn column widths.
                - Added ``index_rep`` parameter to allow customizing index column name with prior behavior set as default representation. Ignored when ``index = False``.
                - Modified to use :meth:`SQLDataModel.to_string()` instead of generating independently formatted repr for more consistency between tabular outputs.
                - Modified to check and escape any invalid LaTeX characters or symbols when generating headers.
            - Version 0.10.4 (2024-07-03):
                - Modified to escape newline characters through :meth:`utils.sqlite_printf_format()` to avoid wrapping table rows.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(filename, str) and filename is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', expected `filename` to be of type 'str' representing a valid file path to write LaTeX")
            raise TypeError(msg)
        if not isinstance(index, bool):
            msg = ErrorFormat(f"TypeError: invalid type '{type(index).__name__}', expected `index` to be of type 'bool' representing whether index should be included in LaTeX output")
            raise TypeError(msg)
        if (not isinstance(min_column_width, int) and (min_column_width is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(min_column_width).__name__}', expected `min_column_width` to be of type 'int' representing minimum column width for table cells")
            raise TypeError(msg)        
        if (not isinstance(max_column_width, int) and (max_column_width is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(max_column_width).__name__}', expected `max_column_width` to be of type 'int' representing maximum column width for table cells")
            raise TypeError(msg)   
        if format_output_as not in ('table', 'document'):
            msg = ErrorFormat(f"ValueError: invalid value '{format_output_as}', expected `format_output_as` to be either 'table' or 'document' representing format for LaTeX output")
            raise ValueError(msg)        
        if (column_alignment is not None) and (column_alignment not in ('left', 'center', 'right', 'dynamic')):
            msg = ErrorFormat(f"ValueError: invalid value '{column_alignment}', expected `column_alignment` to be one of 'left', 'center', 'right', 'dynamic' representing column alignment for LaTeX output")
            raise ValueError(msg)
        display_index = index if index is not None else self.display_index
        min_column_width = self.min_column_width if min_column_width is None else min_column_width
        max_column_width = self.max_column_width if max_column_width is None else max_column_width
        float_precision = self.display_float_precision if float_precision is None else float_precision
        column_alignment = self.column_alignment if column_alignment is None else column_alignment
        index_rep = self.sql_idx if index_rep is None else index_rep
        table_repr = self.to_string(
            index=display_index,
            display_max_rows=self.row_count,
            display_max_width=None,
            min_column_width=min_column_width,
            max_column_width=max_column_width,
            float_precision=float_precision,
            horizontal_ellipses=horizontal_ellipses,
            index_rep=index_rep,
            display_dimensions=False,
            column_alignment=column_alignment,
            table_style='latex'    
        ) 
        # Check for invalid LaTeX to escape and optionally bold headers
        headers_start_pos = table_repr.find("""\\\\""")
        if headers_start_pos > 0:
            latex_headers = table_repr[:headers_start_pos].replace("""_""","""\\_""").replace("""#""","""\\#""")
            if bold_headers:
                latex_headers = latex_headers.replace("""{""","""\\textbf{""")
            table_repr = """""".join((latex_headers, table_repr[headers_start_pos:]))        
        display_headers = [self.sql_idx,*self.headers] if index else self.headers
        latex_column_marker = """|"""
        latex_begin_document_format = """\\documentclass{article}\n\\begin{document}"""
        latex_end_document_format = """\\end{document}"""
        latex_begin_table_format = """\\begin{table}[h]\n\\centering"""
        latex_end_table_format = """\\end{table}"""
        if column_alignment != 'dynamic':
            latex_align_char = column_alignment[:1]
            latex_column_template = latex_column_marker.join([latex_align_char for _ in display_headers])
        else:
            latex_column_template = latex_column_marker.join(['r' if self.header_master[col][1] in ('int','float') else 'l' for col in display_headers])
        latex_begin_tabular_format = f"""\\begin{{tabular}}{{{latex_column_marker}{latex_column_template}{latex_column_marker}}}"""
        latex_end_tabular = """\\end{tabular}"""
        table_repr = f"""{latex_begin_table_format}\n{latex_begin_tabular_format}\n{table_repr}\n{latex_end_tabular}\n{latex_end_table_format}"""
        if format_output_as == 'document':
            table_repr = f"""{latex_begin_document_format}\n{table_repr}\n{latex_end_document_format}"""
        if filename is not None:
            try:
                with open(filename, "w", encoding='utf-8') as f:
                    f.write(table_repr)
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and write LaTeX")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        else:
            return table_repr 

    def to_list(self, index:bool=False, include_headers:bool=False) -> list:
        """
        Returns the current ``SQLDataModel`` data as a 1-dimensional list of values if data dimensions are compatible with flattening, or as a list of lists if data is 2-dimensional.
        Data is returned without index or headers by default, use ``index = True`` or ``include_headers = True`` to modify.

        Parameters:
            ``index`` (bool, optional): If True, includes the index in the result, if False, excludes the index. Default is False.
            ``include_headers`` (bool, optional): If True, includes column headers in the result, if False, excludes headers. Default is False.

        Returns:
            ``list``: The flattened list of values corresponding to the model data.

        Example::
            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('Beth', 27, 172.4),
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)    
            
            # Get all model data as a list of lists
            model_data = df.to_list()

            # Iterate over each row
            for row in model_data:
                print(row)

        This will output:

        ```text
            ['Beth', 27, 172.4]
            ['John', 30, 175.3]
            ['Alice', 28, 162.0]
            ['Travis', 35, 185.8]
        ```

        Data will be flattened into a single dimension if possible, such as when accessing individual columns:

        ```python
            # Get 'Name' column as a list
            col_data = df['Name'].to_list()

            # View output
            print(col_data)
        ```

        This will output a list containing the values from each row for the column:

        ```text
            ['Beth', 'John', 'Alice', 'Travis']
        ```

        Data will also be flattened when accessing individual rows:

        ```python
            # Get first row as a list with index
            row_data = df[0].to_list(index=True)
            
            # View result
            print(row_data)
        ```

        This will output the row's values including the index:

        ```text
            [0, 'Beth', 27, 172.4]
        ```
        
        Note:
            - See :meth:`SQLDataModel.data()` to return the equivalent of ``cursor.fetchall()`` with data as a list of tuples.
            - See :meth:`SQLDataModel.iter_rows()` to generate an iterable over the model data, which is preferred wherever possible.

        Changelog:
            - Version 0.5.0 (2024-05-09):
                - Modified behavior to output 1-dimensional list when possible and a list of lists when not possible.
                - Changed default to ``index = False`` to increase surface for 1-dimensional flattening.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.5 (2023-11-24):
                - New method.
        """
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index))
        data = res.fetchall()
        if len(data) <= 1 and not include_headers:
            return list(data[0])
        else:
            if len(data[0]) == 1:
                data = list(row[0] for row in data)
                return [res.description[0][0],*data] if include_headers else data
            else:
                data = [list(row) for row in data]
                return [[x[0] for x in res.description],*data] if include_headers else data
    
    def to_markdown(self, filename:str=None, index:bool=False, min_column_width:int=None, max_column_width:int=None, float_precision:int=None, horizontal_ellipses:str='⠤⠄', index_rep:str=None, column_alignment:Literal['dynamic', 'left', 'center', 'right']=None) -> str|None:
        """
        Returns the current ``SQLDataModel`` as a markdown table literal if ``filename`` is None, otherwise writes the table to the provided file as markdown.

        Parameters:
            ``filename`` (str, optional): The name of the file to write the Markdown content. If not provided, the Markdown content is returned as a string. Default is None.
            ``index`` (bool, optional): Whether to include the index column in the Markdown output. Default is False.
            ``min_column_width`` (int, optional): The minimum column width for table cells. Default is current value set on :py:attr:`SQLDataModel.min_column_width`.
            ``max_column_width`` (int, optional): The maximum column width for table cells. Default is current value set on :py:attr:`SQLDataModel.max_column_width`.
            ``float_precision`` (int, optional): The precision for floating-point values. Default is current value set on :py:attr:`SQLDataModel.display_float_precision`.
            ``horizontal_ellipses`` (str, optional): Characters to truncate column and cell values when horizontal truncation is required. Default is ``'⠤⠄'``.
            ``index_rep`` (str, optional): String representation for the index. Default is None, using value set on :py:attr:`SQLDataModel.sql_idx` to represent the index column.
                Only used when generating index column, otherwise ignored when ``index = False``.            
            ``column_alignment`` (Literal['dynamic', 'left', 'center', 'right'], optional): The alignment for table columns. Default is current value set on :py:attr:`SQLDataModel.column_alignment`.
                ``'dynamic'``: Dynamically aligns column content, right for numeric types and left for remaining types.
                ``'left'``: Left-aligns all column content.
                ``'center'``: Center-aligns all column content preferring left on uneven splits.
                ``'right'``: Right-aligns all column content.
        
        Raises:
            ``TypeError``: If the ``filename`` argument is not of type 'str', if ``float_precision``, ``min_column_width`` or ``max_column_width`` arguments are not type 'int'.
            ``ValueError``: If the ``column_alignment`` argument is provided and is not one of 'dynamic', 'left', 'center', or 'right'.
            ``Exception``: If there is an OS related error encountered when opening or writing to the provided ``filename``.

        Returns:
            ``str`` or ``None``: If ``filename`` is None, returns the Markdown table as a string, if ``filename`` is provided, writes the Markdown table to the specified file and returns None.

        Column Alignment:
            - ``'dynamic'``: Dynamically aligns column content, right for numeric types and left for remaining types.
            - ``'left'``: Left-aligns all column content.
            - ``'center'``: Center-aligns all column content preferring left on uneven splits.
            - ``'right'``: Right-aligns all column content.

        Examples:
        
        To Markdown Literal
        -------------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Michael', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Generate markdown table literal
            markdown_table = df.to_markdown()

            # View markdown output
            print(markdown_table)
        ```

        This will output:

        ```shell
            | Name    |  Age |  Height |
            |:--------|-----:|--------:|
            | John    |   30 |  175.30 |
            | Alice   |   28 |  162.00 |
            | Michael |   35 |  185.80 |
        ```

        Write to Markdown File
        ----------------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Michael', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Write the output to the file, center-aligning all columns
            df.to_markdown(filename='Table.MD', column_alignment='center')  
        ```

        Contents of ``Table.MD``:

        ```shell            
            | Name    |  Age |  Height |
            |:--------|-----:|--------:|
            | John    |   30 |  175.30 |
            | Alice   |   28 |  162.00 |
            | Michael |   35 |  185.80 |
        ```

        Note:
            - All markdown output will contain the alignment characters ``':'`` as determined by the :py:attr:`SQLDataModel.column_alignment` attribute or parameter.
            - Any exception encountered during file read or writing operations is caught and reraised, see related :meth:`SQLDataModel.from_markdown()`.
            - Use ``index_rep`` to provide a different representation, column name, for the index column if included in output.
            - Unlike other representations, no rowwise or vertical truncation is performed on output content.
        
        Changelog:
            - Version 0.11.0 (2024-07-05):
                - Added ``horizontal_ellipses`` parameter to allow customizing truncation characters used when column or cell values exceed maximumn column widths.
                - Added ``index_rep`` parameter to allow customizing index column name with prior behavior set as default representation. Ignored when ``index = False``.
                - Modified to use :meth:`SQLDataModel.to_string()` instead of generating independently formatted repr for more consistency between tabular outputs.
            - Version 0.10.4 (2024-07-03):
                - Modified to escape newline characters through :meth:`utils.sqlite_printf_format()` to avoid wrapping table rows.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(filename, str) and filename is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', expected `filename` to be of type 'str' representing a valid file path to write markdown")
            raise TypeError(msg)
        if (not isinstance(min_column_width, int) and (min_column_width is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(min_column_width).__name__}', expected `min_column_width` to be of type 'int' representing minimum column width for table cells")
            raise TypeError(msg)        
        if (not isinstance(max_column_width, int) and (max_column_width is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(max_column_width).__name__}', expected `max_column_width` to be of type 'int' representing maximum column width for table cells")
            raise TypeError(msg)   
        if (not isinstance(float_precision, int) and (float_precision is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(float_precision).__name__}', expected `float_precision` to be of type 'int' representing precision to use for values of type 'float'")
            raise TypeError(msg)          
        if (column_alignment is not None) and (column_alignment not in ('dynamic', 'left', 'center', 'right')):
            msg = ErrorFormat(f"ValueError: invalid value '{column_alignment}', argument for `column_alignment` must be one of 'dynamic', 'left', 'center', 'right' representing column alignment for markdown output")
            raise ValueError(msg)
        display_index = index if index is not None else self.display_index
        min_column_width = self.min_column_width if min_column_width is None else min_column_width
        max_column_width = self.max_column_width if max_column_width is None else max_column_width
        float_precision = self.display_float_precision if float_precision is None else float_precision
        column_alignment = self.column_alignment if column_alignment is None else column_alignment
        index_rep = self.sql_idx if index_rep is None else index_rep
        table_repr = self.to_string(
            index=display_index,
            display_max_rows=self.row_count,
            display_max_width=None,
            min_column_width=min_column_width,
            max_column_width=max_column_width,
            float_precision=float_precision,
            horizontal_ellipses=horizontal_ellipses,
            index_rep=index_rep,
            display_dimensions=False,
            column_alignment=column_alignment,
            table_style='markdown'    
        )
        if filename is not None:
            try:
                with open(filename, "w", encoding='utf-8') as f:
                    f.write(table_repr)
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and write markdown")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        else:
            return table_repr  

    def to_numpy(self, index:bool=False, include_headers:bool=False) ->' _np.ndarray':
        """
        Converts ``SQLDataModel`` to a NumPy ``ndarray`` object of shape ``(rows, columns)``.
        Note that the ``numpy`` package must be installed to use this method.

        Parameters:
            ``index`` (bool, optional): If True, includes the model index in the result. Default is False.
            ``include_headers`` (bool, optional): If True, includes column headers in the result. Default is False.
        
        Raises:
            ``ModuleNotFoundError``: If NumPy is not installed.

        Returns:
            ``numpy.ndarray``: The model's data converted into a NumPy array.

        Example::

            import numpy
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the sample model
            df = sdm.SQLDataModel(data, headers)

            # Create the numpy array with default parameters, no indicies or headers
            result_array = df.to_numpy()

            # View array
            print(result_array)
        
        This will output:

        ```shell
            [['John' '30' '175.3']
             ['Alice' '28' '162.0']
             ['Travis' '35' '185.8']]
        ```

        Model headers can also be retained:

        ```python
            # Create the numpy array with with indicies and headers
            result_array = df.to_numpy(index=True, include_headers=True)

            # View array
            print(result_array)
        ```

        This will output:

        ```shell
            [['idx' 'Name' 'Age' 'Height']
             ['0' 'John' '30' '175.3']
             ['1' 'Alice' '28' '162.0']
             ['2' 'Travis' '35' '185.8']]
        ```

        Note:
            - Output will always be a 2-dimensional array of type ``numpy.ndarray``

        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.3 (2023-10-15):
                - New method.
        """
        if not optionals._has_np:
            msg = ErrorFormat(f"""ModuleNotFoundError: required package not found, numpy must be installed in order to use `.to_numpy()` method""")
            raise ModuleNotFoundError(msg)  
        _np = optionals._get_np()          
        fetch_stmt = self._generate_sql_stmt(index=index)
        res = self.sql_db_conn.execute(fetch_stmt)
        if include_headers:
            return _np.vstack([_np.array([x[0] for x in res.description]),[_np.array(x) for x in res.fetchall()]])
        return _np.array([_np.array(x) for x in res.fetchall()])

    def to_pandas(self, index:bool=False, include_headers:bool=True) -> '_pd.DataFrame':
        """
        Converts ``SQLDataModel`` to a Pandas ``DataFrame`` object.
        Note that the ``pandas`` package must be installed to use this method.

        Parameters:
            ``index`` (bool, optional): If True, includes the model index in the result. Default is False.
            ``include_headers`` (bool, optional): If True, includes column headers in the result. Default is True.

        Raises:
            ``ModuleNotFoundError``: If Pandas is not installed.

        Returns:
            ``pandas.DataFrame``: The model's data converted to a Pandas DataFrame.
        
        Example::

            import pandas
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df_sdm = sdm.SQLDataModel(data, headers)

            # Convert the model to a pandas df
            df_pd = df_sdm.to_pandas(include_headers=True, index=True)

            # View result
            print(df_pd)

        This will output:

        ```shell
                Name  Age  Height
            0    John   30   175.3
            1   Alice   28   162.0
            2  Travis   35   185.8
        ```

        Note:
            - SQLDataModel uses different data types than those used in ``pandas``, see :meth:`SQLDataModel.set_column_dtypes()` for more information about casting rules.

        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.3 (2023-10-15):
                - New method.
        """
        if not optionals._has_pd:
            msg = ErrorFormat(f"""ModuleNotFoundError: required package not found, pandas must be installed in order to use `.to_pandas()` method""")
            raise ModuleNotFoundError(msg)
        _pd = optionals._get_pd()
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index))
        raw_data = res.fetchall()
        data = [x[1:] for x in raw_data] if index else [x for x in raw_data]
        indicies = [x[0] for x in raw_data] if index else None
        columns = ([x[0] for x in res.description[1:]] if index else [x[0] for x in res.description]) if include_headers else None
        return _pd.DataFrame(data=data,columns=columns,index=indicies)

    def to_parquet(self, filename:str, index:bool=True, **kwargs) -> None:
        """
        Writes the current SQLDataModel to the specified parquet filename.

        Parameters:
            ``filename`` (str): The file path to save the parquet file, e.g., ``filename = 'user/data/output.parquet'``.
            ``index`` (bool, optional): Whether or not the SQLDataModel index should be included in the export. Default is True.
            ``**kwargs``: Additional keyword arguments to pass to the pyarrow ``write_table`` function.

        Raises:
            ``ModuleNotFoundError``: If the required package ``pyarrow`` is not installed as determined by ``optionals._has_pa`` flag.        
            ``TypeError``: If the ``filename`` argument is not of type 'str' representing a valid parquet file path.
            ``Exception``: If any unexpected exception occurs during the parquet writing process.
        
        Returns:
            ``None``: If successful, a new parquet file ``filename`` is created and ``None`` is returned.

        Example::
            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Rate']
            data = [('Alice', 25, 26.50), ('Bob', 30, 21.25), ('Will', 35, 24.00)]
            
            # Create the model
            df = sdm.SQLDataModel(data,headers, display_index=False)

            # Parquet file
            pq_file = "output.parquet"

            # Write the model as parquet file
            df.to_parquet(pq_file)

            # Confirm result by reading back file
            df_result = sdm.from_parquet(pq_file)

            # View model
            print(df_result)
        
        This will output:
        
        ```shell            
            ┌───────┬──────┬────────┐
            │ Name  │  Age │   Rate │
            ├───────┼──────┼────────┤
            │ Alice │   25 │  26.50 │
            │ Bob   │   30 │  21.25 │
            │ Will  │   35 │  24.00 │
            └───────┴──────┴────────┘
            [3 rows x 3 columns]        
        ```

        Note:
            - The ``pyarrow`` package is required to use this method as well as the :meth:`SQLDataModel.from_parquet()` method.
            - The :meth:`SQLDataModel.to_dict()` method is used prior to writing to parquet to convert the ``SQLDataModel`` into a dictionary suitable for parquet Table format.
            - Exceptions raised by the ``pyarrow`` package and its methods are caught and reraised when encountered to keep with package error formatting.

        Changelog:
            - Version 0.8.2 (2024-06-24):
                - Added ``index`` parameter to toggle inclusion of SQLDataModel ``index`` column for greater flexibility and package consistency to similar methods.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not optionals._has_pa:
            msg = ErrorFormat(f"ModuleNotFoundError: required package not found, pyarrow must be installed in order to use `.to_parquet()` method")
            raise ModuleNotFoundError(msg)        
        _pa, _pq = optionals._get_pa_pq()
        if not isinstance(filename, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid parquet file path")
            raise TypeError(msg)
        try:
            pqtable = _pa.Table.from_pydict(self.to_dict(orient='columns', index=index))
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to write parquet file")
            raise type(e)(msg).with_traceback(e.__traceback__) from None        
        _pq.write_table(pqtable, filename, **kwargs)

    def to_pickle(self, filename:str=None) -> None:
        """
        Save the ``SQLDataModel`` instance to the specified ``filename`` as a pickle object.

        Parameters:
            ``filename`` (str, optional): The file name to save the model to. If None, the invoking Python file's name with a ".sdm" extension will be used.
        
        Raises:
            ``TypeError``: If filename is provided but is not of type 'str' representing a valid pickle filepath.        

        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age']
            data = [
                (0, 'john', 'smith', 27)
                ,(1, 'sarah', 'west', 29)
                ,(2, 'mike', 'harlin', 36)
                ,(3, 'pat', 'douglas', 42)
            ]
            
            # Create the SQLDataModel object
            df = sdm.SQLDataModel(data, headers)

            # Save the model's data as a pickle file "output.sdm"
            df.to_pickle("output.sdm")

            # Alternatively, leave blank to use the current file's name:
            df.to_pickle()

            # This way the same data can be recreated later by calling the from_pickle() method from the same project:
            df = sdm.from_pickle()

        Note:
            - All data, headers, data types and display properties will be saved when pickling.
            - If no ``filename`` argument is provided, then the invoking module's ``__name__`` property will be used by default.
        """
        dtypes = {col:self.header_master[col][1] for col in [self.sql_idx,*self.headers]}
        serialized_sdm = dict(headers=list(dtypes.keys()),dtypes=dtypes, data=self.data(index=True), **self._get_display_args())
        if filename is None:
            filename = os.path.basename(sys.argv[0]).split(".")[0]+'.sdm'
        else:
            if not isinstance(filename, str):
                msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid pickle filepath")
                raise TypeError(msg)
        with open(filename, 'wb') as handle:
            pickle.dump(serialized_sdm, handle, protocol=pickle.HIGHEST_PROTOCOL)

    def to_polars(self, index:bool=False, include_headers:bool=True) -> '_pl.DataFrame':
        """
        Converts ``SQLDataModel`` to a Polars ``DataFrame`` object.
        Note that the ``polars`` package must be installed to use this method.

        Parameters:
            ``index`` (bool, optional): If True, includes the model index in the result. Default is False.
            ``include_headers`` (bool, optional): If True, includes column headers in the result. Default is True.

        Raises:
            ``ModuleNotFoundError``: If Polars is not installed.

        Returns:
            ``polars.DataFrame``: The model's data converted to a Polars DataFrame.
        
        Example::

            import polars
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('Beth', 27, 172.4),
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df_sdm = sdm.SQLDataModel(data, headers)

            # Convert the model to a polars df with the index
            df_pl = df_sdm.to_polars(index=True)

            # View result
            print(df_pl)

        This will output:

        ```shell
            shape: (4, 4)
            ┌─────┬────────┬─────┬────────┐
            │ idx ┆ Name   ┆ Age ┆ Height │
            │ --- ┆ ---    ┆ --- ┆ ---    │
            │ i64 ┆ str    ┆ i64 ┆ f64    │
            ╞═════╪════════╪═════╪════════╡
            │ 0   ┆ Beth   ┆ 27  ┆ 172.4  │
            │ 1   ┆ John   ┆ 30  ┆ 175.3  │
            │ 2   ┆ Alice  ┆ 28  ┆ 162.0  │
            │ 3   ┆ Travis ┆ 35  ┆ 185.8  │
            └─────┴────────┴─────┴────────┘
        ```

        Note:
            - See related :meth:`SQLDataModel.from_polars()` for the inverse method of converting a Polars ``DataFrame`` object into to a ``SQLDataModel``.
            - SQLDataModel uses different data types than those used in ``polars``, see :meth:`SQLDataModel.set_column_dtypes()` for more information about casting rules.
            - Polars does not really have a concept of an index column, therefore when using ``index=True``, the SQLDataModel index is just an additional column in the returned DataFrame object.

        Changelog:
            - Version 1.1.0 (2024-10-22):
                - Added ``orient = 'row'`` argument to explicitly set data orientation when constructing dataframe.
            - Version 0.3.8 (2024-04-12):
                - New method.            
        """
        if not optionals._has_pl:
            msg = ErrorFormat(f"""ModuleNotFoundError: required package not found, polars must be installed in order to use `.to_polars()` method""")
            raise ModuleNotFoundError(msg)
        _pl = optionals._get_pl()
        data = self.data(index=index, include_headers=include_headers)
        return _pl.DataFrame(data=data[1:] if include_headers else data,schema=data[0] if include_headers else None, orient='row')

    def to_pyarrow(self, index:bool=False) -> '_pa.Table':
        """
        Returns the current ``SQLDataModel`` in Apache Arrow columnar format as a ``pyarrow.Table``.

        Parameters:
            ``index`` (bool, optional): Specifies whether to include the index of the SQLDataModel in the resulting Table. Default is to False.

        Raises:
            ``ModuleNotFoundError``: If the required package ``pyarrow`` is not installed.
            ``Exception``: If any unexpected exception occurs during the pyarrow conversion process.

        Returns:
            ``pyarrow.Table``: A table representing the current ``SQLDataModel`` in Apache Arrow columnar format.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Grade']
            data = [('Alice', 25, 3.8), ('Bob', 30, 3.9), ('Charlie', 35, 3.2)]
            
            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Create the pyarrow table
            table = df.to_pyarrow()

            # View result
            print(table)        

        This will output the ``pyarrow`` object details:

        ```shell
            pyarrow.Table
            Name: string
            Age: int64
            Grade: double
            ----
            Name: [["Alice","Bob","Charlie"]]
            Age: [[25,30,35]]
            Grade: [[3.8,3.9,3.2]]
        ```

        Note:
            - Unmodified python types will follow conversion and casting rules specified in ``pyarrow`` implementation, for the modified ``date`` and ``datetime`` types, ``date32[day]`` and ``timestamp[us]`` will be used, respectively.

        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.2.3 (2024-03-28):
                - New method.
        """
        if not optionals._has_pa:
            msg = ErrorFormat(f"ModuleNotFoundError: required package not found, pyarrow must be installed in order to use `.to_pyarrow()` method")
            raise ModuleNotFoundError(msg)    
        _pa, _pq = optionals._get_pa_pq()    
        try:
            table = _pa.Table.from_pydict(self.to_dict(orient='columns', index=index))
        except Exception as e:
            msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to convert to pyarrow format")
            raise type(e)(msg).with_traceback(e.__traceback__) from None        
        return table

    def to_string(self, index:bool=None, display_max_rows:int=None, display_max_width:int=None, min_column_width:int=None, max_column_width:int=None, float_precision:int=None, vertical_ellipses:str='⠒⠂', horizontal_ellipses:str='⠤⠄',  display_dimensions:bool=False, index_rep:str=None, column_alignment:Literal['dynamic', 'left', 'center', 'right']=None, table_style:Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple']=None) -> str:
        """
        Generate a tabular representation of the model based on custom parameters and bounds.
        
        Parameters:
            ``index`` (bool, optional): Whether to include the index column in the output. Default is None, using :py:attr:`SQLDataModel.display_index` value.
            ``display_max_rows`` (int, optional): Maximum number of rows to display. Default is None, using :py:attr:`SQLDataModel.display_max_rows` value.
            ``display_max_width`` (int, optional): Maximum character width of the output table before horizontal truncation occurs. 
                Default is None, generating a full width representation.
            ``min_column_width`` (int, optional): Minimum width of columns. Default is None, using :py:attr:`SQLDataModel.min_column_width` value.
            ``max_column_width`` (int, optional): Maximum width of columns. Default is None, using :py:attr:`SQLDataModel.max_column_width` value with a floor value of 2.
            ``float_precision`` (int, optional): Precision for displaying float values. Default is None, using :py:attr:`SQLDataModel.display_float_precision` value.
            ``vertical_ellipses`` (str, optional): Characters to represent row truncation when vertical truncation is required. Default is ``'⠒⠂'``.
            ``horizontal_ellipses`` (str, optional): Characters to represent column truncation when horizontal truncation is required. Default is ``'⠤⠄'``.
            ``display_dimensions`` (bool, optional): Whether to display the dimensions of the table. Defaults to False.
            ``index_rep`` (str, optional): String representation for the index. Default is None, using a single whitespace character to represent the index column.
                Only used when generating index column, otherwise ignored when ``index = False``.
            ``column_alignment`` (Literal['dynamic', 'left', 'center', 'right'], optional): Alignment for columns. Default is None, using :py:attr:`SQLDataModel.column_alignment` value.
            ``table_style`` (Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple'], optional): Table style. Default is None, using :py:attr:`SQLDataModel.table_style` value.
        
        Returns:
            ``str``: A string representing the tabular output of the model with the restrictions and styling applied.
        
        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Gender', 'City']
            data = [
                ('Alice', 38, 'Female', 'Milwaukee'),
                ('Sarah', None, 'Female', 'Houston'),
                ('Michael', 42, 'Male', 'Atlanta'),
                ('John', None, 'Male', 'Boston'),
                ('Bobby', 25, 'Male', 'Chicago'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Generate a Markdown style representation using 'ID' to represent the index
            markdown_repr = df.to_string(table_style='markdown', index_rep='ID')

        This will generate a 'Markdown' styled representation:

        ```text
            | ID | Name    | Age | Gender | City      |
            |----|---------|-----|--------|-----------|
            |  0 | Alice   |  38 | Female | Milwaukee |
            |  1 | Sarah   |     | Female | Houston   |
            |  2 | Michael |  42 | Male   | Atlanta   |
            |  3 | John    |     | Male   | Boston    |
            |  4 | Bobby   |  25 | Male   | Chicago   |
        ```

        Vertical and horizontal limits can also be applied:

        ```python
            # Set vertical and horizontal limits with custom styling
            truncated_repr = df.to_string(
                table_style='polars', 
                display_max_rows=4, 
                display_max_width=36,
                horizontal_ellipses='..',
                vertical_ellipses='...',
                index=False
            )

            # View output
            print(truncated_repr)
        ```

        This will output a vertically and horizontally truncated representation that fits within the bounds provided:

        ```text
            ┌───────┬─────┬────┬───────────┐
            │ Name  ┆ Age ┆ .. ┆ City      │
            ╞═══════╪═════╪════╪═══════════╡
            │ Alice ┆  38 ┆ .. ┆ Milwaukee │
            │ Sarah ┆     ┆ .. ┆ Houston   │
            │  ...  ┆ ... ┆ .. ┆    ...    │
            │ John  ┆     ┆ .. ┆ Boston    │
            │ Bobby ┆  25 ┆ .. ┆ Chicago   │
            └───────┴─────┴────┴───────────┘
        ```

        Note:
            - Table styles reflect style similarity only, format specifc methods should be used for generating complete and valid output.
            - Vertical truncation characters are applied to column wide truncation and horizontal truncation characters are applied at row and cell level.
            - When a discrepancy exists between minimum and maximum column widths, conflict is resolved by setting max width equal to ``max(min_column_width, max_column_width)``.
            - See :meth:`SQLDataModel.to_text()` for writing textual representation directly to '.txt' files.
            - See :meth:`SQLDataModel.set_table_style()` for available style options and output examples.

        Changelog:
            - Version 0.11.0 (2024-07-05):
                - New method.
        """
        if (column_alignment is not None) and (column_alignment not in ('dynamic', 'left', 'center', 'right')):
            msg = ErrorFormat(f"ValueError: invalid value '{column_alignment}', argument for `column_alignment` must be one of 'dynamic', 'left', 'center', 'right' representing column alignment for text output")
            raise ValueError(msg)
        if (table_style is not None) and (table_style not in ('ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple')):
            msg = ErrorFormat(f"ValueError: invalid value '{table_style}', argument for `table_style` must be one of 'ascii', 'bare', 'dash', 'default', 'double', 'list', 'markdown', 'outline', 'pandas', 'polars', 'postgresql' or 'round'")
            raise ValueError(msg)
        display_index = self.display_index if index is None else index
        display_max_rows = display_max_rows if display_max_rows is not None else self.display_max_rows if self.display_max_rows is not None else self.row_count
        min_column_width = self.min_column_width  if min_column_width is None else min_column_width
        max_column_width = self.max_column_width if max_column_width is None else max_column_width
        max_column_width = max(2, min_column_width, max_column_width) # Guard against min and max descrepancies, resetting max to use larger of the two with a floor width of 2
        display_float_precision = self.display_float_precision if float_precision is None else float_precision
        column_alignment = column_alignment if column_alignment is not None else self.column_alignment
        column_alignment = None if column_alignment == 'dynamic' else '<' if column_alignment == 'left' else '^' if column_alignment == 'center' else '>'
        index_rep = index_rep if index_rep else ' ' # Default index representation

        vertical_truncation_required = True if display_max_rows < self.row_count else False
        split_row = display_max_rows // 2 if vertical_truncation_required else None

        header_py_dtype_dict = self.dtypes
        header_length_dict = self._calculate_col_widths(
            index=display_index, 
            min_column_width=min_column_width, 
            max_column_width=max_column_width, 
            float_precision=display_float_precision,
            split_row=split_row,
            index_rep=index_rep
        )
        display_headers = list(header_length_dict.keys())
        header_align_dict = {col:column_alignment for col in display_headers} if column_alignment is not None else {col:self.header_master[col][-1] for col in display_headers}

        table_style = self.table_style if table_style is None else table_style
        table_format = self._generate_table_style(style=table_style)
        top_lh, top_hbar, top_sep, top_rh = table_format[0]
        mid_lh, mid_hbar, mid_sep, mid_rh = table_format[1]
        row_lh, row_sep, row_rh = table_format[2]
        low_lh, low_hbar, low_sep, low_rh = table_format[3]
        row_lh_width, row_sep_width, row_rh_width = len(row_lh), len(row_sep), len(row_rh)

        horizontal_sep_marker = """__horiz_sep__""" # never displayed
        horizontal_ellipses_width = len(horizontal_ellipses)
        # Guard against absurd horizontal ellipses chars that exceed max col width, truncating ellipses chars instead
        if horizontal_ellipses_width > max_column_width:
            horizontal_ellipses = horizontal_ellipses[:max_column_width]
            horizontal_ellipses_width = len(horizontal_ellipses)

        total_required_width = row_lh_width + sum((row_sep_width + length) for length in header_length_dict.values()) + row_rh_width - row_sep_width
        horizontal_truncation_required = False if display_max_width is None or (display_max_width > total_required_width) else True        

        # Horizontal truncation required, expect to really only be used by repr method itself
        if horizontal_truncation_required:
            display_max_width -= horizontal_ellipses_width
            horiz_max_width = row_lh_width + row_rh_width
            lh_headers, rh_headers = [], []
            header_length_dict.update({horizontal_sep_marker:(horizontal_ellipses_width)}) # required to backport element and width for ensuing string formatting
            for i in range(len(display_headers) // 2):
                lh_header, rh_header = display_headers[i], display_headers[-(i + 1)]
                lh_width, rh_width = (header_length_dict[lh_header] + 3), (header_length_dict[rh_header] + 3)
                if horiz_max_width < display_max_width:                
                    horiz_max_width += lh_width
                    if horiz_max_width > display_max_width:
                        break
                    lh_headers.append(lh_header)
                    horiz_max_width += rh_width
                    if horiz_max_width > display_max_width:
                        break                    
                    rh_headers.append(rh_header)
            display_headers = [*lh_headers,horizontal_sep_marker,*rh_headers[::-1]]        
        table_repr = """""" # big things...
        table_bare_newline = """\n"""
        row_sep_concat = f"""|| '{row_sep}' ||"""
        fetch_idx = utils.sqlite_printf_format(self.sql_idx,"index",header_length_dict[self.sql_idx]) + row_sep_concat if display_index else ""
        header_fmt_str = row_sep_concat.join([f"""{utils.sqlite_printf_format(col,header_py_dtype_dict[col],header_length_dict[col],self.display_float_precision,alignment=column_alignment,escape_newline=True,truncation_chars=horizontal_ellipses)}""" if col != horizontal_sep_marker else f"""{utils.sqlite_printf_format(horizontal_ellipses,'custom',horizontal_ellipses_width,self.display_float_precision,alignment=column_alignment)}""" for col in display_headers if col != self.sql_idx])
        if vertical_truncation_required:
            vertical_sep_fmt_str = f'''{row_lh}{row_sep.join([f"""{vertical_ellipses:^{max(0,header_length_dict[col]+1)}}"""[:header_length_dict[col]] for col in display_headers])}{row_rh}{table_bare_newline}'''
            fetch_fmt_stmt = f"""
            with "_repr" as (
                select "{self.sql_idx}" as "_row" from "{self.sql_model}" where "{self.sql_idx}" in 
                    (select "{self.sql_idx}" from "{self.sql_model}" order by "{self.sql_idx}" asc limit {split_row}+1)
                        or "{self.sql_idx}" in
                    (select "{self.sql_idx}" from "{self.sql_model}" order by "{self.sql_idx}" desc limit {split_row})
                order by "{self.sql_idx}" asc limit {display_max_rows}+1)
                ,"_trigger" as (select "{self.sql_idx}" as "_sep" from "{self.sql_model}" order by "{self.sql_idx}" asc limit 1 offset {split_row})
            select CASE WHEN "{self.sql_idx}" <> (select "_sep" from "_trigger") THEN "_full_row" 
            ELSE '{vertical_sep_fmt_str}' 
            END from (select "{self.sql_idx}",'{row_lh}' || {fetch_idx}{header_fmt_str}||'{row_rh}{table_bare_newline}' as "_full_row" from "{self.sql_model}" where "{self.sql_idx}" in (select "_row" from "_repr") order by "{self.sql_idx}" asc)"""
        else:
            fetch_fmt_stmt = f"""select '{row_lh}' || {fetch_idx}{header_fmt_str}||'{row_rh}{table_bare_newline}' as "_full_row" from "{self.sql_model}" order by "{self.sql_idx}" asc limit {display_max_rows}"""

        formatted_response = self.sql_db_conn.execute(fetch_fmt_stmt)
        if table_style == 'latex': # Exception for LaTeX style to wrap headers in curly braces: {header}
            formatted_headers = [(f"""{f'{{{col}}}':{header_align_dict[col]}{header_length_dict[col]}}""" if len(col) <= header_length_dict[col] else f"""{f'{{{col}}}'[:(header_length_dict[col]-horizontal_ellipses_width)]}{horizontal_ellipses}""") if col not in (self.sql_idx, horizontal_sep_marker) else f"""{f'{{{index_rep}}}':>{header_length_dict[col]}}""" if col != horizontal_sep_marker else f"""{horizontal_ellipses:>{horizontal_ellipses_width}}""" for col in display_headers]
        else:
            formatted_headers = [(f"""{col:{header_align_dict[col]}{header_length_dict[col]}}""" if len(col) <= header_length_dict[col] else f"""{col[:(header_length_dict[col]-horizontal_ellipses_width)]}{horizontal_ellipses}""") if col not in (self.sql_idx, horizontal_sep_marker) else f"""{index_rep:>{header_length_dict[col]}}""" if col != horizontal_sep_marker else f"""{horizontal_ellipses:>{horizontal_ellipses_width}}""" for col in display_headers]
        col_lengths = [header_length_dict[col] for col in display_headers]
        table_top_bar = "".join([top_lh, top_sep.join([top_hbar * length for length in col_lengths]), top_rh, table_bare_newline])
        table_top_bar = table_top_bar if len(table_top_bar.strip()) >=1 else """"""
        table_repr = "".join([table_repr, table_top_bar])
        table_repr = "".join([table_repr, row_lh, row_sep.join(formatted_headers), row_rh, table_bare_newline])
        if table_style == 'markdown': # Exception for Markdown column alignment
            table_mid_bar = "".join(["|", "|".join([f"""{':' if header_align_dict[col] in ('^','<') else mid_hbar}{mid_hbar*header_length_dict[col]}{':' if header_align_dict[col] in ('^','>') else mid_hbar}""" for col in display_headers]), "|", table_bare_newline])
        else:
            table_mid_bar = "".join([mid_lh, mid_sep.join([mid_hbar * length for length in col_lengths]), mid_rh, table_bare_newline])
        table_mid_bar = table_mid_bar if len(table_mid_bar.strip()) >=1 else """"""
        table_repr = "".join([table_repr, table_mid_bar])
        table_repr = "".join([table_repr,*[row[0] for row in formatted_response]])
        table_low_bar = "".join([low_lh, low_sep.join([low_hbar * length for length in col_lengths]), low_rh, table_bare_newline])
        table_low_bar = table_low_bar if len(table_low_bar.strip()) >=1 else """"""
        table_repr = "".join([table_repr, table_low_bar])
        table_caption = f"""[{self.row_count} rows x {self.column_count} columns]""" if display_dimensions else """"""
        table_repr = "".join([table_repr, table_caption]).rstrip()
        return table_repr 

    def to_sql(self, table:str, con:sqlite3.Connection|Any, *, schema:str=None, if_exists:Literal['fail','replace','append']='fail', index:bool=True, primary_key:str|int=None) -> None:
        """
        Insert the ``SQLDataModel`` into the specified table using the provided database connection.
        
        Supported Connection APIs:
            - SQLite using ``sqlite3`` or url with format ``'file:///path/to/database.db'``
            - PostgreSQL using ``psycopg2`` or url with format ``'postgresql://user:pass@hostname:port/db'``
            - SQL Server ODBC using ``pyodbc`` or url with format ``'mssql://user:pass@hostname:port/db'``
            - Oracle using ``cx_Oracle`` or url with format ``'oracle://user:pass@hostname:port/db'``
            - Teradata using ``teradatasql`` or url with format ``'teradata://user:pass@hostname:port/db'``

        Parameters:
            ``table`` (str): The name of the table where data will be inserted.
            ``con`` (sqlite3.Connection | Any): The database connection object or connection url. Supported connection APIs are ``sqlite3``, ``psycopg2``, ``pyodbc``, ``cx_Oracle``, ``teradatasql``
            ``schema`` (str, optional): The schema to use for PostgreSQL and ODBC SQL Server connections, ignored otherwise. Default is None.
            ``if_exists`` (Literal['fail', 'replace', 'append'], optional): Action to take if the table already exists. If ``fail`` an error is raised if table exists and no inserts occur. If ``replace`` any existing table is dropped prior to inserts. If ``append`` existing table is appended to by subsequent inserts.
            ``index`` (bool, optional): If the model index should be included in the target table. Default is True.
            ``primary_key`` (str | int, optional): Column name or index to use as table primary key. Default is None, using the index column as the primary key when ``index=True``.

        Raises:
            ``SQLProgrammingError``: If an error occurs during cursor accessing, table creation or data insertion into the database.
            ``ModuleNotFoundError``: If ``con`` is provided as a connection url and the specified scheme driver module is not found.
            ``ValueError``: If specified ``table`` already exists when using ``if_exists='fail'`` or if ``con`` is not one of the currently supported connection modules.
            ``IndexError``: If ``primary_key`` is provided as an ``int`` representing a column index but is out of range of the current model :py:attr:`SQLDataModel.column_count`.
            ``TypeError``: If ``primary_key`` argument provided is not of type 'str' or 'int' representing a valid column name or index to use as the primary key column for the target table.

        Returns:
            ``None``

        Example::

            import sqlite3
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Grade']
            data = [('Alice', 25, 3.8), ('Bob', 30, 3.9), ('Charlie', 35, 3.2), ('David', 28, 3.4)]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Create connection object
            sqlite_db_conn = sqlite3.connect('students.db')

            # Basic usage, creating a new table
            df.to_sql('users', sqlite_db_conn)

        This will create a new table ``users``, or fail if one already exists:

        ```text
            sqlite> select * from users;

            idx  Name     Age  Grade
            ---  -------  ---  -----
            0    Alice    25   3.8
            1    Bob      30   3.9
            2    Charlie  35   3.2
            3    David    28   3.4
        ```

        Connect to PostgreSQL, SQL Server, Oracle or Teradata:

        ```python
            import psycopg2
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Grade']
            data = [('Alice', 25, 3.8), ('Bob', 30, 3.9), ('Charlie', 35, 3.2), ('David', 28, 3.4)]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Setup the connection, whether using psycopg2 or other supported modules like pyodbc
            con = psycopg2.connect(...)

            # Create or replace existing table in database
            df.to_sql('users', con, if_exists='replace', index=False)
        ```

        This will result in a new table ``users`` in our PostgreSQL database:

        ```text
            => select * from users;

            Name    | Age | Grade |
            --------+-----+-------+
            Alice   |  25 |   3.8 |
            Bob     |  30 |   3.9 |
            Charlie |  35 |   3.2 |
            David   |  28 |   3.4 |
        ```
        
        For SQL Server connections using ``pyodbc``, the example would be almost identical except for which ``con`` object we use:

        ```python
            import pyodbc

            # For SQL Server ODBC connections using pyodbc
            con = pyodbc.connect(...)
        ```
            
        The same is true for Oracle and other connections:

        ```python
            import cx_Oracle

            # For Oracle connections using cx_Oracle
            con = cx_Oracle.connect(...)
        ```
        
        Using a Primary Key

        ```python
            import sqldatamodel as sdm

            # Sample data
            headers = ['ID', 'User']
            data = [(1001, 'Alice'), (1002, 'Bob'), (1003, 'Charlie'), (1004, 'David')]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Create connection object
            sqlite_db_conn = sqlite3.connect('students.db')

            # Create the table using the 'ID' column as the primary key
            df.to_sql('users', sqlite_db_conn, if_exists='replace', index=False, primary_key='ID')
        ```

        This will create a ``users`` table with the schema: 

        ```text
            sqlite> .schema users

            CREATE TABLE "users" ( "ID" INTEGER PRIMARY KEY,  "User" TEXT);
        ```

        With the ``ID`` column as its primary key:

        ```text
            sqlite> select * from users;

            ID    User   
            ----  -------
            1001  Alice
            1002  Bob
            1003  Charlie
            1004  David
        ```

        If table creation is necessary, column types will be mapped according to the destination database by the following conversion:
        
        ```text
            ┌─────────────────┬─────────┬─────────┬────────┬─────────┬────────────────┬──────┬───────────┐
            │ Database \\ Type │ NULL    │ INTEGER │ REAL   │ TEXT    │ BLOB           │ DATE │ TIMESTAMP │
            ├─────────────────┼─────────┼─────────┼────────┼─────────┼────────────────┼──────┼───────────┤
            │ PostgreSQL      │ UNKNOWN │ INTEGER │ FLOAT  │ TEXT    │ BYTEA          │ DATE │ TIMESTAMP │
            │ SQL Server ODBC │ UNKNOWN │ INTEGER │ FLOAT  │ TEXT    │ VARBINARY(MAX) │ DATE │ DATETIME  │
            │ Oracle          │ UNKNOWN │ NUMBER  │ NUMBER │ VARCHAR │ BLOB           │ DATE │ DATETIME  │
            │ Teradata        │ UNKNOWN │ INTEGER │ FLOAT  │ VARCHAR │ BYTE           │ DATE │ DATETIME  │
            │ SQLite          │ NULL    │ INTEGER │ REAL   │ TEXT    │ BLOB           │ DATE │ TIMESTAMP │
            └─────────────────┴─────────┴─────────┴────────┴─────────┴────────────────┴──────┴───────────┘
            [5 rows x 8 columns]
        ```
        
        Note:
            - When providing a ``primary_key`` column it will be assumed unique and the model will not perform any unique-ness constraints.
            - When ``con`` is provided as a string a connection will be attempted using :meth:`utils._create_connection()` if the path does not exist, otherwise a ``sqlite3`` local connection will be attempted.
            - When ``con`` is provided as an object a connection is assumed to be open and valid, if a cursor cannot be created from the object an exception will be raised. 
            - Connections with write access can be used in the :meth:`SQLDataModel.to_sql()` method for writing to the same connection types, be careful.
            - ValueError will be raised if ``table`` already exists, use ``if_exists = 'replace'`` or ``if_exists = 'append'`` to instead replace or append to the table.
            - See relevant module documentation for additional details or information pertaining to specific database or connection dialect being used.
            - See related :meth:`SQLDataModel.from_sql()` for creating ``SQLDataModel`` from existing SQL database connections.
            - See utility methods :meth:`utils._parse_connection_url()` and :meth:`utils._create_connection()` for implementation on creating database connections from urls.

        Changelog:
            - Version 0.9.1 (2024-06-27):
                - Modified handling of ``con`` parameter to allow database connection url to also be provided as ``'scheme://user:pass@host:port/db'``
            - Version 0.8.2 (2024-06-24):
                - Modified handling of ``con`` parameter to allow providing SQLite database filepath directly as string to instantiate connection.
            - Version 0.3.0 (2024-03-31):
                - Renamed arguments ``extern_con``: ``con``, ``replace_existing``: ``if_exists``, ``include_index``: ``index``.
                - Added ``primary_key`` argument for specifying a primary key column for table schema.
                - Added ``schema`` argument for specifying a target schema for the table.        
        """    
        if isinstance(con, str):
            if os.path.exists(con): # Connection provided as SQLite database filepath
                try:
                    con = sqlite3.connect(con)
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open database connection '{con}'")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None     
            else: # Connection provided as url with format 'scheme://user:pass@host:port/path'
                con = utils._create_connection(con)
        try:
            ext_c = con.cursor()
        except Exception as e:
            msg = ErrorFormat(f"""SQLProgrammingError: provided SQL connection is not open, please reopen the database connection or resolve "{e}" """)
            raise SQLProgrammingError(msg) from None
        if if_exists not in ('fail', 'replace', 'append'):
            msg = ErrorFormat(f"ValueError: invalid value '{if_exists}', argument for `if_exists` must be one of 'fail', 'replace' or 'append' to determine appropriate action if table already exists")
            raise ValueError(msg)
        if primary_key is not None and not isinstance(primary_key, (str,int)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(primary_key).__name__}', argument for `primary_key` must be of type 'str' or 'int' representing a column name or index to use as the table primary key")
            raise TypeError(msg)
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index))
        model_data = [x for x in res.fetchall()]
        model_headers = [x[0] for x in res.description]    
        if isinstance(primary_key, int):
            try:
                primary_key = model_headers[primary_key]
            except IndexError:
                msg = ErrorFormat(f"IndexError: invalid column index '{primary_key}', index for `primary_key` is outside of current model range '0:{self.column_count}', use `get_headers()` to view current valid arguments")
                raise IndexError(msg) from None
        if isinstance(primary_key, str):
            if primary_key not in model_headers:
                msg = ErrorFormat(f"ValueError: column not found '{primary_key}', a valid column is required when providing a `primary_key` argument, use `get_headers()` to view current valid columns")
                raise ValueError(msg)
        primary_key = self.sql_idx if primary_key is None else primary_key
        dialect = type(con).__module__.split('.')[0].lower() # 'sqlite3', 'psycopg2', 'pyodbc', 'cx_oracle', 'teradatasql'
        dyn_table_label = f'"{table}"' if (schema is None or dialect not in ('psycopg2', 'pyodbc')) else f'"{schema}"."{table}"'
        check_exists = True if if_exists == 'fail' else False
        if dialect == 'sqlite3':
            sdm_dialect_map = {'NULL': 'NULL','INTEGER': 'INTEGER','REAL': 'REAL','TEXT': 'TEXT','BLOB': 'BLOB', 'DATE': 'DATE', 'TIMESTAMP': 'TIMESTAMP'} # sqlite / else
            if check_exists:
                check_stmt, check_params = "SELECT name FROM sqlite_master WHERE type='table' AND lower(name)=lower(?)", (table,)
        elif dialect == 'psycopg2':
            sdm_dialect_map = {'NULL': 'UNKNOWN','INTEGER': 'INTEGER','REAL': 'FLOAT','TEXT': 'TEXT','BLOB': 'BYTEA', 'DATE': 'DATE', 'TIMESTAMP': 'TIMESTAMP'} # pgsql
            if check_exists:
                check_stmt = "SELECT table_name from information_schema.tables WHERE lower(table_name) = lower(%s)" if schema is None else "SELECT table_name FROM information_schema.tables WHERE lower(table_schema) = lower(%s) AND lower(table_name) = lower(%s)"
                check_params = (table,) if schema is None else (schema, table)
        elif dialect == 'pyodbc':
            sdm_dialect_map = {'NULL': 'UNKNOWN','INTEGER': 'INTEGER','REAL': 'FLOAT','TEXT': 'TEXT','BLOB': 'VARBINARY(MAX)', 'DATE': 'DATE', 'TIMESTAMP': 'DATETIME'} # odbc
            if check_exists:
                check_stmt = "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE lower(TABLE_NAME) = lower(?)" if schema is None else "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE lower(TABLE_SCHEMA) = lower(?) AND lower(TABLE_NAME) = lower(?)"
                check_params = (table,) if schema is None else (schema, table)
        elif dialect == 'cx_oracle':
            sdm_dialect_map = {'NULL': 'UNKNOWN','INTEGER': 'NUMBER','REAL': 'NUMBER','TEXT': 'VARCHAR','BLOB': 'BLOB', 'DATE': 'DATE', 'TIMESTAMP': 'DATETIME'} # oracle
            if check_exists:
                check_stmt, check_params = "SELECT TABLE_NAME FROM all_tables WHERE lower(TABLE_NAME)= lower(:1)", (table,)
        elif dialect == 'teradatasql':
            sdm_dialect_map = {'NULL': 'UNKNOWN','INTEGER': 'INTEGER','REAL': 'FLOAT','TEXT': 'VARCHAR','BLOB': 'BYTE', 'DATE': 'DATE', 'TIMESTAMP': 'DATETIME'} # teradata
            if check_exists:
                check_stmt, check_params = "SELECT TableName FROM DBC.Tables WHERE lower(TableName) = lower(?)", (table,)
        else:
            msg = ErrorFormat(f"ValueError: unsupported connection '{dialect}', currently only 'sqlite3', 'psycopg2', 'pyodbc', 'cx_oracle' and 'teradatasql' connections are supported")
            raise ValueError(msg)
        if if_exists == 'fail':
            ext_c.execute(check_stmt, check_params)
            if ext_c.fetchone() is not None:
                msg = ErrorFormat(f"ValueError: table '{table}' already exists, use `if_exists = 'replace'` or `if_exists = 'append'` to overwrite or modify existing table instead")
                raise ValueError(msg)
        elif if_exists == 'replace':
            ext_c.execute(f"""drop table if exists {dyn_table_label}""")
            con.commit()
        created_header_dict = {col:f"{sdm_dialect_map[self.header_master[col][0]]}" if col != primary_key else f"{sdm_dialect_map[self.header_master[col][0]]} PRIMARY KEY" for col in model_headers}    
        sql_dtypes_stmt = ", ".join(f""" "{header}" {created_header_dict[header]}""" for header in model_headers) # generates sql create table statement using type mapping dict
        sql_create_stmt = f"""create table if not exists {dyn_table_label} ({sql_dtypes_stmt})""" if dialect != 'pyodbc' else f"""if object_id(N'{table}', N'U') is null create table "{table}" ({sql_dtypes_stmt})""" if schema is None else f"""if object_id(N'{schema}.{table}', N'U') is null create table {dyn_table_label} ({sql_dtypes_stmt})"""
        dyn_bind = '?' if dialect != 'psycopg2' else '%s'
        val_params = ','.join([dyn_bind for _ in model_headers]) if dialect != 'cx_oracle' else ','.join([f":{i+1}" for i in range(len(model_headers))])
        sql_insert_stmt = f"""insert into {dyn_table_label} ({','.join([f'"{col}"' for col in model_headers])}) values ({val_params})""" # changed to string formatter
        try:
            ext_c.execute(sql_create_stmt)
            con.commit()
        except Exception as e:
            con.rollback()
            msg = ErrorFormat(f"SQLProgrammingError: {e} encountered when trying to create and define table schema")
            raise SQLProgrammingError (msg) from None             
        try:
            ext_c.executemany(sql_insert_stmt,model_data)
            con.commit()
        except Exception as e:
            con.rollback()
            msg = ErrorFormat(f"SQLProgrammingError: {e} encountered when trying to insert model data into provided connection")
            raise SQLProgrammingError (msg) from None   
        return

    def to_text(self, filename:str=None, index:bool=None, min_column_width:int=None, max_column_width:int=None, float_precision:int=None, horizontal_ellipses:str='⠤⠄', index_rep:str=' ', display_dimensions:bool=False, column_alignment:Literal['dynamic', 'left', 'center', 'right']=None, table_style:Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple']=None) -> str|None:
        """
        Returns a textual representation of the current ``SQLDataModel`` as a string literal or by writing to file if a ``filename`` is provided.

        Parameters:
            ``filename`` (str, optional): The name of the file to write the text content. If provided, writes the text to the specified file. Default is None.
            ``index`` (bool, optional): Whether to include the index column in the text output. Default is value set on :py:attr:`SQLDataModel.display_index`.
            ``min_column_width`` (int, optional): The minimum column width for table cells. Default is value set on :py:attr:`SQLDataModel.min_column_width`.
            ``max_column_width`` (int, optional): The maximum column width for table cells. Default is value set on :py:attr:`SQLDataModel.max_column_width`.
            ``float_precision`` (int, optional): The precision for floating-point values. Default is value set on :py:attr:`SQLDataModel.display_float_precision`.
            ``horizontal_ellipses`` (str, optional): Characters to truncate column and cell values when horizontal truncation is required. Default is ``'⠤⠄'``.
            ``index_rep`` (str, optional): String representation for the index. Default is None, using a single whitespace character to represent the index column.
                Only used when generating index column, otherwise ignored when ``index = False``.
            ``display_dimensions`` (bool, optional): Whether to include the model dimensions ``[N rows x N cols]`` in the text output. Default is False.
            ``column_alignment`` (Literal['dynamic', 'left', 'center', 'right'], optional): Column alignment. Default is value at :py:attr:`SQLDataModel.column_alignment`.
                ``'dynamic'``: Dynamically aligns column content, right for numeric types and left for remaining types.
                ``'left'``: Left-aligns all column content.
                ``'center'``: Center-aligns all column content preferring left on uneven splits.
                ``'right'``: Right-aligns all column content.                
            ``table_style`` (Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple'], optional): 
                The table styling to use. Default is value set on :py:attr:`SQLDataModel.table_style`.

        Raises:
            ``TypeError``: If arguments are provided but are not the correct types: ``filename`` (str), ``index`` (bool), ``min_column_width`` (int), ``max_column_width`` (int), ``float_precision`` (int).
            ``ValueError``: If the ``column_alignment`` argument is provided and is not one of 'dynamic', 'left', 'center', or 'right'.
            ``Exception``: If there is an OS related error encountered when opening or writing to the provided ``filename``.

        Returns:
            ``str`` or ``None``: If ``filename`` is None, returns the textual representation as a string. If ``filename`` is provided, writes the textual representation to the specified file and returns None.

        Examples:

        Returning Text Literal
        ----------------------

        ```python
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Michael', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Generate text table literal
            text_table = df.to_text()

            # View output
            print(text_table)
        ```

        This will output:

        ```shell
            ┌─────────┬──────┬────────┐
            │ Name    │  Age │ Height │
            ├─────────┼──────┼────────┤
            │ John    │   30 │  175.3 │
            │ Alice   │   28 │  162.0 │
            │ Michael │   35 │  185.8 │
            └─────────┴──────┴────────┘
        ```

        Write to File
        -------------

        ```python            
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Michael', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data=data, headers=headers)

            # Write the output to the file, center-aligning all columns
            df.to_text(filename='Table.txt', column_alignment='center')        
        ```

        Contents of ``Table.txt``:
        
        ```shell            
            ┌───┬─────────┬──────┬────────┐
            │   │  Name   │ Age  │ Height │
            ├───┼─────────┼──────┼────────┤
            │ 0 │  John   │  30  │ 175.3  │
            │ 1 │  Alice  │  28  │ 162.0  │
            │ 2 │ Michael │  35  │ 185.8  │
            └───┴─────────┴──────┴────────┘
        ```
        
        Important:
            Unlike output from ``print(df)`` or other calls to :meth:`SQLDataModel.__repr__()`, the output from this method includes the full ``SQLDataModel`` and is not restricted by current terminal size or the value set at :py:attr:`SQLDataModel.display_max_rows`. As such, horizontal truncation only occurs on cell values as determined by ``max_column_width`` and no other horizontal or vertical table-wide truncation is performed.

        Note:
            - If ``filename`` is provided, the method writes the text to the specified file; otherwise, it returns the textual representation as a string.
            - If ``index`` is ``None``, the method uses the current value on :py:attr:`SQLDataModel.display_index`.
            - If ``min_column_width`` is ``None``, the method uses the current value on :py:attr:`SQLDataModel.min_column_width`.
            - If ``max_column_width`` is ``None``, the method uses the current value on :py:attr:`SQLDataModel.max_column_width`.
            - If ``float_precision`` is ``None``, the method uses the current value on :py:attr:`SQLDataModel.display_float_precision`.
            - If ``column_alignment`` is ``None``, the method uses the current value on :py:attr:`SQLDataModel.column_alignment`.
            - If ``table_style`` is ``None``, the method uses the current value on :py:attr:`SQLDataModel.table_style`.
            - See :meth:`SQLDataModel.set_table_style()` for modifying table format and available styles.
            - See :meth:`SQLDataModel.to_string()` for greater flexibility and control over generated string representations.

        Changelog:
            - Version 0.11.0 (2024-07-05):
                - Added ``horizontal_ellipses`` parameter to allow customizing truncation characters used when column or cell values exceed maximumn column widths.
                - Added ``index_rep`` parameter to allow customizing index column name with prior behavior set as default representation. Ignored when ``index = False``.
                - Modified to use :meth:`SQLDataModel.to_string()` instead of generating independently formatted repr for more consistency between tabular outputs.
            - Version 0.10.4 (2024-07-03):
                - Modified to escape newline characters through :meth:`utils.sqlite_printf_format()` to avoid wrapping table rows.
            - Version 0.9.3 (2024-06-28):
                - Added additional options 'rst-simple' and 'rst-grid' for ``table_style`` parameter. 
            - Version 0.3.10 (2024-04-16):
                - Added ``table_style`` parameter and updated output to reflect new formatting styles introduced in version 0.3.9.
                - Added ``display_dimensions`` parameter to allow toggling display of table dimensions in output.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
        """        
        if not isinstance(filename, str) and filename is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', expected `filename` to be of type 'str' representing a valid file path to write text")
            raise TypeError(msg)
        if (not isinstance(min_column_width, int) and (min_column_width is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(min_column_width).__name__}', expected `min_column_width` to be of type 'int' representing minimum column width for table cells")
            raise TypeError(msg)        
        if (not isinstance(max_column_width, int) and (max_column_width is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(max_column_width).__name__}', expected `max_column_width` to be of type 'int' representing maximum column width for table cells")
            raise TypeError(msg)   
        if (not isinstance(float_precision, int) and (float_precision is not None)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(float_precision).__name__}', expected `float_precision` to be of type 'int' representing precision to use for values of type 'float'")
            raise TypeError(msg)          
        if (column_alignment is not None) and (column_alignment not in ('dynamic', 'left', 'center', 'right')):
            msg = ErrorFormat(f"ValueError: invalid value '{column_alignment}', argument for `column_alignment` must be one of 'dynamic', 'left', 'center', 'right' representing column alignment for text output")
            raise ValueError(msg)
        if (table_style is not None) and (table_style not in ('ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple')):
            msg = ErrorFormat(f"ValueError: invalid value '{table_style}', argument for `table_style` must be one of 'ascii', 'bare', 'dash', 'default', 'double', 'latex', 'list', 'markdown', 'outline', 'pandas', 'polars', 'postgresql', 'round', 'rst-grid' or 'rst-simple'.")
            raise ValueError(msg)
        display_index = self.display_index if index is None else index
        min_column_width = self.min_column_width  if min_column_width is None else min_column_width
        max_column_width = self.max_column_width if max_column_width is None else max_column_width
        max_column_width = max_column_width if max_column_width >= 2 else 2 # minimum required width
        display_float_precision = self.display_float_precision if float_precision is None else float_precision
        column_alignment = self.column_alignment if column_alignment is None else column_alignment
        table_repr = self.to_string(
            index=display_index,
            display_max_rows=self.row_count,
            display_max_width=None,
            min_column_width=min_column_width,
            max_column_width=max_column_width,
            float_precision=display_float_precision,
            horizontal_ellipses = horizontal_ellipses,
            index_rep = index_rep,
            display_dimensions = display_dimensions,
            column_alignment=column_alignment,
            table_style=table_style    
        )
        if filename is not None:
            try:
                with open(filename, "w", encoding='utf-8') as f:
                    f.write(table_repr)
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to open and write text to '{filename}'")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
        else:
            return table_repr
        
    def to_local_db(self, filename:str) -> None:
        """
        Writes the ``SQLDataModel`` in-memory database to disk as a SQLite database file using the specified filename.

        Parameters:
            ``filename`` (str): The filename or filepath to use when writing the model to disk.

        Raises:
            ``TypeError``: If ``filename`` is provided and is not of type 'str' representing a valid sqlite database save path.
            ``sqlite3.Error``: If there is an issue with the SQLite database operations during backup.

        Returns:
            ``None``
            
        Example::

            import sqlite3
            import sqldatamodel as sdm

            # Sample data
            data = [('Alice', 20, 'F'), ('Billy', 25, 'M'), ('Chris', 30, 'M')]

            # Create the model
            df = sdm.SQLDataModel(data, headers=['Name','Age','Sex'])

            # Filename to use for database
            db_file = "model.db"

            # Write the in-memory database model to disk
            df.to_local_db(db_file)

            # Loading the model back from disk can now be done at anytime
            df = sdm.from_sql("sdm", sqlite3.connect(db_file))

            # View restored model
            print(df)

        This will output the model we originally created:

        ```text
            ┌───┬───────┬─────┬─────┐
            │   │ Name  │ Age │ Sex │
            ├───┼───────┼─────┼─────┤
            │ 0 │ Alice │  20 │ F   │
            │ 1 │ Billy │  25 │ M   │
            │ 2 │ Chris │  30 │ M   │
            └───┴───────┴─────┴─────┘
            [3 rows x 3 columns]
        ```

        Note:
            - Use any compatible SQL API to load the resulting database file or use :meth:`SQLDataModel.from_sql()` to reload it back into a ``SQLDataModel``.
            - Table name is determined by value at :py:attr:`SQLDataModel.sql_model` which is set to ``'sdm'`` by default, use :meth:`SQLDataModel.set_model_name()` to modify.

        Changelog:
            - Version 0.5.2 (2024-05-13):
                - Renamed ``db`` parameter to ``filename`` for package consistency and to avoid confusion between similarily named database objects.
                - Changed ``filename`` from keyword to positional argument making it a required parameter to avoid accidental overwriting.
            - Version 0.1.5 (2023-11-24):
                - New method.            
        """
        if not isinstance(filename, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(filename).__name__}', argument for `filename` must be of type 'str' representing a valid sqlite database save path")
            raise TypeError(msg)
        self.sql_db_conn.commit()
        with sqlite3.connect(filename) as target:
            self.sql_db_conn.backup(target)

##################################################################################################################
##################################### arithmetic & bitwise operators methods #####################################
##################################################################################################################

    def __bool__(self) -> bool:
        """
        Implements logical boolean operator for ``SQLDataModel`` using the current row count.

        Returns:
            ``bool``: True if :py:attr:`SQLDataModel.row_count` != 0, False otherwise.
        
        Example::

            import sqldatamodel as sdm

            # Create an empty model
            df = sdm.SQLDataModel(headers=['Stage', 'Match', 'Result'])

            # Use boolean method to avoid duplicating result
            if not df:
                df[0] = ['Group', 1, 'Scotland Win']
            else:
                print('Match result already stored')
        
        Note:
            - This method is equivalent to ``df.row_count != 0``
            - See :meth:`SQLDataModel.__eq__()` and related comparison methods for more details.

        Changelog:
            - Version 0.7.1 (2024-06-09):
                - New method.
        """
        return self.row_count != 0

    def __lt__(self, other) -> set[int]:
        """
        Implements the less than operator ``<`` for comparing ``SQLDataModel`` against ``other`` and performing the equivalent set operation against the model's current indicies.

        Parameters:
            ``other``: The ``SQLDataModel`` or scalar (``int``, ``str``, ``float``) to compare with.

        Returns:
            ``set[int]``: The set of row indicies resulting from the operation that satisfy the condition.

        Example::
            
            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers) 

            # Filter by 'Age' column
            df = df[df['Age'] < 40]
                    
            # View result
            print(df)

        This will output:

        ```shell
            ┌───┬───────┬────────┬──────┬─────────┬────────────┬────────┐
            │   │ First │ Last   │  Age │ Service │ Hired      │ Gender │
            ├───┼───────┼────────┼──────┼─────────┼────────────┼────────┤
            │ 0 │ John  │ Smith  │   27 │    1.22 │ 2023-02-01 │ Male   │
            │ 1 │ Kelly │ Lee    │   32 │    8.00 │ 2016-09-18 │ Female │
            │ 2 │ Mike  │ Harlin │   36 │    3.90 │ 2020-08-27 │ Male   │
            └───┴───────┴────────┴──────┴─────────┴────────────┴────────┘
            [3 rows x 6 columns]
        ```

        Note:
            - For scalar ``other`` (int, str, or float), compares each element with the scalar and returns the row indicies evaluating to ``True``.
            - For SQLDataModel ``other``, compares each element across X rows for Y columns for all (X_i, Y_j) in range of ``row_count`` and ``column_count`` and returns those row indicies evaluating to ``True``.
            - All the equality operations return a python ``set`` object containing the row indicies which were returned from the evaluation.
            - All operations on standard types like ``int``, ``float`` or ``str`` follow standard behavior and are not modified by performing the operations.
            - Operations can be chained using standard ``set`` operators like ``&`` and ``|`` to allow complex filtering, multiple operations require parenthesis.
            
        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """           
        self_data = self.data(strict_2d=True)
        i_dim, j_dim = len(self_data), len(self_data[0])
        if isinstance(other, SQLDataModel):
            other_data = other.data(strict_2d=True)
            return set(i for i in range(i_dim) if all(self_data[i][j] < other_data[i][j] for j in range(j_dim)))
        elif isinstance(other, (int,str,float,datetime.date)):
            return set(i for j in range(j_dim) for i in range(i_dim) if self_data[i][j] < other)
        else:
            return set()
    
    def __le__(self, other) -> set[int]:
        """
        Implements the less than or equal to operator ``<=`` for comparing ``SQLDataModel`` against ``other`` and performing the equivalent set operation against the model's current indicies.

        Parameters:
            ``other``: The ``SQLDataModel`` or scalar (``int``, ``str``, ``float``) to compare with.

        Returns:
            ``set[int]``: The set of row indicies resulting from the operation that satisfy the condition.

        Example::
        
            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers) 

            # Filter by 'Age' column
            df = df[df['Age'] <= 40]
                    
            # View result
            print(df)

        This will output:

        ```shell
            ┌───┬───────┬────────┬──────┬─────────┬────────────┬────────┐
            │   │ First │ Last   │  Age │ Service │ Hired      │ Gender │
            ├───┼───────┼────────┼──────┼─────────┼────────────┼────────┤
            │ 0 │ John  │ Smith  │   27 │    1.22 │ 2023-02-01 │ Male   │
            │ 1 │ Kelly │ Lee    │   32 │    8.00 │ 2016-09-18 │ Female │
            │ 2 │ Mike  │ Harlin │   36 │    3.90 │ 2020-08-27 │ Male   │
            └───┴───────┴────────┴──────┴─────────┴────────────┴────────┘
            [3 rows x 6 columns]
        ```

        Note:
            - For scalar ``other`` (int, str, or float), compares each element with the scalar and returns the row indicies evaluating to ``True``.
            - For SQLDataModel ``other``, compares each element across X rows for Y columns for all (X_i, Y_j) in range of ``row_count`` and ``column_count`` and returns those row indicies evaluating to ``True``.
            - All the equality operations return a python ``set`` object containing the row indicies which were returned from the evaluation.
            - All operations on standard types like ``int``, ``float`` or ``str`` follow standard behavior and are not modified by performing the operations.
            - Operations can be chained using standard ``set`` operators like ``&`` and ``|`` to allow complex filtering, multiple operations require parenthesis.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """          
        self_data = self.data(strict_2d=True)
        i_dim, j_dim = len(self_data), len(self_data[0])
        if isinstance(other, SQLDataModel):
            other_data = other.data(strict_2d=True)
            return set(i for i in range(i_dim) if all(self_data[i][j] <= other_data[i][j] for j in range(j_dim)))
        elif isinstance(other, (int,str,float,datetime.date)):
            return set(i for j in range(j_dim) for i in range(i_dim) if self_data[i][j] <= other)
        else:
            return set()
    
    def __eq__(self, other) -> set[int]:
        """
        Implements the is equal to operator ``==`` for comparing ``SQLDataModel`` against ``other`` and performing the equivalent set operation against the model's current indicies.

        Parameters:
            ``other``: The ``SQLDataModel`` or scalar (``int``, ``str``, ``float``) to compare with.

        Returns:
            ``set[int]``: The set of row indicies resulting from the operation that satisfy the condition.

        Example::

            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers) 

            # Filter by 'Gender' column
            df = df[df['Gender'] == 'Female']
                    
            # View result
            print(df)

        This will output:
        
        ```shell            
            ┌───┬───────┬──────┬──────┬─────────┬────────────┬────────┐
            │   │ First │ Last │  Age │ Service │ Hired      │ Gender │
            ├───┼───────┼──────┼──────┼─────────┼────────────┼────────┤
            │ 0 │ Kelly │ Lee  │   32 │    8.00 │ 2016-09-18 │ Female │
            │ 1 │ Sarah │ West │   51 │    0.70 │ 2023-10-01 │ Female │
            └───┴───────┴──────┴──────┴─────────┴────────────┴────────┘
            [2 rows x 6 columns]
        ```

        Note:
            - For scalar ``other`` (int, str, or float), compares each element with the scalar and returns the row indicies evaluating to ``True``.
            - For SQLDataModel ``other``, compares each element across X rows for Y columns for all (X_i, Y_j) in range of ``row_count`` and ``column_count`` and returns those row indicies evaluating to ``True``.
            - All the equality operations return a python ``set`` object containing the row indicies which were returned from the evaluation.
            - All operations on standard types like ``int``, ``float`` or ``str`` follow standard behavior and are not modified by performing the operations.
            - Operations can be chained using standard ``set`` operators like ``&`` and ``|`` to allow complex filtering, multiple operations require parenthesis.
        """        
        self_data = self.data(strict_2d=True)
        i_dim, j_dim = len(self_data), len(self_data[0])
        if isinstance(other, SQLDataModel):
            other_data = other.data(strict_2d=True)
            return set(i for i in range(i_dim) if all(self_data[i][j] == other_data[i][j] for j in range(j_dim)))
        elif isinstance(other, (int,str,float,datetime.date)):
            return set(i for j in range(j_dim) for i in range(i_dim) if self_data[i][j] == other)
        else:
            return set()

    def __ne__(self, other) -> set[int]:
        """
        Implements the not equal to operator ``!=`` for comparing ``SQLDataModel`` against ``other`` and performing the equivalent set operation against the model's current indicies.

        Parameters:
            ``other``: The ``SQLDataModel`` or scalar (``int``, ``str``, ``float``) to compare with.

        Returns:
            ``set[int]``: The set of row indicies resulting from the operation that satisfy the condition.

        Example::
        
            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers) 

            # Filter by 'First' column
            df = df[df['First'] != 'John']
                    
            # View result
            print(df)

        This will output:
        
        ```shell            
            ┌───┬───────┬─────────┬──────┬─────────┬────────────┬────────┐
            │   │ First │ Last    │  Age │ Service │ Hired      │ Gender │
            ├───┼───────┼─────────┼──────┼─────────┼────────────┼────────┤
            │ 0 │ Kelly │ Lee     │   32 │    8.00 │ 2016-09-18 │ Female │
            │ 1 │ Mike  │ Harlin  │   36 │    3.90 │ 2020-08-27 │ Male   │
            │ 2 │ Sarah │ West    │   51 │    0.70 │ 2023-10-01 │ Female │
            │ 3 │ Pat   │ Douglas │   42 │   11.50 │ 2015-11-06 │ Male   │
            └───┴───────┴─────────┴──────┴─────────┴────────────┴────────┘
            [4 rows x 6 columns]
        ```

        Note:
            - For scalar ``other`` (int, str, or float), compares each element with the scalar and returns the row indicies evaluating to ``True``.
            - For SQLDataModel ``other``, compares each element across X rows for Y columns for all (X_i, Y_j) in range of ``row_count`` and ``column_count`` and returns those row indicies evaluating to ``True``.
            - All the equality operations return a python ``set`` object containing the row indicies which were returned from the evaluation.
            - All operations on standard types like ``int``, ``float`` or ``str`` follow standard behavior and are not modified by performing the operations.
            - Operations can be chained using standard ``set`` operators like ``&`` and ``|`` to allow complex filtering, multiple operations require parenthesis.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """          
        self_data = self.data(strict_2d=True)
        i_dim, j_dim = len(self_data), len(self_data[0])
        if isinstance(other, SQLDataModel):
            other_data = other.data(strict_2d=True)
            return set(i for i in range(i_dim) if all(self_data[i][j] != other_data[i][j] for j in range(j_dim)))
        elif isinstance(other, (int,str,float,datetime.date)):
            return set(i for j in range(j_dim) for i in range(i_dim) if self_data[i][j] != other)
        else:
            return set()

    def __gt__(self, other) -> set[int]:
        """
        Implements the greater than operator ``>`` for comparing ``SQLDataModel`` against ``other`` and performing the equivalent set operation against the model's current indicies.

        Parameters:
            ``other``: The ``SQLDataModel`` or scalar (`int`, ``str``, ``float``) to compare with.

        Returns:
            ``set[int]``: The set of row indicies resulting from the operation that satisfy the condition.

        Example::
        
            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers) 

            # Filter by 'Service' column
            df = df[df['Service'] > 5.0]
                    
            # View result
            print(df)

        This will output:

        ```shell            
            ┌───┬───────┬─────────┬──────┬─────────┬────────────┬────────┐
            │   │ First │ Last    │  Age │ Service │ Hired      │ Gender │
            ├───┼───────┼─────────┼──────┼─────────┼────────────┼────────┤
            │ 0 │ Kelly │ Lee     │   32 │    8.00 │ 2016-09-18 │ Female │
            │ 1 │ Pat   │ Douglas │   42 │   11.50 │ 2015-11-06 │ Male   │
            └───┴───────┴─────────┴──────┴─────────┴────────────┴────────┘
            [2 rows x 6 columns]
        ```

        Note:
            - For scalar ``other`` (int, str, or float), compares each element with the scalar and returns the row indicies evaluating to ``True``.
            - For SQLDataModel ``other``, compares each element across X rows for Y columns for all (X_i, Y_j) in range of ``row_count`` and ``column_count`` and returns those row indicies evaluating to ``True``.
            - All the equality operations return a python ``set`` object containing the row indicies which were returned from the evaluation.
            - All operations on standard types like ``int``, ``float`` or ``str`` follow standard behavior and are not modified by performing the operations.
            - Operations can be chained using standard ``set`` operators like ``&`` and ``|`` to allow complex filtering, multiple operations require parenthesis.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """          
        self_data = self.data(strict_2d=True)
        i_dim, j_dim = len(self_data), len(self_data[0])
        if isinstance(other, SQLDataModel):
            other_data = other.data(strict_2d=True)
            return set(i for i in range(i_dim) if all(self_data[i][j] > other_data[i][j] for j in range(j_dim)))
        elif isinstance(other, (int,str,float,datetime.date)):
            return set(i for j in range(j_dim) for i in range(i_dim) if self_data[i][j] > other)
        else:
            return set()

    def __ge__(self, other) -> set[int]:
        """
        Implements the greater than or equal to operator ``>=`` for comparing ``SQLDataModel`` against ``other`` and performing the equivalent set operation against the model's current indicies.

        Parameters:
            ``other``: The ``SQLDataModel`` or scalar (`int`, ``str``, ``float``) to compare with.

        Returns:
            ``set[int]``: The set of row indicies resulting from the operation that satisfy the condition.

        Example::
        
            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers) 

            # Filter by 'Hired' column
            df = df[df['Hired'] >= datetime.date(2020,1,1)]

            # View result
            print(df)

        This will output:

        ```shell
            ┌───┬───────┬────────┬──────┬─────────┬────────────┬────────┐
            │   │ First │ Last   │  Age │ Service │ Hired      │ Gender │
            ├───┼───────┼────────┼──────┼─────────┼────────────┼────────┤
            │ 0 │ John  │ Smith  │   27 │    1.22 │ 2023-02-01 │ Male   │
            │ 1 │ Mike  │ Harlin │   36 │    3.90 │ 2020-08-27 │ Male   │
            │ 2 │ Sarah │ West   │   51 │    0.70 │ 2023-10-01 │ Female │
            └───┴───────┴────────┴──────┴─────────┴────────────┴────────┘
            [3 rows x 6 columns]
        ```

        Note:
            - For scalar ``other`` (int, str, or float), compares each element with the scalar and returns the row indicies evaluating to ``True``.
            - For SQLDataModel ``other``, compares each element across X rows for Y columns for all (X_i, Y_j) in range of ``row_count`` and ``column_count`` and returns those row indicies evaluating to ``True``.
            - All the equality operations return a python ``set`` object containing the row indicies which result from the evaluation.
            - All operations on standard types like ``int``, ``float`` or ``str`` follow standard behavior and are not modified by performing the operations.
            - Operations can be chained using standard ``set`` operators like ``&`` and ``|`` to allow complex filtering, multiple operations require parenthesis.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """          
        self_data = self.data(strict_2d=True)
        i_dim, j_dim = len(self_data), len(self_data[0])
        if isinstance(other, SQLDataModel):
            other_data = other.data(strict_2d=True)
            return set(i for i in range(i_dim) if all(self_data[i][j] >= other_data[i][j] for j in range(j_dim)))
        elif isinstance(other, (int,str,float,datetime.date)):
            return set(i for j in range(j_dim) for i in range(i_dim) if self_data[i][j] >= other)
        else:
            return set()
        
    def __add__(self, value:str|int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``+`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (str | int | float | SQLDataModel): The value to be added to each element in the SQLDataModel.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (str, int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as addition) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the addition operation.

        Example::

            import sqldatamodel as sdm
            
            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar addition
            df['x + 100'] = df['x'] + 100

            # Perform vector addition using another column
            df['x + y'] = df['x'] + df['y']

            # View both results
            print(df)

        This will output:

        ```shell
            ┌─────┬─────┬─────────┬───────┐
            │   x │   y │ x + 100 │ x + y │
            ├─────┼─────┼─────────┼───────┤
            │   2 │  10 │     102 │    12 │
            │   4 │  20 │     104 │    24 │
            │   8 │  30 │     108 │    38 │
            │  16 │  40 │     116 │    56 │
            │  32 │  50 │     132 │    82 │
            └─────┴─────┴─────────┴───────┘
            [5 rows x 4 columns]
        ```
        
        We can also use addition to concatenate strings:

        ```python
            import sqldatamodel as sdm

            # Sample data
            headers = ['First', 'Last']
            data = [['Alice', 'Smith'],['Bob', 'Johnson'],['Charlie', 'Hall'],['David', 'Brown']]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Concatenate scalar character
            df['Loud First'] = df['First'] + '!'

            # Concatenate scalar and vector using existing columns
            df['Full Name'] = df['First'] + ' ' + df['Last']

            # View it
            print(df)
        ```

        This will output:

        ```shell
            ┌─────────┬─────────┬────────────┬──────────────┐
            │ First   │ Last    │ Loud First │ Full Name    │
            ├─────────┼─────────┼────────────┼──────────────┤
            │ Alice   │ Smith   │ Alice!     │ Alice Smith  │
            │ Bob     │ Johnson │ Bob!       │ Bob Johnson  │
            │ Charlie │ Hall    │ Charlie!   │ Charlie Hall │
            │ David   │ Brown   │ David!     │ David Brown  │
            └─────────┴─────────┴────────────┴──────────────┘
            [4 rows x 4 columns]
        ```

        Note:
            - Mixing summands such as ``int + float`` will work, however an exception will be raised when attempting to perform addition on incompatible types such as ``str + float``.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(value, (str,int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', addition operations can only be performed on types 'str', 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.get_shape()
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.get_shape()):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(), value.data()
                new_data = [tuple(base_data[i][j] + value_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data
        if isinstance(value, (str,int,float)):
            return self.apply(lambda x: x + value)

    def __radd__(self, value:str|int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the right side operand for ``+`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (str | int | float | SQLDataModel): The value to be added to each element in the SQLDataModel.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (str, int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as addition) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the addition operation.

        Example::

            import sqldatamodel as sdm
            
            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar addition
            df['100 + x'] = 100 + df['x']

            # Perform vector addition using another column
            df['y + x'] = df['y'] + df['x']

            # View both results
            print(df)

        This will output:

        ```shell
            ┌─────┬─────┬─────────┬───────┐
            │   x │   y │ 100 + x │ y + x │
            ├─────┼─────┼─────────┼───────┤
            │   2 │  10 │     102 │    12 │
            │   4 │  20 │     104 │    24 │
            │   8 │  30 │     108 │    38 │
            │  16 │  40 │     116 │    56 │
            │  32 │  50 │     132 │    82 │
            └─────┴─────┴─────────┴───────┘
            [5 rows x 4 columns]
        ```
        
        We can also use addition to concatenate strings:

        ```python
            import sqldatamodel as sdm

            # Sample data
            headers = ['First', 'Last']
            data = [['Alice', 'Smith'],['Bob', 'Johnson'],['Charlie', 'Hall'],['David', 'Brown']]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Concatenate scalar character
            df['Prefixed First'] = 'Name: ' + df['First']

            # Concatenate scalar and vector using existing columns
            df['Full Name'] = df['First'] + ' ' + df['Last']

            # View it
            print(df)
        ```

        This will output:

        ```shell
            ┌─────────┬─────────┬────────────────┬──────────────┐
            │ First   │ Last    │ Prefixed First │ Full Name    │
            ├─────────┼─────────┼────────────────┼──────────────┤
            │ Alice   │ Smith   │ Name: Alice    │ Alice Smith  │
            │ Bob     │ Johnson │ Name: Bob      │ Bob Johnson  │
            │ Charlie │ Hall    │ Name: Charlie  │ Charlie Hall │
            │ David   │ Brown   │ Name: David    │ David Brown  │
            └─────────┴─────────┴────────────────┴──────────────┘
            [4 rows x 4 columns]
        ```

        Note:
            - Mixing summands such as ``int + float`` will work, however an exception will be raised when attempting to perform addition on incompatible types such as ``str + float``.
            - See :meth:`SQLDataModel.__add__()` for left side operand addition or :meth:`SQLDataModel.__iadd__()` for in-place addition.

        Changelog:
            - Version 0.7.3 (2024-06-12):
                - New method.
        """
        if not isinstance(value, (str,int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', addition operations can only be performed on types 'str', 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.get_shape()
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.get_shape()):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                new_data = [tuple(value_data[i][j] + base_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data
        if isinstance(value, (str,int,float)):
            return self.apply(lambda x: value + x)

    def __sub__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``-`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to subtract from each element in the SQLDataModel.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int or float).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as subtraction) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the subtraction operation.
            
        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar subtraction
            df['x - 100'] = df['x'] - 100

            # Perform vector subtraction using another column
            df['x - y'] = df['x'] - df['y']

            # View both results
            print(df)

        This will output:

        ```shell
            ┌─────┬─────┬─────────┬───────┐
            │   x │   y │ x - 100 │ x - y │
            ├─────┼─────┼─────────┼───────┤
            │   2 │  10 │     -98 │    -8 │
            │   4 │  20 │     -96 │   -16 │
            │   8 │  30 │     -92 │   -22 │
            │  16 │  40 │     -84 │   -24 │
            │  32 │  50 │     -68 │   -18 │
            └─────┴─────┴─────────┴───────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing subtractors such as ``int + float`` will work, however an exception will be raised when attempting to perform subtraction on incompatible types such as ``str - float``.
            - See :meth:`SQLDataModel.__rsub__()` for right side operand subtraction operations.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', subtraction operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else:
                if value_shape != (model_shape := self.get_shape()):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(), value.data()
                new_data = [tuple(base_data[i][j] - value_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data
        if isinstance(value, (int,float)):
            return self.apply(lambda x: x - value)

    def __rsub__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the right side operand for ``-`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to subtract from each element in the SQLDataModel.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int or float).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as subtraction) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the subtraction operation.
            
        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar subtraction
            df['100 - x'] = 100 - df['x']

            # Perform vector subtraction using another column
            df['y - x'] = df['y'] - df['x']

            # View both results
            print(df)

        This will output:

        ```shell
            ┌─────┬─────┬─────────┬───────┐
            │   x │   y │ 100 - x │ y - x │
            ├─────┼─────┼─────────┼───────┤
            │   2 │  10 │      98 │     8 │
            │   4 │  20 │      96 │    16 │
            │   8 │  30 │      92 │    22 │
            │  16 │  40 │      84 │    24 │
            │  32 │  50 │      68 │    18 │
            └─────┴─────┴─────────┴───────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing subtractors such as ``int + float`` will work, however an exception will be raised when attempting to perform subtraction on incompatible types such as ``str - float``.
            - See :meth:`SQLDataModel.__sub__()` for left side operand subtraction or :meth:`SQLDataModel.__isub__()` for in-place subtraction.

        Changelog:
            - Version 0.7.3 (2024-06-12):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', subtraction operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else:
                if value_shape != (model_shape := self.get_shape()):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                new_data = [tuple(value_data[i][j] - base_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data
        if isinstance(value, (int,float)):
            return self.apply(lambda x: value - x)

    def __mul__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``*`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to multiply each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as multiplication) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the multiplication operation.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdmSQLDataModel(data, headers)

            # Perform scalar multiplication
            df['x * 10'] = df['x'] * 10

            # Perform vector multiplication using another column
            df['x * y'] = df['x'] * df['y']

            # View results
            print(df)
        
        This will output:

        ```shell
            ┌─────┬─────┬────────┬───────┐
            │   x │   y │ x * 10 │ x * y │
            ├─────┼─────┼────────┼───────┤
            │   2 │  10 │     20 │    20 │
            │   4 │  20 │     40 │    80 │
            │   8 │  30 │     80 │   240 │
            │  16 │  40 │    160 │   640 │
            │  32 │  50 │    320 │  1600 │
            └─────┴─────┴────────┴───────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing multipliers such as ``int * float`` will work, however an exception will be raised when attempting to perform multiplication on incompatible types such as ``str * float``.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', multiplication operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                new_data = [tuple(base_data[i][j] * value_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data        
        if isinstance(value, (int,float)):
            return self.apply(lambda x: x * value)

    def __rmul__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the right side operand for ``*`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to multiply each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as multiplication) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the multiplication operation.
        
        Note:
            - See :meth:`SQLDataModel.__mul__()` for additional details and usage examples.
            - This function simply wraps the primary multiplication method after swapping the order of the arguments.

        Changelog:
            - Version 0.7.3 (2024-06-12):
                - New method.
        """
        return self.__mul__(value)

    def __truediv__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``/`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to divide each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as division) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.
            ``ZeroDivisionError``: If ``value`` is 0.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the division operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar division
            df['y / 10'] = df['y'] / 10

            # Perform vector division using another column
            df['y / x'] = df['y'] / df['x']

            # View both results
            print(df)
        
        This will output:

        ```shell
            ┌─────┬─────┬────────┬───────┐
            │   x │   y │ y / 10 │ y / x │
            ├─────┼─────┼────────┼───────┤
            │   2 │  10 │   1.00 │  5.00 │
            │   4 │  20 │   2.00 │  5.00 │
            │   8 │  30 │   3.00 │  3.75 │
            │  16 │  40 │   4.00 │  2.50 │
            │  32 │  50 │   5.00 │  1.56 │
            └─────┴─────┴────────┴───────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing divisor types such as ``int / float`` will work, however an exception will be raised when attempting to perform division on incompatible types such as ``str / float``.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', division operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                new_data = [tuple(base_data[i][j] / value_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data 
        if value == 0:
            msg = ErrorFormat(f"ZeroDivisionError: invalid argument '{value}', division operations cannot be performed with a divisor of zero")
            raise ZeroDivisionError(msg)
        if isinstance(value, (int,float)):
            return self.apply(lambda x: x / value)

    def __rtruediv__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the right side operand ``/`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to divide each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as division) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.
            ``ZeroDivisionError``: If ``value`` is 0.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the division operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar division
            df['10 / y'] = 10 / df['y']

            # Perform vector division using another column
            df['x / y'] = df['x'] / df['y']

            # View both results
            print(df)
        
        This will output:

        ```shell
            ┌─────┬─────┬────────┬───────┐
            │   x │   y │ 10 / y │ x / y │
            ├─────┼─────┼────────┼───────┤
            │   2 │  10 │   1.00 │  0.20 │
            │   4 │  20 │   0.50 │  0.20 │
            │   8 │  30 │   0.33 │  0.27 │
            │  16 │  40 │   0.25 │  0.40 │
            │  32 │  50 │   0.20 │  0.64 │
            └─────┴─────┴────────┴───────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing divisor types such as ``int / float`` will work, however an exception will be raised when attempting to perform division on incompatible types such as ``str / float``.
            - See :meth:`SQLDataModel.__truediv__()` for left side operand division operations.

        Changelog:
            - Version 0.7.3 (2024-06-12):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', division operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                new_data = [tuple(value_data[i][j] / base_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                return new_data 
        if value == 0:
            msg = ErrorFormat(f"ZeroDivisionError: invalid argument '{value}', division operations cannot be performed with a divisor of zero")
            raise ZeroDivisionError(msg)
        if isinstance(value, (int,float)):
            return self.apply(lambda x: value / x)       

    def __floordiv__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``//`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to divide each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as division) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.
            ``ZeroDivisionError``: If ``value`` is 0.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the floor division operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,10], [4,20], [8,30], [16,40], [32,50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar floor division
            df['y // 10'] = df['y'] // 10

            # Perform vector floor division using another column
            df['y // x'] = df['y'] // df['x']

            # View both results
            print(df)
        
        This will output:

        ```shell
            ┌─────┬─────┬─────────┬────────┐
            │   x │   y │ y // 10 │ y // x │
            ├─────┼─────┼─────────┼────────┤
            │   2 │  10 │       1 │      5 │
            │   4 │  20 │       2 │      5 │
            │   8 │  30 │       3 │      3 │
            │  16 │  40 │       4 │      2 │
            │  32 │  50 │       5 │      1 │
            └─────┴─────┴─────────┴────────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing divisor types such as ``int // float`` will work, however an exception will be raised when attempting to perform division on incompatible types such as ``str // float``.

        Changelog:
            - Version 0.2.2 (2024-03-26):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', floor division operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                try:
                    new_data = [tuple(base_data[i][j] // value_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: '{e}' encountered when trying to perform floor division operations")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None                
                return new_data 
        if value == 0:
            msg = ErrorFormat(f"ZeroDivisionError: invalid argument '{value}', floor division operations cannot be performed with a divisor of zero")
            raise ZeroDivisionError(msg)
        if isinstance(value, (int,float)):
            return self.apply(lambda x: x // value)

    def __rfloordiv__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the right side operand ``//`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to divide each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as division) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.
            ``ZeroDivisionError``: If ``value`` is 0.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the floor division operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,8], [4,16], [8,32], [32,64], [32,128]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar floor division
            df['128 // y'] = 128 // df['y']

            # Perform vector floor division using another column
            df['y // x'] = df['y'] // df['x']

            # View both results
            print(df)
        
        This will output:

        ```shell
            ┌─────┬─────┬──────────┬────────┐
            │   x │   y │ 128 // y │ y // x │
            ├─────┼─────┼──────────┼────────┤
            │   2 │   8 │       16 │      4 │
            │   4 │  16 │        8 │      4 │
            │   8 │  32 │        4 │      4 │
            │  32 │  64 │        2 │      2 │
            │  32 │ 128 │        1 │      4 │
            └─────┴─────┴──────────┴────────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing divisor types such as ``int // float`` will work, however an exception will be raised when attempting to perform division on incompatible types such as ``str // float``.
            - See :meth:`SQLDataModel.__floordiv__()` for standard left side operand implementation of floor division operations.

        Changelog:
            - Version 0.7.7 (2024-06-17):
                - New method.
        """        
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', floor division operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                try:
                    new_data = [tuple(value_data[i][j] // base_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: '{e}' encountered when trying to perform floor division operations")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None
                return new_data 
        if isinstance(value, (int,float)):
            return self.apply(lambda x: value // x)
                
    def __pow__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``**`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The exponent value to raise each element in the SQLDataModel to.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as exponentiation) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the exponential operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,1], [4,2], [8,3], [16,4], [32,5]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar exponentiation
            df['y ** 2'] = df['y'] ** 2

            # Perform vector exponentiation using another column
            df['x ** y'] = df['x'] ** df['y']

            # View results
            print(df)

        This will output:

        ```shell
            ┌─────┬─────┬────────┬──────────┐
            │   x │   y │ y ** 2 │   x ** y │
            ├─────┼─────┼────────┼──────────┤
            │   2 │   1 │      1 │        2 │
            │   4 │   2 │      4 │       16 │
            │   8 │   3 │      9 │      512 │
            │  16 │   4 │     16 │    65536 │
            │  32 │   5 │     25 │ 33554432 │
            └─────┴─────┴────────┴──────────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing exponent types such as ``int ** float`` will work, however an exception will be raised when attempting to exponentiate incompatible types such as ``str ** float``.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', exponential operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                try:
                    new_data = [tuple(base_data[i][j] ** value_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: '{e}' encountered when trying to perform exponential operations")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None                
                return new_data        
        if isinstance(value, (int,float)):
            return self.apply(lambda x: x ** value)      

    def __rpow__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the right side operand ``**`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The exponent value to raise each element in the SQLDataModel to.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``DimensionError``: Raised when the dimensions of the provided ``value`` are incompatible with the current model's dimensions. For example, attempting to perform an operation (such as exponentiation) on data of shape ``(4, 1)`` with values of shape ``(3, 2)`` will raise this exception.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the exponential operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['x', 'y']
            data = [[2,1], [4,2], [6,3], [8,4], [10,5]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Perform scalar exponentiation
            df['2 ** y'] = 2 ** df['y']

            # Perform vector exponentiation using another column
            df['y ** x'] = df['y'] ** df['x']

            # View results
            print(df)  

        This will output:

        ```shell
            ┌─────┬─────┬────────┬─────────┐
            │   x │   y │ 2 ** y │  y ** x │
            ├─────┼─────┼────────┼─────────┤
            │   2 │   1 │      2 │       1 │
            │   4 │   2 │      4 │      16 │
            │   6 │   3 │      8 │     729 │
            │   8 │   4 │     16 │   65536 │
            │  10 │   5 │     32 │ 9765625 │
            └─────┴─────┴────────┴─────────┘
            [5 rows x 4 columns]
        ```

        Note:
            - Mixing exponent types such as ``int ** float`` will work, however an exception will be raised when attempting to exponentiate incompatible types such as ``str ** float``.
            - See :meth:`SQLDataModel.__pow__()` for standard left side operand implementation of exponential operations.

        Changelog:
            - Version 0.7.7 (2024-06-17):
                - New method.
        """
        if not isinstance(value, (int,float,SQLDataModel)):
            msg = ErrorFormat(f"TypeError: unsupported operand type '{type(value).__name__}', exponential operations can only be performed on types 'int', 'float' or 'SQLDataModel'")
            raise TypeError(msg)
        if isinstance(value, SQLDataModel):
            value_shape = value.shape
            if value_shape == (1,1):
                value = value.data()
            else: 
                if value_shape != (model_shape := self.shape):
                    msg = ErrorFormat(f"DimensionError: shape mismatch '{model_shape} != {value_shape}', model dim '{model_shape}' is not compatible with values dim '{value_shape}' for performing vectorized operations")
                    raise DimensionError(msg)
                base_data, value_data = self.data(strict_2d=True), value.data(strict_2d=True)
                try:
                    new_data = [tuple(value_data[i][j] ** base_data[i][j] for j in range(self.column_count)) for i in range(self.row_count)]
                except Exception as e:
                    msg = ErrorFormat(f"{type(e).__name__}: '{e}' encountered when trying to perform exponential operations")
                    raise type(e)(msg).with_traceback(e.__traceback__) from None                  
                return new_data        
        if isinstance(value, (int,float)):
            return self.apply(lambda x: value ** x) 

    def __iadd__(self, value:str|int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``+=`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (str | int | float | SQLDataModel): The value to be added to each element in the SQLDataModel.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (str, int, float, or SQLDataModel).

        Returns:
            ``SQLDataModel``: The modified SQLDataModel after the addition operation.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['idx', 'first', 'last', 'age', 'service']
            data = [
                (0, 'john', 'smith', 27, 1.22),
                (1, 'sarah', 'west', 39, 0.7),
                (2, 'mike', 'harlin', 36, 3),
                (3, 'pat', 'douglas', 42, 11.5)
            ]     

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Modifying first name column with a bang!
            df['first'] += '!'

            # View model
            print(df)

        This will output:
        
        ```shell            
            ┌───┬────────┬─────────┬────────┬─────────┐
            │   │ first  │ last    │    age │ service │
            ├───┼────────┼─────────┼────────┼─────────┤
            │ 0 │ john!  │ smith   │     27 │    1.22 │
            │ 1 │ sarah! │ west    │     39 │    0.70 │
            │ 2 │ mike!  │ harlin  │     36 │    3.00 │
            │ 3 │ pat!   │ douglas │     42 │   11.50 │
            └───┴────────┴─────────┴────────┴─────────┘
            [4 rows x 4 columns]
        ```  

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        return self.__add__(value)
    
    def __isub__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``-=`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to subtract from each element in the SQLDataModel.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float, or SQLDataModel).

        Returns:
            ``SQLDataModel``: The modified SQLDataModel after the subtraction operation.

        Example::     

            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age', 'service']
            data = [
                (0, 'john', 'smith', 27, 1.22),
                (1, 'sarah', 'west', 39, 0.7),
                (2, 'mike', 'harlin', 36, 3),
                (3, 'pat', 'douglas', 42, 11.5)
            ]     

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Modifying age column in the best direction
            df['age'] -= 10

            # View model
            print(df)

        This will output:
        
        ```shell
            ┌───┬────────┬─────────┬────────┬─────────┐
            │   │ first  │ last    │    age │ service │
            ├───┼────────┼─────────┼────────┼─────────┤
            │ 0 │ john   │ smith   │     17 │    1.22 │
            │ 1 │ sarah  │ west    │     29 │    0.70 │
            │ 2 │ mike   │ harlin  │     26 │    3.00 │
            │ 3 │ pat    │ douglas │     32 │   11.50 │
            └───┴────────┴─────────┴────────┴─────────┘
            [4 rows x 4 columns]
        ```
                
        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        return self.__sub__(value)    
    
    def __imul__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``*=`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to multiply each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int or float).

        Returns:
            ``SQLDataModel``: The modified SQLDataModel after the multiplication operation.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Salary'])

            # Give raises to all!
            df['Salary'] *= 12

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        return self.__mul__(value)    

    def __idiv__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``/=`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to divide each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int, float or SQLDataModel).
            ``ZeroDivisionError``: If ``value`` of divisor is 0.

        Returns:
            ``SQLDataModel``: The modified SQLDataModel after the division operation.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Budget'])

            # Adjust existing column
            df['Budget'] /= 52
        
        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        return self.__truediv__(value)

    def __ifloordiv__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``//=`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to divide each element in the SQLDataModel by.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int or float).
            ``ZeroDivisionError``: If ``value`` is 0.

        Returns:
            ``SQLDataModel``: A new SQLDataModel resulting from the floor division operation.

        Example::
        
            import sqldatamodel as sdm
            
            # Sample data
            headers = ['x']
            data = [[10],[20],[30],[40],[50]]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Modify the existing column
            df['x'] //= 3
            
            # View result
            print(df)

        This will output:
        
        ```shell
            ┌───┬──────┐
            │   │    x │
            ├───┼──────┤
            │ 0 │    3 │
            │ 1 │    6 │
            │ 2 │   10 │
            │ 3 │   13 │
            │ 4 │   16 │
            └───┴──────┘
            [5 rows x 1 columns]
        ```

        Changelog:
            - Version 0.2.2 (2024-03-26):
                - New method.
        """
        return self.__floordiv__(value)

    def __ipow__(self, value:int|float|SQLDataModel) -> SQLDataModel:
        """
        Implements the ``**=`` operator functionality for compatible ``SQLDataModel`` operations.

        Parameters:
            ``value`` (int | float | SQLDataModel): The value to raise each element in the SQLDataModel to.

        Raises:
            ``TypeError``: If the provided ``value`` is not a valid type (int or float).

        Returns:
            ``SQLDataModel``: The modified SQLDataModel after the exponential operation.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Salary'])

            # More raises!
            df['Salary'] **= 2
        
        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        return self.__pow__(value)
    
    def __and__(self, other:SQLDataModel) -> set[int]:
        """
        Implements the bitwise AND operator ``&`` for combining the result sets of ``self`` and ``other``.

        Parameters:
            ``other``: The ``SQLDataModel`` to combine with.

        Returns:
            ``set[int]``: A set of indices representing the intersection of the result rows from both ``SQLDataModel`` instances.

        Example::
            
            import sqldatamodel as sdm

            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the sample model
            df = sdm.SQLDataModel(data, headers)

            # Apply some filtering conditions to both models
            filter_1 = df[df['Age'] <= 40]
            filter_2 = df[df['Service'] > 2]

            # Perform a bitwise AND operation to return a new model
            result = df[filter_1 & filter_2]

            # View result
            print(result)

        This will output the result of filtering by 'Age' and 'Service':

        ```shell
            ┌───┬───────┬────────┬─────┬─────────┬────────────┬────────┐
            │   │ First │ Last   │ Age │ Service │ Hired      │ Gender │
            ├───┼───────┼────────┼─────┼─────────┼────────────┼────────┤
            │ 1 │ Kelly │ Lee    │  32 │    8.00 │ 2016-09-18 │ Female │
            │ 2 │ Mike  │ Harlin │  36 │    3.90 │ 2020-08-27 │ Male   │
            └───┴───────┴────────┴─────┴─────────┴────────────┴────────┘
            [2 rows x 6 columns]
        ```

        Note:
            - If ``other`` is not an instance of ``SQLDataModel``, a ``NotImplementedError`` is raised to be consistent with current conventions.
            - See :meth:`SQLDataModel.__or__()` for bitwise OR operation.

        Changelog:
            - Version 0.7.4 (2024-06-13):
                - New method.
        """        
        if not isinstance(other, SQLDataModel):
            msg = ErrorFormat(f"NotImplementedError: unsupported type '{type(other).__name__}', operand `&` is only supported for type `SQLDataModel`")
            raise NotImplementedError(msg)
        return (set(self.indicies) & set(other.indicies))

    def __or__(self, other:SQLDataModel) -> set[int]:
        """
        Implements the bitwise OR operator ``|`` for combining the result sets of ``self`` and ``other``.

        Parameters:
            ``other``: The ``SQLDataModel`` to combine with.

        Returns:
            ``set[int]``: A set of indices representing the union of the result rows from both ``SQLDataModel`` instances.

        Example::
            
            import sqldatamodel as sdm
            
            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the sample model
            df = sdm.SQLDataModel(data, headers)

            # Apply some filtering conditions to both models
            filter_1 = df[df['Age'] > 40]
            filter_2 = df[df['Gender'] == 'Male']

            # Perform a bitwise OR operation to return a new model
            result = df[filter_1 | filter_2]

            # View result
            print(result) 

        This will output the result of filtering by 'Age' or 'Gender':

        ```shell
            ┌───┬───────┬─────────┬─────┬─────────┬────────────┬────────┐
            │   │ First │ Last    │ Age │ Service │ Hired      │ Gender │
            ├───┼───────┼─────────┼─────┼─────────┼────────────┼────────┤
            │ 0 │ John  │ Smith   │  27 │    1.22 │ 2023-02-01 │ Male   │
            │ 2 │ Mike  │ Harlin  │  36 │    3.90 │ 2020-08-27 │ Male   │
            │ 3 │ Sarah │ West    │  51 │    0.70 │ 2023-10-01 │ Female │
            │ 4 │ Pat   │ Douglas │  42 │   11.50 │ 2015-11-06 │ Male   │
            └───┴───────┴─────────┴─────┴─────────┴────────────┴────────┘
            [4 rows x 6 columns]
        ```

        Note:
            - If ``other`` is not an instance of ``SQLDataModel``, a ``NotImplementedError`` is raised to be consistent with current conventions.
            - See :meth:`SQLDataModel.__and__()` for bitwise AND operation.

        Changelog:
            - Version 0.7.4 (2024-06-13):
                - New method.
        """        
        if not isinstance(other, SQLDataModel):
            msg = ErrorFormat(f"NotImplementedError: unsupported type '{type(other).__name__}', operand `|` is only supported for type `SQLDataModel`")
            raise NotImplementedError(msg)
        return (set(self.indicies) | set(other.indicies))
    
##################################################################################################################
############################################## other dunder methods ##############################################
##################################################################################################################

    def __iter__(self) -> Iterator[tuple]:
        """
        Returns an iterator over the current range of rows in the ``SQLDataModel`` starting from the first row.

        Raises:
            ``StopIteration``: When there are no more rows to return.

        Yields:
            ``tuple``: Next row fetched from the current ``SQLDataModel``.
        
        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Iterate through rows
            for row in df:
                print(row)

        This will output:

        ```text
            (0, 'John', 30, 175.3)
            (1, 'Alice', 28, 162.0)
            (2, 'Travis', 35, 185.8)
        ```

        Note:
            - This iterator fetches rows from the ``SQLDataModel`` using a SQL statement generated by the :meth:`SQLDataModel._generate_sql_stmt()` method.
            - The iteration starts from the first row, index 0, and continues until :py:attr:`SQLDataModel.row_count` is reached.
            - See :meth:`SQLDataModel.iter_rows()` for iterating over rows with custom start and stop indicies.
            - See :meth:`SQLDataModel.iter_tuples()` for iterating over rows as named tuples.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """  
        yield from (self.sql_db_conn.execute(self._generate_sql_stmt_fetchall(index=True)))

    def __getitem__(self, target_indicies) -> SQLDataModel:
        """
        Retrieves a subset of the SQLDataModel based on the specified indices.

        Parameters:
            ``slc``: Indices specifying the rows and columns to be retrieved. This can be an integer, a tuple, a slice, or a combination of these.

        Raises:
            ``ValueError``: if there are issues with the specified indices, such as invalid row or column names.
            ``TypeError``: if the ``slc`` type is not compatible with indexing SQLDataModel.
            ``IndexError``: if the ``slc`` includes a range or int that is outside of the current row count or column count.

        Returns:
            ``SQLDataModel``: An instance of SQLDataModel containing the selected subset of data.

        Example::

            import sqldatamodel as sdm

            # Create a sample model
            headers = ['First', 'Last', 'Age', 'Service', 'Hired', 'Gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Sarah', 'West', 51, 0.7, '2023-10-01', 'Female'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
            ]  

            # Create the sample model
            df = sdm.SQLDataModel(data, headers)

            # Retrieve a specific row by index
            subset_model = df[3]

            # Retrieve multiple rows and specific columns using a tuple
            subset_model = df[(1, 2, 4), ["First", "Service", "Age"]]

            # Retrieve a range of rows and all columns using a slice
            subset_model = df[1:4]

            # Retrieve a single column by name
            subset_model = df["First"]
        
        Changelog:
            - Version 0.5.0 (2024-05-09):
                - Modified index retention behavior to pass through row indicies and avoid resetting view order.

        Note:
            - The ``slc`` parameter can be an integer, a tuple of disconnected row indices, a slice representing a range of rows, a string or list of strings representing column names, or a tuple combining row and column indices.
            - The returned SQLDataModel instance will contain the specified subset of rows and columns, retaining the row indicies of the original view.

        Changelog:
            - Version 0.5.0 (2024-05-09):
                - Modified index retention behavior to pass through row indicies and avoid resetting view order.
        """         
        validated_rows, validated_columns = self._validate_indicies(target_indicies)
        sql_stmt_generated = self._generate_sql_stmt(rows=validated_rows,columns=validated_columns,index=True) # NOTE: toggle to retain prior indicies after getitem slicing, changed in version 0.5.0 to True
        return self.execute_fetch(sql_stmt_generated)

    def __setitem__(self, target_indicies, update_values) -> None:
        """
        Updates specified rows and columns in the SQLDataModel with the provided values.

        Parameters:
            ``target_indicies``: Indices specifying the rows and columns to be updated. This can be an integer, a tuple, a slice, or a combination of these.
            ``update_values``: The values to be assigned to the corresponding model records. It can be of types: str, int, float, bool, bytes, list, tuple, or another SQLDataModel object.

        Raises:
            ``TypeError``: If the ``update_values`` type is not compatible with SQL datatypes.
            ``DimensionError``: If there is a shape mismatch between targeted indicies and provided update values.
            ``ValueError``: If there are issues with the specified indices, such as invalid row or column names.
        
        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Job']
            data = [
                ('Billy', 30, 'Barber'), 
                ('Alice', 28, 'Doctor'), 
                ('John', 25, 'Technician'), 
                ('Travis', 35, 'Musician'),
                ('William', 15, 'Student')
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Update a specific row with new values
            df[2] = ("John", 25, "Engineer")

            # See result
            print(df)

        This will output:

        ```shell
            ┌───┬─────────┬──────┬──────────┐
            │   │ Name    │  Age │ Job      │
            ├───┼─────────┼──────┼──────────┤
            │ 0 │ Billy   │   30 │ Barber   │
            │ 1 │ Alice   │   28 │ Doctor   │
            │ 2 │ John    │   25 │ Engineer │
            │ 3 │ Travis  │   35 │ Musician │
            │ 4 │ William │   15 │ Student  │
            └───┴─────────┴──────┴──────────┘
            [5 rows x 3 columns]
        ```

        Conditional updates can also be made using multiple columns:

        ```python
            import sqldatamodel as sdm

            headers = ['Employee', 'Base', 'Salary']
            data = [
                ('Alice', '58,500', '62,250'),
                ('Bobby', '60,750',  None),
                ('Chloe', '58,500', '63,125'),
                ('David', '65,000',  None),
                ('Ellie', '65,000',  None),
                ('Fiona', '65,000', '71,450'),
            ]

            # Create sample model
            df = sdm.SQLDataModel(data, headers)

            # Selectively update values based on conditions
            df[df['Salary'].isna(), 'Salary'] = df['Base']

            # View updates
            print(df)
        ```

        This will output the resulting model where 'Salary' was updated with values from 'Base' only if missing:

        ```shell
            ┌───┬──────────┬────────┬────────┐
            │   │ Employee │ Base   │ Salary │
            ├───┼──────────┼────────┼────────┤
            │ 0 │ Alice    │ 58,500 │ 62,250 │
            │ 1 │ Bobby    │ 60,750 │ 60,750 │
            │ 2 │ Chloe    │ 58,500 │ 63,125 │
            │ 3 │ David    │ 65,000 │ 65,000 │
            │ 4 │ Ellie    │ 65,000 │ 65,000 │
            │ 5 │ Fiona    │ 65,000 │ 71,450 │
            └───┴──────────┴────────┴────────┘
            [6 rows x 3 columns]
        ```

        Values for multiple columns can also be set:

        ```python
            # Update multiple rows and columns with a list of values
            df[1:5, ["Name", "Age", "Job"]] = [("Alice", 30, "Manager"), ("Bob", 28, "Developer"), ("Charlie", 35, "Designer"), ("David", 32, "Analyst")]

            # See result
            print(df)
        ```

        This will output:

        ```shell
            ┌───┬─────────┬──────┬───────────┐
            │   │ Name    │  Age │ Job       │
            ├───┼─────────┼──────┼───────────┤
            │ 0 │ Billy   │   30 │ Barber    │
            │ 1 │ Alice   │   30 │ Manager   │
            │ 2 │ Bob     │   28 │ Developer │
            │ 3 │ Charlie │   35 │ Designer  │
            │ 4 │ David   │   32 │ Analyst   │
            └───┴─────────┴──────┴───────────┘
            [5 rows x 3 columns]
        ```

        Values can also be set along the row axes:

        ```python
            # Create a new column "Hobby" and set the values
            df["Hobby"] = [('Fishing',), ('Biking',), ('Computers',), ('Photography',), ('Studying',)]

            # See result
            print(df)
        ```

        This will output:

        ```shell            
            ┌───┬─────────┬──────┬───────────┬─────────────┐
            │   │ Name    │  Age │ Job       │ Hobby       │
            ├───┼─────────┼──────┼───────────┼─────────────┤
            │ 0 │ Billy   │   30 │ Barber    │ Fishing     │
            │ 1 │ Alice   │   30 │ Manager   │ Biking      │
            │ 2 │ Bob     │   28 │ Developer │ Computers   │
            │ 3 │ Charlie │   35 │ Designer  │ Photography │
            │ 4 │ David   │   32 │ Analyst   │ Studying    │
            └───┴─────────┴──────┴───────────┴─────────────┘
            [5 rows x 4 columns]            
        ```

        Note:
            - If ``update_values`` is another ``SQLDataModel`` object, its data will be normalized using the :meth:`SQLDataModel.data()` method.
            - The ``target_indicies`` parameter can be an integer, a tuple of disconnected row indices, a slice representing a range of rows, a string or list of strings representing column names, or a tuple combining row and column indices.
            - Values can be single values or iterables matching the specified rows and columns.
            - See :meth:`SQLDataModel.apply()` for setting values using a function.

        Changelog:
            - Version 0.7.5 (2024-06-14):
                - Added row indicies masking to allow selective updating when ``update_values`` is also an instance of ``SQLDataModel`` using ``target_indicies`` as mask.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """            
        # first check if target is new column that needs to be created, if so create it and return so long as the target values aren't another sqldatamodel object:
        if isinstance(update_values, SQLDataModel):
            other_is_sdm = True
            update_values = update_values.data(strict_2d=True, index=True) # normalize data input
        else:
            other_is_sdm = False
        if not isinstance(update_values, (str,int,float,bool,bytes,list,tuple,datetime.date)) and (update_values is not None):
            msg = ErrorFormat(f"TypeError: invalid values type '{type(update_values).__name__}', update values must be compatible with SQL datatypes such as <'str', 'int', 'float', 'datetime', 'bool', 'bytes'>")
            raise TypeError(msg)
        # short circuit remaining operations and proceed to insert row if target_indicies is int and equals current row count
        if isinstance(target_indicies, int) and target_indicies == self.row_count:
            update_values = update_values[0][1:] if other_is_sdm else update_values
            try:
                self.append_row(update_values)
                return
            except TypeError as e:
                msg = ErrorFormat(f"{e}")
                raise TypeError(msg) from None                
            except DimensionError as e:
                msg = ErrorFormat(f"{e}")
                raise DimensionError(msg) from None
        # normal update values process where target update values is not another SQLDataModel object:
        if isinstance(target_indicies, str) and target_indicies not in self.headers:
            validated_rows, validated_columns = self.indicies, [target_indicies]
        else:
            validated_rows, validated_columns = self._validate_indicies(target_indicies)
        # convert various row options to be tuple or int
        if other_is_sdm:
            update_values = [row[1:] for row in update_values if row[0] in validated_rows]
        self._update_rows_and_columns_with_values(rows_to_update=validated_rows,columns_to_update=validated_columns,values_to_update=update_values)
        return
    
    def __len__(self) -> int:
        """
        Returns the :py:attr:`SQLDataModel.row_count` property for the current ``SQLDataModel`` which represents the current number of rows in the model.

        Returns:
            ``int``: The total number of rows in the SQLDataModel.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Get current length
            num_rows = len(df)

            # View number
            print(num_rows)

        This will output:

        ```shell        
            1000
        ```

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        return self.row_count

    def set_table_style(self, style:Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple']='default') -> None:
        """
        Sets the table style used for string representations of ``SQLDataModel``.
        
        Parameters:
            ``style`` (Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple']): The table styling to set. 
                Setting to ``'default'`` style will return the style representation to the original format.

        Raises:
            ``ValueError``: If ``style`` provided is not one of the currently supported options 'ascii', 'bare', 'dash', 'default', 'double', 'latex', 'list', 'markdown', 'outline', 'pandas', 'polars', 'postgresql' or 'round'.
        
        Returns:
            ``None``

        Examples::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height', 'Birthday']
            data = [
                ('Alice', 28, 162.08, '1996-11-20'), 
                ('Bobby', 30, 175.36, '1994-06-15'), 
                ('Craig', 37, 185.82, '1987-01-07'),
                ('David', 32, 179.75, '1992-12-28')
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Lets try the round style
            df.set_table_style('round')

            # View it
            print(df)

        This outputs the ``'round'`` table style:

        ```shell
            ╭───────┬─────┬─────────┬────────────╮
            │ Name  │ Age │  Height │ Birthday   │
            ├───────┼─────┼─────────┼────────────┤
            │ Alice │  28 │  162.08 │ 1996-11-20 │
            │ Bobby │  30 │  175.36 │ 1994-06-15 │
            │ Craig │  37 │  185.82 │ 1987-01-07 │
            │ David │  32 │  179.75 │ 1992-12-28 │
            ╰───────┴─────┴─────────┴────────────╯
        ```

        Alternatively, set ``style = 'ascii'`` to format ``SQLDataModel`` in the ASCII style, the OG of terminal tables:        

        ```shell
            +-------+-----+---------+------------+
            | Name  | Age |  Height | Birthday   |
            +-------+-----+---------+------------+
            | Alice |  28 |  162.08 | 1996-11-20 |
            | Bobby |  30 |  175.36 | 1994-06-15 |
            | Craig |  37 |  185.82 | 1987-01-07 |
            | David |  32 |  179.75 | 1992-12-28 |
            +-------+-----+---------+------------+
        ```

        Set ``style = 'bare'`` to format ``SQLDataModel`` in the following style:

        ```shell
            Name   Age   Height  Birthday
            -------------------------------
            Alice   28   162.08  1996-11-20
            Bobby   30   175.36  1994-06-15
            Craig   37   185.82  1987-01-07
            David   32   179.75  1992-12-28
        ```

        Set ``style = 'dash'`` to format ``SQLDataModel`` with dashes for internal borders:

        ```shell
            ┌───────┬─────┬─────────┬────────────┐
            │ Name  ╎ Age ╎  Height ╎ Birthday   │
            ├╴╴╴╴╴╴╴┼╴╴╴╴╴┼╴╴╴╴╴╴╴╴╴┼╴╴╴╴╴╴╴╴╴╴╴╴┤
            │ Alice ╎  28 ╎  162.08 ╎ 1996-11-20 │
            │ Bobby ╎  30 ╎  175.36 ╎ 1994-06-15 │
            │ Craig ╎  37 ╎  185.82 ╎ 1987-01-07 │
            │ David ╎  32 ╎  179.75 ╎ 1992-12-28 │
            └───────┴─────┴─────────┴────────────┘
        ```        

        Set ``style = 'default'`` to format ``SQLDataModel`` in the following style, which also happens to be the default styling applied:

        ```shell
            ┌───────┬─────┬─────────┬────────────┐
            │ Name  │ Age │  Height │ Birthday   │
            ├───────┼─────┼─────────┼────────────┤
            │ Alice │  28 │  162.08 │ 1996-11-20 │
            │ Bobby │  30 │  175.36 │ 1994-06-15 │
            │ Craig │  37 │  185.82 │ 1987-01-07 │
            │ David │  32 │  179.75 │ 1992-12-28 │
            └───────┴─────┴─────────┴────────────┘
            [4 rows x 4 columns]
        ```        
        
        Set ``style = 'list'`` to format ``SQLDataModel`` as a list of values, similar to the SQLite CLI representation:

        ```shell
            Name   Age   Height  Birthday  
            -----  ---  -------  ----------
            Alice   28   162.08  1996-11-20
            Bobby   30   175.36  1994-06-15
            Craig   37   185.82  1987-01-07
            David   32   179.75  1992-12-28
        ```

        Set ``style = 'double'`` to format ``SQLDataModel`` using double line borders:

        ```shell
            ╔═══════╦═════╦═════════╦════════════╗
            ║ Name  ║ Age ║  Height ║ Birthday   ║
            ╠═══════╬═════╬═════════╬════════════╣
            ║ Alice ║  28 ║  162.08 ║ 1996-11-20 ║
            ║ Bobby ║  30 ║  175.36 ║ 1994-06-15 ║
            ║ Craig ║  37 ║  185.82 ║ 1987-01-07 ║
            ║ David ║  32 ║  179.75 ║ 1992-12-28 ║
            ╚═══════╩═════╩═════════╩════════════╝
        ```

        Set ``style = 'markdown'`` to format ``SQLDataModel`` in the Markdown style:

        ```shell
            | Name  | Age |  Height | Birthday   |
            |-------|-----|---------|------------|
            | Alice |  28 |  162.08 | 1996-11-20 |
            | Bobby |  30 |  175.36 | 1994-06-15 |
            | Craig |  37 |  185.82 | 1987-01-07 |
            | David |  32 |  179.75 | 1992-12-28 |
        ```

        Set ``style = 'outline'`` to format ``SQLDataModel`` in the following style:

        ```shell
            ┌─────────────────────────────────┐
            │ Name   Age   Height  Birthday   │
            ├─────────────────────────────────┤
            │ Alice   28   162.08  1996-11-20 │
            │ Bobby   30   175.36  1994-06-15 │
            │ Craig   37   185.82  1987-01-07 │
            │ David   32   179.75  1992-12-28 │
            └─────────────────────────────────┘
        ```

        Set ``style = 'pandas'`` to format ``SQLDataModel`` in the style used by Pandas DataFrames:

        ```shell
            Name   Age   Height  Birthday
            Alice   28   162.08  1996-11-20
            Bobby   30   175.36  1994-06-15
            Craig   37   185.82  1987-01-07
            David   32   179.75  1992-12-28
        ```

        Set ``style = 'polars'`` to format ``SQLDataModel`` in the style used by Polars DataFrames:

        ```shell
            ┌───────┬─────┬─────────┬────────────┐
            │ Name  ┆ Age ┆  Height ┆ Birthday   │
            ╞═══════╪═════╪═════════╪════════════╡
            │ Alice ┆  28 ┆  162.08 ┆ 1996-11-20 │
            │ Bobby ┆  30 ┆  175.36 ┆ 1994-06-15 │
            │ Craig ┆  37 ┆  185.82 ┆ 1987-01-07 │
            │ David ┆  32 ┆  179.75 ┆ 1992-12-28 │
            └───────┴─────┴─────────┴────────────┘
        ```

        Set ``style = 'postgresql'`` to format ``SQLDataModel`` in the style used by PostgreSQL:

        ```shell
            Name  | Age |  Height | Birthday
            ------+-----+---------+-----------
            Alice |  28 |  162.08 | 1996-11-20
            Bobby |  30 |  175.36 | 1994-06-15
            Craig |  37 |  185.82 | 1987-01-07
            David |  32 |  179.75 | 1992-12-28
        ```
        
        Set ``style = 'rst-grid'`` to format ``SQLDataModel`` in the style required for Sphinx and reStructured text grid tables:

        ```shell
            +-------+-----+---------+------------+
            | Name  | Age |  Height | Birthday   |
            +=======+=====+=========+============+
            | Alice |  28 |  162.08 | 1996-11-20 |
            | Bobby |  30 |  175.36 | 1994-06-15 |
            | Craig |  37 |  185.82 | 1987-01-07 |
            | David |  32 |  179.75 | 1992-12-28 |
            +-------+-----+---------+------------+
        ```

        Set ``style = 'rst-simple'`` to format ``SQLDataModel`` in the style required for Sphinx and reStructured simple tables:

        ```shell
            =====  ===  =======  ==========
            Name   Age   Height  Birthday
            =====  ===  =======  ==========
            Alice   28   162.08  1996-11-20
            Bobby   30   175.36  1994-06-15
            Craig   37   185.82  1987-01-07
            David   32   179.75  1992-12-28
            =====  ===  =======  ==========
        ```

        Set ``style = 'latex'`` to format ``SQLDataModel`` in the style of a LaTeX table:

        ```text
            \\hline
                Name  & Age &  Height & Birthday   \\\\
            \\hline
                Alice &  28 &  162.08 & 1996-11-20 \\\\
                Bobby &  30 &  175.36 & 1994-06-15 \\\\
                Craig &  37 &  185.82 & 1987-01-07 \\\\
                David &  32 &  179.75 & 1992-12-28 \\\\
            \\hline
        ```

        However, :meth:`SQLDataModel.to_latex()` should be used to format complete table elements for LaTeX files.

        Note:
            - The labels given to certain styles are entirely subjective and do not in any way express original design or ownership of the styling used.
            - Legacy character sets on older terminals may not support all the character encodings required for some styles.
            - See :meth:`SQLDataModel._generate_table_style()` for implementation details related to each format.

        Changelog:
            - Version 0.11.0 (2024-07-05):
                - Added style ``'latex'`` to generate LaTeX style tables.
            - Version 0.9.3 (2024-06-28):
                - Added styles ``'rst-grid'`` and ``'rst-simple'`` to allow ``SQLDataModel`` to generate table formats used by Sphinx and reStructured Text.
            - Version 0.3.11 (2024-04-18):
                - Removed ``'thick'`` style and added ``'list'`` style for greater variety of available formats.
            - Version 0.3.8 (2024-04-12):
                - New method.            
        """
        if style not in ('ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple'):
            msg = ErrorFormat(f"ValueError: invalid value '{style}', argument for `style` must be one of 'ascii', 'bare', 'dash', 'default', 'double', 'list', 'markdown', 'outline', 'pandas', 'polars', 'postgresql', 'rst-grid', 'rst-simple' or 'round'")
            raise ValueError(msg)
        self.table_style = style

    def __repr__(self) -> str:
        """
        Returns a pretty printed string representation of ``SQLDataModel`` formatted to the current terminal size.

        Returns:
            ``str``: The string representation of the SQLDataModel instance output using display and format values set on instance.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['idx', 'first', 'last', 'age']
            data = [
                (0, 'john', 'smith', 27)
                ,(1, 'sarah', 'west', 29)
                ,(2, 'mike', 'harlin', 36)
                ,(3, 'pat', 'douglas', 42)
            ]

            # Create the model
            df = sdm.SQLDataModel(data,headers)

            # Display the string representation
            print(df)

        This will output the default alignment, dynamically aligning columns based on their dtype, right-aligned for numeric, left otherwise:

        ```shell
            ┌───┬────────┬─────────┬────────┐
            │   │ first  │ last    │    age │
            ├───┼────────┼─────────┼────────┤
            │ 0 │ john   │ smith   │     27 │
            │ 1 │ sarah  │ west    │     29 │
            │ 2 │ mike   │ harlin  │     36 │
            │ 3 │ pat    │ douglas │     42 │
            └───┴────────┴─────────┴────────┘  
            [4 rows x 3 columns]      
        ```

        Using ``'left'`` column alignment:

        ```python        
            # Using left alignment instead
            df.set_column_alignment("left")

            # See difference
            print(df)
        ```

        This will output:

        ```shell            
            ┌───┬────────┬─────────┬────────┐
            │   │ first  │ last    │ age    │
            ├───┼────────┼─────────┼────────┤
            │ 0 │ john   │ smith   │ 27     │
            │ 1 │ sarah  │ west    │ 29     │
            │ 2 │ mike   │ harlin  │ 36     │
            │ 3 │ pat    │ douglas │ 42     │
            └───┴────────┴─────────┴────────┘
            [4 rows x 3 columns]
        ```
        
        Using ``'center'`` column alignment:

        ```python        
            # Using center alignment instead
            df.set_column_alignment("center")

            # See difference
            print(df)
        ```

        This will output:

        ```shell            
            ┌───┬────────┬─────────┬────────┐
            │   │ first  │  last   │  age   │
            ├───┼────────┼─────────┼────────┤
            │ 0 │  john  │  smith  │   27   │
            │ 1 │ sarah  │  west   │   29   │
            │ 2 │  mike  │ harlin  │   36   │
            │ 3 │  pat   │ douglas │   42   │
            └───┴────────┴─────────┴────────┘
            [4 rows x 3 columns]
        ```

        Using ``'right'`` column alignment:

        ```python        
            # Using right alignment instead
            df.set_column_alignment("right")

            # See difference
            print(df)
        ```

        This will output:

        ```shell
            ┌───┬────────┬─────────┬────────┐
            │   │  first │    last │    age │
            ├───┼────────┼─────────┼────────┤
            │ 0 │   john │   smith │     27 │
            │ 1 │  sarah │    west │     29 │
            │ 2 │   mike │  harlin │     36 │
            │ 3 │    pat │ douglas │     42 │
            └───┴────────┴─────────┴────────┘
            [4 rows x 3 columns]        
        ```

        Note:
            - Use :meth:`SQLDataModel.set_display_max_rows()` to explicitly set vertical height and modify vertical truncation behavior, which uses current terminal height by default.
            - Use :meth:`SQLDataModel.set_min_column_width()` and :meth:`SQLDataModel.set_max_column_width()` to adjust column widths and modify horizontal truncation behavior.
            - Use :meth:`SQLDataModel.set_column_alignment()` to modify column alignment, available options are dynamic alignment based on dtype, left, center or right alignment.
            - Use :meth:`SQLDataModel.set_display_color()` to modify the table color, by default no color is applied with characters drawn using platform specific settings.
            - Use :meth:`SQLDataModel.set_table_style()` to modify the table style format and box characters used to draw the table.

        Changelog:
            - Version 0.12.0 (2024-07-06):
                - Changed default behavior to display a minimum of 4 rows when ``display_max_rows = None`` to retain data visibility when terminal size is below threshold.
            - Version 0.10.4 (2024-07-03):
                - Modified to escape newline characters through :meth:`utils.sqlite_printf_format()` to avoid wrapping table rows.
            - Version 0.7.0 (2024-06-08):
                - Modified horizontal truncation behavior to alternate column selection between table start and table end instead of sequential left to right ordering.
        """
        total_available_width, total_available_height = shutil.get_terminal_size()
        # Vertical offset of 6 rows/newlines required for header top bar, headers, header sub bar, table bottom bar and table dimensions
        display_max_rows = self.display_max_rows if self.display_max_rows is not None else max(total_available_height-6, 4) # Minimum rows displayed
        table_repr = self.to_string(
            index=self.display_index,
            display_max_rows=display_max_rows,
            display_max_width=total_available_width,
            min_column_width=self.min_column_width,
            max_column_width=self.max_column_width,
            float_precision=self.display_float_precision,
            column_alignment=self.column_alignment,
            table_style=self.table_style,    
            horizontal_ellipses = '⠤⠄',
            vertical_ellipses = '⠒⠂',
            display_dimensions = True,
            index_rep = ' '
        )
        return table_repr if self.display_color is None else self.display_color.wrap(table_repr)     
    
##################################################################################################################
############################################## sqldatamodel methods ##############################################
##################################################################################################################

    def append_row(self, values:list|tuple=None) -> None:
        """
        Appends ``values`` as a new row in the ``SQLDataModel`` at the next available index based on the current max row index from :py:attr:`SQLDataModel.indicies`. If ``values = None``, an empty row with SQL ``null`` values will be used.

        Parameters:
            ``values`` (list or tuple, optional): The values to be inserted into the row. If not provided or set to None, an empty row with SQL ``null`` values will be inserted.

        Raises:
            ``TypeError``: If ``values`` is provided and is not of type list or tuple.
            ``DimensionError``: If the number of values provided does not match the current column count.
            ``SQLProgrammingError``: If there is an issue with the SQL execution during the insertion.
        
        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Create a rowless model
            df = sdm.SQLDataModel(headers=['Name', 'Age'])

            # Append a row with values
            df.append_row(['Alice', 31])

            # Append another row
            df.append_row(['John', 48])

            # View result
            print(df)
        
        This will output:

        ```text
            ┌───┬───────┬──────┐
            │   │ Name  │ Age  │
            ├───┼───────┼──────┤
            │ 0 │ Alice │ 31   │
            │ 1 │ John  │ 48   │
            └───┴───────┴──────┘
            [2 rows x 2 columns]
        ```

        Note:
            - If no values are provided, ``None`` or SQL 'null' will be used for the values.
            - Rows will be appended to the bottom of the model at one index greater than the current max index.

        Changelog:
            - Version 0.6.0 (2024-05-14):
                - Mirrors previous behavior of :meth:`SQLDataModel.insert_row()` for versions <= 0.5.2.
                - New method. 
        """
        if values is not None:
            if not isinstance(values, (list,tuple)):
                msg = ErrorFormat(f"TypeError: invalid type '{type(values).__name__}', insert values must be of type 'list' or 'tuple'")
                raise TypeError(msg) from None
            if isinstance(values,list):
                values = tuple(values)
            if (len_val := len(values)) != self.column_count:
                msg = ErrorFormat(f"DimensionError: invalid dimensions '{len_val} != {self.column_count}', the number of values provided: '{len_val}' must match the current column count '{self.column_count}'")
                raise DimensionError(msg) from None
        else:
            values = tuple([None for _ in range(self.column_count)])
        insert_cols = ",".join([f'"{col}"' for col in self.headers])
        insert_vals = ",".join(["?" if not isinstance(val,datetime.date) else "datetime(?)" if isinstance(val, datetime.datetime) else "date(?)" for val in values])
        insert_stmt = f"""insert into {self.sql_model}({insert_cols}) values ({insert_vals})"""
        sql_cur = self.sql_db_conn.cursor()
        try:
            sql_cur.execute(insert_stmt, values)
            self.sql_db_conn.commit()
            self._update_indicies_deterministic(sql_cur.lastrowid)
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f'SQLProgrammingError: unable to update values, SQL execution failed with: "{e}"')
            raise SQLProgrammingError(msg) from None
        self._update_model_metadata(update_row_meta=False)  

    def concat(self, other:SQLDataModel|list|tuple, inplace:bool=True) -> None|SQLDataModel:
        """
        Concatenates the provided data to ``SQLDataModel`` along the row axis, returning a new model or modifying the existing instance inplace.

        Parameters:
            ``other`` (SQLDataModel | list | tuple): The SQLDataModel, list, or tuple to concatenate or append.
            ``inplace`` (bool, optional): If True (default), performs the concatenation in-place, modifying the current model. If False, returns a new ``SQLDataModel`` instance with the concatenated result.

        Returns:
            ``None`` or ``SQLDataModel``: ``None`` when ``inplace = True`` and ``SQLDataModel`` when ``in_place = False``

        Raises:
            ``TypeError``: If the ``other`` argument is not one of type ``SQLDataModel``, ``list``, or ``tuple``.
            ``ValueError``: If ``other`` is a list or tuple with insufficient data where the column dimension is < 1.
            ``DimensionError``: If the column count of the current model does not match the column count of the ``other`` model or tuple.

        Example::

            import sqldatamodel as sdm

            # Datasets a and b
            data_a = (['A', 1], ['B', 2])
            data_b = (['C', 3], ['D', 4])

            # Create the models
            df_a = SQLDataModel(data_a, headers=['letter', 'number'])
            df_b = SQLDataModel(data_b, headers=['letter', 'number'])

            # Concatenate the two models
            df_ab = df_a.concat(df_b, inplace=False)

            # View result
            print(df_ab)

        This will output:
        
        ```shell            
            ┌────────┬────────┐
            │ letter │ number │
            ├────────┼────────┤
            │ A      │ 1      │
            │ B      │ 2      │
            │ C      │ 3      │
            │ D      │ 4      │
            └────────┴────────┘
            [4 rows x 2 columns]
        ```

        Concatenation can be done using other objects as well:

        ```python            
            # List or tuples can also be used directly
            data_e = ['E', 5]

            # Append in place
            df_ab.concat(data_e)

            # View result
            print(df_ab)
        ```

        This will output:

        ```shell            
            ┌───┬────────┬────────┐
            │   │ letter │ number │
            ├───┼────────┼────────┤
            │ 0 │ A      │      1 │
            │ 1 │ B      │      2 │
            │ 2 │ C      │      3 │
            │ 3 │ D      │      4 │
            │ 4 │ E      │      5 │
            └───┴────────┴────────┘
            [5 rows x 2 columns]        
        ```

        Note:
            - Models must be of compatible dimensions with equal ``column_count`` or equivalent dimension if ``list`` or ``tuple``
            - Headers are inherited from the model calling the :meth:`SQLDataModel.concat()` method whether done inplace or being returned as new instance.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        if not isinstance(other, (SQLDataModel,list,tuple)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(other).__name__}', argument for ``other`` must be of type 'SQLDataModel' to concatenate compatible models")
            raise TypeError(msg)
        if isinstance(other, SQLDataModel):
            num_cols_other = other.column_count
            num_rows_other = other.row_count
            other = other.data()
        elif isinstance(other, (list,tuple)):
            if len(other) < 1:
                msg = ErrorFormat(f"ValueError: insufficient data length '{len(other)}', argument ``other`` must have length >= 1 or contain at least 1 row to concatenate")
                raise ValueError(msg)
            if not isinstance(other[0], (list,tuple)):
                other = [other]
            num_cols_other = len(other[0])
            num_rows_other = len(other)
        if self.column_count != num_cols_other:
            msg = ErrorFormat(f"DimensionError: incompatible column count '{num_cols_other}', cannot concatenate model shape '{self.get_shape()}' to model shape '({num_rows_other}, {num_cols_other})' as same number of columns are required")
            raise DimensionError(msg)
        if inplace:
            sql_insert_stmt = f"""insert into "{self.sql_model}" ({','.join([f'"{col}"' for col in self.headers])}) values ({','.join(['?' if self.header_master[col][1] not in ('datetime','date') else "datetime(?)" if self.header_master[col][1] == 'datetime' else "date(?)" for col in self.headers])})"""
            self.sql_db_conn.executemany(sql_insert_stmt, other)
            self._update_model_metadata(update_row_meta=True)
            return
        else:
            return type(self)((*self.data(),*other), self.headers, display_max_rows=self.display_max_rows, min_column_width=self.min_column_width, max_column_width=self.max_column_width, column_alignment=self.column_alignment, display_color=self.display_color, display_index=self.display_index, display_float_precision=self.display_float_precision)

    def copy(self, data_only:bool=False) -> SQLDataModel:
        """
        Returns a deep copy of the current model as a new ``SQLDataModel``.

        Parameters:
            ``data_only`` (bool): If True, only the data is copied, otherwise display and styling parameters are included. Default is False.

        Returns:
            ``SQLDataModel``: A cloned copy from the original as a new ``SQLDataModel``.  
        
        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height']
            data = [
                ('John', 30, 175.3), 
                ('Alice', 28, 162.0), 
                ('Travis', 35, 185.8)
            ]    

            # Create the original model with list styling
            df = sdm.SQLDataModel(data, headers, table_style='list')

            # Create two copies, one full and one with data only
            copy_full = df.copy()
            copy_data = df.copy(data_only=True)

            # View both copies
            print(copy_full)
            print(copy_data)   

        This will output both copies, with ``copy_full`` including any styling parameters such as ``table_style='list'``:

        ```shell
               Name    Age   Height
            -  ------  ---  -------
            0  John     30   175.30
            1  Alice    28   162.00
            2  Travis   35   185.80
        ```

        With the output for ``copy_data`` containing only the original model's data:

        ```shell
            ┌───┬────────┬─────┬─────────┐
            │   │ Name   │ Age │  Height │
            ├───┼────────┼─────┼─────────┤
            │ 0 │ John   │  30 │  175.30 │
            │ 1 │ Alice  │  28 │  162.00 │
            │ 2 │ Travis │  35 │  185.80 │
            └───┴────────┴─────┴─────────┘
        ```

        Note:
            - Model headers and dtypes are considered part of the model data and are included when ``data_only=True``.
            - Default behavior, ``data_only=False``, includes the following additional display parameters:

              - :py:attr:`SQLDataModel.display_max_rows`: The maximum number of rows to display.
              - :py:attr:`SQLDataModel.min_column_width`: The minimum width of columns when displaying the model.
              - :py:attr:`SQLDataModel.max_column_width`: The maximum width of columns when displaying the model.
              - :py:attr:`SQLDataModel.column_alignment`: The alignment of columns ('left', 'center', 'right' or 'dynamic').
              - :py:attr:`SQLDataModel.display_color`: The color to use when displaying the table, None by default.
              - :py:attr:`SQLDataModel.display_index`: True if displaying index column, False otherwise.
              - :py:attr:`SQLDataModel.display_float_precision`: The precision for displaying floating-point numbers.
              - :py:attr:`SQLDataModel.table_style`: The table styling format to use for strng representations of the model.

        Changelog:
            - Version 0.4.2 (2024-05-03):
                - New method.
        """ 
        if data_only:
            return type(self)(self.data(index=True), headers=[self.sql_idx, *self.headers], dtypes=self.dtypes)
        else:
            return type(self)(self.data(index=True), headers=[self.sql_idx, *self.headers], dtypes=self.dtypes, **self._get_display_args())   

    def count(self) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` containing the counts of non-null values for each column in a row-wise orientation.

        Returns:
            ``SQLDataModel``: A new SQLDataModel containing the counts of non-null values in each column.

        Example::
            
            import sqldatamodel as sdm

            # Sample data with missing values
            headers = ['Name', 'Age', 'Gender', 'Tenure']
            data = [
                ('Alice', 25, 'Female', 1.0),
                ('Bob', None, 'Male', 2.7),
                ('Charlie', 30, 'Male', None),
                ('David', None, 'Male', 3.8)
            ]   

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Get counts
            counts = df.count()

            # View result
            print(counts)

        This will output the count of all non-null values for each column:

        ```shell
            ┌──────┬─────┬────────┬────────┐
            │ Name │ Age │ Gender │ Tenure │
            ├──────┼─────┼────────┼────────┤
            │    4 │   2 │      4 │      3 │
            └──────┴─────┴────────┴────────┘
            [1 rows x 4 columns]
        ```

        Note:
            - See :meth:`SQLDataModel.count_unique()` for column-wise count of unique, null and total values for each column.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        fetch_stmt = " ".join(("select",",".join([f"""sum(case when "{col}" is null then 0 else 1 end) as "{col}" """ for col in self.headers]),f'from "{self.sql_model}"'))
        return self.execute_fetch(fetch_stmt, dtypes={col:'int' for col in self.headers})

    def count_unique(self) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` containing the total counts and unique values for each column in the model for both null and non-null values.

        Metrics:
            - ``'column'`` contains the names of the columns counted.
            - ``'na'`` contains the total number of null values in the column.
            - ``'unique'`` contains the total number of unique values in the column.
            - ``'count'`` contains the total number of non-null values in the column.
            - ``'total'`` contains the total number of all null and non-null values in the column.
            
        Returns:
            ``SQLDataModel``: A new SQLDataModel containing columns 'column', 'unique', and 'count' representing the column name, total unique values, and total values count, respectively.
        
        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Gender']
            data = [
                ('Alice', 25, 'Female'), 
                ('Bob', 30, None), 
                ('Alice', 25, 'Female')
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Get the value count information
            count_model = df.count_unique()

            # View the count information
            print(count_model)

        This will output:
        
        ```shell            
            ┌────────┬──────┬────────┬───────┬───────┐
            │ column │   na │ unique │ count │ total │
            ├────────┼──────┼────────┼───────┼───────┤
            │ Name   │    0 │      2 │     3 │     3 │
            │ Age    │    0 │      2 │     3 │     3 │
            │ Gender │    1 │      1 │     2 │     3 │
            └────────┴──────┴────────┴───────┴───────┘
            [3 rows x 5 columns]
        ```

        Note:
            - See :meth:`SQLDataModel.count()` for the count of non-null values for each column in a row-wise orientation.

        Changelog:
            - Version 0.3.2 (2024-04-02):
                - Renamed method from ``counts`` to ``count_unique`` for more precise definition.
                - New method.
        """
        fetch_stmt = " UNION ALL ".join([f"""select '{col}' as 'column', sum(case when "{col}" is null then 1 else 0 end) as 'na', count(distinct "{col}") as 'unique', count("{col}") as 'count',sum(case when "{col}" is null then 1 else 1 end) as 'total' from "{self.sql_model}" """ for col in self.headers])
        return self.execute_fetch(fetch_stmt)

    def deduplicate(self, subset:list[str]=None, reset_index:bool=True, keep_first:bool=True, inplace:bool=True) -> None|SQLDataModel:
        """
        Removes duplicate rows from the SQLDataModel based on the specified subset of columns. Deduplication occurs inplace by default, otherwise use ``inplace=False`` to return a new ``SQLDataModel``.

        Parameters:
            ``subset`` (list[str], optional): List of columns to consider when identifying duplicates. If None, all columns are considered. Defaults to None.
            ``reset_index`` (bool, optional): If True, resets the index after deduplication starting at 0; otherwise retains current indicies.
            ``keep_first`` (bool, optional): If True, keeps the first occurrence of each duplicated row; otherwise, keeps the last occurrence. Defaults to True.
            ``inplace`` (bool, optional): If True, modifies the current SQLDataModel in-place; otherwise, returns a new SQLDataModel without duplicates. Defaults to True.

        Raises:
            ``ValueError``: If a column specified in ``subset`` is not found in the SQLDataModel.

        Returns:
            ``None`` or ``SQLDataModel``: If ``inplace = True`` the method modifies the current SQLDataModel in-place return ``None``, otherwise if ``inplace = False`` a new ``SQLDataModel`` is returned.

        Examples:

        Based on Single Column
        ----------------------

        ```python
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Deduplicate based on a specific column
            df.deduplicate(subset='ID', keep_first=True, inplace=True)

        ```
        Based on Multiple Columns
        -------------------------

        ```python
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['ID', 'Name', 'Value'])

            # Deduplicate based on multiple columns and save to keep both models
            df_deduped = df.deduplicate(subset=['ID', 'Name'], keep_first=False, inplace=False)
        ```

        Note:
            - Ordering for ``keep_first`` is determined by the current :py:attr:`SQLDataModel.sql_idx` order of the instance.
            - For multiple columns ordering is done sequentially favoring first index in ``subset``, then i+1, ..., to ``i+len(subset)``

        Changelog:
            - Version 0.1.5 (2023-11-24):
                - New method.
        """        
        dyn_keep_order = 'min' if keep_first else 'max'
        if subset is not None:
            if isinstance(subset, str):
                subset = [subset]
            for col in subset:
                if col not in self.headers:
                    msg = ErrorFormat(f"ValueError: column not found '{col}', provided columns in `subset` must be valid columns, use `get_headers()` to view current valid column names")
                    raise ValueError(msg)
        else:
            subset = self.headers
        if inplace:
            sql_stmt = f"""delete from "{self.sql_model}" where rowid not in (select {dyn_keep_order}(rowid) from "{self.sql_model}" group by {','.join(f'"{col}"' for col in subset)})"""
            self.sql_db_conn.execute(sql_stmt)
            self.sql_db_conn.commit()
            if reset_index:
                self.reset_index()
                return
            self._update_model_metadata(update_row_meta=True)
            return
        else:
            index_str = f'"{self.sql_idx}",' if not reset_index else ''
            headers_str = ",".join([f'"{col}"' for col in self.headers])
            sql_stmt = f"""select {index_str}{headers_str} from "{self.sql_model}" where rowid in (select {dyn_keep_order}(rowid) from "{self.sql_model}" group by {','.join(f'"{col}"' for col in subset)})"""
            return self.execute_fetch(sql_stmt)

    def fillna(self, value, strictly_null:bool=False, inplace:bool=True) -> None|SQLDataModel:
        """
        Fills missing (na or nan) values in the current ``SQLDataModel`` with the provided ``value`` inplace or as a new instance.

        Parameters:
            ``value``: The scalar value to fill missing values with. Should be of type 'str', 'int', 'float', 'bytes', or 'bool'.
            ``inplace`` (bool): If True, modifies the current instance in-place. If False, returns a new instance with missing values filled.
            ``strictly_null`` (bool): If True, only strictly null values are filled. If False, values like ``'NA'``, ``'NaN'``, ``'n/a'``, ``'na'``, and whitespace only strings are also filled.

        Raises:
            ``TypeError``: If ``value`` is not a scalar type or is incompatible with SQLite's type system.

        Returns:
            ``None`` or ``SQLDataModel``: When ``inplace=True`` modifies model inplace, returning ``None``, when ``inplace=False`` a new ``SQLDataModel`` is returned.

        Example::
            
            import sqldatamodel as sdm

            # Create sample data
            data = [('Alice', 25, None), ('Bob', None, 'N/A'), ('Charlie', 'NaN', ' '), ('David', 30, 'NA')]

            # Create the model
            df = SQLDataModel(data, headers=['Name', 'Age', 'Status'])

            # Fill missing values with 0
            df_filled = df.fillna(value=0, strictly_null=False, inplace=False)

            # View filled model
            print(df_filled)
        
        This will output:
        
        ```shell            
            ┌───┬─────────┬──────┬────────┐
            │   │ Name    │  Age │ Status │
            ├───┼─────────┼──────┼────────┤
            │ 0 │ Alice   │   25 │      0 │
            │ 1 │ Bob     │    0 │      0 │
            │ 2 │ Charlie │    0 │      0 │
            │ 3 │ David   │   30 │      0 │
            └───┴─────────┴──────┴────────┘
            [4 rows x 3 columns]    
        ```

        Note:
            - The method supports filling missing values with various scalar types which are then adapted to the columns set dtype.
            - The ``strictly_null`` parameter controls whether additional values like ``('NA', 'NAN', 'n/a', 'na', '')`` with last being an empty string, are treated as null.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(value, (str,int,float,bytes,bool)) and value is not None:
            msg = ErrorFormat(f"TypeError: invalid type '{type(value).__name__}', argument for `value` must be scalar type or one of 'str', 'int', 'bytes', 'bool' or 'float'")
            raise TypeError(msg)
        na_values = ('none','null','nan','n/a','na','')
        if inplace:
            for col in self.headers:
                null_predicate = "" if strictly_null else f'or (trim(lower("{col}")) in {na_values})'
                fillna_col_stmt = f"""update "{self.sql_model}" set "{col}" = ? where "{col}" is null {null_predicate}"""
                self.sql_db_conn.execute(fillna_col_stmt,(value,))
            return
        fetch_na_stmt = ",".join([f"""case when lower(trim("{col}")) in {na_values} or "{col}" is null then ? else "{col}" end as '{col}'""" if not strictly_null else f"""case when "{col}" is null then ? else "{col}" end as '{col}'""" for col in self.headers])
        fetch_na_stmt, fetch_na_params = f"""select {fetch_na_stmt} from "{self.sql_model}" """, tuple(value for _ in range(self.column_count))
        return self.execute_fetch(fetch_na_stmt,fetch_na_params)

    def group_by(self, columns:str|list[str], order_by_count:bool=True) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` after performing a group by operation on specified columns.

        Parameters:
            ``columns`` (str, list, tuple): Columns to group by. Accepts either individual strings or a list/tuple of strings.
            ``order_by_count`` (bool, optional): If True (default), orders the result by count. If False, orders by the specified columns.

        Raises:
            ``TypeError``: If the columns argument is not of type str, list, or tuple.
            ``ValueError``: If any specified column does not exist in the current model.
            ``SQLProgrammingError``: If any specified columns or aggregate keywords are invalid or incompatible with the current model.
        
        Returns:
            ``SQLDataModel``: A new ``SQLDataModel`` instance containing the result of the group by operation.

        Example::

            import sqldatamodel as sdm

            headers = ['first', 'last', 'age', 'service', 'hire_date', 'gender']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01', 'Male'),
                ('Sarah', 'West', 39, 0.7, '2023-10-01', 'Female'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27', 'Male'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06', 'Male'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18', 'Female')
            ]        
            # Create the model
            df = sdm.SQLDataModel(data, headers, display_float_precision=2, display_index=True)

            # Group by 'gender' column
            df_gender = df.group_by("gender")

            # View model
            print(df_gender)
        
        This will output:

        ```shell            
            ┌───┬────────┬───────┐
            │   │ gender │ count │
            ├───┼────────┼───────┤
            │ 0 │ Male   │     3 │
            │ 1 │ Female │     2 │
            └───┴────────┴───────┘
            [2 rows x 2 columns]     
        ```

        Multiple columns can also be used to group by:

        ```python
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('data.csv')

            # Group by multiple columns
            df.group_by(["country", "state", "city"])
        ```

        Note:
            - Use ``order_by_count=False`` to change ordering from count to column arguments provided.
            - See :meth:`SQLDataModel.describe()` for generating descriptive statistics by column data type.
            - See :meth:`SQLDataModel.pivot()` for creating a pivot table using categorization and aggregate functions.

        Changelog:
            - Version 0.8.0 (2024-06-21):
                - Modified to allow ``columns`` to be referenced by their integer index as well as directly to allow broader inputs and reflect similar access patterns across package.
        """
        columns = self._validate_column(columns, unmodified=False) # +VALCOL
        columns_group_by = ",".join(f'"{col}"' for col in columns)
        order_by = "count(*)" if order_by_count else columns_group_by
        group_by_stmt = f"""select {columns_group_by}, count(*) as count from "{self.sql_model}" group by {columns_group_by} order by {order_by} desc"""
        return self.execute_fetch(group_by_stmt)

    def head(self, n_rows:int=5) -> SQLDataModel:
        """
        Returns the first ``n_rows`` of the current ``SQLDataModel``.

        Parameters:
            ``n_rows`` (int, optional): Number of rows to return. Defaults to 5.

        Raises:
            ``TypeError``: If ``n_rows`` argument is not of type 'int' representing the number of rows to return from the head of the model.

        Returns:
            ``SQLDataModel``: A new ``SQLDataModel`` instance containing the specified number of rows.

        Example::
        
            import sqldatamodel as sdm

            # Countries data available for sample dataset
            url = 'https://developers.google.com/public-data/docs/canonical/countries_csv'

            # Create the model
            df = sdm.from_html(url)

            # Get head of model
            df_head = df.head()

            # View it
            print(df_head)

        This will grab the top 5 rows by default:

        ```shell
            ┌───┬─────────┬──────────┬───────────┬────────────────┐
            │   │ country │ latitude │ longitude │ name           │
            ├───┼─────────┼──────────┼───────────┼────────────────┤
            │ 0 │ AF      │  33.9391 │   67.7100 │ Afghanistan    │
            │ 1 │ AL      │  41.1533 │   20.1683 │ Albania        │
            │ 2 │ DZ      │  28.0339 │    1.6596 │ Algeria        │
            │ 3 │ AS      │ -14.2710 │ -170.1322 │ American Samoa │
            │ 4 │ AD      │  42.5462 │    1.6016 │ Andorra        │
            └───┴─────────┴──────────┴───────────┴────────────────┘
            [5 rows x 4 columns]
        ```

        Note:
            - See related :meth:`SQLDataModel.tail()` for the opposite, grabbing the bottom ``n_rows`` from the current model.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(n_rows, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(n_rows).__name__}', argument for `n_rows` must be of type 'int' representing the number of rows to return from the head of the model")
            raise TypeError(msg)
        n_rows = min(self.row_count, max(1, n_rows))        
        row_indicies = self.indicies[:n_rows]
        return self.execute_fetch(self._generate_sql_stmt(rows=row_indicies))
                    
    def hstack(self, *other:SQLDataModel, inplace:bool=False) -> SQLDataModel:
        """
        Horizontally stacks one or more ``SQLDataModel`` objects to the current model.

        Parameters:
            ``other`` (SQLDataModel or sequence of): The SQLDataModel objects to horizontally stack.
            ``inplace`` (bool, optional): If True, performs the horizontal stacking in-place, modifying the current model. Defaults to False, returning a new ``SQLDataModel``.

        Returns:
            ``SQLDataModel``: The horizontally stacked SQLDataModel instance when inplace is False.

        Raises:
            ``ValueError``: If no additional SQLDataModels are provided for horizontal stacking.
            ``TypeError``: If any argument in 'other' is not of type SQLDataModel, list, or tuple.
            ``SQLProgrammingError``: If an error occurs when updating the model values in place.            

        Example::

            import sqldatamodel as sdm

            # Create models A and B
            df_a = sdm.SQLDataModel([('A', 'B'), ('1', '2')], headers=['A1', 'A2'])
            df_b = sdm.SQLDataModel([('C', 'D'), ('3', '4')], headers=['B1', 'B2'])

            # Horizontally stack B onto A
            df_ab = df_a.hstack(df_b)

            # View stacked model
            print(df_ab)

        This will output the result of stacking B onto A, using each model's headers and dtypes:    

        ```shell
            ┌─────┬─────┬─────┬─────┐
            │ A1  │ A2  │ B1  │ B2  │
            ├─────┼─────┼─────┼─────┤
            │ A   │ B   │ C   │ D   │
            │ 1   │ 2   │ 3   │ 4   │
            └─────┴─────┴─────┴─────┘
            [2 rows x 4 columns]
        ```

        Multiple models can be stacked simultaneously, here we stack a total of 3 models:

        ```python
            # Create a third model C
            df_c = sdm.SQLDataModel([('E', 'F'), ('5', '6')], headers=['C1', 'C2'])

            # Horizontally stack three models
            df_abc = df_a.hstack([df_b, df_c])

            # View stacked result
            print(df_abc)
        ```

        This will output the result of stacking C and B onto A:

        ```shell
            ┌─────┬─────┬─────┬─────┬─────┬─────┐
            │ A1  │ A2  │ B1  │ B2  │ C1  │ C2  │
            ├─────┼─────┼─────┼─────┼─────┼─────┤
            │ A   │ B   │ C   │ D   │ E   │ F   │
            │ 1   │ 2   │ 3   │ 4   │ 5   │ 6   │
            └─────┴─────┴─────┴─────┴─────┴─────┘
            [2 rows x 6 columns]
        ```

        Note:
            - Model dimensions will be truncated or padded to coerce compatible dimensions when stacking, use :meth:`SQLDataModel.merge()` for strict SQL joins instead of hstack.
            - Headers and data types are inherited from all the models being stacked, this requires aliasing duplicate column names if present, see :meth:`utils.alias_duplicates()` for aliasing rules.
            - Use ``setitem`` syntax such as ``sdm['New Column'] = values`` to create new columns directly into the current model instead of stacking or see :meth:`SQLDataModel.add_column_with_values()` for convenience method accomplishing the same.
            - See :meth:`SQLDataModel.vstack()` for vertical stacking.

        Changelog:
            - Version 0.3.4 (2024-04-05):
                - New method.
        """
        other = list(other[0]) if len(other) == 1 and isinstance(other[0], (list,tuple)) else list(other)
        if len(other) < 1:
            msg = ErrorFormat(f"ValueError: insufficient model count '{len(other)}', at least 1 additional 'SQLDataModel' is required to horizontally stack against")
            raise ValueError(msg)
        if not all(isinstance(sdm, SQLDataModel) for sdm in other):
            msg = ErrorFormat(f"TypeError: invalid type encountered in '{[type(other[n]).__name__ for n in other]}', arguments for `other` must all be of type 'SQLDataModel' to horizontally stack")
            raise TypeError(msg)
        other_headers, other_dtypes = zip(*[(col[0], col[1]) for sdm in other for col in sdm.dtypes.items()])
        other_headers, other_dtypes = list(utils.alias_duplicates([*self.headers,*other_headers])), [*self.dtypes.values(), *other_dtypes]
        other_data = [(sdm[:self.row_count].data(strict_2d=True)) + [tuple(None for _ in range(sdm.column_count)) for _ in range(self.row_count - sdm.row_count)] for sdm in other]
        other_data = [tuple(item for sublist in row for item in sublist) for row in list(zip(*other_data))]
        if inplace:
            other_headers = other_headers[self.column_count:] # since in place remove current model's headers
            other_dtypes = other_dtypes[self.column_count:] # since in place remove current model's dtypes
            update_sql_script = ";".join([f"""alter table "{self.sql_model}" add column "{col_name}" {self.static_py_to_sql_map_dict.get(col_type, 'TEXT')}""" for col_name, col_type in zip(other_headers, other_dtypes)])
            col_val_param = ','.join([f""" "{col}" = {utils.sqlite_cast_type_format(param='?', dtype=dtype)} """ for col,dtype in zip(other_headers, other_dtypes)])
            update_stmt = f"""update "{self.sql_model}" set {col_val_param} where {self.sql_idx} = ?"""
            update_params = [(*other_data[i], row) for i,row in enumerate(self.indicies)]
            try:
                self.execute_transaction(update_sql_script)
                self.sql_db_conn.executemany(update_stmt, update_params)
                self.sql_db_conn.commit()
            except sqlite3.ProgrammingError as e:
                self.sql_db_conn.rollback()
                msg = ErrorFormat(f"SQLProgrammingError: invalid update values, SQL execution failed with '{e}'")
                raise SQLProgrammingError(msg) from None
            self._update_model_metadata()
            return
        else:
            other_data = [(t1 + t2) for t1, t2 in zip(self.data(index=False, include_headers=False, strict_2d=True), other_data)]
            dtype_dict = dict(zip(other_headers, other_dtypes))
            return type(self)(data=other_data, headers=other_headers, dtypes=dtype_dict, **self._get_display_args())

    def insert_row(self, index:int, values:list|tuple, on_conflict:Literal['replace','ignore']='replace') -> None:
        """
        Inserts a new row into the ``SQLDataModel`` at the specified ``index`` with the provided ``values``.

        Parameters:
            ``index`` (int): The position at which to insert the row.
            ``values`` (list or tuple): The values to be inserted into the row.
            ``on_conflict`` (Literal['replace', 'ignore'], optional): Specifies the action to take if the index already exists. Default is 'replace'.

        Raises:
            ``TypeError``: If ``index`` is not an integer or ``values`` is not a list or tuple.
            ``ValueError``: If ``on_conflict`` is not ``'replace'`` or ``'ignore'``.
            ``DimensionError``: If the dimensions of the provided ``values`` are incompatible with the current model dimensions.
            ``SQLProgrammingError``: If there is an issue with the SQL execution during the insertion.

        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Sample data
            data = [('Alice', 20, 'F'), ('Billy', 25, 'M'), ('Chris', 30, 'M')]

            # Create the model
            df = sdm.SQLDataModel(data, headers=['Name','Age','Sex'])    

            # Insert a new row at index 3
            df.insert_row(3, ['David', 35, 'M'])

            # Insert or replace row at index 1
            df.insert_row(1, ['Beth', 27, 'F'], on_conflict='replace')
            
            # View result
            print(df)

        This will output the modified model:

        ```text
            ┌───┬───────┬─────┬─────┐
            │   │ Name  │ Age │ Sex │
            ├───┼───────┼─────┼─────┤
            │ 0 │ Alice │  20 │ F   │
            │ 1 │ Beth  │  27 │ F   │
            │ 2 │ Chris │  30 │ M   │
            │ 3 │ David │  35 │ M   │
            └───┴───────┴─────┴─────┘
            [4 rows x 3 columns]
        ```

        Note:
            - Use ``on_conflict = 'ignore'`` to take no action if row already exists, and ``on_conflict = 'replace'`` to replace it.
            - See :meth:`SQLDataModel.append_row()` for appending rows at the next available index instead of insertion at index.

        Changelog:
            - Version 0.6.0 (2024-05-14):
                - Added ``index`` and ``on_conflict`` parameters for greater specificity and to align with broader conventions surrounding insert methods.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        if not isinstance(index, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(index).__name__}', argument for `index` must be of type 'int' representing the row position to insert into")
            raise TypeError(msg)
        if not isinstance(values, (list,tuple)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(values).__name__}', insert values must be of type 'list' or 'tuple'")
            raise TypeError(msg)
        if on_conflict not in ('replace','ignore'):
            msg = ErrorFormat(f"ValueError: invalid value '{on_conflict}', argument for `on_conflict` must be either 'replace' or 'ignore' representing the action to take if specified index already exists")
            raise ValueError(msg)
        if isinstance(values,list):
            values = tuple(values)
        if (len_val := len(values)) != self.column_count:
            msg = ErrorFormat(f"DimensionError: invalid dimensions '{len_val} != {self.column_count}', the number of values provided: '{len_val}' must match the current column count '{self.column_count}'")
            raise DimensionError(msg)
        values = (index,*values)
        insert_cols = self.headers if index is None else [self.sql_idx, *self.headers]
        insert_cols = ",".join([f'"{col}"' for col in insert_cols])
        insert_vals = ",".join(["?" if not isinstance(val,datetime.date) else "datetime(?)" if isinstance(val, datetime.datetime) else "date(?)" for val in values])
        insert_stmt = f"""insert or {on_conflict} into {self.sql_model}({insert_cols}) values ({insert_vals})"""
        sql_cur = self.sql_db_conn.cursor()
        try:
            sql_cur.execute(insert_stmt, values)
            self.sql_db_conn.commit()
            self._update_indicies_deterministic(sql_cur.lastrowid)
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f'SQLProgrammingError: unable to update values, SQL execution failed with: "{e}"')
            raise SQLProgrammingError(msg) from None
        self._update_model_metadata(update_row_meta=False)

    def iter_rows(self, min_row:int=None, max_row:int=None, index:bool=True, include_headers:bool=False) -> Iterator[tuple]:
        """
        Returns an iterator over the specified rows in the current ``SQLDataModel``.

        Parameters:
            ``min_row`` (int, optional): The minimum row index to start iterating from (inclusive). Defaults to None.
            ``max_row`` (int, optional): The maximum row index to iterate up to (exclusive). Defaults to None.
            ``index`` (bool, optional): Whether to include the row index in the output. Defaults to True.
            ``include_headers`` (bool, optional): Whether to include headers as the first row. Defaults to False.

        Yields:
            ``Iterator[tuple]``: An iterator containing the rows from the specified range with headers as the first row if specified.
        
        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['First', 'Last', 'Salary'])

            # Iterate over the rows
            for row in df.iter_rows(min_row=2, max_row=4):
                pass # Do stuff
        
        Note:
            - Rows are referenced by their index and not their value. E.g., ``min_row = 0`` and ``max_row = -1`` will reference the first and last rows, respectively.
            - See :meth:`SQLDataModel.iter_tuples()` for iterating over rows as named tuples.    

        Changelog:
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        min_row, max_row = min_row if min_row is not None else 0, max_row if max_row is not None else self.row_count
        res = self.sql_db_conn.execute(self._generate_sql_stmt(index=index, rows=slice(min_row,max_row)))
        if include_headers:
            yield tuple(x[0] for x in res.description)
        yield from (res)
    
    def iter_tuples(self, index:bool=False) -> Iterator[NamedTuple]:
        """
        Returns an iterator of rows from the current ``SQLDataModel`` as namedtuples using headers as field names.

        Parameters:
            ``index`` (bool, optional): Whether to include the index column in the namedtuples. Default is False.

        Raises:
            ``ValueError``: Raised if headers are not valid Python identifiers. Use :meth:`SQLDataModel.normalize_headers()` method to fix.

        Yields:
            ``Iterator[NamedTuple]``: An iterator of namedtuples for each row using current headers for field names.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['First', 'Last', 'Salary'])

            # Iterate over the namedtuples
            for row_tuple in df.iter_tuples(index=True):
                pass # Do stuff with namedtuples

        Note:
            - See :meth:`SQLDataModel.iter_rows()` for iterating over rows with custom start and stop indicies.
        
        Changelog:
            - Version 0.10.0 (2024-06-29):
                - Renamed ``include_idx_col`` parameter to ``index`` for package consistency.
                - Modified to use :meth:`SQLDataModel._generate_sql_stmt_fetchall()` to leverage deterministic behavior of method.        
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        try:
            Row = namedtuple('Row', [self.sql_idx,*self.headers] if index else self.headers)
        except ValueError as e:
            msg = ErrorFormat(f'ValueError: {e}, rename header or use `normalize_headers()` method to fix')
            raise ValueError(msg) from None
        res = self.sql_db_conn.execute(self._generate_sql_stmt_fetchall(index=index))
        yield from (Row(*x) for x in res.fetchall())
    
    def mean(self) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` containing the mean value of all viable columns in the current model. Calculated by ``sum(x_i, ..., x_n) * (1 / N)``

        Returns:
            ``SQLDataModel``: A new SQLDataModel containing the mean values of each column.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Birthday', 'Height', 'Date of Hire']
            data = [
                ('John', 30, '1994-06-15', 175.3, '2018-03-03 11:20:19'), 
                ('Alice', 28, '1996-11-20', 162.0, '2023-04-24 08:45:30'), 
                ('Travis', 37, '1987-01-07', 185.8, '2012-10-06 15:30:40')
            ]

            # Create the model and infer correct types
            df = sdm.SQLDataModel(data, headers, infer_dtypes=True)

            # View full model
            print(df)
        
        This will output the sample model we'll be using to calculate mean values for:

        ```shell
            ┌────────┬─────┬────────────┬─────────┬─────────────────────┐
            │ Name   │ Age │ Birthday   │  Height │ Date of Hire        │
            ├────────┼─────┼────────────┼─────────┼─────────────────────┤
            │ John   │  30 │ 1994-06-15 │  175.30 │ 2018-03-03 11:20:19 │
            │ Alice  │  28 │ 1996-11-20 │  162.00 │ 2023-04-24 08:45:30 │
            │ Travis │  37 │ 1987-01-07 │  185.80 │ 2012-10-06 15:30:40 │
            └────────┴─────┴────────────┴─────────┴─────────────────────┘
            [3 rows x 5 columns]        
        ```

        Now let's find the mean values:

        ```python
            # Calculate the mean values
            df_mean = df.mean()

            # View result
            print(df_mean)
        ```

        This will output the mean values for the "Age", "Birthday", "Height" and "Date of Hire" columns:

        ```shell
            ┌──────┬────────┬────────────┬─────────┬─────────────────────┐
            │ Name │    Age │ Birthday   │  Height │ Date of Hire        │
            ├──────┼────────┼────────────┼─────────┼─────────────────────┤
            │ NaN  │  31.67 │ 1992-10-14 │  174.37 │ 2018-01-30 11:52:09 │
            └──────┴────────┴────────────┴─────────┴─────────────────────┘
            [1 rows x 5 columns]
        ```

        Note:
            - Only non-null values are included in the calculation of the sum and the total number of values in the column, use :meth:`SQLDataModel.fillna()` to fill null values.
            - For ``date`` and ``datetime`` columns values are converted to julian days prior to calculation and recast into original data type, some imprecision may occur as a result.
            - See :meth:`SQLDataModel.min()` for returning the minimum value, :meth:`SQLDataModel.max()` for maximum value, and :meth:`SQLDataModel.describe()` for descriptive statical values.

        Changelog:
            - Version 0.3.7 (2024-04-10):
                - New method.
        """
        fetch_stmt = ",".join([f"""avg("{col}") as "{col}" """ if dtype in ('int', 'float','bool') else f"""{dtype}(avg(julianday("{col}"))) as "{col}" """ if dtype in ('date','datetime') else f"""'NaN' as "{col}" """ for col,dtype in self.dtypes.items()])
        fetch_stmt = " ".join(["select", fetch_stmt, f"""from "{self.sql_model}" """])
        dtypes = {col:dtype if dtype not in ('int','bytes') else 'float' if dtype == 'int' else 'str' for col,dtype in self.dtypes.items()}
        return self.execute_fetch(fetch_stmt, dtypes=dtypes)

    def min(self) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` containing the minimum value of all non-null values for each column in a row-wise orientation.

        Returns:
            ``SQLDataModel``: A new SQLDataModel containing the minimum non-null value for each column.

        Example::
            
            import sqldatamodel as sdm

            # Sample data with missing values
            headers = ['Name', 'Age', 'Gender', 'Tenure']
            data = [
                ('Alice', 25, 'Female', 1.0),
                ('Bob', None, 'Male', 2.7),
                ('Charlie', 30, 'Male', None),
                ('David', None, 'Male', 3.8)
            ]   

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Get minimum values
            min_values = df.min()

            # View result
            print(min_values)

        This will output the minimum value of all non-null values for each column:

        ```shell
            ┌───────┬─────┬────────┬────────┐
            │ Name  │ Age │ Gender │ Tenure │
            ├───────┼─────┼────────┼────────┤
            │ Alice │  25 │ Female │   1.00 │
            └───────┴─────┴────────┴────────┘
            [1 rows x 4 columns]
        ```

        Note:
            - See :meth:`SQLDataModel.count_unique()` for column-wise count of unique, null and total values for each column.
            - See :meth:`SQLDataModel.max()` for returning the maximum values in each column.

        Changelog:
            - Version 0.3.1 (2024-04-01):
                - New method.
        """
        fetch_stmt = " ".join(("select",",".join([f"""min("{col}") as "{col}" """ for col in self.headers]),f'from "{self.sql_model}"'))
        return self.execute_fetch(fetch_stmt) 

    def max(self) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` containing the maximum value of all non-null values for each column in a row-wise orientation.

        Returns:
            ``SQLDataModel``: A new SQLDataModel containing the maximum non-null value for each column.

        Example::
            
            import sqldatamodel as sdm

            # Sample data with missing values
            headers = ['Name', 'Age', 'Gender', 'Tenure']
            data = [
                ('Alice', 25, 'Female', 1.0),
                ('Bob', None, 'Male', 2.7),
                ('Charlie', 30, 'Male', None),
                ('David', None, 'Male', 3.8)
            ]   

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Get maximum values
            min_values = df.min()

            # View result
            print(min_values)

        This will output the maximum value of all non-null values for each column:

        ```shell
            ┌───────┬─────┬────────┬────────┐
            │ Name  │ Age │ Gender │ Tenure │
            ├───────┼─────┼────────┼────────┤
            │ David │  30 │ Male   │   3.80 │
            └───────┴─────┴────────┴────────┘
            [1 rows x 4 columns]
        ```

        Note:
            - See :meth:`SQLDataModel.count_unique()` for column-wise count of unique, null and total values for each column.
            - See :meth:`SQLDataModel.min()` for returning the minimum values in each column.

        Changelog:
            - Version 0.3.1 (2024-04-01):
                - New method.
        """
        fetch_stmt = " ".join(("select",",".join([f"""max("{col}") as "{col}" """ for col in self.headers]),f'from "{self.sql_model}"'))
        return self.execute_fetch(fetch_stmt)     

    def merge(self, merge_with:SQLDataModel, how:Literal["left","right","inner","full outer","cross"]="left", left_on:str=None, right_on:str=None, include_join_column:bool=False) -> SQLDataModel:
        """
        Merges two ``SQLDataModel`` instances based on specified columns and merge type, ``how``, returning the result as a new instance. 
        If the join column shares the same name in both models, ``left_on`` and ``right_on`` column arguments are not required and will be inferred. Otherwise, explicit arguments for both are required.

        Parameters:
            ``merge_with`` (SQLDataModel): The SQLDataModel to merge with the current model.
            ``how`` (Literal["left", "right", "inner", "full outer", "cross"]): The type of merge to perform.
            ``left_on`` (str): The column name from the current model to use as the left join key.
            ``right_on`` (str): The column name from the ``merge_with`` model to use as the right join key.
            ``include_join_column`` (bool): If the shared column being used as the join key should be included from both tables. Default is False.
        
        Raises:
            ``TypeError``: If ``merge_with`` is not of type ``SQLDataModel``.
            ``SQLProgrammingError``: If sqlite3 version < 3.39.0 and join type is one of 'right' or 'full outer' which were unsupported.
            ``DimensionError``: If no shared column exists, and explicit ``left_on`` and ``right_on`` arguments are not provided.
            ``ValueError``: If the specified ``left_on`` or ``right_on`` column is not found in the respective models.

        Returns:
            ``SQLDataModel``: A new SQLDataModel containing the product of the merged result.

        Example::

            import sqldatamodel as sdm

            # Left table data with ID column
            left_headers = ["Name", "Age", "ID"]
            left_data = [        
                ["Bob", 35, 1],
                ["Alice", 30, 5],
                ["David", 40, None],
                ["Charlie", 25, 2]
            ]
            # Right table data with shared ID column
            right_headers = ["ID", "Country"]
            right_data = [
                [1, "USA"],
                [2, "Germany"],
                [3, "France"],
                [4, "Latvia"]
            ]

            # Create the left and right tables
            df_left = sdm.SQLDataModel(left_data, left_headers)
            df_right = sdm.SQLDataModel(right_data, right_headers)

        Here are the left and right tables we will be joining:
            
        ```shell
            Left Table:                     Right Table:
            ┌─────────┬──────┬──────┐       ┌──────┬─────────┐
            │ Name    │  Age │   ID │       │   ID │ Country │
            ├─────────┼──────┼──────┤       ├──────┼─────────┤
            │ Bob     │   35 │    1 │       │    1 │ USA     │
            │ Alice   │   30 │    5 │       │    2 │ Germany │
            │ David   │   40 │      │       │    3 │ France  │
            │ Charlie │   25 │    2 │       │    4 │ Latvia  │
            └─────────┴──────┴──────┘       └──────┴─────────┘
            [4 rows x 3 columns]            [4 rows x 2 columns]
        ```

        Left Join
        ---------

        ```python
            # Create a model by performing a left join with the tables
            df_joined = df_left.merge(df_right, how="left")

            # View result
            print(df_joined)
        ```

        This will output:

        ```shell
            Left Join:
            ┌─────────┬──────┬──────┬─────────┐
            │ Name    │  Age │   ID │ Country │
            ├─────────┼──────┼──────┼─────────┤
            │ Bob     │   35 │    1 │ USA     │
            │ Alice   │   30 │    5 │         │
            │ David   │   40 │      │         │
            │ Charlie │   25 │    2 │ Germany │
            └─────────┴──────┴──────┴─────────┘
            [4 rows x 4 columns]    
        ```

        Right Join
        ----------

        ```python
            # Create a model by performing a right join with the tables
            df_joined = df_left.merge(df_right, how="right")

            # View result
            print(df_joined)
        ```

        This will output:

        ```shell
            Right Join:
            ┌─────────┬──────┬──────┬─────────┐
            │ Name    │  Age │   ID │ Country │
            ├─────────┼──────┼──────┼─────────┤
            │ Bob     │   35 │    1 │ USA     │
            │ Charlie │   25 │    2 │ Germany │
            │         │      │      │ France  │
            │         │      │      │ Latvia  │
            └─────────┴──────┴──────┴─────────┘
            [4 rows x 4 columns] 
        ```  

        Inner Join
        ----------  

        ```python
            # Create a model by performing an inner join with the tables
            df_joined = df_left.merge(df_right, how="inner")

            # View result
            print(df_joined)
        ```

        This will output:

        ```shell
            Inner Join:
            ┌─────────┬──────┬──────┬─────────┐
            │ Name    │  Age │   ID │ Country │
            ├─────────┼──────┼──────┼─────────┤
            │ Bob     │   35 │    1 │ USA     │
            │ Charlie │   25 │    2 │ Germany │
            └─────────┴──────┴──────┴─────────┘
            [2 rows x 4 columns]
        ``` 

        Full Outer Join
        ---------------  

        ```python
            # Create a model by performing a full outer join with the tables
            df_joined = df_left.merge(df_right, how="full outer")

            # View result
            print(df_joined)
        ```

        This will output:

        ```shell
            Full Outer Join:
            ┌─────────┬──────┬──────┬─────────┐
            │ Name    │  Age │   ID │ Country │
            ├─────────┼──────┼──────┼─────────┤
            │ Bob     │   35 │    1 │ USA     │
            │ Alice   │   30 │    5 │         │
            │ David   │   40 │      │         │
            │ Charlie │   25 │    2 │ Germany │
            │         │      │      │ France  │
            │         │      │      │ Latvia  │
            └─────────┴──────┴──────┴─────────┘
            [6 rows x 4 columns]
        ```  

        Cross Join
        ----------  

        ```python
            # Create a model by performing a cross join with the tables
            df_joined = df_left.merge(df_right, how="cross")

            # View result
            print(df_joined)
        ```

        This will output:

        ```shell
            Cross Join:
            ┌─────────┬──────┬──────┬─────────┐
            │ Name    │  Age │   ID │ Country │
            ├─────────┼──────┼──────┼─────────┤
            │ Bob     │   35 │    1 │ USA     │
            │ Bob     │   35 │    1 │ Germany │
            │ Bob     │   35 │    1 │ France  │
            │ Bob     │   35 │    1 │ Latvia  │
            │ Alice   │   30 │    5 │ USA     │
            │ Alice   │   30 │    5 │ Germany │
            │ Alice   │   30 │    5 │ France  │
            │ Alice   │   30 │    5 │ Latvia  │
            │ David   │   40 │      │ USA     │
            │ David   │   40 │      │ Germany │
            │ David   │   40 │      │ France  │
            │ David   │   40 │      │ Latvia  │
            │ Charlie │   25 │    2 │ USA     │
            │ Charlie │   25 │    2 │ Germany │
            │ Charlie │   25 │    2 │ France  │
            │ Charlie │   25 │    2 │ Latvia  │
            └─────────┴──────┴──────┴─────────┘
            [16 rows x 4 columns]
        ```

        Note:
            - If ``include_join_column=False`` then only the ``left_on`` join column is included in the result, with the ``right_on`` column removed to avoid redundant shared key values.
            - If ``include_join_column=True`` then all the columns from both models are included in the result, with aliasing to avoid naming conflicts, see :meth:`utils.alias_duplicates()` for details.
            - The resulting ``SQLDataModel`` is created based on the ``sqlite3`` join definition and specified columns and merge type, for details see ``sqlite3`` documentation.
            - See :meth:`SQLDataModel.hstack()` for horizontally stacking SQLDataModel using shared row dimensions.
            - See :meth:`SQLDataModel.vstack()` for vertically stacking SQLDataModel using shared column dimensions.
        
        Changelog:
            - Version 0.10.2 (2024-06-30):
                - Changed ``merge_with`` from keyword argument to positional argument to reflect argument is required and not optional.
            - Version 0.10.1 (2024-06-29):
                - Modified to raise ``SQLProgrammingError`` if available sqlite3 version < 3.39.0 and join type is one of 'right' or 'full outer', which was not supported by older versions.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        # Add check for prior sqlite3 versions without support for right and full outer joins
        if (sqlite3.sqlite_version_info < (3,39,0)) and how in ("right", "full outer"):
            msg = ErrorFormat(f"SQLProgrammingError: incompatible sqlite3 version, available version '{sqlite3.sqlite_version}' < 3.39.0 which does not support right or full outer joins")
            raise SQLProgrammingError(msg)
        if not isinstance(merge_with, SQLDataModel):
            msg = ErrorFormat(f"TypeError: invalid merge type '{type(merge_with).__name__}', argument `merge_with` must be another instance of type ``SQLDataModel``")
            raise TypeError(msg)
        if left_on is None and right_on is None:
            shared_column = set(self.headers) & set(merge_with.headers)
            if len(shared_column) != 1:
                msg = ErrorFormat(f"DimensionError: no shared column exists, a shared column name is required to merge without explicit `left_on` and `right_on` arguments")
                raise DimensionError(msg)
            shared_column = next(iter(shared_column))
            left_on, right_on = shared_column, shared_column
        else:
            if left_on not in self.headers:
                msg = ErrorFormat(f"ValueError: column not found '{left_on}', a valid `left_on` column is required, use `get_headers()` to view current valid arguments")
                raise ValueError(msg)
            if right_on not in merge_with.headers:
                msg = ErrorFormat(f"ValueError: column not found '{right_on}', a valid `right_on` column is required, use `get_headers()` to view current valid arguments")
                raise ValueError(msg)            
        tmp_table_name = "_merge_with"
        merge_with.to_sql(tmp_table_name, self.sql_db_conn, if_exists='replace')
        left_headers, right_headers = self.headers, merge_with.headers if include_join_column else [x for x in merge_with.headers if x != right_on]
        all_cols = [*left_headers, *right_headers]
        headers_str = ",".join([f'a."{col}" as "{alias}"' if i < self.column_count else f'b."{col}" as "{alias}"' for i, (col, alias) in enumerate(zip(all_cols,utils.alias_duplicates(all_cols)))])
        join_stmt = f"""on a."{left_on}" = b."{right_on}" """ if how != 'cross' else """"""
        fetch_stmt = " ".join(("select",headers_str,f"""from "{self.sql_model}" a {how} join "{tmp_table_name}" b {join_stmt}"""))
        return self.execute_fetch(fetch_stmt)
      
    def pivot(self, pivot_column:str, category_column:str, amount_column:str|list[str]=None, fill_value:Any=None, agg_func:Literal['sum','avg','min','max']='sum') -> SQLDataModel:
        """
        Create a pivot table using the columns specified and return the result as a new SQLDataModel.
        
        The pivot method transforms the data in the SQLDataModel into a pivot table format summarizing the values of one column ``amount_column`` 
        based on unique values from two other columns, the ``pivot_column`` and the ``category_column``
        
        Parameters:
            ``pivot_column`` (str): Column to pivot on. The unique values in this column will form the rows of the pivot table.
            ``category_column`` (str): Column to categorize the data. The unique values in this column will form the columns of the pivot table.
            ``amount_column`` (str, list, optional): Column(s) to aggregate. Accepts either a single string or a list of strings. Defaults to all numeric columns if not provided.
            ``fill_value`` (Any, optional): Value to fill when there is no data for a particular category. Defaults to None.
            ``agg_func`` (Literal['sum', 'avg', 'min', 'max'], optional): Aggregate function to use. Defaults to 'sum'.

        Raises:
            ``TypeError``: If arguments for columns are of type 'str', 'int' or list, representing column name or integer index.
            ``ValueError``: If there are insufficient numeric columns for aggregation, invalid aggregate function, or insufficient distinct values in the category column.

        Returns:
            ``SQLDataModel``: A new SQLDataModel instance containing the result of the pivot operation.
        
        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Product', 'Units', 'Qtr','Sales']
            data = [
                ('Chainsaw', 75, 'Q1', 1500),
                ('Chainsaw', 80, 'Q2', 1600),
                ('Chainsaw', 78, 'Q3', 1550),
                ('Chainsaw', 79, 'Q4', 1580),
                ('Hammer', 40, 'Q1', 800),
                ('Hammer', 42, 'Q2', 850),
                ('Hammer', 41, 'Q3', 820),
                ('Hammer', 42, 'Q4', 830),
                ('Drill', 50, 'Q1', 1000),
                ('Drill', 55, 'Q2', 1100),
                ('Drill', 52, 'Q3', 1050),
                ('Drill', 54, 'Q4', 1080)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Pivot 'Product' by quarterly sales
            quarterly_sales = df.pivot('Product', 'Qtr', 'Sales')

            # View result
            print(quarterly_sales)
        
        This will output a wide pivot table summing up 'Sales' by 'Qtr' for each 'Product':

        ```text
            ┌───┬──────────┬──────┬──────┬──────┬──────┐
            │   │ Product  │   Q1 │   Q2 │   Q3 │   Q4 │
            ├───┼──────────┼──────┼──────┼──────┼──────┤
            │ 0 │ Chainsaw │ 4810 │ 7800 │ 1550 │ 1580 │
            │ 1 │ Drill    │ 1000 │ 1100 │ 1050 │ 1080 │
            │ 2 │ Hammer   │  800 │  850 │  820 │  830 │
            └───┴──────────┴──────┴──────┴──────┴──────┘
            [3 rows x 5 columns]
        ```  

        Multiple columns can be aggregated to produce a wider pivot

        ```python
            # This time pivot by 'Sales' and 'Units'
            quarterly_metrics = df.pivot('Product', 'Qtr', ['Units','Sales'])

            # View new pivot
            print(quarterly_metrics)
        ```

        When multiple aggregates are used, columns are labeled '<category_column> <amount_column>':

        ```text
            ┌───┬──────────┬──────────┬──────────┬──────────┬──────────┬──────────┬──────────┬──────────┬──────────┐
            │   │ Product  │ Q1 Units │ Q1 Sales │ Q2 Units │ Q2 Sales │ Q3 Units │ Q3 Sales │ Q4 Units │ Q4 Sales │
            ├───┼──────────┼──────────┼──────────┼──────────┼──────────┼──────────┼──────────┼──────────┼──────────┤
            │ 0 │ Chainsaw │      225 │     4810 │      210 │     7800 │       78 │     1550 │       79 │     1580 │
            │ 1 │ Drill    │       50 │     1000 │       55 │     1100 │       52 │     1050 │       54 │     1080 │
            │ 2 │ Hammer   │       40 │      800 │       42 │      850 │       41 │      820 │       42 │      830 │
            └───┴──────────┴──────────┴──────────┴──────────┴──────────┴──────────┴──────────┴──────────┴──────────┘
            [3 rows x 9 columns]
        ```
        
        Note:
            - When no ``amount_column`` is provided, all numeric column types will be used.
            - All column arguments can also be specified by their integer index as well as their value.
            - Use ``fill_value`` to provide a default value when there is no data for a particular category.
            - The aggregate function specified in ``agg_func`` must be one of 'sum', 'avg', 'min', or 'max', with sum used by default.
            - See :meth:`SQLDataModel.transpose()` for transposing rows and columns directly.
            - See :meth:`SQLDataModel.group_by()` for regular aggregation.

        Changelog:
            - Version 0.10.3 (2024-07-01):
                - New method.
        """
        if amount_column:
            amount_column = [amount_column] if isinstance(amount_column, (str,int)) else amount_column
        else: # Use all numeric columns if amount column is not provided
            amount_column = [col for col, dtype in self.dtypes.items() if dtype in ('int','float')]
        if (num_cols := len(amount_column)) < 1:
            msg = ErrorFormat(f"ValueError: insufficient numeric columns '{num_cols}', at least 1 numeric column is required for aggregating a pivot result")
            raise ValueError(msg)
        if agg_func.lower() not in ('sum','avg','min','max'):
            msg = ErrorFormat(f"ValueError: invalid aggregate function '{agg_func}', argument for `agg_func` must be one of 'sum', 'avg', 'min' or 'max'")
            raise ValueError(msg)
        pivot_column, category_column, *amount_column = self._validate_column([pivot_column,category_column,*[col for col in amount_column]], unmodified=False)
        unique_category_values = [r[0] for r in self.sql_db_conn.execute(f"""SELECT DISTINCT "{category_column}" FROM "{self.sql_model}" """).fetchall()]
        if (num_category := len(unique_category_values)) < 1:
            msg = ErrorFormat(f"ValueError: insufficient distinct values '{num_category}', category column '{category_column}' requires at least 1 unique value for aggregating a pivot result")
            raise ValueError(msg)
        if len(amount_column) == 1:
            agg_stmt = ','.join(f'''IFNULL({agg_func}(CASE WHEN "{category_column}" = '{category_val}' THEN {amount_column[0]} END),{f'{fill_value}' if fill_value else 'NULL'}) AS "{category_val}"''' for category_val in unique_category_values)
        else:
            agg_stmt = ','.join(f'''IFNULL({agg_func}(CASE WHEN "{category_column}" = '{category_val}' THEN {amt} END),{f'{fill_value}' if fill_value else 'NULL'}) AS "{category_val} {amt}"''' for category_val in unique_category_values for amt in amount_column)
        pivot_stmt = f"""SELECT "{pivot_column}", {agg_stmt} FROM "{self.sql_model}" GROUP BY "{pivot_column}" ORDER BY "{pivot_column}";"""
        return self.execute_fetch(pivot_stmt)

    def reset_index(self, start_index:int=0) -> None:
        """
        Resets the index of the ``SQLDataModel`` instance inplace to zero-based sequential autoincrement, or to specified ``start_index`` base with sequential incrementation.

        Parameters:
            ``start_index`` (int, optional): The starting index for the reset operation. Defaults to 0.

        Raises:
            ``TypeError``: If provided ``start_index`` argument is not of type ``int``
            ``ValueError``: If the specified ``start_index`` is greater than the minimum index in the current model.
            ``SQLProgrammingError``: If reset index execution results in constraint violation or programming error.

        Example::

            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age', 'service']
            data = [
                (0, 'john', 'smith', 27, 1.22),
                (1, 'sarah', 'west', 39, 0.7),
                (2, 'mike', 'harlin', 36, 3),
                (3, 'pat', 'douglas', 42, 11.5)
            ]        

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # View current state
            print(df)
        
        This will output:

        ```shell
            ┌─────┬────────┬─────────┬────────┬─────────┐
            │     │ first  │ last    │    age │ service │
            ├─────┼────────┼─────────┼────────┼─────────┤
            │ 994 │ john   │ smith   │     27 │    1.22 │
            │ 995 │ sarah  │ west    │     39 │    0.70 │
            │ 996 │ mike   │ harlin  │     36 │    3.00 │
            │ 997 │ pat    │ douglas │     42 │   11.50 │
            └─────┴────────┴─────────┴────────┴─────────┘
            [4 rows x 4 columns]        
        ```

        Now reset the index column:

        ```python            
            # Reset the index with default start value
            df.reset_index()

            # View updated model
            print(df)
        ```

        This will output:

        ```shell            
            ┌───┬────────┬─────────┬────────┬─────────┐
            │   │ first  │ last    │    age │ service │
            ├───┼────────┼─────────┼────────┼─────────┤
            │ 0 │ john   │ smith   │     27 │    1.22 │
            │ 1 │ sarah  │ west    │     39 │    0.70 │
            │ 2 │ mike   │ harlin  │     36 │    3.00 │
            │ 3 │ pat    │ douglas │     42 │   11.50 │
            └───┴────────┴─────────┴────────┴─────────┘
            [4 rows x 4 columns]        
        ```

        Reset the index to a custom value:

        ```python            
            # Reset the index with a different value
            df.reset_index(start_index = -3)

            # View updated model
            print(df)
        ```

        This will output:

        ```shell            
            ┌────┬────────┬─────────┬────────┬─────────┐
            │    │ first  │ last    │    age │ service │
            ├────┼────────┼─────────┼────────┼─────────┤
            │ -3 │ john   │ smith   │     27 │    1.22 │
            │ -2 │ sarah  │ west    │     39 │    0.70 │
            │ -1 │ mike   │ harlin  │     36 │    3.00 │
            │  0 │ pat    │ douglas │     42 │   11.50 │
            └────┴────────┴─────────┴────────┴─────────┘
            [4 rows x 4 columns]        
        ```

        Note:
            - The current index should be viewed more as a soft row number, to assign hard indicies use :meth:`SQLDataModel.freeze_index()` method.
            - Setting ``start_index`` too a very large negative or positive integer made lead to unpredictable behavior.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(start_index,int):
            msg = ErrorFormat(f"""TypeError: invalid start index type '{type(start_index).__name__}', start index must be type 'int'""")
            raise TypeError(msg)            
        if start_index > 0:
            msg = ErrorFormat(f"""ValueError: invalid start index '{start_index}', constraints require start index <= minimum index, which is '0' in the current model""")
            raise ValueError(msg)
        start_index = start_index - 1
        tmp_table_name = f"_temp_{self.sql_model}"
        created_headers = [self.sql_idx,*self.headers]
        headers_str = ",".join([f'"{col}"' for col in created_headers[1:]])
        headers_dtypes_str = ",".join([f'"{col}" {self.header_master[col][0]}' if self.header_master[col][2] else f'"{col}" {self.header_master[col][0]} PRIMARY KEY' for col in created_headers])
        reset_idx_stmt = f"""drop table if exists "{tmp_table_name}";create table "{tmp_table_name}"({headers_dtypes_str}); 
        insert into "{tmp_table_name}"("{self.sql_idx}",{headers_str})
        select row_number() over (order by "{self.sql_idx}" asc) + {start_index} as "{self.sql_idx}",{headers_str} from "{self.sql_model}" order by "{self.sql_idx}" asc;
        drop table if exists "{self.sql_model}"; alter table "{tmp_table_name}" rename to "{self.sql_model}";"""
        self.execute_transaction(reset_idx_stmt, update_row_meta=True)
        return

    def set_display_color(self, color:str|tuple=None, rand_color:bool=False) -> None:
        """
        Sets the table string representation color when ``SQLDataModel`` is displayed in the terminal, selecting a random color if ``rand_color = True``.

        Parameters:
            ``color`` (str or tuple): Color to set. Accepts hex value (e.g., ``'#A6D7E8'``) or tuple of RGB values e.g., ``(166, 215, 232)``.
                When not provided or ``color = None``, the terminal default color will be used.
            ``rand_color`` (bool, optional): Set a random color from a preselected pool of options. 
                When True ``color`` will be ignored and a random color selected instead.

        Returns:
            ``None``: The color value is set at :py:attr:`SQLDataModel.display_color` and nothing is returned.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['Name', 'Age', 'Salary'])

            # Set color using hex value
            df.set_display_color('#A6D7E8')
            
            # Set color using rgb value
            df.set_display_color((166, 215, 232))

        If you're unsure of which color to use, have one selected for you:

        ```python
            # Surprise me! Use a random color
            df.set_display_color(rand_color=True)

            # View the value set
            print(df.display_color)
        ```

        In this case we got a nice 'plum' color:

        ```text
            ANSIColor('#F6A8CC')
        ```

        To reset to the default terminal color:

        ```python
            # Set color to None
            df.set_display_color(color=None)

            # View the value set
            print(df.display_color) # None   
        ```

        This will return None, signifying the default terminal color will be used.

        Note:
            - By default, no color styling is applied and the native terminal color is used.
            - To use rgb values, ensure a single tuple is provided as an argument.
            - When ``rand_color = True`` a random color is selected from a preexisting pool, see :mod:`SQLDataModel.ANSIColor.ANSIColor.rand_color()` for more details.

        Changelog:
            - Version 0.12.0 (2024-07-06):
                - Added ``rand_color`` argument to require explicit selection for random color and return ``color = None`` to instead reset color to terminal default.
            - Version 0.10.2 (2024-06-30):
                - Modified to randomly select a color from preselected pool when ``color = None`` for demonstration purposes, see :mod:`SQLDataModel.ANSIColor` for more details.
            - Version 0.7.0 (2024-06-08):
                - Removed warning message and modified to raise exception on failure to create display color pen.
            - Version 0.1.5 (2023-11-24):
                - New method.
        """
        self.display_color = ANSIColor.rand_color() if rand_color else ANSIColor(color) if color is not None else color

    def sort(self, by:str|int|Iterable[str|int]=None, asc:bool=True) -> SQLDataModel:
        """
        Sort columns in the dataset by the specified ordering. If no value is specified, the current :py:attr:`SQLDataModel.sql_idx` column is used with the default ordering ``asc = True``.

        Parameters:
            ``by`` (str | int | Iterable[str | int], optional): The column or list of columns by which to sort the dataset. Defaults to sorting by the dataset's index.
            ``asc`` (bool, optional): If True, sort in ascending order; if False, sort in descending order. Defaults to ascending order.

        Raises:
            ``TypeError``: If value for ``by`` argument is not one of type 'str', 'int' or 'list'.
            ``ValueError``: If a specified column in ``by`` is not found in the current dataset or is an invalid column.
            ``IndexError``: If columns are indexed by integer but are outside of the current model range.

        Returns:
            ``SQLDataModel``: A new instance of SQLDataModel with columns sorted according to the specified ordering.

        Example::

            import sqldatamodel as sdm

            headers = ['first', 'last', 'age', 'service', 'hire_date']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01'),
                ('Sarah', 'West', 39, 0.7, '2023-10-01'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18')
            ]     

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Sort by last name column
            sorted_df = df.sort('last')

            # View sorted model
            print(sorted_df)
        
        This will output:

        ```shell            
            ┌───┬───────┬─────────┬──────┬─────────┬────────────┐
            │   │ first │ last    │  age │ service │ hire_date  │
            ├───┼───────┼─────────┼──────┼─────────┼────────────┤
            │ 0 │ Pat   │ Douglas │   42 │   11.50 │ 2015-11-06 │
            │ 1 │ Mike  │ Harlin  │   36 │    3.90 │ 2020-08-27 │
            │ 2 │ Kelly │ Lee     │   32 │    8.00 │ 2016-09-18 │
            │ 3 │ John  │ Smith   │   27 │    1.22 │ 2023-02-01 │
            │ 4 │ Sarah │ West    │   39 │    0.70 │ 2023-10-01 │
            └───┴───────┴─────────┴──────┴─────────┴────────────┘
            [5 rows x 5 columns]
        ```
        
        Sort by multiple columns:

        ```python            
            # Sort by multiple columns in descending order
            sorted_df = df.sort(['age','hire_date'], asc=False)

            # View sorted
            print(sorted_df)
        ```

        This will output:

        ```shell            
            ┌───┬───────┬─────────┬──────┬─────────┬────────────┐
            │   │ first │ last    │  age │ service │ hire_date  │
            ├───┼───────┼─────────┼──────┼─────────┼────────────┤
            │ 0 │ Pat   │ Douglas │   42 │   11.50 │ 2015-11-06 │
            │ 1 │ Sarah │ West    │   39 │    0.70 │ 2023-10-01 │
            │ 2 │ Mike  │ Harlin  │   36 │    3.90 │ 2020-08-27 │
            │ 3 │ Kelly │ Lee     │   32 │    8.00 │ 2016-09-18 │
            │ 4 │ John  │ Smith   │   27 │    1.22 │ 2023-02-01 │
            └───┴───────┴─────────┴──────┴─────────┴────────────┘
            [5 rows x 5 columns]
        ```

        Note:
            - Standard sorting process for ``sqlite3`` is used, whereby the ordering prefers the first column mentioned to the last.
            - Ascending and descending ordering follows this order of operations for multiple columns as well.

        Changelog:
            - Version 0.8.0 (2024-06-21):
                - Modified to allow mixed integer and value indexing for columns sort order in ``by`` argument to reflect similar flexibility for column input across package.
            - Version 0.5.1 (2024-05-10):
                - Modified to allow integer indexing for column sort order in ``by`` argument.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if by is not None:
            by = self._validate_column(by, unmodified=False) # +REVCOL
        else:
            by = [self.sql_idx]
        sort_ord = "asc" if asc else "desc"
        sort_by_str = ",".join([f'"{x}" {sort_ord}' for x in by])
        headers_str = ",".join([f'"{col}"' for col in self.headers])
        sort_stmt = " ".join(("select",headers_str,f'from "{self.sql_model}" order by {sort_by_str}'))
        return self.execute_fetch(sort_stmt)

    def strip(self, characters:str=None, str_dtype_only:bool=True, inplace:bool=False) -> SQLDataModel|None:
        """
        Removes the specified characters from the beginning and end of each value in the current ``SQLDataModel`` removing leading and trailing whitespace characters by default.

        Parameters:
            ``characters`` (str, optional): The characters to remove from both ends of the value. Default is None, removing whitespace (``' '``, ``'\\t'``, ``'\\n'``, ``'\\r'``).
            ``str_dtype_only`` (bool, optional): If True, only columns with dtype = 'str' are stripped, otherwise all columns are stripped. Default is True.
            ``inplace`` (bool, optional): If True, modifies the current SQLDataModel instance in-place. Default is False.
        
        Raises:
            ``TypeError``: If ``characters`` argument is provided and is not of type ``'str'`` representing unordered characters to remove.

        Returns:
            ``SQLDataModel``: If ``inplace=False``, returns a new SQLDataModel with the stripped values. Otherwise modifies the current instance in-place returning None.

        Example::
            
            import sqldatamodel as sdm

            # Create a single item model
            df = sdm.SQLDataModel([[' Hello, World! ']])

            # Strip whitespace and print
            print(df.strip())

        This will output the model after stripping the leading and trailing whitespace characters:

        ```shell
            ┌───┬───────────────┐
            │   │ 0             │
            ├───┼───────────────┤
            │ 0 │ Hello, World! │
            └───┴───────────────┘
            [1 rows x 1 columns]
        ```

        Non-whitespace characters can also be stripped:

        ```python
            import sqldatamodel as sdm

            headers = ['Col A', 'Col B', 'Col C']
            data = [
                ['A1', 'B1', 'C1'],
                ['A2', 'B2', 'C2'],
                ['A3', 'B3', 'C3']
            ]

            # Create the sample model
            df = sdm.SQLDataModel(data, headers)

            # Strip leading and trailing 'A' character
            df_stripped = df.strip('A')

            # View result
            print(df_stripped)
        ```

        This will output a new model where any leading and trailing 'A' characters have been removed:

        ```shell
            ┌───────┬───────┬───────┐
            │ Col A │ Col B │ Col C │
            ├───────┼───────┼───────┤
            │ 1     │ B1    │ C1    │
            │ 2     │ B2    │ C2    │
            │ 3     │ B3    │ C3    │
            └───────┴───────┴───────┘
            [3 rows x 3 columns]
        ```

        Multiple characters can be stripped, and the model modified inplace:

        ```python
            # Strip multiple characters and this time modify model inplace
            df.strip('123', inplace=True)

            # View result
            print(df)
        ```
        
        This will output the modified model after stripping leading and trailing '123' characters:

        ```shell
            ┌───────┬───────┬───────┐
            │ Col A │ Col B │ Col C │
            ├───────┼───────┼───────┤
            │ A     │ B     │ C     │
            │ A     │ B     │ C     │
            │ A     │ B     │ C     │
            └───────┴───────┴───────┘
            [3 rows x 3 columns]
        ```

        Note:
            - For string replacement instead of string removal, see :meth:`SQLDataModel.replace()`.
            - When using ``str_dtype_only = False``, numeric values may be modified due to SQLite's type affinity rules.
            - This method is equivalent to the SQLite ``trim(string, character)`` function, wrapping and passing the equivalent arguments.

        Changelog:
            - Version 0.4.3 (2024-05-07):
                - New method.
        """
        if characters is None:
            trim_arg = """"""
        else:
            if not isinstance(characters, str):
                msg = ErrorFormat(f"TypeError: invalid type '{type(characters).__name__}', argument for `characters` must be of type 'str' representing an unordered set of characters to remove")
                raise TypeError(msg)
            trim_arg = f""",'{characters}'"""
        if inplace:
            trim_cols = ",".join([f""" "{col}"=trim("{col}" {trim_arg}) """ for col in self.headers if self.dtypes[col] == 'str']) if str_dtype_only else ",".join([f""" "{col}"=trim("{col}" {trim_arg}) """ for col in self.headers])
            trim_stmt = " ".join((f"""update "{self.sql_model}" set""", trim_cols))
            self.sql_db_conn.execute(trim_stmt)
            return
        else:
            if str_dtype_only:
                trim_cols = ",".join([f""" trim("{col}" {trim_arg}) as "{col}" """ if self.dtypes[col] == 'str' else f""" "{col}" as "{col}" """ for col in self.headers])
            else:
                trim_cols = ",".join([f""" trim("{col}" {trim_arg}) as "{col}" """ for col in self.headers])
            trim_stmt = " ".join(("select", trim_cols, f'from "{self.sql_model}"'))
            return self.execute_fetch(trim_stmt)

    def tail(self, n_rows:int=5) -> SQLDataModel:
        """
        Returns the last ``n_rows`` of the current ``SQLDataModel``.

        Parameters:
            ``n_rows`` (int, optional): Number of rows to return. Defaults to 5.

        Raises:
            ``TypeError``: If ``n_rows`` argument is not of type 'int' representing the number of rows to return from the tail of the model.            
            
        Returns:
            ``SQLDataModel``: A new ``SQLDataModel`` instance containing the specified number of rows.

        Example::
        
            import sqldatamodel as sdm

            # Countries data available for sample dataset
            url = 'https://developers.google.com/public-data/docs/canonical/countries_csv'

            # Create the model
            df = sdm.from_html(url)

            # Get tail of model
            df_tail = df.tail()

            # View it
            print(df_tail)

        This will grab the bottom 5 rows by default:

        ```shell
            ┌─────┬─────────┬──────────┬───────────┬───────────────────┐
            │     │ country │ latitude │ longitude │ name              │
            ├─────┼─────────┼──────────┼───────────┼───────────────────┤
            │ 240 │ WF      │ -13.7688 │ -177.1561 │ Wallis and Futuna │
            │ 241 │ EH      │  24.2155 │  -12.8858 │ Western Sahara    │
            │ 242 │ YE      │  15.5527 │   48.5164 │ Yemen             │
            │ 243 │ ZM      │ -13.1339 │   27.8493 │ Zambia            │
            │ 244 │ ZW      │ -19.0154 │   29.1549 │ Zimbabwe          │
            └─────┴─────────┴──────────┴───────────┴───────────────────┘
            [5 rows x 4 columns]        
        ```    

        Note:
            - See related :meth:`SQLDataModel.head()` for the opposite, grabbing the top ``n_rows`` from the current model.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(n_rows, int):
            msg = ErrorFormat(f"TypeError: invalid type '{type(n_rows).__name__}', argument for `n_rows` must be of type 'int' representing the number of rows to return from the tail of the model")
            raise TypeError(msg)
        n_rows = min(self.row_count, max(1, n_rows))
        row_indicies = self.indicies[-n_rows:]
        return self.execute_fetch(self._generate_sql_stmt(rows=row_indicies))   
  
    def transpose(self, infer_types:bool=True, include_headers:bool=False) -> SQLDataModel:
        """
        Transposes the model and returns as a new ``SQLDataModel``.

        Parameters:
            ``infer_types`` (bool, optional): If types should be inferred after the transposition. Defaults to True.
            ``include_headers`` (bool, optional): If headers are included in the transposed data. Defaults to False.

        Returns:
            ``SQLDataModel``: The transposition of the model as a new SQLDataModel instance.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.SQLDataModel([('A1', 'A2'), ('B1', 'B2'), ('C1', 'C2')])

            # Transpose it
            df_transposed = df.transpose()

            # View original
            print(f"Original:\\n{df}")

            # Along with transposed
            print(f"Transposed:\\n{df_transposed}")

        This will output the result of the transposition:

        ```shell
            Original:
            ┌───┬─────┬─────┐
            │   │ 0   │ 1   │
            ├───┼─────┼─────┤
            │ 0 │ A1  │ A2  │
            │ 1 │ B1  │ B2  │
            │ 2 │ C1  │ C2  │
            └───┴─────┴─────┘
            [3 rows x 2 columns]
            
            Transposed:
            ┌───┬─────┬─────┬─────┐
            │   │ 0   │ 1   │ 2   │
            ├───┼─────┼─────┼─────┤
            │ 0 │ A1  │ B1  │ C1  │
            │ 1 │ A2  │ B2  │ C2  │
            └───┴─────┴─────┴─────┘
            [2 rows x 3 columns]            
        ```

        Note:
            - When ``infer_types=False``, the first row of the transposed result will be used to set the ``dtypes`` of the new model. This is generally a poor choice considering the nature of transposing data.
            - If ``include_headers=True``, the headers will be included as the first row in the transposed data.
            - Running this method sequentially should return the original model, ``sdm == sdm.transpose().transpose()``

        Changelog:
            - Version 0.3.5 (2024-04-08):
                - New method.
        """
        return type(self)(data=[row for row in zip(*self.data(include_headers=include_headers, index=False))], infer_types=infer_types, **self._get_display_args())

    def unique(self, ignore_index:bool=True) -> SQLDataModel:
        """
        Returns a new model using the unique values of the current model, keeping the first by order of appearance.

        Parameters:
            ``ignore_index`` (bool, optional): If True, the original index of the unique values is ignored. If False, the original index is kept. Default is True.

        Returns:
            ``SQLDataModel``: A new model consisting of the unique values contained in the original model.

        Example:

        ```python
            import sqldatamodel as sdm

            # Sample data
            data = [
                ('Bob', 'Chicago'),
                ('Bob', 'Chicago'),
                ('Bob', 'Chicago'),
                ('Alice', 'New York'),
                ('Alice', 'New York'),
                ('Charles', 'Boston')
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers=['Name', 'City'])

            # Create a new model from only unique rows
            df_unique = df.unique()

            # View it
            print(df_unique)    
        ```

        This will output the first unique rows, ignoring the original indicies:

        ```shell
            ┌───┬─────────┬──────────┐
            │   │ Name    │ City     │
            ├───┼─────────┼──────────┤
            │ 0 │ Bob     │ Chicago  │
            │ 1 │ Alice   │ New York │
            │ 2 │ Charles │ Boston   │
            └───┴─────────┴──────────┘
            [3 rows x 2 columns]
        ```

        Alternatively, the original index for each unique row can be retained

        ```python
            # Do not ignore the indicies
            df_unique_with_idx = df.unique(ignore_index=False)

            # View it
            print(df_unique_with_idx)
        ```

        This will output a similar result, but the original indicies from the rows kept is retained:

        ```shell
            ┌───┬─────────┬──────────┐
            │   │ Name    │ City     │
            ├───┼─────────┼──────────┤
            │ 0 │ Bob     │ Chicago  │
            │ 3 │ Alice   │ New York │
            │ 5 │ Charles │ Boston   │
            └───┴─────────┴──────────┘
            [3 rows x 2 columns]
        ```

        This method is particularly useful when needing to extract subsets of data, for example:

        ```python
            # Sample data
            headers = ['Name', 'Age', 'Department']
            data = [
                ('Alice', 38, 'HR'),
                ('Carol', 37, 'HR'),
                ('Billy', 23, 'Marketing'),
                ('Nate',  28, 'Sales'),
                ('Jill',  27, 'Sales'),
                ('John',  31, 'Engineering'),
                ('Kyle',  32, 'Engineering'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter rows by 'Age' and return unique 'Department' values
            under_30_depts = df[df['Age'] < 30, 'Department'].unique()

            # View it
            print(under_30_depts)
        ```

        This will output the unique 'Department' values for those rows matching the 'Age' filter:

        ```shell
            ┌───┬────────────┐
            │   │ Department │
            ├───┼────────────┤
            │ 0 │ Marketing  │
            │ 1 │ Sales      │
            └───┴────────────┘
            [2 rows x 1 columns]
        ```

        Note:
            - Null values are considered a unique value by SQLite and are treated accordingly when present.
            - See :meth:`SQLDataModel.deduplicate()` to drop duplicates or modify the model in place based on duplicate values.
            - See :meth:`SQLDataModel.count_unique()` to generate counts of unique values by column.
            - See :meth:`SQLDataModel.fillna()` to fill missing or null values in the model if needed.

        Changelog:
            - Version 1.3.0 (2025-02-09):
                - New method.
        """
        columns = ', '.join(f'"{col}"' for col in self.headers)
        fetch_stmt = f'SELECT DISTINCT {columns} FROM "{self.sql_model}"' if ignore_index else f'SELECT MIN("{self.sql_idx}") AS "{self.sql_idx}", {columns} FROM "{self.sql_model}" GROUP BY {columns} '
        return self.execute_fetch(fetch_stmt)

    def vstack(self, *other:SQLDataModel, inplace:bool=False) -> SQLDataModel:
        """
        Vertically stacks one or more ``SQLDataModel`` objects to the current model.

        Parameters:
            ``other`` (SQLDataModel or sequence of): The SQLDataModel objects to vertically stack.
            ``inplace`` (bool, optional): If True, performs the vertical stacking in-place, modifying the current model. Defaults to False, returning a new ``SQLDataModel``.

        Returns:
            ``SQLDataModel``: The vertically stacked SQLDataModel instance when inplace is False.

        Raises:
            ``ValueError``: If no additional SQLDataModels are provided for vertical stacking.
            ``TypeError``: If any argument in 'other' is not of type SQLDataModel, list, or tuple.
            ``SQLProgrammingError``: If an error occurs when updating the model values in place.

        Example::

            import sqldatamodel as sdm

            # Create models A and B
            df_a = sdm.SQLDataModel([('A', 1), ('B', 2)], headers=['A1', 'A2'])
            df_b = sdm.SQLDataModel([('C', 3), ('D', 4)], headers=['B1', 'B2'])

            # Vertically stack B onto A
            df_ab = df_a.vstack(df_b)

            # View stacked model
            print(df_ab)

        This will output the result of stacking B onto A, using the base model columns and dtypes:

        ```shell
            ┌─────┬─────┐
            │ A1  │  A2 │
            ├─────┼─────┤
            │ A   │   1 │
            │ B   │   2 │
            │ C   │   3 │
            │ D   │   4 │
            └─────┴─────┘
            [4 rows x 2 columns]
        ```

        Multiple models can be stacked simultaneously, here we vertically stack 3 models:

        ```python
            # Create a third model C
            df_c = sdm.SQLDataModel([('E', 5), ('F', 6)], headers=['C1', 'C2'])

            # Vertically stack all three models
            df_abc = df_a.vstack([df_b, df_c])

            # View stacked result
            print(df_abc)
        ```

        This will output the result of stacking C and B onto A:

        ```shell
            ┌─────┬─────┐
            │ A1  │  A2 │
            ├─────┼─────┤
            │ A   │   1 │
            │ B   │   2 │
            │ C   │   3 │
            │ D   │   4 │
            │ E   │   5 │
            │ F   │   6 │
            └─────┴─────┘
            [6 rows x 2 columns]
        ```

        Note:
            - Headers and data types are inherited from the model calling the :meth:`SQLDataModel.vstack()` method, casting stacked values corresponding to the base model types.
            - Model dimensions will be truncated or padded to coerce compatible dimensions when stacking, use :meth:`SQLDataModel.concat()` for strict concatenation instead of vstack.
            - See :meth:`SQLDataModel.insert_row()` for inserting new values or types other than ``SQLDataModel`` directly into the current model.
            - See :meth:`SQLDataModel.hstack()` for horizontal stacking.

        Changelog:
            - Version 0.3.4 (2024-04-05):
                - New method.
        """
        other = list(other[0]) if len(other) == 1 and isinstance(other[0], (list,tuple)) else list(other)
        if len(other) < 1:
            msg = ErrorFormat(f"ValueError: insufficient model count '{len(other)}', at least 1 additional 'SQLDataModel' is required to vertically stack against")
            raise ValueError(msg)
        if not all(isinstance(sdm, SQLDataModel) for sdm in other):
            msg = ErrorFormat(f"TypeError: invalid type encountered in '{[type(other[n]).__name__ for n in other]}', arguments for `other` must all be of type 'SQLDataModel' to vertically stack")
            raise TypeError(msg)
        other_data = [[cell for cell in sublist] + [None] * (self.column_count - len(sublist)) for sublist in [item for sublist in [sdm[:,:self.column_count].data(index=False, include_headers=False, strict_2d=True) for sdm in other] for item in sublist]]
        if inplace:
            sql_insert_stmt = f"""insert into "{self.sql_model}" ({','.join([f'"{col}"' for col in self.headers])}) values ({",".join([utils.sqlite_cast_type_format(dtype=self.header_master[col][1], as_binding=True) for col in self.headers])})"""
            try:
                self.sql_db_conn.executemany(sql_insert_stmt, other_data)
                self.sql_db_conn.commit()
            except sqlite3.ProgrammingError as e:
                self.sql_db_conn.rollback()
                msg = ErrorFormat(f"SQLProgrammingError: invalid update values, SQL execution failed with '{e}'")
                raise SQLProgrammingError(msg) from None            
            self._update_model_metadata(update_row_meta=True)
            return
        else:
            other_data = (*self.data(index=False, include_headers=False, strict_2d=True),*other_data)
            return type(self)(data=other_data, headers=self.headers, dtypes=self.dtypes, **self._get_display_args())

    def where(self, predicate:str) -> SQLDataModel:
        """
        Filters the rows of the current ``SQLDataModel`` object based on the specified SQL predicate and returns a
        new ``SQLDataModel`` containing only the rows that satisfy the condition. Only the predicates are needed as the statement prepends the select clause as "select [current model columns] where [`predicate`]", see below for detailed examples.

        Parameters:
            ``predicate`` (str): The SQL predicate used for filtering rows that follows the 'where' keyword in a normal SQL statement.

        Raises:
            ``TypeError``: If the provided ``predicate`` argument is not of type ``str``.
            ``SQLProgrammingError``: If the provided string is invalid or malformed SQL when executed against the model

        Returns:
            ``SQLDataModel``: A new ``SQLDataModel`` containing rows that satisfy the specified predicate.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Job']
            data = [
                ('Billy', 30, 'Barber'), 
                ('Alice', 28, 'Doctor'), 
                ('John', 25, 'Technician'), 
                ('Travis', 35, 'Musician'),
                ('William', 15, 'Student')
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter model by 'Age' > 30
            df_filtered = df.where('Age > 20')

            # View result
            print(df_filtered)

        This will output:

        ```shell
            ┌───┬────────┬──────┬────────────┐
            │   │ Name   │  Age │ Job        │
            ├───┼────────┼──────┼────────────┤
            │ 0 │ Billy  │   30 │ Barber     │
            │ 1 │ Alice  │   28 │ Doctor     │
            │ 2 │ John   │   25 │ Technician │
            │ 3 │ Travis │   35 │ Musician   │
            └───┴────────┴──────┴────────────┘
            [4 rows x 3 columns]
        ```
        
        Filter by multiple parameters:

        ```python        
            # Filter by 'Job' and 'Age'
            df_filtered = df.where("Job = 'Student' and Age < 18")

            # View result
            print(df_filtered)
        ```

        This will output:

        ```shell
            ┌───┬─────────┬──────┬─────────┐
            │   │ Name    │  Age │ Job     │
            ├───┼─────────┼──────┼─────────┤
            │ 4 │ William │   15 │ Student │
            └───┴─────────┴──────┴─────────┘
            [1 rows x 3 columns]
        ```

        Note:
            - ``predicate`` can be any valid SQL, for example ordering can be acheived without any filtering by simple using the argument ``'(1=1) order by "age" asc'``
            - If ``predicate`` is not of type ``str``, a ``TypeError`` is raised, if it is not valid SQL, ``SQLProgrammingError`` will be raised.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        if not isinstance(predicate, str):
            msg = ErrorFormat(f"TypeError: invalid predicate type '{type(predicate).__name__}' received, argument must be of type 'str'")
            raise TypeError(msg)
        cols_str = ",".join((f'"{col}" as "{col}"' for col in [self.sql_idx, *self.headers]))
        fetch_stmt = f""" select {cols_str} from "{self.sql_model}" where {predicate} """
        return self.execute_fetch(fetch_stmt)

##############################################################################################################
################################################ sql commands ################################################
##############################################################################################################

    def astype(self, dtype:Callable|Type|Literal['bool','bytes','date','datetime','float','int','None','str']) -> SQLDataModel:
        """
        Casts the model data into the specified python ``dtype``.

        Parameters:
            ``dtype`` (Callable|Type|Literal['bool', 'bytes', 'datetime', 'float', 'int', 'None', 'str']): The target python data type to cast the values to.

        Raises:
            ``ValueError``: If ``dtype`` is a string and not one of 'bool', 'bytes', 'datetime', 'float', 'int', 'None', 'str'.
            ``TypeError``: If ``dtype`` is a ``Type`` object that does not map to the current values, such as trying to convert a string column using the built-in ``float`` type.

        Returns:
            ``SQLDataModel``: The data casted as the specified type as a new ``SQLDataModel``.

        Warning:
            - Type casting will coerce any nonconforming values to the ``dtype`` being set, this means data will be lost if casting values to incompatible types.
        
        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Height', 'Hired']
            data = [
                ('John', 30, 175.3, 'True'), 
                ('Alice', 28, 162.0, 'True'), 
                ('Travis', 35, 185.8, 'False')
            ]    

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # See what we're working with
            print(df)

        This will output:

        ```shell
            ┌────────┬──────┬─────────┬───────┐
            │ Name   │  Age │  Height │ Hired │
            ├────────┼──────┼─────────┼───────┤
            │ John   │   30 │  175.30 │ True  │
            │ Alice  │   28 │  162.00 │ True  │
            │ Travis │   35 │  185.80 │ False │
            └────────┴──────┴─────────┴───────┘
            [3 rows x 4 columns]
        ```
        
        We can return the values as new types or save them to a column:

        ```python
            # Convert the string based 'Hired' column to boolean values
            df['Hired'] = df['Hired'].astype('bool')

            # Let's also create a new 'Height' column, this time as an integer
            df['Height int'] = df['Height'].astype('int')

            # See the new values and their types
            print(df)
        ```
        
        This will output:

        ```shell
            ┌────────┬──────┬─────────┬───────┬────────────┐
            │ Name   │  Age │  Height │ Hired │ Height int │
            ├────────┼──────┼─────────┼───────┼────────────┤
            │ John   │   30 │  175.30 │ 1     │        175 │
            │ Alice  │   28 │  162.00 │ 1     │        162 │
            │ Travis │   35 │  185.80 │ 0     │        185 │
            └────────┴──────┴─────────┴───────┴────────────┘
            [3 rows x 5 columns]
        ```        
        
        Types can also be passed directly to ``dtype``:

        ```python
            # Convert 'Age' directly to float using the built-in type:
            df['Age float'] = df['Age'].astype(float)

            # View updated model
            print(df)
        ```

        This will output the result of mapping the built-in ``float`` type to 'Age' as a new column:

        ```shell
            ┌────────┬─────┬─────────┬───────┬────────────┬───────────┐
            │ Name   │ Age │  Height │ Hired │ Height int │ Age float │
            ├────────┼─────┼─────────┼───────┼────────────┼───────────┤
            │ John   │  30 │  175.30 │ 1     │        175 │     30.00 │
            │ Alice  │  28 │  162.00 │ 1     │        162 │     28.00 │
            │ Travis │  35 │  185.80 │ 0     │        185 │     35.00 │
            └────────┴─────┴─────────┴───────┴────────────┴───────────┘
            [3 rows x 6 columns]
        ```
        
        Note:
            - Unless the returned values are saved as a new column, using this method does not change the underlying column's type currently assigned to it, to modify the column type use :meth:`SQLDataModel.set_column_dtypes()` instead.
            - Any ``None`` or ``null`` values encountered will not be coerced to the specified ``dtype``, see :meth:`SQLDataModel.fillna()` for handling and filling null values appropriately.
            - When passing a type directly, ``dtype=Type``, the type must be a ``Callable`` that can be mapped directly to a value like the built-in ``str``, ``int``, ``float`` and ``bool`` types.

        Changelog:
            - Version 0.7.6 (2024-06-16):
                - Modified to allow ``Callable`` or ``Type`` to be provided directly for ``dtype`` argument to map to data and return as new model for broader type conversion.
            - Version 0.2.1 (2024-03-24):
                - New method.
        """
        if dtype in ('bool','bytes','date','datetime','float','int','None','str'):
            str_col_cast = ",".join([utils.sqlite_cast_type_format(param=col, dtype=dtype, as_binding=False, as_alias=True) for col in self.headers])        
            sql_stmt = " ".join(("select",str_col_cast,f'from "{self.sql_model}"'))
            dtype_dict = {col:dtype for col in self.headers}
            return self.execute_fetch(sql_stmt, dtypes=dtype_dict)                
        elif isinstance(dtype, (Callable,Type)):
            try:
                data = [tuple([dtype(val) for val in row]) for row in self.data(strict_2d=True, include_headers=False)]
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e} encountered when trying to return as type '{dtype}'")
                raise type(e)(msg).with_traceback(e.__traceback__) from None
            return type(self)(data, headers=self.headers, infer_types=False, **self._get_display_args())
        else:
            msg = ErrorFormat(f"ValueError: invalid value '{dtype}', argument for `dtype` must be a `Callable` type or one of 'bool','bytes','date','datetime','float','int','None' or 'str'")
            raise ValueError(msg)

    def apply(self, func:Callable) -> SQLDataModel:
        """
        Applies ``func`` to the current ``SQLDataModel`` object and returns a modified ``SQLDataModel`` by passing its
        current values to the argument of ``func`` updated with the output.

        Parameters:
            ``func`` (Callable): A callable function to apply to the ``SQLDataModel``.

        Raises:
            ``TypeError``: If the provided argument for ``func`` is not a valid callable.
            ``SQLProgrammingError``: If the provided function is not valid based on the current SQL datatypes.

        Returns:
            ``SQLDataModel``: A modified ``SQLDataModel`` resulting from the application of ``func``.

        Examples:

        Applying to Single Column
        -------------------------

        ```python
            import sqldatamodel as sdm

            # Create the SQLDataModel:
            df = sdm.from_csv('employees.csv', headers=['First Name', 'Last Name', 'City', 'State'])

            # Create the function:
            def uncase_name(x):
                return x.lower()
            
            # Apply to existing column:
            df['First Name'] = df['First Name'].apply(uncase_name) # existing column will be updated with new values

            # Or create new one by passing in a new column name:
            df['New Column'] = df['First Name'].apply(uncase_name) # new column will be created with returned values
        ```

        Applying to Multiple Columns
        ----------------------------

        ```python
            import sqldatamodel as sdm
        
            # Create the function, note that ``func`` must have the same number of args as the model ``.apply()`` is called on:
            def summarize_employee(first, last, city, state)
                summary = f"{first} {last} is from {city}, {state}"
            
            # Create a new 'Employee Summary' column for the returned values:
            df['Employee Summary'] = df.apply(summarize_employee)
        ```

        Applying a Built-in Function
        ----------------------------

        ```python            
            import math
            import sqldatamodel as sdm

            # Create the SQLDataModel:
            df = sdm.from_csv('number-data.csv', headers=['Number'])

            # Apply the math.sqrt function to the original 'Number' column:
            df_sqrt = df.apply(math.sqrt)
        ```

        Applying a Lambda Function
        --------------------------

        ```python
            import sqldatamodel as sdm

            # Create the SQLDataModel:
            df = sdm.from_csv('example.csv', headers=['Column1', 'Column2'])
            
            # Create a new 'Column3' using the values returned from the lambda function:
            df['Column3'] = df.apply(lambda x, y: x + y)

            # Alternatively, an existing column can be updated in place:
            df['Column1'] = df['Column1'].apply(lambda x: x // 4)
        ```

        Note:
            - The number of ``args`` in the inspected signature of ``func`` must equal the current number of ``SQLDataModel`` columns.
            - The number of ``func`` args must match the current number of columns in the model, or an ``Exception`` will be raised.
            - Use :meth:`SQLDataModel.generate_apply_function_stub()` method to return a preconfigured template using current ``SQLDataModel`` columns and dtypes to assist.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        ### get column name from str or index ###
        if not isinstance(func, Callable):
            msg = ErrorFormat(f'TypeError: invalid argument for ``func``, expected type "Callable" but type "{type(func).__name__}" was provided, please provide a valid python "Callable"...')
            raise TypeError(msg)
        try:
            func_name = func.__name__.replace('<','').replace('>','')
            func_argcount = func.__code__.co_argcount
            self.sql_db_conn.create_function(func_name, func_argcount, func)
        except Exception as e:
            msg = ErrorFormat(f'SQLProgrammingError: unable to create function with provided callable "{func}", SQL process failed with: {e}')
            raise SQLProgrammingError(msg) from None
        input_columns = ",".join([f'"{col}"' for col in self.headers])
        derived_query = f"""select {func_name}({input_columns}) as "{func_name}" from "{self.sql_model}" """
        return self.execute_fetch(derived_query)

    def get_column_dtypes(self, columns:str|int|list=None, dtypes:Literal["python","sql"]="python") -> dict:
        """
        Get the data types of specified columns as either Python or SQL datatypes as a ``dict`` in the format of ``{'column': 'dtype'}``.

        Parameters: 
            ``columns`` (str | int | list): The column or columns for which to retrieve data types. Defaults to all columns.
            ``dtypes`` (Literal["python", "sql"]): The format in which to retrieve data types. Defaults to "python".

        Raises:
            ``TypeError``: If ``columns`` is not of type ``str``, ``int``, or ``list``.
            ``IndexError``: If ``columns`` is of type ``int`` and the index is outside the valid range.
            ``ValueError``: If a specified column in ``columns`` is not found in the current dataset. Use :meth:`SQLDataModel.get_headers()` to view valid columns.

        Returns:
            ``dict``: A dictionary mapping column names to their data types.

        Example::
        
            import sqldatamodel as sdm

            # Sample data
            headers = ['first', 'last', 'age', 'service', 'hire_date']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01'),
                ('Sarah', 'West', 39, 0.7, '2023-10-01'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18')
            ]  

            # Create the model
            df = sdm.SQLDataModel(data, headers)
            
            # Get all column python dtypes
            df_dtypes = df.get_column_dtypes()

            # View dict items
            for col, dtype in df_dtypes.items():
                print(f"{col}: {dtype}")
        
        This will output:

        ```shell            
            first: str
            last: str
            age: int
            service: float
            hire_date: date
        ```

        Get SQL data types as well:

        ```python            
            # Get specific column sql dtypes
            df_dtypes = df.get_column_dtypes(columns=['first','age','service'], dtypes="sql")

            # View dict items
            for col, dtype in df_dtypes.items():
                print(f"{col}: {dtype}")
        ```

        This will output:
    
        ```shell
            first: TEXT
            age: INTEGER
            service: REAL
        ```

        Note:
            - SQLDataModel index column is not included, only columns specified in the :py:attr:`SQLDataModel.headers` attribute are in scope.
            - Only the dtypes are returned, any primary key references are removed to ensure compatability with external calls.
            - Python datatypes are returned in lower case, while SQL dtypes are returned in upper case to reflect convention.
            - See :py:attr:`SQLDataModel.dtypes` for direct mapping from column to Python data type returned as ``{'col': 'dtype'}``.

        Changelog:
            - Version 0.8.0 (2024-06-21):
                - Modified to allow ``columns`` argument to be provided as an any valid reference including integer indexes or an iterable sequence of indexes to reflect similar flexibility surrounding column referencing across package.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        dtypes = 1 if dtypes == "python" else 0
        if columns is None:
            return {col:self.header_master[col][dtypes] for col in self.headers}
        columns = self._validate_column(columns, unmodified=False) # +VALCOL
        return {col:self.header_master[col][dtypes] for col in columns}

    def set_column_dtypes(self, column:str|int|dict, dtype:Literal['bool','bytes','date','datetime','float','int','None','str']=None) -> None:
        """
        Casts the specified ``column`` into the provided python ``dtype`` using the equivalent SQL data type.

        Parameters:
            ``column`` (str or int or dict): The name or index of the column to be cast, or a dictionary mapping column names to dtypes. 
                    If a dictionary, keys are column names or indices and values are the dtypes.
            ``dtype`` (Literal['bool', 'bytes', 'date', 'datetime', 'float', 'int', 'None', 'str']): The target Python data type for the specified column. 
                    Ignored if ``column`` is a dictionary.

        Raises:
            ``TypeError``: If ``column`` is not of type 'str', 'int', or 'dict', or if any dtype is invalid.
            ``IndexError``: If ``column`` is an integer and the index is outside of the current model range.
            ``ValueError``: If ``column`` is a string and the column is not found in the current model.

        Returns:
            ``None``: The model's data types are successfully casted to the new type and nothing is returned.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['idx', 'First', 'Last', 'Age']
            data = [
                (0, 'John', 'Smith', 27)
                (1, 'Sarah', 'West', 29),
                (2, 'Mike', 'Harlin', 36),
                (3, 'Pat', 'Douglas', 42),
            ]
            
            # Create the model
            df = sdm.SQLDataModel(data, headers)        

            # Original dtype for comparison
            old_dtype = df.get_column_dtypes('Age')

            # Set the data type of the 'Age' column to 'float'
            df.set_column_dtypes('Age', 'float')

            # Confirm column dtype
            new_dtype = df.get_column_dtypes('Age')

            # View result
            print(f"Age dtype: {old_dtype} -> {new_dtype}")
        
        This will output:

        ```shell
            Age dtype: int -> float
        ```

        Warning:
            - Type casting will coerce any nonconforming values to the ``dtype`` being set, this means data will be lost if casted incorrectly.

        Note:
            - Column data types are mapped to SQL types and not Python class types, see ``sqlite3`` docs for additional information.
            - See :meth:`SQLDataModel.infer_dtypes()` to automatically infer the correct column data types using random sampling.

        Changelog:
            - Version 0.7.9 (2024-06-20):
                - Modified to allow ``column`` argument to be provided as a dictionary mapping column names to dtypes to reflect current structure at :py:attr:`SQLDataModel.dtypes`.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if isinstance(column, dict):
            validated_args = {self._validate_column(k, unmodified=False)[0]:v for k,v in column.items()}
            for col_dtype in validated_args.values():
                if col_dtype not in ('bool','bytes','date','datetime','float','int','None','str'):
                    msg = ErrorFormat(f"TypeError: invalid argument '{col_dtype}', `column` dictionary values must be one of 'bool','bytes','date','datetime','float','int','None','str' use `get_column_dtypes()` to view current column datatypes")
                    raise TypeError(msg) 
            update_col_sql = """"""
            for val_column, val_dtype in validated_args.items():
                col_sql_dtype = self.static_py_to_sql_map_dict[val_dtype]
                dyn_dtype_cast = utils.sqlite_cast_type_format(param=val_column, dtype=val_dtype, as_binding=False, as_alias=False)
                update_col_sql = """;""".join((update_col_sql,f"""alter table "{self.sql_model}" add column "{val_column}_x" {col_sql_dtype}; update "{self.sql_model}" set "{val_column}_x" = {dyn_dtype_cast}; alter table "{self.sql_model}" drop column "{val_column}"; alter table "{self.sql_model}" rename column "{val_column}_x" to "{val_column}";"""))
            self.execute_transaction(update_col_sql)
        else:
            if dtype not in ('bool','bytes','date','datetime','float','int','None','str') or dtype is None:
                msg = ErrorFormat(f"TypeError: invalid argument '{dtype}', `dtype` must be one of 'bool','bytes','date','datetime','float','int','None','str' use `get_column_dtypes()` to view current dtypes")
                raise TypeError(msg)        
            column = self._validate_column(column, unmodified=False)
            update_col_sql = """"""
            for val_column in column:
                col_sql_dtype = self.static_py_to_sql_map_dict[dtype]
                dyn_dtype_cast = utils.sqlite_cast_type_format(param=val_column, dtype=dtype, as_binding=False, as_alias=False)
                update_col_sql = """;""".join((update_col_sql,f"""alter table "{self.sql_model}" add column "{val_column}_x" {col_sql_dtype}; update "{self.sql_model}" set "{val_column}_x" = {dyn_dtype_cast}; alter table "{self.sql_model}" drop column "{val_column}"; alter table "{self.sql_model}" rename column "{val_column}_x" to "{val_column}";"""))
            self.execute_transaction(update_col_sql)
        
    def get_model_name(self) -> str:
        """
        Returns the ``SQLDataModel`` table name currently being used by the model as an alias for any SQL queries executed by the user and internally.

        Returns:
            ``str``: The current ``SQLDataModel`` table name set by value of attribute :py:attr:`SQLDataModel.model_name`.
        
        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['Column1', 'Column2'])

            # Get the current name
            model_name = df.get_model_name()

            # View it
            print(f'The model is currently using the table name: {model_name}')
        
        Note:
            - Use :meth:`SQLDataModel.set_model_name()` to modify the table name used internally to represent the ``SQLDataModel`` instance.

        Changelog:
            - Version 0.1.5 (2023-11-24):
                - New method.            
        """
        return self.sql_model
    
    def set_model_name(self, new_name:str) -> None:
        """
        Sets the new ``SQLDataModel`` table name that will be used as an alias for any SQL queries executed by the user or internally.

        Parameters:
            ``new_name`` (str): The new table name for the ``SQLDataModel``.

        Raises:
            ``SQLProgrammingError``: If unable to rename the model table due to SQL execution failure.

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['Column1', 'Column2'])

            # Rename the model
            df.set_model_name('custom_table')

        Note:
            - The provided value must be a valid SQL table name.
            - This alias will be reset to the default value for any new ``SQLDataModel`` instances: ``'sdm'``.

        Changelog:
            - Version 0.1.5 (2023-11-24):
                - New method.                
        """
        full_stmt = f"""begin transaction; alter table "{self.sql_model}" rename to {new_name}; end transaction;"""
        try:
            self.sql_db_conn.executescript(full_stmt)
            self.sql_db_conn.commit()
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f"SQLProgrammingError: unable to rename model table, SQL execution failed with: '{e}'")
            raise SQLProgrammingError(msg) from None
        self.sql_model = new_name
     
    def execute_fetch(self, sql_query:str, sql_params:tuple=None, **kwargs) -> SQLDataModel:
        """
        Returns a new ``SQLDataModel`` object, including display and style properties, after executing the provided SQL query using the current ``SQLDataModel``. 
        This method is called by other methods which expect results to be returned from their execution.

        Parameters:
            ``sql_query`` (str): The SQL query to execute with the expectation of rows returned.
            ``sql_params`` (tuple, optional): The SQL parameters to provide for parameterized queries.
            ``**kwargs`` (optional): Additional keyword args to pass to ``SQLDataModel`` constructor

        Raises:
            ``SQLProgrammingError``: If the provided SQL query is invalid or malformed.
            ``ValueError``: If the provided SQL query was valid but returned 0 rows, which is insufficient to return a new model.

        Returns:
            ``SQLDataModel``: A new ``SQLDataModel`` instance containing the result of the SQL query.

        Example::
        
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('example.csv', headers=['Column1', 'Column2'])

            # Create the SQL query to execute
            query = 'SELECT * FROM sdm WHERE Column1 > 10'

            # Fetch and save the result to a new instance
            result_model = df.execute_fetch(query)

            # Create a parameterized SQL query to execute
            query = 'SELECT * FROM sdm WHERE Column1 > ? OR Column2 < ?'
            params = (10, 20)

            # Provide the SQL and the statement parameters
            result_parameterized = df.execute_fetch(query, params)

        Important:
            - The default table name is ``'sdm'``, you can use :meth:`SQLDataModel.set_model_name()` to modify the name used by ``SQLDataModel``.

        Note:
            - Use :meth:`SQLDataModel.set_model_name()` to modify the table name used by the model, default name set as ``'sdm'``.
            - Display properties such as float precision, index column or table styling are also passed to the new instance when not provided in ``kwargs``.
        
        Changelog:
            - Version 0.6.2 (2024-05-15):
                - Inclusion of :py:attr:`SQLDataModel.table_style` argument in returned ``SQLDataModel`` to inherit all display properties in result.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        try:
            res = self.sql_db_conn.execute(sql_query) if sql_params is None else self.sql_db_conn.execute(sql_query, sql_params) 
        except Exception as e:
            msg = ErrorFormat(f'SQLProgrammingError: invalid or malformed SQL, provided query failed with error "{e}"')
            raise SQLProgrammingError(msg) from None
        try:
            fetch_result = res.fetchall()
        except sqlite3.OperationalError as e:
            msg = ErrorFormat(f"SQLProgrammingError: '{e}' encountered when trying to fetch and parse SQL query results")
            raise SQLProgrammingError(msg) from None            
        fetch_headers = [x[0] for x in res.description]
        if (rows_returned := len(fetch_result)) < 1:
            msg = ErrorFormat(f"ValueError: nothing to return, provided query returned '{rows_returned}' rows which is insufficient to return or generate a new model from")
            raise ValueError(msg)
        sdm_args = self._get_display_args(include_dtypes=True)
        if kwargs:
            sdm_args.update({k:v for k,v in kwargs.items()})
        return type(self)(fetch_result, headers=fetch_headers, **sdm_args)

    def execute_statement(self, sql_stmt:str, sql_params:tuple=None, update_row_meta:bool=True) -> None:
        """
        Executes an arbitrary SQL query against the current model without the expectation of selection or returned rows.

        Parameters:
            ``sql_stmt`` (str): The SQL query to execute.
            ``sql_params`` (tuple, optional): The SQL parameters to provide for parameterized queries.
            ``update_row_meta`` (bool, optional): Whether the row count metadata should be updated after executing the statement. Default is True, using :meth:`SQLDataModel._update_model_metadata()` to ensure any schema modifications remain in sync.

        Raises:
            ``SQLProgrammingError``: If the SQL execution fails.
        
        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('data.csv')

            # Execute statement without results, modifying column in place
            df.execute_statement('UPDATE table SET column = value WHERE condition')

            # Execute a parameterized with statement by providing values
            df.execute_statement('DELETE FROM table WHERE idx = ? or name = ?', (7,'Bob'))

        Note:
            - To execute a query with the expectation of results, see :meth:`SQLDataModel.execute_fetch()` method.
            - To execute multiple queries within a single transaction, see :meth:`SQLDataModel.execute_transaction()` method.

        Changelog:
            - Version 0.8.0 (2024-06-21):
                - Added ``update_row_meta`` parameter to speed up transactions that are guaranteed to have no effect on the current model :py:attr:`SQLDataModel.indicies` metadata. 
                  A shallower and computationally cheaper check will still occur to ensure :py:attr:`SQLDataModel.header_master` remains in sync.
            - Version 0.7.4 (2024-06-13):
                - Added ``sql_params`` parameter to allow parameterized statements similar to other SQL execution methods.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        try:
            self.sql_db_conn.execute(sql_stmt, sql_params) if sql_params else self.sql_db_conn.execute(sql_stmt)
            self.sql_db_conn.commit()
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f'SQLProgrammingError: invalid or malformed SQL, unable to execute provided SQL query with error "{e}"...')
            raise SQLProgrammingError(msg) from None
        self._update_model_metadata(update_row_meta=update_row_meta)
    
    def execute_transaction(self, sql_script:str, update_row_meta:bool=True) -> None:
        """
        Executes a prepared SQL script wrapped in a transaction against the current model without the expectation of selection or returned rows.

        Parameters:
            ``sql_script`` (str): The SQL script to execute within a transaction.
            ``update_row_meta`` (bool, optional): Whether the row count metadata should be updated after executing the transaction. Default is True, using :meth:`SQLDataModel._update_model_metadata()` to ensure any schema modifications remain in sync.

        Raises:
            ``SQLProgrammingError``: If the provided ``sql_script`` cannot be executed or the SQL execution fails.

        Returns:
            ``None``

        Example::
            
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('data.csv')

            # Script to update columns with predicate
            transaction_script = '''
                UPDATE table1 SET column1 = value1 WHERE condition1;
                UPDATE table2 SET column2 = value2 WHERE condition2;
            '''

            # Execute the script
            df.execute_transaction(transaction_script)

        Note:
            - Use :meth:`SQLDataModel.execute_fetch()` method if the SQL script is expected to return a selection or result set upon execution.
            - Use :meth:`SQLDataModel.execute_statement()` method if the SQL script is not expected to return a selection, but parameter bindings and values are needed.
            - Many other methods heavily rely on the :meth:`SQLDataModel.execute_transaction` method, therefore modifying it may adversly affect many other methods.
        
        Changelog:
            - Version 0.8.0 (2024-06-21):
                - Added ``update_row_meta`` parameter to speed up transactions that are guaranteed to have no effect on the current model :py:attr:`SQLDataModel.indicies` metadata. 
                  A shallower and computationally cheaper check will still occur to ensure :py:attr:`SQLDataModel.header_master` remains in sync.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        full_stmt = f"""begin transaction; {sql_script}; end transaction;"""
        try:
            self.sql_db_conn.executescript(full_stmt)
            self.sql_db_conn.commit()
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f'SQLProgrammingError: unable to execute provided transaction, SQL execution failed with: "{e}"')
            raise SQLProgrammingError(msg) from None
        self._update_model_metadata(update_row_meta=update_row_meta)        

    def freeze_index(self, column_name:str=None) -> None:
        """
        Freeze the current index as a new column, expanding it into the current model. The new column is unaffected by any future changes to the primary index column.

        Parameters:
            ``column_name`` (str, optional): The name for the new frozen index column. If not provided, a default name 'frzn_id' will be used.

        Raises:
            ``TypeError``: If the provided ``column_name`` is not of type 'str'.

        Returns:
            ``None``

        Example::
        
            import sqldatamodel as sdm

            headers = ['first', 'last', 'age', 'service', 'hire_date']
            data = [
                ('John', 'Smith', 27, 1.22, '2023-02-01'),
                ('Sarah', 'West', 39, 0.7, '2023-10-01'),
                ('Mike', 'Harlin', 36, 3.9, '2020-08-27'),
                ('Pat', 'Douglas', 42, 11.5, '2015-11-06'),
                ('Kelly', 'Lee', 32, 8.0, '2016-09-18')
            ]   
                
            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Freeze index as new column 'id'
            df.freeze_index("id")    

            # View model
            print(df)
        
        This will output:
        
        ```shell            
            ┌───┬───────┬─────────┬──────┬─────────┬────────────┬──────┐
            │   │ first │ last    │  age │ service │ hire_date  │   id │
            ├───┼───────┼─────────┼──────┼─────────┼────────────┼──────┤
            │ 0 │ John  │ Smith   │   27 │    1.22 │ 2023-02-01 │    0 │
            │ 1 │ Sarah │ West    │   39 │    0.70 │ 2023-10-01 │    1 │
            │ 2 │ Mike  │ Harlin  │   36 │    3.90 │ 2020-08-27 │    2 │
            │ 3 │ Pat   │ Douglas │   42 │   11.50 │ 2015-11-06 │    3 │
            │ 4 │ Kelly │ Lee     │   32 │    8.00 │ 2016-09-18 │    4 │
            └───┴───────┴─────────┴──────┴─────────┴────────────┴──────┘
            [5 rows x 6 columns]
        ```

        Note:
            - Freezing the index will assign the current :py:attr:`SQLDataModel.sql_idx` for each row as a new column, leaving the current index in place.
            - To modify the actual :py:attr:`SQLDataModel.sql_idx` value, use the :meth:`SQLDataModel.reset_index()` method instead.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        if not isinstance(column_name, str):
            msg = ErrorFormat(f"TypeError: invalid type '{type(column_name).__name__}', argument for ``column_name`` must be of type 'str'")
            raise TypeError(msg)
        if column_name is None:
            column_name = "frzn_id"
        column_name = tuple(utils.alias_duplicates([self.sql_idx,*self.headers,column_name]))[-1]
        sql_script = f"""alter table "{self.sql_model}" add column "{column_name}" integer;update "{self.sql_model}" set "{column_name}" = "{self.sql_idx}";"""
        self.execute_transaction(sql_script)

    def add_column_with_values(self, column_name:str, value=None) -> None:
        """
        Adds a new column with the specified ``column_name`` to the ``SQLDataModel``. The new column is populated with the values provided in the ``value`` argument. If ``value`` is not provided (default), the new column is populated with NULL values.

        Parameters:
            ``column_name`` (str): The name of the new column to be added.
            ``value``: The value to populate the new column. If None (default), the column is populated with NULL values. If a valid column name is provided, the values of that column will be used to fill the new column.

        Raises:
            ``DimensionError``: If the length of the provided values does not match the number of rows in the model.
            ``TypeError``: If the data type of the provided values is not supported or translatable to an SQL data type.

        Example::

            import sqldatamodel as sdm

            # Create model from data
            df = sdm.from_csv('data.csv')

            # Add new column with default value 42
            df.add_column_with_values('new_column', value=42)

            # Add new column by copying values from an existing column
            df.add_column_with_values('new_column', value='existing_column')
        
        Note:
            - Many other methods, including :meth:`SQLDataModel.__setitem__` rely on this method, therefore modifying it may cause unpredictable behavior.
            - Determination for when to copy existing versus when to assign string is value is done by :meth:`SQLDataModel.__eq__` against both values

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        create_col_stmt = f"""alter table "{self.sql_model}" add column "{column_name}" """
        if (value is not None) and (value in self.headers):
            dyn_dtype_default_value = f"""{self.header_master[value][0]}"""
            dyn_copy_existing = f"""update "{self.sql_model}" set "{column_name}" = "{value}";"""
            sql_script = f"""{create_col_stmt} {dyn_dtype_default_value};{dyn_copy_existing};"""
            self.execute_transaction(sql_script)
            return
        if value is None:
            sql_script = create_col_stmt
            self.execute_transaction(sql_script)
            return
        if isinstance(value, (str,int,float,bool)):
            value = f"'{value}'" if isinstance(value,str) else value
            dyn_dtype_default_value = f"""{self.static_py_to_sql_map_dict[type(value).__name__]} not null default {value}""" if value is not None else "TEXT"
            sql_script = f"""{create_col_stmt} {dyn_dtype_default_value};"""
            self.execute_transaction(sql_script)
            return
        if isinstance(value, datetime.date):
            dyn_dtype_default_value = f"""{self.static_py_to_sql_map_dict[type(value).__name__]}"""
            create_col_stmt = f"""{create_col_stmt} {dyn_dtype_default_value};"""
            update_val_stmt = f"""update "{self.sql_model}" set "{column_name}" = {f"datetime('{value}')" if isinstance(value, datetime.datetime) else f"date('{value}')"};"""
            sql_script = f"""{create_col_stmt} {update_val_stmt}"""
            self.execute_transaction(sql_script)
            return 
        if isinstance(value, bytes):
            dyn_dtype_default_value = f"""{self.static_py_to_sql_map_dict[type(value).__name__]} not null default X'{value.hex()}'""" if value is not None else "TEXT"
            sql_script = f"""{create_col_stmt} {dyn_dtype_default_value};"""
            self.execute_transaction(sql_script)
            return                 
        if isinstance(value, (list,tuple)):
            if (len_values := len(value)) != self.row_count:
                msg = ErrorFormat(f"DimensionError: invalid dimensions '{len_values} != {self.row_count}', provided values have length '{len_values}' while current row count is '{self.row_count}'")
                raise DimensionError(msg)
            try:
                seq_dtype = self.static_py_to_sql_map_dict[type(value[0]).__name__]
            except:
                msg = ErrorFormat(f'TypeError: invalid datatype "{type(value[0]).__name__}", please provide a valid and SQL translatable datatype...')
                raise TypeError(msg) from None
            sql_script = f"""{create_col_stmt} {seq_dtype};"""
            all_model_idxs = tuple(range(self.row_count))
            for i,val in enumerate(value):
                sql_script += f"""update {self.sql_model} set "{column_name}" = "{val}" where {self.sql_idx} = {all_model_idxs[i]};"""
            self.execute_transaction(sql_script)
            return

    def apply_function_to_column(self, func:Callable, column:str|int) -> None:
        """
        Applies the specified callable function (`func`) to the provided ``SQLDataModel`` column. The function's output is used to update the values in the column. For broader uses or more input flexibility, see related method `apply()`.

        Parameters:
            ``func`` (Callable): The callable function to apply to the column.
            ``column`` (str | int): The name or index of the column to which the function will be applied.

        Raises:
            ``TypeError``: If the provided column argument is not a valid type (str or int).
            ``IndexError``: If the provided column index is outside the valid range of column indices.
            ``ValueError``: If the provided column name is not valid for the current model.
            ``SQLProgrammingError``: If the provided function return types or arg count is invalid or incompatible to SQL types.
        
        Returns:
            ``None``

        Example::
        
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('data.csv')

            # Apply upper() method using lambda function to column ``name``
            df.apply_function_to_column(lambda x: x.upper(), column='name')

            # Apply addition through lambda function to column at index 1
            df.apply_function_to_column(lambda x, y: x + y, column=1)

        Note:
            - This method is a simplified version of the :meth:`SQLDataModel.apply()` method, which can be used for arbitrary function params and inputs.
            - If providing a function name, ensure it can be used a valid ``sqlite3`` identifier for the instance's connection otherwise ``SQLProgrammingError`` will be raised.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        ### get column name from str or index ###
        if (not isinstance(column, int)) and (not isinstance(column, str)):
            msg = ErrorFormat(f"TypeError: invalid column argument, '{type(column).__name__}' is not a valid target, provide column index or column name as a string")
            raise TypeError(msg)
        if isinstance(column, int):
            try:
                column = self.headers[column]
            except IndexError as e:
                msg = ErrorFormat(f"IndexError: invalid column index provided, '{column}' is not a valid column index, use `.column_count` property to get valid range")
                raise IndexError(msg) from None
        if isinstance(column, str):
            if column not in self.headers:
                msg = ErrorFormat(f"ValueError: column not found '{column}', use `.get_headers()` to view current valid model headers")
                raise ValueError(msg)
            else:
                column = column
        target_column = column
        try:
            func_name = func.__name__
            func_argcount = func.__code__.co_argcount
            self.sql_db_conn.create_function(func_name, func_argcount, func)
        except Exception as e:
            msg = ErrorFormat(f"SQLProgrammingError: unable to create function with provided callable '{func}', SQL process failed with: '{e}'")
            raise SQLProgrammingError(msg) from None
        if func_argcount == 1:
            input_columns = target_column
        elif func_argcount == self.column_count:
            input_columns = ",".join([f'"{col}"' for col in self.headers])
        else:
            msg = ErrorFormat(f"ValueError: invalid function arg count: '{func_argcount}', input args to '{func_name}' must be 1 or '{self.column_count}' based on the current models structure, e.g.,\n{self.generate_apply_function_stub()}")
            raise ValueError(msg)
        sql_apply_update_stmt = f"""update {self.sql_model} set {target_column} = {func_name}({input_columns})"""
        full_stmt = f"""begin transaction; {sql_apply_update_stmt}; end transaction;"""
        try:
            self.sql_db_conn.executescript(full_stmt)
            self.sql_db_conn.commit()
        except Exception as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f"SQLProgrammingError: unable to apply function, SQL execution failed with: '{e}'")
            raise SQLProgrammingError(msg) from None
        self._update_model_metadata()

    def generate_apply_function_stub(self) -> str:
        """
        Generates a function template using the current ``SQLDataModel`` to format function arguments for the :meth:`SQLDataModel.apply_function_to_column()` method.

        Returns:
            ``str``: A string representing the function template.

        Example::
        
            import sqldatamodel as sdm

            # Create the model
            df = sdm.from_csv('data.csv')

            # Create the stub
            stub = df.generate_apply_function_stub()
            
            # View it
            print(stub)

        This will output:

        ```text
            def func(user_name:str, user_age:int, user_salaray:float):
                # apply logic and return value
                return
        ```

        Containing all the required inputs and column names needed to generate a compatible function to apply to the model and can be copy pasted into existing code.

        Note:
            - This method is to meant as a general informative tool or for debugging assistance if needed
            - See :meth:`SQLDataModel.apply()` method for usage and implementation of functions in SQLDataModel using ``sqlite3`` 

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        func_signature = ", ".join([f"""{k.replace(" ","_")}:{v[1]}""" for k,v in self.header_master.items() if k != self.sql_idx])
        return f"""def func({func_signature}):\n    # apply logic and return value\n    return"""
    
    def update_index_at(self, row_index:int, column_index:int|str, value:Any=None) -> None:
        """
        Updates a specific cell in the ``SQLDataModel`` at the given row and column indices with the provided value.

        Parameters:
            ``row_index`` (int): The index of the row to be updated.
            ``column_index`` (int or str): The index or name of the column to be updated.
            ``value`` (Any, optional): The new value to be assigned to the specified cell.

        Raises:
            ``TypeError``: If ``row_index`` is not of type 'int' or if ``column_index`` is not of type 'int' or 'str'.
            ``IndexError``: If row or column provided as an 'int' but is outside of the current model row or column range.
            ``ValueError``: If column provided as a 'str' but is not found in the current model headers.
            ``SQLProgrammingError``: If there is an issue with the SQL execution during the update.
        
        Returns:
            ``None``

        Example::

            import sqldatamodel as sdm

            # Create an initial 3x3 model filled with dashes
            df = sdm.from_shape((3,3), fill='---', headers=['A', 'B', 'C'])

            # Update cell based on integer indicies
            df.update_index_at(0, 0, 'Top Left')
            df.update_index_at(0, 2, 'Top Right')

            # Update cell based on row index and column name
            df.update_index_at(2, 'A', 'Bottom Left')
            df.update_index_at(2, 'C', 'Bottom Right')

            # Update based on negative row and column indexing
            df.update_index_at(-2, -2, 'Center')

            # View result
            print(df)

        This will output cumulative result of our updates:

        ```text
            ┌───┬─────────────┬────────┬──────────────┐
            │   │ A           │ B      │ C            │
            ├───┼─────────────┼────────┼──────────────┤
            │ 0 │ Top Left    │ ---    │ Top Right    │
            │ 1 │ ---         │ Center │ ---          │
            │ 2 │ Bottom Left │ ---    │ Bottom Right │
            └───┴─────────────┴────────┴──────────────┘
            [3 rows x 3 columns]
        ```

        Important:
            Indexing is done using zero-based integers and not done by index value. Most of the time this distinction is irrelevant as the row index at position '0' will have an index value of '0', however this can change after transformation operations like filter or sort. To reset and realign the index value use :meth:`SQLDataModel.reset_index()` or use :py:attr:`SQLDataModel.indicies` to view the current row indicies.
        
        Note:
            - This method only updates individual cells in the current model based on integer indexing for both rows and columns using their (row, column) position.
            - To broadcast updates across row and column dimensions use the syntax of ``sdm[row, column] = value`` or see :meth:`SQLDataModel.__setitem__()` for more details.

        Changelog:
            - Version 0.8.0 (2024-06-21):
                - Modified to allow ``row_index`` and ``column_index`` arguments the same input type flexibility found across package, allowing both to be referenced directly or by their integer index.
            - Version 0.5.2 (2024-05-13):
                - Modified ``row_index`` parameter to use :py:attr:`SQLDataModel.indicies` to index into rows in lieu of row index value equality.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        row_index = self._validate_row(row_index, unmodified=False)[0] # +VALROW
        column_index = self._validate_column(column_index, unmodified=False)[0] # +VALCOL
        if not isinstance(value,tuple):
            value = (value,)
        if isinstance(value[0], datetime.datetime):
            val_binding = "datetime(?)"
        elif isinstance(value[0], datetime.date):
            val_binding = "date(?)"
        else:
            val_binding = "?"
        update_stmt = f"""update "{self.sql_model}" set "{column_index}" = {val_binding} where "{self.sql_idx}" = {row_index}"""
        self.execute_statement(update_stmt, sql_params=value, update_row_meta=False)

    def _calculate_col_widths(self, index:bool=None, min_column_width:int=None, max_column_width:int=None, float_precision:int=None, split_row:int=None, index_rep:str=None) -> dict[str, int]:
        """
        Calculate the maximum column widths for each header column based on the provided conditions to assist with representation methods.
        
        Parameters:
            ``index`` (bool, optional): Indicates whether to include the index column in the calculations. Default is None, using :py:attr:`SQLDataModel.display_index` value.
            ``min_column_width`` (int, optional): Minimum width for columns. Default is None, using :py:attr:`SQLDataModel.minimum_column_width` value.
            ``max_column_width`` (int, optional): Maximum width for columns. Default is None, using :py:attr:`SQLDataModel.maximum_column_width` value.
            ``float_precision`` (int, optional): Precision for displaying float values. Default is None, using :py:attr:`SQLDataModel.display_float_precision` value.
            ``split_row`` (int, optional): Row index to determine vertical truncation. If None, no vertical truncation is applied.
            ``index_rep`` (str, optional): String representation for the index. If None, uses a single whitespace character ``' '`` to represent the index column.

        Returns:
            ``dict[str, int]``: Dictionary mapping each header column to its maximum calculated width as ``{'column': width}``.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['User', 'Key', 'Value', 'Active']
            data = [
                ('Allison', 130, 237.03, True),
                ('Bobby', -400, 723.41, False),
                ('Connor', 698, 154.70, False),
                ('Dimitry', 287, 409.14, True)
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Calculate the max column widths
            col_widths = df._calculate_col_widths()

            # View result
            print(col_widths)

        This will output the maximum widths for each column calculated by the column name and the corresponding values:

        ```text
            {'idx': 1, 'User': 7, 'Key': 4, 'Value': 7, 'Active': 6}        
        ```

        Note:
            - When ``index_rep`` is provided, the length of the index representation will be used when calculating the maximum column width.
            - When ``split_row`` is provided, width calculation checks are restricted to only the top and bottom N number of rows specified.
            - Used by :meth:`SQLDataModel.to_string()` to determine appropriate column representation widths.

        Changelog:
            - Version 0.11.0 (2024-07-05):
                - New method.
        """
        display_index = self.display_index if index is None else index
        min_column_width = self.min_column_width  if min_column_width is None else min_column_width
        max_column_width = self.max_column_width if max_column_width is None else max_column_width
        max_column_width = max_column_width if max_column_width >= 2 else 2 # minimum required width
        display_float_precision = self.display_float_precision if float_precision is None else float_precision
        display_headers = [self.sql_idx,*self.headers] if display_index else self.headers
        if split_row:
            check_width_top, check_width_bottom = self.indicies[split_row], self.indicies[-split_row]
            check_width_scope = f'where ("{self.sql_idx}" < {check_width_top} or "{self.sql_idx}" >= {check_width_bottom})'
        else:
            check_width_scope = ''   
        header_py_dtype_dict = self.dtypes
        header_printf_modifiers_dict = {col:(f"'% .{display_float_precision}f'" if dtype == 'float' else "'%!s'" if dtype != 'bytes' else "'b''%!s'''") for col,dtype in header_py_dtype_dict.items()}
        # Required to escape newlines when calculating column widths
        headers_sub_select = " ".join(("select",f"""max(length("{self.sql_idx}")) as "{self.sql_idx}",""" if display_index else "",",".join([f"""max(max(length(printf({header_printf_modifiers_dict[col]},"{col}"))),length('{col}')) as "{col}" """ if header_py_dtype_dict[col] != 'str' else ("".join((f"""max(max(length(printf({header_printf_modifiers_dict[col]},REPLACE("{col}", """, """'\n','\\n' )""", f"""))),length('{col}')) as "{col}" """ ))) for col in display_headers if col != self.sql_idx]),f'from "{self.sql_model}" {check_width_scope} order by "{self.sql_idx}" asc'))
        headers_parse_lengths_select = " ".join(("select",",".join([f"""min(max(ifnull("{col}",length('{col}')),{min_column_width}),{max_column_width})""" if col != self.sql_idx else f"""max(ifnull("{col}",1),{0 if not index_rep else len(index_rep)})""" for col in display_headers]),"from"))
        headers_full_select = f"""{headers_parse_lengths_select}({headers_sub_select})"""
        length_meta = self.sql_db_conn.execute(headers_full_select).fetchone()
        header_length_dict = dict(zip(display_headers,length_meta))
        return header_length_dict

    def _update_model_metadata(self, update_row_meta:bool=False) -> None:
        """
        Generates and updates metadata information about the columns and optionally the rows in the ``SQLDataModel`` instance based on the current model. 

        Attributes updated:
            - :py:attr:`SQLDataModel.header_master`: Master dictionary of column metadata.
            - :py:attr:`SQLDataModel.headers`: List of current model headers, order retained.
            - :py:attr:`SQLDataModel.column_count`: Number of columns in current model.
            - :py:attr:`SQLDataModel.shape`: The current ``(rows, cols)`` dimensions of the model.
            - :py:attr:`SQLDataModel.dtype`: The current ``{'col': 'dtype'}`` mapping of the model.

              - :py:attr:`SQLDataModel.indicies`: Optionally updated, represents current valid row indicies.
              - :py:attr:`SQLDataModel.row_count`: Optionally updated, represents current row count.

        Parameters:
            ``update_row_meta`` (bool, optional): If True, updates row metadata information; otherwise, retrieves column metadata only (default).
            
        Returns:
            ``None``

        Example::
        
            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age', 'service_time']
            data = [
                (0, 'john', 'smith', 27, 1.22),
                (1, 'sarah', 'west', 39, 0.7),
                (2, 'mike', 'harlin', 36, 3),
                (3, 'pat', 'douglas', 42, 11.5)
            ]

            # Create the model with sample data
            df = sdm.SQLDataModel(data, headers)

            # View header master
            print(df.header_master)

        This will output:

        ```shell
            {'first': ('TEXT', 'str', True, '<'),
             'last': ('TEXT', 'str', True, '<'),
             'age': ('INTEGER', 'int', True, '>'),
             'service_time': ('REAL', 'float', True, '>'),
             'idx': ('INTEGER', 'int', False, '>')}
        ```

        Example Attributes Modified:

        ```python
            import sqldatamodel as sdm

            headers = ['idx', 'first', 'last', 'age', 'service_time']
            data = [
                (0, 'john', 'smith', 27, 1.22),
                (1, 'sarah', 'west', 0.7),
                (2, 'mike', 'harlin', 3),
                (3, 'pat', 'douglas', 11.5)
            ]

            # Create the model with sample data
            df = sdm.SQLDataModel(data, headers)

            # Get current column count
            num_cols_before = df.column_count

            # Add new column
            df['new_column'] = 'empty'

            # Method is called behind the scenes
            df._update_model_metadata()

            # Get new column count
            num_cols_after = df.column_count

            # View difference
            print(f"cols before: {num_cols_before}, cols after: {num_cols_after}")
        ```   

        Note:
            - This method is called after operations that may modify the current model's structure and require synchronization.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """        
        fetch_metadata = f"""select "name" as "_ordered_name","type" as "_ordered_type","pk" as "_is_regular_column",case when ("type"='INTEGER' or "type"='REAL') then '>' else '<' end as "_def_alignment" from pragma_table_info('{self.sql_model}') order by {",".join([f'''"_ordered_name"='{col}' desc''' for col in self.headers])}"""
        metadata = self.sql_db_conn.execute(fetch_metadata).fetchall()
        header_master = {m[0]:(m[1], self.static_sql_to_py_map_dict[m[1]],True if m[2] == 0 else False,m[3]) for m in metadata}
        self.headers = list(dict.fromkeys([k for k,v in header_master.items() if v[2]]))
        self.column_count = len(self.headers)
        # format: 'column_name': ('sql_dtype', 'py_dtype', is_regular_column, 'default_alignment')
        self.header_master = header_master 
        """``dict[str, tuple]``: Maps the current model's column metadata in the format of ``'column_name': ('sql_dtype', 'py_dtype', is_regular_column, 'default_alignment')``, updated by :meth:`SQLDataModel._update_model_metadata`."""
        self.dtypes = {k:v[1] for k,v in self.header_master.items() if v[2]}
        self.shape = (self.shape[0], self.column_count)
        if update_row_meta:
            self._update_indicies()

    def _generate_sql_stmt(self, columns:list[str]=None, rows:int|slice|tuple|str=None, index:bool=True, na_rep:str=None) -> str:
        """
        Generate an SQL statement for fetching specific columns and rows from the model, duplicate column references are aliased in order of appearance.

        Parameters:
            ``columns`` (list of str, optional): The list of columns to include in the SQL statement. If not provided, all columns from the model will be included.
            ``rows`` (int, slice, tuple, optional): The rows to include in the SQL statement. It can be an integer for a single row, a slice for a range of rows, or a tuple for specific row indices. If not provided, all rows will be included.
            ``index`` (bool, optional): If True, include the primary index column in the SQL statement.
            ``na_rep`` (str, optional): If provided, all null or empty string values are replaced with value.

        Returns:
            ``str``: The generated SQL statement.   

        Note:
            - No validation is performed on row or column indicies, see :meth:`SQLDataModel._validate_indicies()` for implementation and usage.
            - See :meth:`SQLDataModel._generate_sql_stmt_fetchall()` for fetching all model data without predicates or filters.
        
        Changelog:
            - Version 0.5.1 (2024-05-10):
                - Modified to allow ``rows`` argument to be provided directly as a string predicate to bypass numeric range-based selections.
            - Version 0.4.0 (2024-04-23):
                - Added ``nap_rep`` parameter to fill null or missing fields with provided value.
            - Version 0.3.0 (2024-03-31):
                - Renamed ``include_index`` parameter to ``index`` for package consistency.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        columns = columns if columns else self.headers
        columns = [self.sql_idx,*columns] if index else columns
        if na_rep is None:
            headers_selection_str = ",".join([f'''"{col}" as "{col_alias}"''' for col,col_alias in zip(columns,utils.alias_duplicates(columns))])
        else:
            headers_selection_str = ",".join([f'''ifnull("{col}",'{na_rep}') as "{col_alias}"''' for col,col_alias in zip(columns,utils.alias_duplicates(columns))])
        if isinstance(rows, int):
            row_selection_str = f"""where "{self.sql_idx}" = {rows}"""
        elif isinstance(rows, slice):
            rows = self.indicies[rows]
            row_selection_str = f"""where "{self.sql_idx}" in {f'{rows}' if len(rows) != 1 else f'({rows[0]})'}"""
        elif isinstance(rows, tuple):
            row_selection_str = f"""where "{self.sql_idx}" in {f'{rows}' if len(rows) != 1 else f'({rows[0]})'}"""
        elif isinstance(rows, str):
            row_selection_str = rows
        else:
            row_selection_str = """"""
        order_by_str = f"""order by "{self.sql_idx}" asc"""
        fetch_stmt = f"""select {headers_selection_str} from "{self.sql_model}" {row_selection_str} {order_by_str}"""
        return fetch_stmt

    def _generate_sql_stmt_fetchall(self, index:bool=True) -> str:
        """
        Generates an SQL statement for fetching all current rows and columns in ``SQLDataModel``.

        Parameters:
            ``index`` (bool, optional): Whether or not to include index column in the SQL statement. Default is True, including the index.

        Returns:
            ``str``: The generated SQL statement selecting all rows and columns.

        Example::

            import sqldatamodel as sdm

            # Create a sample model
            df = sdm.from_shape(shape=(10,3), headers=['Name','Age','Sex'])

            # Generate an SQL statement for all data
            sql_stmt = df._generate_sql_stmt_fetchall(index=False) 

            # View it
            print(sql_stmt)

        This will output statement required to fetch all the data:

        ```text
            SELECT 
                "Name" AS "Name",
                "Age" AS "Age",
                "Sex" AS "Sex" 
            FROM 
                "sdm" 
            ORDER BY 
                "idx"
        ```

        Note:
            - Used internally for methods selecting all the current rows and columns
            - See :meth:`SQLDataModel._generate_sql_stmt()` for generating statements for specified rows and columns only.
                
        Changelog:
            - Version 0.10.0 (2024-06-29):
                - New method.
        """
        columns = [self.sql_idx,*self.headers] if index else self.headers
        headers_selection_str = ",".join((f'"{col}" AS "{col}"' for col in columns))
        return " ".join(("SELECT",headers_selection_str,f'FROM "{self.sql_model}" ORDER BY "{self.sql_idx}"'))

    def _generate_table_style(self, style:Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple']=None) -> tuple[tuple[str]]:
        """
        Generates the character sets required for formatting ``SQLDataModel`` according to the value currently set at :py:attr:`SQLDataModel.table_style`.
        
        Parameters:
            ``style`` (Literal['ascii','bare','dash','default','double','latex','list','markdown','outline','pandas','polars','postgresql','round','rst-grid','rst-simple'], optional): The table style to return. Default is value set on :py:attr:`SQLDataModel.table_style`.

        Returns:
            ``tuple[tuple[str]]``: A 4-tuple containing the characters required for top, middle, row and lower table sections.
                    
        Note:
            - This method is called by :meth:`SQLDataModel.__repr__()` to parse the characters necessary for constructing the tabular representation of the ``SQLDataModel``, any modifications or changes to this method may result in unexpected behavior.

        Changelog:
            - Version 0.11.0 (2024-07-05):
                - Added 'latex' style format.
            - Version 0.9.3 (2024-06-28):
                - Added 'rst-simple' and 'rst-grid' style formats.
            - Version 0.3.10 (2024-04-16):
                - Added ``style`` parameter to allow use by :meth:`SQLDataModel.to_text()` to generate new formatting styles introduced in version 0.3.9.
            - Version 0.3.8 (2024-04-12):
                - New method.            
        """
        style = self.table_style if style is None else style
        if style == 'ascii':    
            return  (('+-','-','-+-','-+') 
                    ,('+-','-','-+-','-+') 
                    ,('| ',    ' | ',' |') 
                    ,('+-','-','-+-','-+'))
        if style == 'bare':     
            return  (('','','','')
                    ,('','-','--','')      
                    ,('',    '  ','')      
                    ,('','','',''))        
        if style == 'dash': 
            return  (('┌─','─','─┬─','─┐')
                    ,('├╴','╴','╴┼╴','╴┤')
                    ,('│ ',    ' ╎ ',' │')
                    ,('└─','─','─┴─','─┘'))            
        if style == 'default':  
            return  (('┌─','─','─┬─','─┐') 
                    ,('├─','─','─┼─','─┤') 
                    ,('│ ',    ' │ ',' │') 
                    ,('└─','─','─┴─','─┘'))
        if style == 'double':   
            return  (('╔═','═','═╦═','═╗') 
                    ,('╠═','═','═╬═','═╣') 
                    ,('║ ',    ' ║ ',' ║') 
                    ,('╚═','═','═╩═','═╝'))
        if style == 'latex':     
            return  (('\\hline','','','') 
                    ,('\\hline','','','')
                    ,('    ',   ' & ',  ' \\\\') 
                    ,('\\hline','','',''))
        if style == 'list':     
            return  (('',' ','  ','')
                    ,('','-','  ','')
                    ,('',    '  ','')
                    ,('',' ','  ',''))        
        if style == 'markdown': 
            return  (('',    '',   '',  '')
                    ,('|-' ,'-','-|-','-|')
                    ,('| ' ,    ' | ',' |')
                    ,('','', '',        ''))
        if style == 'outline':
            return  (('┌─','─','──','─┐')
                    ,('├─','─','──','─┤')
                    ,('│ ',    '  ',' │')
                    ,('└─','─','──','─┘'))
        if style == 'pandas':
            return  (('','','','')
                    ,('','','','')
                    ,('', '  ','')
                    ,('','','',''))
        if style == 'polars':
            return  (('┌─','─','─┬─','─┐')
                    ,('╞═','═','═╪═','═╡')
                    ,('│ ',    ' ┆ ',' │')
                    ,('└─','─','─┴─','─┘'))
        if style == 'postgresql':
            return  (('', '',     '','')
                    ,('','-',  '-+-','')
                    ,('',      ' | ','')
                    ,('', '',     '',''))
        if style == 'round':
            return  (('╭─','─','─┬─','─╮')
                    ,('├─','─','─┼─','─┤')
                    ,('│ ',    ' │ ',' │')
                    ,('╰─','─','─┴─','─╯'))
        if style == 'rst-grid':
            return  (('+-','-','-+-','-+') 
                    ,('+=','=','=+=','=+') 
                    ,('| ',    ' | ',' |') 
                    ,('+-','-','-+-','-+'))
        if style == 'rst-simple':
            return  (('','=','  ','') 
                    ,('','=','  ','') 
                    ,('',    '  ','') 
                    ,('','=','  ',''))
        else: # default styling
            return  (('┌─','─','─┬─','─┐') 
                    ,('├─','─','─┼─','─┤') 
                    ,('│ ',    ' │ ',' │') 
                    ,('└─','─','─┴─','─┘'))   

    def _update_indicies(self) -> None:
        """
        Updates the :py:attr:`SQLDataModel.indicies` and :py:attr:`SQLDataModel.row_count` properties of the ``SQLDataModel`` instance representing the current valid row indicies and count.

        Returns:
            ``None``

        Note:
            - This method is called internally any time the :py:attr:`SQLDataModel.row_count` property is subject to change, or data manipulation requires updating the current values.
            - There is no reason to call this method manually unless the model has been changed outside of the standard instance methods.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        fetch_stmt = f"""select "{self.sql_idx}" from "{self.sql_model}" order by "{self.sql_idx}" asc"""
        self.indicies = tuple([x[0] for x in self.sql_db_conn.execute(fetch_stmt).fetchall()])
        self.row_count = len(self.indicies)
        self.shape = (self.row_count,self.shape[1])
    
    def _update_indicies_deterministic(self, row_index:int) -> None:
        """
        Quick implementation to update the :py:attr:`SQLDataModel.indicies` and :py:attr:`SQLDataModel.row_count` properties of the ``SQLDataModel`` instance representing the current valid row indicies and count based on the last inserted rowid.

        Returns:
            ``None``
            
        Note:
            - This method is called internally any time the :py:attr:`SQLDataModel.row_count` property is subject to deterministic change to avoid the more expensive call to :meth:`SQLDataModel._update_indicies()`

        Changelog:
            - Version 0.6.0 (2024-05-14):
                - Improves performance for updating row indicies when update is deterministic.
                - New method.
        """
        if row_index is None:
            return
        self.indicies = tuple(sorted(set((*self.indicies,row_index))))
        if not self.row_count and self.indicies == (1,):
            self.indicies = (0,) # required to account for insert trigger that adjusts first rowid from 1 to 0 when model originally contained zero rows and gets first insert
        self.row_count = len(self.indicies)
        self.shape = (self.row_count,self.shape[1])    

    def get_indicies(self) -> tuple:
        """
        Returns the current valid row indicies for the ``SQLDataModel`` instance.

        Returns:
            ``tuple``: A tuple of the current values for :py:attr:`SQLDataModel.sql_idx` in ascending order.

        Example::

            import sqldatamodel as sdm

            headers = ['Name', 'Age', 'Height']
            data = [('John', 30, 175.3), ('Alice', 28, 162.0), ('Travis', 35, 185.8)]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Get current valid indicies
            valid_indicies = df.get_indicies()

            # View results
            print(valid_indicies)
        
        This will output:

        ```text
            (0, 1, 2)
        ```

        Notes
            - Primary use is to confirm valid model indexing when starting index != 0 or filtering changes minimum/maximum indexes.

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        return self.indicies

    def _update_rows_and_columns_with_values(self, rows_to_update:tuple[int]=None, columns_to_update:list[str]=None, values_to_update:list[tuple]=None) -> None:
        """
        Generates and executes a SQL update statement to modify specific rows and columns with provided values in the SQLDataModel.

        Parameters:
            ``rows_to_update``: A tuple of row indices to be updated. If set to None, it defaults to all rows in the SQLDataModel.
            ``columns_to_update``: A list of column names to be updated. If set to None, it defaults to all columns in the SQLDataModel.
            ``values_to_update``: A list of tuples representing values to update in the specified rows and columns.

        Raises:
            ``TypeError``: If the ``values_to_update`` parameter is not a list or tuple.
            ``DimensionError``: If the shape of the provided values does not match the specified rows and columns.
            ``SQLProgrammingError``: If the ``values_to_update`` parameter contains invalid or SQL incompatible data.

        Example::

            import sqldatamodel as sdm

            # Create the model with some sample data
            df = sdm.SQLDataModel(
                data=[(23, 'W'), (24, 'X'), (25, 'Y'), (26, 'Z')], 
                headers=['column1', 'column2']
            )

            # Update specific rows and columns with provided values
            df._update_rows_and_columns_with_values(
                rows_to_update=(1, 2, 3),
                columns_to_update=["column1", "column2"],
                values_to_update=[(10, 'A'), (20, 'B'), (30, 'C')]
            )

            # Create a new column named "new_column" with default values
            df._update_rows_and_columns_with_values(
                columns_to_update=["new_column"],
                values_to_update=[(None,)] * df.row_count
            )

        Note:
            - Used by :meth:`SQLDataModel.__setitem__()` to broadcast updates across row and column index ranges.
            - To create a new column, pass a single header item in a list to the ``columns_to_update`` parameter.
            - To copy an existing column, pass the corresponding data is a list of tuples to the ``values_to_update`` parameter.        

        Changelog:
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        update_sql_script = None
        # this is the problem, even if the indicies are 2-7, this will generate 0:5 since the rowcount is 5 regardless of min and max indicies
        rows_to_update = rows_to_update if rows_to_update is not None else self.indicies # use all rows if none specified, courtesy of new `indicies` property
        columns_to_update = columns_to_update if columns_to_update is not None else self.headers
        if not isinstance(values_to_update, (tuple,list)):
            values_to_update = (values_to_update,)
            rowwise_update = False
        else:
            rowwise_update = True
        if isinstance(values_to_update, Iterable) and len(values_to_update) < 1:
            return None # Nothing to update
        if isinstance(values_to_update, list):
            if not isinstance(values_to_update[0], (tuple,list)):
                values_to_update = tuple(values_to_update)
        if isinstance(values_to_update, tuple):
            values_to_update = [values_to_update]
        num_rows_to_update = len(rows_to_update)
        num_columns_to_update = len(columns_to_update)
        num_value_rows_to_update = len(values_to_update)
        num_value_columns_to_update = len(values_to_update[0])
        create_new_column = True if (num_columns_to_update == 1 and columns_to_update[0] not in self.headers) else False
        if create_new_column:
            new_column = columns_to_update[0]
            new_column_py_dtype = type(values_to_update[0][0]).__name__ # should always have an item
            new_column_sql_dtype = self.static_py_to_sql_map_dict.get(new_column_py_dtype, 'TEXT')
            update_sql_script = f"""alter table "{self.sql_model}" add column "{new_column}" {new_column_sql_dtype};"""
        if not rowwise_update:
            values_to_update = values_to_update[0][0]
            values_to_update = "null" if values_to_update is None else f"""{values_to_update}""" if not isinstance(values_to_update, (str,bytes,datetime.date)) else f"datetime('{values_to_update}')" if isinstance(values_to_update,datetime.datetime) else f"date('{values_to_update}')" if isinstance(values_to_update,datetime.date) else f"""'{values_to_update.replace("'","''")}'""" if not isinstance(values_to_update,bytes) else f"""X'{values_to_update.hex()}'"""
            col_val_param = ','.join([f""" "{column}" = {values_to_update} """ for column in columns_to_update]) 
            if update_sql_script is None:
                update_sql_script = f"""update "{self.sql_model}" set {col_val_param} where {self.sql_idx} in {f'{rows_to_update}' if num_rows_to_update != 1 else f'({rows_to_update[0]})'};"""
            else:
                update_sql_script += f"""update "{self.sql_model}" set {col_val_param} where {self.sql_idx} in {f'{rows_to_update}' if num_rows_to_update != 1 else f'({rows_to_update[0]})'};"""
            self.execute_transaction(update_sql_script, update_row_meta=False)
            return            
        if num_rows_to_update != num_value_rows_to_update:
            msg = ErrorFormat(f"DimensionError: shape mismatch '{num_rows_to_update} != {num_value_rows_to_update}', number of rows to update '{num_rows_to_update}' must match provided number of value rows to update '{num_value_rows_to_update}'")
            raise DimensionError(msg)
        if num_columns_to_update != num_value_columns_to_update:
            msg = ErrorFormat(f"DimensionError: shape mismatch '{num_columns_to_update} != {num_value_columns_to_update}', number of columns to update '{num_columns_to_update}' must match provided number of value columns to update '{num_value_columns_to_update}'")
            raise DimensionError(msg)
        if update_sql_script is not None:
            self.execute_transaction(update_sql_script, update_row_meta=False)
        col_val_param = ','.join([f""" "{column}" = ? """ for column in columns_to_update]) 
        update_stmt = f"""update "{self.sql_model}" set {col_val_param} where {self.sql_idx} = ?"""
        update_params = [(*values_to_update[i], row) for i,row in enumerate(rows_to_update)]
        try:
            self.sql_db_conn.executemany(update_stmt, update_params)
            self.sql_db_conn.commit()
        except sqlite3.ProgrammingError as e:
            self.sql_db_conn.rollback()
            msg = ErrorFormat(f"SQLProgrammingError: invalid update values, SQL execution failed with '{e}'")
            raise SQLProgrammingError(msg) from None
        self._update_model_metadata(update_row_meta=False)
        return

    def _get_sql_create_stmt(self) -> str:
        """
        Retrieves the SQL create table statement used to create the current SQLDataModel.

        Returns:
            ``str``: The SQL create table statement for the SQLDataModel.

        Example::
        
            import sqldatamodel as sdm
            
            headers = ['idx', 'first', 'last', 'age']
            data = [
                (0, 'john', 'smith', 27)
                ,(1, 'sarah', 'west', 29)
                ,(2, 'mike', 'harlin', 36)
                ,(3, 'pat', 'douglas', 42)
            ]

            # Create the sample model
            df = sdm.SQLDataModel(data,headers)

            # Retrieve the create statement for the SQLDataModel
            create_stmt = df._get_sql_create_stmt()

            # Print the returned statement
            print(create_stmt)

        This will output:

        ```shell
            CREATE TABLE "sdm" ("idx" INTEGER PRIMARY KEY,"first" TEXT,"last" TEXT,"age" INTEGER)
        ```

        Changelog:
            - Version 0.1.5 (2023-11-24):
                - New method.
        """
        return self.sql_db_conn.execute("select sql from sqlite_master").fetchone()[0]

    def _get_display_args(self, include_dtypes:bool=False) -> dict:
        """
        Retrieves the current display configuration settings of the ``SQLDataModel`` with the correct ``kwargs`` for the class :meth:`SQLDataModel.__init__()` method.

        Parameters:
            ``include_dtypes`` (bool, optional): Whether :py:attr:`SQLDataModel.dtypes` should be included in the result. Default is False, including only display arguments.

        Returns:
            ``dict``: A dictionary containing the display configuration settings in the format ``{'setting': 'value'}``.

        Display Properties:
            - :py:attr:`SQLDataModel.display_max_rows`: The maximum number of rows to display.
            - :py:attr:`SQLDataModel.min_column_width`: The minimum width of columns when displaying the model.
            - :py:attr:`SQLDataModel.max_column_width`: The maximum width of columns when displaying the model.
            - :py:attr:`SQLDataModel.column_alignment`: The alignment of columns ('left', 'center', 'right' or 'dynamic').
            - :py:attr:`SQLDataModel.display_color`: The color to use when displaying the table, None by default.
            - :py:attr:`SQLDataModel.display_index`: True if displaying index column, False otherwise.
            - :py:attr:`SQLDataModel.display_float_precision`: The precision for displaying floating-point numbers.
            - :py:attr:`SQLDataModel.table_style`: The table styling format to use for strng representations of the model.
        
        Dtype Property:
            - :py:attr:`SQLDataModel.dtypes`: A dictionary mapping the current model's columns to their corresponding Python data type.

        Changelog:
            - Version 0.6.2 (2024-05-15):
                - Added ``include_dtypes`` parameter for use by methods such as :meth:`SQLDataModel.min()` and :meth:`SQLDataModel.max()` for operations that require returning the results of SQL fetch statements.
            - Version 0.1.9 (2024-03-19):
                - New method.
        """
        args = {"display_max_rows":self.display_max_rows, "min_column_width":self.min_column_width, "max_column_width":self.max_column_width, "column_alignment":self.column_alignment, "display_color":self.display_color, "display_index":self.display_index, "display_float_precision":self.display_float_precision, "table_style":self.table_style}
        if include_dtypes:
            args['dtypes'] = self.dtypes
        return args

    def _validate_row(self, row:int|slice|Iterable[int], unmodified:bool=False) -> tuple[int]:
        """
        Utility function used to validate row selection and return parsed values.

        Parameters:
            ``row`` (int|slice|Iterable[int]): The row selection to validate, argument should reflect the integer indexes of the rows to select.
            ``unmodified`` (bool, optional): Whether ``row`` should be returned as originally indexed. Default is False, returning as tuple.

        Raises:
            ``TypeError``: If ``row`` is not one of type 'int', 'slice' or 'Iterable' representing the integer index of row(s) to select.
            ``IndexError``: If ``row`` is outside of current model range bounded by :py:attr:`SQLDataModel.row_count` whether positively or negatively indexed.
            
        Returns:
            ``tuple[int]``: A tuple containing the validated row values resulting from the selection.

        Example::

            import sqldatamodel as sdm

            # Create a 10 rows x 3 column model
            df = sdm.from_shape(shape=(10, 3), headers=['A','B','C'])
            
            # Various row index types
            row_indicies = [
                2,
                -3,
                {4,5,8},
                [-1,5,0],
                slice(2, 5),
                slice(-8, -1, 2),
            ]

            # Loop over the indicies
            for row_index in row_indicies:
                # Print original and validated indexes
                print(f"{row_index} --> {sdm._validate_row(row_index)}")

        This will output the original and validated row indexes:

        ```text
            2 --> (2,)
            -3 --> (7,)
            {8, 4, 5} --> (8, 4, 5)
            [-1, 5, 0] --> (9, 5, 0)
            slice(2, 5, None) --> (2, 3, 4)
            slice(-8, -1, 2) --> (2, 4, 6, 8)
        ```

        Note:
            - Rows are referenced by their integer index and not their value, as such ``row = 0`` and ``row = -1`` will always return the first and last rows, respectively.
            - An input of ``row`` == :py:attr:`SQLDataModel.row_count` is allowed to accomodate the append row syntax of ``sdm[sdm.row_count] = (values)``.
            - See :meth:`SQLDataModel._validate_column()` for validating column indicies and returning the corresponding headers.

        Changelog:
            - Version 0.7.9 (2024-06-20):
                - New method.
        """
        if not isinstance(row, (int,slice,Iterable)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(row).__name__}', rows must be referenced by their integer index or by an iterable of indexes as type 'int'")
            raise TypeError(msg)        
        if isinstance(row, int):
            try:
                validated_row = self.indicies[row]
            except IndexError:
                msg = ErrorFormat(f"IndexError: invalid row index '{row}', index must be within current model row range of '0:{self.row_count}' ")
                raise IndexError(msg) from None
            return (validated_row,) if not unmodified else row
        elif isinstance(row, slice):
            validated_row = self.indicies[row]
            if (num_rows_in_scope := len(validated_row)) < 1:
                msg = ErrorFormat(f"IndexError: insufficient rows '{num_rows_in_scope}', provided row slice returned no valid row indicies within current model range of '0:{self.row_count}'")
                raise IndexError(msg)
            return validated_row if not unmodified else row
        else: # Iterable[int]
            try:
                validated_row = tuple([self.indicies[rid] for rid in row])
            except Exception as e:
                msg = ErrorFormat(f"{type(e).__name__}: {e}, rows must be referenced by their integer index within current model row range of '0:{self.row_count}'")
                raise type(e)(msg).with_traceback(e.__traceback__) from None                
            return validated_row if not unmodified else row
        
    def _validate_column(self, column:str|int|slice|Iterable, unmodified:bool=False) -> list[str]:
        """
        Utility function used to validate column selection and return parsed values.

        Parameters:
            ``column`` (str|int|slice|Iterable): The column selection to validate, argument should reflect the integer indexes or column names.
            ``unmodified`` (bool, optional): Whether ``column`` should be returned as originally indexed. Default is False, returning as list.
        
        Raises:
            ``TypeError``: If ``column`` is not one of type 'str', 'int', 'slice' or 'Iterable' representing the integer index or values of column(s) to select.
            ``IndexError``: If ``column`` is outside of current model range bounded by :py:attr:`SQLDataModel.column_count` whether positively or negatively indexed.
            ``DimensionError``: If ``len(column)`` is greater than :py:attr:`SQLDataModel.column_count` when provided as an iterable or sequence.
            
        Returns:
            ``list[str]``: A list containing the validated column values resulting from the selection.

        Example::

            import sqldatamodel as sdm

            # Create a 10 rows x 6 column model
            df = sdm.from_shape((10, 6), headers=['A','B','C','D','E','F'])
            
            # Various column index types
            column_indicies = [
                2,
                -3,
                ['A','B'],
                (4, 5, 1),
                [-1, 2, 0],
                slice(1, 3),
                slice(-6, -1, 2),
            ]

            # Loop over the indicies
            for column_index in column_indicies:
                # Print original and validated indexes
                print(f"{column_index} --> {sdm._validate_column(column_index)}")

        This will output the original and validated column indexes:

        ```text
            2 --> ['C']
            -3 --> ['D']
            ['A', 'B'] --> ['A', 'B']
            (4, 5, 1) --> ['E', 'F', 'B']
            [-1, 2, 0] --> ['F', 'C', 'A']
            slice(1, 3, None) --> ['B', 'C']
            slice(-6, -1, 2) --> ['A', 'C', 'E']
        ```

        Note:
            - Columns are referenced by their integer index or directly by their value as a column name, when using integers ``column = 0`` and ``column = -1`` will always return the first and last columns, respectively.
            - Validated column outputs will be returned as a list containing the results of the indexed columns found at :py:attr:`SQLDataModel.headers` with original ordering intact.
            - See :meth:`SQLDataModel._validate_row()` for validating row indicies and returning the corresponding values.        

        Changelog:
            - Version 0.7.9 (2024-06-20):
                - New method.
        """
        if not isinstance(column, (str,int,slice,Iterable)):
            msg = ErrorFormat(f"TypeError: invalid type '{type(column).__name__}', columns must be referenced by name as type 'str' or by their index as type 'int'")
            raise TypeError(msg)
        if isinstance(column, (str,int,slice)):
            if isinstance(column, int):
                try:
                    validated_column = self.headers[column]
                except IndexError:
                    msg = ErrorFormat(f"IndexError: invalid column index '{column}', column index is outside of current model range '0:{self.column_count}', use `get_headers()` to veiw current valid arguments")
                    raise IndexError(msg) from None
                return [validated_column] if not unmodified else column
            elif isinstance(column, str):
                if column not in self.headers:
                    msg = ErrorFormat(f"ValueError: column not found '{column}', column must be in current model, use `get_headers()` to view valid arguments")
                    raise ValueError(msg)
                return [column] if not unmodified else column
            else: # Slice
                validated_column = self.headers[column]
                if len(validated_column) < 1:
                    msg = ErrorFormat(f"ValueError: no columns selected, at least 1 valid column selection is required when specifying column indicies")
                    raise ValueError(msg)
                return validated_column if not unmodified else column
        validated_column = []
        # Iterable[str|int]
        if len(column) > self.column_count:
            msg = ErrorFormat(f"DimensionError: invalid column dimensions '{len(column)} > {self.column_count}', provided column count exceeds current model dimensions")
            raise DimensionError(msg)
        for col in column: 
            if isinstance(col, int):
                try:
                    validated_column.append(self.headers[col])
                except IndexError:
                    msg = ErrorFormat(f"IndexError: invalid column index '{col}', column index is outside of current model range '0:{self.column_count}', use `get_headers()` to veiw current valid arguments")
                    raise IndexError(msg) from None
            elif isinstance(col, str):
                if col not in self.headers:
                    msg = ErrorFormat(f"ValueError: column not found '{col}', column must be in current model, use `get_headers()` to view valid arguments")
                    raise ValueError(msg)
                validated_column.append(col)
            else:
                msg = ErrorFormat(f"TypeError: invalid type '{type(col).__name__}', columns must be referenced by name as type 'str' or by their index as type 'int'")
                raise TypeError(msg)
        return validated_column if not unmodified else column

    def _validate_indicies(self, indicies) -> tuple[tuple[int], list[str]]:
        """
        Validates and returns a predictable notation form of indices for accessing rows and columns in the ``SQLDataModel`` from varying 2-dimensional ``(row, column)`` indexing input types.

        Two dimensional indexing:
            - ``tuple[row_index, column_index]``: Where row_index and column_index are defined below.

        Row indexing:
            - ``int``: Single integer index. E.g., ``sdm[0]`` or ``sdm[-1]``
            - ``slice``: Range of row indices. E.g., ``sdm[2:5]`` or ``sdm[-8:-1]``
            - ``set[int]``: Discontiguous row indicies. E.g., ``sdm[{13, 7, 42}]``
            - ``tuple[int]``: Like set, discontiguous row indices. E.g., ``sdm[(-1, 9, 11)]``
        
        Column indexing:
            - ``int``: Single integer index. E.g., ``sdm[:, 0]`` or ``sdm[:, -1]``
            - ``str``: Single column name. E.g., ``sdm['Col A']`` or ``sdm['Name']``
            - ``list[str]``: List of column names. E.g., ``sdm[:,['A', 'B', 'F']]``
            - ``list[int]``: List of column indicies. E.g., ``sdm[:,[0, 3, 4, 9, -2]]``

        Parameters:
            ``indicies``: Specifies the indices for rows and columns.

        Raises:
            ``TypeError``: If the type of indices is invalid such as a float for row index or a boolean for a column name index.
            ``ValueError``: If the indices are outside the current model range or if a column is not found in the current model headers when indexed by column name as ``str``.
            ``IndexError``: If the column indices are outside the current column range or if a column is not found in the current model headers when indexed by ``int``.

        Returns:
            ``tuple[tuple[int], list[str]]``: A tuple containing validated row indices as a tuple and validated column indices as a list of column names.

        Example::

            import sqldatamodel as sdm

            # Create a 10 rows by 4 columns model
            df = sdm.from_shape(shape=(10,4), headers=['A','B','C','D'])

            # Index pairs to validate
            input_idx = [
                (0, 'A'),
                (-1, ['B','D']),
                ({2,-7,-2}, (-2,-1)),
                (slice(-6,-1,2), slice(0,3))
            ]

            # Store validated pairs
            valid_idx = []

            # Loop over the [row, col] pairs
            for row, col in input_idx:
                # Validated and store the pairs
                valid_idx.append(df._validate_indicies((row, col)))
            
            # View input and validated pairs
            for original, validated in zip(input_idx, valid_idx):
                print(f"{original} --> {validated}") 

        This will output both the input and validated row, column index pairs:

        ```text
            (0, 'A') --> ((0,), ['A'])
            (-1, ['B', 'D']) --> ((9,), ['B', 'D'])
            ({-7, 2, -2}, (-2, -1)) --> ((3, 2, 8), ['C', 'D'])
            (slice(-6, -1, 2), slice(0, 3, None)) --> ((4, 6, 8), ['A', 'B', 'C'])
        ```

        Note:
            - This method expects indicies to be provided as a two dimensional pair of ``(row, column)`` indicies, with exceptions made for single row integer indexes or single column names.
            - Use empty slice notation to include all indicies from a given dimension, for example ``sdm[:, :]`` will always return the full model by accessing all rows and all columns.
            - See :meth:`SQLDataModel.__getitem__()` and :meth:`SQLDataModel.__setitem__()` for implementations relying on this method.
            - See :meth:`SQLDataModel._validate_row()` for one dimensional validation against a single ``row`` index.
            - See :meth:`SQLDataModel._validate_column()` for one dimensional validation against a single ``column`` index.

        Changelog:
            - Version 0.8.1 (2024-06-23):
                - Modified implementation to leverage new utility methods :meth:`SQLDataModel._validate_row()` and :meth:`SQLDataModel._validate_column()` to improve performance.
                - New method.
        """ 
        # Single row index
        if isinstance(indicies, (int, slice, set)):
            validated_rows = self._validate_row(indicies)
            return (validated_rows, self.headers)        
        # Single col index
        if isinstance(indicies, (list, str)):
            validated_columns = self._validate_column(indicies)
            return (self.indicies, validated_columns)
        if not isinstance(indicies, tuple):
            msg = ErrorFormat(f"TypeError: invalid indexing type '{type(indicies).__name__}', indexing args must be compatible with two-dimensional `sdm[rows, columns]` parameters with correct types")
            raise TypeError(msg)
        if (arg_length := len(indicies)) != 2:
            msg = ErrorFormat(f"ValueError: invalid indexing args, expected no more than 2 indicies for `sdm[row, column]` but '{arg_length}' were received")
            raise ValueError(msg)
        # Two dimensional row, column indicies
        input_rows, input_cols = indicies
        validated_rows = self._validate_row(input_rows)
        validated_cols = self._validate_column(input_cols)
        return (validated_rows, validated_cols) 

###############################################################################################################
##################################### string methods and value comparison #####################################
###############################################################################################################

    def contains(self, pat:str|Iterable[str], case:bool=True) -> set[int]:
        """
        Return the row indices that contain the specified pattern(s) in any column from the model, converting to ``str(value)`` for comparison.

        Parameters:
            ``pat`` (str | Iterable[str]): The pattern or iterable of patterns to search for within the data.
            ``case`` (bool, optional): If True (default), the search is case-sensitive. If False, the search is case-insensitive.

        Raises:
            ``TypeError``: If argument for ``pat`` is not of type 'str' or an iterable of type 'str' representing the substring pattern(s).

        Returns:
            ``set[int]``: Set of row indices containing values that match the pattern(s).

        Example::

            import sqldatamodel as sdm
            
            # Sample data
            headers = ['Name', 'Age', 'Sex', 'City']
            data = [
                ('Mike', 31, 'M', 'Chicago'),
                ('John', 25, 'M', 'Dayton'),
                ('Alice', 27, 'F', 'Boston'),
                ('Sarah', 35, 'F', 'Houston'),
                ('Bobby', 42, 'M', 'Chicago'),
                ('Steve', 28, 'F', 'Austin'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter for rows containing the string 'Chicago'
            matching_indicies = df['City'].contains('Chicago')

            # Apply filter to model
            df_chicago = df[matching_indicies]

            # View result
            print(df_chicago)

        This will output the result of applying the filter to the model:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┐
            │   │ Name  │ Age │ Sex │ City    │
            ├───┼───────┼─────┼─────┼─────────┤
            │ 0 │ Mike  │  31 │ M   │ Chicago │
            │ 4 │ Bobby │  42 │ M   │ Chicago │
            └───┴───────┴─────┴─────┴─────────┘
            [2 rows x 4 columns]
        ```

        Instead of searching a single column, the entire model can be searched:

        ```python
            # Method can also search all columns, and be applied directly
            df_with_e = df[df.contains('E', case=False)]
            
            # View result
            print(df_with_e)
        ```

        This will output the result of a case-insensitive search:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┐
            │   │ Name  │ Age │ Sex │ City    │
            ├───┼───────┼─────┼─────┼─────────┤
            │ 0 │ Mike  │  31 │ M   │ Chicago │
            │ 2 │ Alice │  27 │ F   │ Boston  │
            │ 5 │ Steve │  28 │ F   │ Austin  │
            └───┴───────┴─────┴─────┴─────────┘
            [3 rows x 4 columns]
        ```

        This can be used in combination with the setitem syntax to selectively update values as well:

        ```python
            # Create a 'State' column with a default value
            df['State'] = None
            
            # Filter and set the values that contain the pattern
            df[df.contains('Chicago'), 'State'] = 'Illinois'

            # Multiple conditions can be used
            tx_1 = df.contains('Houston')
            tx_2 = df.contains('Austin')

            # Then chained together using set notation
            df[(tx_1 | tx_2), 'State'] = 'Texas'

            # Alternatively, an iterable of patterns can be provided
            df[df.contains(['Houston','Austin']), 'State'] = 'Texas'
        ```

        Note:
            - Any non-string values are converted using ``str(value)`` for comparisons only.
            - See :meth:`SQLDataModel.__eq__()` for strict equality comparison operations.
            - See :meth:`SQLDataModel.__and__()` for more details on bitwise and set operations.
            - See :meth:`SQLDataModel.__setitem__()` for more details on syntax ``df[row, column] = value`` and correct usage.
            - See :meth:`SQLDataModel.startswith()` and :meth:`SQLDataModel.endswith()` for additional string methods.

        Changelog:
            - Version 0.7.8 (2024-06-18):
                - New method.
        """
        if isinstance(pat, str):
            str_func = lambda x: pat in x if case else pat.lower() in x.lower()
        elif isinstance(pat, Iterable):
            str_func = lambda x: any(p in x for p in pat) if case else any(p.lower() in x.lower() for p in pat)
        else:
            msg = ErrorFormat(f"TypeError: invalid type '{type(pat).__name__}', argument for `pat` must be of type 'str' or an iterable of 'str' representing the substring pattern(s)")
            raise TypeError(msg)
        str_data = [tuple([str(x) for x in row]) for row in self.data(strict_2d=True)]
        return set(i for i in range(len(str_data)) if any(str_func(str_data[i][j]) for j in range(len(str_data[0]))))

    def startswith(self, pat:str|Iterable[str], case:bool=True) -> set[int]:
        """
        Return the row indices that start with the specified pattern(s) in any column from the model, converting to ``str(value)`` for comparison.

        Parameters:
            ``pat`` (str | Iterable[str]): The pattern or iterable of patterns to search for within the data.
            ``case`` (bool, optional): If True (default), the search is case-sensitive. If False, the search is case-insensitive.

        Raises:
            ``TypeError``: If argument for ``pat`` is not of type 'str' or an iterable of type 'str' representing the substring pattern(s).

        Returns:
            ``set[int]``: Set of row indices containing values that match the pattern(s).

        Example::

            import sqldatamodel as sdm
            
            # Sample data
            headers = ['Name', 'Age', 'Sex', 'City']
            data = [
                ('Mike', 31, 'M', 'Chicago'),
                ('John', 25, 'M', 'Dayton'),
                ('Alice', 27, 'F', 'Boston'),
                ('Sarah', 35, 'F', 'Houston'),
                ('Bobby', 42, 'M', 'Chicago'),
                ('Steve', 28, 'F', 'Austin'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter for rows where any column starts with the string 'Chi'
            matching_indices = df['City'].startswith('Chi')

            # Apply filter to model
            df_city = df[matching_indices]

            # View result
            print(df_city)

        This will output the result of applying the filter to the model:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┐
            │   │ Name  │ Age │ Sex │ City    │
            ├───┼───────┼─────┼─────┼─────────┤
            │ 0 │ Mike  │  31 │ M   │ Chicago │
            │ 4 │ Bobby │  42 │ M   │ Chicago │
            └───┴───────┴─────┴─────┴─────────┘
            [2 rows x 4 columns]
        ```

        Instead of searching a single column, the entire model can be searched:

        ```python
            # Method can also search all columns, and be applied directly
            df_prefix = df[df.startswith('A', case=False)]
            
            # View result
            print(df_prefix)
        ```

        This will output the result of a case-insensitive search:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┐
            │   │ Name  │ Age │ Sex │ City    │
            ├───┼───────┼─────┼─────┼─────────┤
            │ 2 │ Alice │  27 │ F   │ Boston  │
            │ 5 │ Steve │  28 │ F   │ Austin  │
            └───┴───────┴─────┴─────┴─────────┘
            [2 rows x 4 columns]
        ```

        This can be used in combination with the setitem syntax to selectively update values as well:

        ```python
            # Create a 'State' column with a default value
            df['State'] = None
            
            # Filter and set the values that start with the pattern
            df[df.startswith('Chi'), 'State'] = 'Illinois'

            # Multiple conditions can be used
            tx_1 = df.startswith('Hou')
            tx_2 = df.startswith('Aus')

            # Then chained together using set notation
            df[(tx_1 | tx_2), 'State'] = 'Texas'

            # Alternatively, an iterable of patterns can be provided
            df[df.startswith(['Hou','Aus']), 'State'] = 'Texas'
        ```

        Note:
            - Any non-string values are converted using ``str(value)`` for comparisons only.
            - See :meth:`SQLDataModel.__eq__()` for strict equality comparison operations.
            - See :meth:`SQLDataModel.__and__()` for more details on bitwise and set operations.
            - See :meth:`SQLDataModel.__setitem__()` for more details on syntax ``df[row, column] = value`` and correct usage.
            - See :meth:`SQLDataModel.contains()` and :meth:`SQLDataModel.endswith()` for additional string methods.

        Changelog:
            - Version 0.7.8 (2024-06-18):
                - New method.
        """
        if isinstance(pat, str):
            str_func = lambda x: x.startswith(pat) if case else x.lower().startswith(pat.lower())
        elif isinstance(pat, Iterable):
            str_func = lambda x: any(x.startswith(str(p)) for p in pat) if case else any(x.lower().startswith(str(p).lower()) for p in pat)
        else:
            msg = ErrorFormat(f"TypeError: invalid type '{type(pat).__name__}', argument for `pat` must be of type 'str' or an iterable of 'str' representing the substring pattern(s)")
            raise TypeError(msg)
        str_data = [tuple([str(x) for x in row]) for row in self.data(strict_2d=True)]
        return set(i for i in range(len(str_data)) if any(str_func(str_data[i][j]) for j in range(len(str_data[0]))))
    
    def endswith(self, pat:str|Iterable[str], case:bool=True) -> set[int]:
        """
        Return the row indices that end with the specified pattern(s) in any column from the model, converting to ``str(value)`` for comparison.

        Parameters:
            ``pat`` (str | Iterable[str]): The pattern or iterable of patterns to search for within the data.
            ``case`` (bool, optional): If True (default), the search is case-sensitive. If False, the search is case-insensitive.

        Raises:
            ``TypeError``: If argument for ``pat`` is not of type 'str' or an iterable of type 'str' representing the substring pattern(s).

        Returns:
            ``set[int]``: Set of row indices containing values that match the pattern(s).

        Example::

            import sqldatamodel as sdm
            
            # Sample data
            headers = ['Name', 'Age', 'Sex', 'City']
            data = [
                ('Mike', 31, 'M', 'Chicago'),
                ('John', 25, 'M', 'Dayton'),
                ('Alice', 27, 'F', 'Boston'),
                ('Sarah', 35, 'F', 'Houston'),
                ('Bobby', 42, 'M', 'Chicago'),
                ('Steve', 28, 'F', 'Austin'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter for rows where any column ends with the string 'ston'
            matching_indices = df['City'].endswith('ston')

            # Apply filter to model
            df_suffix = df[matching_indices]

            # View result
            print(df_suffix)  

        This will output the result of applying the filter to the model:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┐
            │   │ Name  │ Age │ Sex │ City    │
            ├───┼───────┼─────┼─────┼─────────┤
            │ 2 │ Alice │  27 │ F   │ Boston  │
            │ 3 │ Sarah │  35 │ F   │ Houston │
            └───┴───────┴─────┴─────┴─────────┘
            [2 rows x 4 columns]
        ```

        Instead of searching a single column, the entire model can be searched:

        ```python
            # Method can also search all columns, and be applied directly
            df_n = df[df.endswith('N', case=False)]
            
            # View result
            print(df_n)
        ```

        This will output the result of a case-insensitive search:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┐
            │   │ Name  │ Age │ Sex │ City    │
            ├───┼───────┼─────┼─────┼─────────┤
            │ 1 │ John  │  25 │ M   │ Dayton  │
            │ 2 │ Alice │  27 │ F   │ Boston  │
            │ 3 │ Sarah │  35 │ F   │ Houston │
            │ 5 │ Steve │  28 │ F   │ Austin  │
            └───┴───────┴─────┴─────┴─────────┘
            [4 rows x 4 columns]
        ```

        This can be used in combination with the setitem syntax to selectively update values as well:

        ```python
            # Create a new column 'Parity' with a default value
            df['Parity'] = None
            
            # Create patterns for even or odd suffixes
            even_suffixes = [0,2,4,6,8]
            odd_suffixes = [1,3,5,7,9]

            # Create the filters for both outcomes
            even_filter = df.endswith(even_suffixes)
            odd_filter = df.endswith(odd_suffixes)

            # Update values based on filters using setitem syntax
            df[even_filter, 'Parity'] = 'Even'
            df[odd_filter, 'Parity'] = 'Odd'

            # View result
            print(df)
        ```

        This will output the result of selectively applying updates based on our filters:

        ```shell
            ┌───┬───────┬─────┬─────┬─────────┬────────┐
            │   │ Name  │ Age │ Sex │ City    │ Parity │
            ├───┼───────┼─────┼─────┼─────────┼────────┤
            │ 0 │ Mike  │  31 │ M   │ Chicago │ Odd    │
            │ 1 │ John  │  25 │ M   │ Dayton  │ Odd    │
            │ 2 │ Alice │  27 │ F   │ Boston  │ Odd    │
            │ 3 │ Sarah │  35 │ F   │ Houston │ Odd    │
            │ 4 │ Bobby │  42 │ M   │ Chicago │ Even   │
            │ 5 │ Steve │  28 │ F   │ Austin  │ Even   │
            └───┴───────┴─────┴─────┴─────────┴────────┘
            [6 rows x 5 columns]
        ```

        Note:
            - Any non-string values are converted using ``str(value)`` for comparisons only.
            - See :meth:`SQLDataModel.__eq__()` for strict equality comparison operations.
            - See :meth:`SQLDataModel.__and__()` for more details on bitwise and set operations.
            - See :meth:`SQLDataModel.__setitem__()` for more details on syntax ``df[row, column] = value`` and correct usage.        
            - See :meth:`SQLDataModel.contains()` and :meth:`SQLDataModel.startswith()` for additional string methods.

        Changelog:
            - Version 0.7.8 (2024-06-18):
                - New method.
        """
        if isinstance(pat, str):
            str_func = lambda x: x.endswith(pat) if case else x.lower().endswith(pat.lower())
        elif isinstance(pat, Iterable):
            str_func = lambda x: any(x.endswith(str(p)) for p in pat) if case else any(x.lower().endswith(str(p).lower()) for p in pat)
        else:
            msg = ErrorFormat(f"TypeError: invalid type '{type(pat).__name__}', argument for `pat` must be of type 'str' or an iterable of 'str' representing the substring pattern(s)")
            raise TypeError(msg)
        str_data = [tuple([str(x) for x in row]) for row in self.data(strict_2d=True)]
        return set(i for i in range(len(str_data)) if any(str_func(str_data[i][j]) for j in range(len(str_data[0]))))

    def isna(self) -> set[int]:
        """
        Return the row indicies containing null values from the current model.
        
        Returns:
            ``set[int]``: Set of row indicies containing null values.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Gender', 'City']
            data = [
                ('Sarah', 35, 'Female', 'Houston'),
                ('Alice', None, 'Female', 'Milwaukee'),
                ('Mike', None, 'Male', 'Atlanta'),
                ('John', 25, 'Male', 'Boston'),
                ('Bob', None, 'Male', 'Chicago'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter for rows where 'Age' is null
            df = df[df['Age'].isna()]

            # View result
            print(df)

        This will output the result containing the rows where 'Age' was null:

        ```text
            ┌───────┬─────┬────────┬───────────┐
            │ Name  │ Age │ Gender │ City      │
            ├───────┼─────┼────────┼───────────┤
            │ Alice │     │ Female │ Milwaukee │
            │ Mike  │     │ Male   │ Atlanta   │
            │ Bob   │     │ Male   │ Chicago   │
            └───────┴─────┴────────┴───────────┘
            [3 rows x 4 columns]    
        ```

        This can be used in combination with the setitem syntax to selectively update values as well:

        ```python
            # Filter and set the null values
            df[df['Age'].isna(), 'Age'] = 'Missing'
        ```

        Note:
            - Null or na like is determined by satisfying the SQL NULL value or the Python equivalent ``None`` for all values in the row.
            - See related :meth:`SQLDataModel.notna()` to filter for rows containing values that are not null.
            - See :meth:`SQLDataModel.fillna()` to fill all missing or null values in the model.

        Changelog:
            - Version 0.7.2 (2024-06-11):
                - New method.
        """
        self_data = self.data(strict_2d=True)
        return set(i for i in range(len(self_data)) if all(self_data[i][j] is None for j in range(len(self_data[0]))))
    
    def notna(self) -> set[int]:
        """
        Return the row indicies that do not contain null values from the current model.
        
        Returns:
            ``set[int]``: Set of row indicies containing values that are not null.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Gender', 'City']
            data = [
                ('Sarah', 35, 'Female', 'Houston'),
                ('Alice', None, 'Female', 'Milwaukee'),
                ('Mike', None, 'Male', 'Atlanta'),
                ('John', 25, 'Male', 'Boston'),
                ('Bob', None, 'Male', 'Chicago'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Filter for rows where 'Age' is not null
            df = df[df['Age'].notna()]

            # View result
            print(df)

        This will output the result containing the rows where 'Age' was not null:

        ```text
            ┌───────┬─────┬────────┬─────────┐
            │ Name  │ Age │ Gender │ City    │
            ├───────┼─────┼────────┼─────────┤
            │ Sarah │  35 │ Female │ Houston │
            │ John  │  25 │ Male   │ Boston  │
            └───────┴─────┴────────┴─────────┘
            [2 rows x 4 columns]
        ```

        This can be used in combination with the setitem syntax to selectively update values as well:

        ```python
            # Create a 'Notes' column with a default value
            df['Notes'] = 'Missing'
            
            # Filter and set the values that are not null
            df[df['Age'].notna(), 'Notes'] = 'Valid'    
        ```

        Note:
            - Null or na like is determined by satisfying the SQL NOT NULL value or the Python equivalent ``None`` for any values in the row.
            - See related :meth:`SQLDataModel.isna()` to filter for rows containing values that are null.
            - See :meth:`SQLDataModel.fillna()` to fill all missing or null values in the model.

        Changelog:
            - Version 0.7.2 (2024-06-11):
                - New method.
        """
        self_data = self.data(strict_2d=True)
        return set(i for i in range(len(self_data)) if any(self_data[i][j] is not None for j in range(len(self_data[0]))))

    def dropna(self, axis:Literal['rows','columns']='columns', how:Literal['any','all']='all', strictly_null:bool=True, ignore_index:bool=True, inplace:bool=True) -> None|SQLDataModel:
        """
        Drop rows or columns with NA values from the SQLDataModel.

        Parameters:
            ``axis`` (Literal['rows', 'columns'], optional): The axis along which to drop NA values as either ``'rows'`` or ``'columns'``. 
                Default is ``'columns'``, dropping columns with all NA values.
            ``how`` (Literal['any', 'all'], optional): Determines when to drop NA values, ``'any'`` drops if any NA values are present, ``'all'`` drops only if all values are NA.
                Default is ``'all'``, dropping only when all the values are NA along a the specified axis.
            ``strictly_null`` (bool, optional): If True, only strictly NULL values are considered NA. If False, additional representations of NA (e.g., 'NaN', 'n/a') are also considered.
            ``ignore_index`` (bool, optional): If True, the index column is not considered when dropping rows. Ignored if when axis is set to ``'columns'``.
            ``inplace`` (bool, optional): If True, perform the operation in place and modify the SQLDataModel. If False, return a new SQLDataModel with the NA values dropped.

        Returns:
            ``None`` or ``SQLDataModel``: If ``inplace=False`` returns a new SQLDataModel with the NA values dropped. Otherwise, modifies the current SQLDataModel in place and returns ``None``.

        Raises:
            ``ValueError``: If ``axis`` is not one of ('rows', 'columns') or ``how`` is not one of ``'any'`` or ``'all'``.
            ``DimensionError``: If all columns are to be dropped when ``axis='columns'`` resulting in an invalid model schema.

        Example::

            import sqldatamodel as sdm

            # Sample data
            headers = ['Name', 'Age', 'Gender', 'City']
            data = [
                ('Sarah', 35, 'Female', 'Houston'),
                ('Alice', None, 'Female', 'Milwaukee'),
                ('Mike', None, 'Male', 'Atlanta'),
                ('John', 25, 'Male', 'Boston'),
                ('Bob', None, 'Male', 'Chicago'),
            ]

            # Create the model
            df = sdm.SQLDataModel(data, headers)

            # Drop columns with any NA values in place
            df.dropna(axis='columns', how='any', inplace=True)

            # View result
            print(df)

        This will output the updated model after dropping the 'Age' column:

        ```text
            ┌───────┬────────┬───────────┐
            │ Name  │ Gender │ City      │
            ├───────┼────────┼───────────┤
            │ Sarah │ Female │ Houston   │
            │ Alice │ Female │ Milwaukee │
            │ Mike  │ Male   │ Atlanta   │
            │ John  │ Male   │ Boston    │
            │ Bob   │ Male   │ Chicago   │
            └───────┴────────┴───────────┘
        ```

        Rows can also be used as the axis to check against

        ```python
            # Drop rows with any NA values
            df = df.dropna(axis='rows', how='any')

            # View result
            print(df)
        ```

        This will output the result containing only the rows where no NA values are present:

        ```text
            ┌───────┬─────┬────────┬─────────┐
            │ Name  │ Age │ Gender │ City    │
            ├───────┼─────┼────────┼─────────┤
            │ Sarah │  35 │ Female │ Houston │
            │ John  │  25 │ Male   │ Boston  │
            └───────┴─────┴────────┴─────────┘
            [2 rows x 4 columns]
        ```

        Note:
            - Null or na like is determined by satisfying the SQL NULL value or 'null like' values when ``strictly_null = False`` in the specified axis.
            - See :meth:`SQLDataModel.isna()` or :meth:`SQLDataModel.notna()` to filter for rows containing null values.
            - See :meth:`SQLDataModel.fillna()` to fill all missing or null values in the model.

        Changelog:
            - Version 1.0.0 (2024-08-09):
                - Changed default to ``inplace = True`` to align more with similar ``SQLDataModel`` methods.
            - Version 0.12.3 (2024-07-11):
                - New method.
        """
        if axis not in ('rows', 'columns'):
            msg = ErrorFormat(f"ValueError: invalid value '{axis}', argument for `axis` must be one of 'rows' or 'columns' representing along which axes to drop NA values")
            raise ValueError(msg)
        if how not in ('any', 'all'):
            msg = ErrorFormat(f"ValueError: invalid value '{how}', argument for `how` must be one of 'any' or 'all' representing when to drop NA values")
            raise ValueError(msg)    
        na_values = ('none','null','nan','n/a','na','')
        if axis == 'rows':
            cols_in_scope = self.headers if ignore_index else [self.sql_idx, *self.headers]
            logic_op = """ OR """ if how == 'any' else """ AND """
            logical_negate = """""" if inplace else """NOT"""
            action_take_str = """ """.join(("""SELECT""", """,""".join((f'"{col}"' for col in cols_in_scope)))) if not inplace else """DELETE"""
            logic_op = """ AND """ if (logic_op == ' OR ' and not inplace) else """ OR """ if (logic_op == ' AND ' and not inplace) else logic_op
            drop_where_str = logic_op.join((f'({logical_negate} ("{col}" IS NULL))' if strictly_null else f'({logical_negate} ("{col}" IS NULL OR (LOWER("{col}") IN {na_values})))' for col in cols_in_scope if col != self.sql_idx))
            sql_stmt = f'{action_take_str} FROM "{self.sql_model}" WHERE ({drop_where_str})'
            if inplace:
                self.execute_statement(sql_stmt=sql_stmt, update_row_meta=True)
                if ignore_index:
                    self.reset_index()
                return
            else:
                return self.execute_fetch(sql_stmt)
        else: # 'columns'
            headers_str = """,""".join([f'''count(*)-(SUM(CASE WHEN "{header}" IS NULL THEN 1 ELSE 0 END)) AS "{header}"'''  if strictly_null else f'''count(*)-(SUM(CASE WHEN "{header}" IS NULL OR (LOWER("{header}") IN {na_values}) THEN 1 ELSE 0 END)) AS "{header}"''' for header in self.headers])
            sql_stmt = """ """.join(("SELECT",headers_str,f'FROM "{self.sql_model}"'))
            null_counts = dict(zip(self.headers,self.sql_db_conn.execute(sql_stmt).fetchone()))
            cols_to_drop = [col for col, nulls in null_counts.items() if nulls == 0] if how == 'all' else [col for col, nulls in null_counts.items() if nulls != self.row_count]
            if (num_cols := len(cols_to_drop)) == self.column_count:
                msg = ErrorFormat(f"DimensionError: cannot drop all columns, the number of columns to drop '{num_cols}' cannot be equal to the current column count '{self.column_count}'")
                raise DimensionError(msg)
            if inplace:
                self.drop_column(cols_to_drop)
                return
            else:
                cols_to_keep = [col for col in self.headers if col not in cols_to_drop]
                keep_stmt = self._generate_sql_stmt(columns=cols_to_keep)
                return self.execute_fetch(keep_stmt)